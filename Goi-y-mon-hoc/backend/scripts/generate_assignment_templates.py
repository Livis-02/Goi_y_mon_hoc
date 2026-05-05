"""
Generate mau_phan_cong_day_du.xlsx và mau_phan_cong_toi_gian.xlsx
vào docs/templates/.

Chạy: python -m backend.scripts.generate_assignment_templates
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).parent.parent.parent / "data" / "test_scenarios"
OUT_DIR.mkdir(parents=True, exist_ok=True)

HEADER_FILL   = PatternFill("solid", fgColor="5B21B6")   # violet-800
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL      = PatternFill("solid", fgColor="EDE9FE")   # violet-100
BORDER_SIDE   = Side(style="thin", color="C4B5FD")
CELL_BORDER   = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                       top=BORDER_SIDE, bottom=BORDER_SIDE)
NOTE_FILL     = PatternFill("solid", fgColor="FEF3C7")   # amber-100
NOTE_FONT     = Font(color="92400E", size=10)             # amber-800


def _apply_header(ws, cols: list[str]):
    ws.append(cols)
    for col_idx in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER


def _style_data_row(ws, row_idx: int, n_cols: int, alt: bool):
    fill = ALT_FILL if alt else None
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fill:
            cell.fill = fill


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Dữ liệu mẫu ──────────────────────────────────────────────────────────────

FULL_ROWS = [
    # (Mã SV, Mã GV)  — 5 SV có trong hệ thống
    ("1000000001", "GV0001"),
    ("1100000001", "GV0001"),
    ("1200000001", "GV0002"),
    ("1300000001", "GV0002"),
    ("1400000001", "GV0003"),
    # 5 SV giả định chưa có trong hệ thống
    ("2500000001", "GV0001"),
    ("2500000002", "GV0001"),
    ("2500000003", "GV0002"),
    ("2500000004", "GV0003"),
    ("2500000005", "GV0003"),
]

SIMPLE_ROWS = [
    ("1000000001", "GV0001"),
    ("1100000001", "GV0002"),
    ("1200000001", "GV0003"),
]

GUIDE_RULES = [
    ("QUY TẮC TẠO FILE PHÂN CÔNG",),
    ("",),
    ("1.", "Cột \"Mã SV\": 10 chữ số, 2 số đầu là khóa (VD: 11xxxxxxxx = K11)"),
    ("2.", "Cột \"Mã GV\": format GV + 4 chữ số (VD: GV0001)"),
    ("3.", "Hàng đầu tiên là header, không chứa dữ liệu"),
    ("4.", "Không để ô trống"),
    ("5.", "Mỗi Mã SV chỉ xuất hiện 1 lần trong file"),
    ("6.", "Nếu Mã GV chưa có trong hệ thống, admin sẽ được hỏi có muốn tạo mới không"),
    ("7.", "Nếu Mã SV chưa có, dòng đó sẽ bị bỏ qua (báo lỗi trong kết quả)"),
    ("",),
    ("LƯU Ý:", "File hỗ trợ định dạng .xlsx và .csv"),
    ("",       "Mã cột có thể viết hoa hoặc thường đều được"),
    ("",       "Tên cột: \"Mã SV\" và \"Mã GV\" (hoặc \"ma_sv\" / \"ma_gv\")"),
]


def _build_guide_sheet(wb):
    ws = wb.create_sheet("Hướng dẫn")
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 72

    title_fill = PatternFill("solid", fgColor="4C1D95")
    title_font = Font(color="FFFFFF", bold=True, size=13)

    ws.append(("QUY TẮC TẠO FILE PHÂN CÔNG",))
    ws.merge_cells("A1:B1")
    cell = ws["A1"]
    cell.fill = title_fill
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.append(())  # blank row

    rules = [
        ("1.", "Cột \"Mã SV\": 10 chữ số, 2 số đầu là khóa (VD: 11xxxxxxxx = K11)"),
        ("2.", "Cột \"Mã GV\": format GV + 4 chữ số (VD: GV0001)"),
        ("3.", "Hàng đầu tiên là header, không chứa dữ liệu"),
        ("4.", "Không để ô trống"),
        ("5.", "Mỗi Mã SV chỉ xuất hiện 1 lần trong file"),
        ("6.", "Nếu Mã GV chưa có trong hệ thống, admin được hỏi có muốn tạo mới không"),
        ("7.", "Nếu Mã SV chưa có trong hệ thống, dòng đó bị bỏ qua (báo lỗi)"),
    ]
    note_fill = PatternFill("solid", fgColor="F5F3FF")
    num_font  = Font(bold=True, color="5B21B6", size=10)
    txt_font  = Font(color="1E1B4B", size=10)
    for num, rule in rules:
        row_idx = ws.max_row + 1
        ws.append((num, rule))
        ws.cell(row=row_idx, column=1).font = num_font
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2).font = txt_font
        for col in (1, 2):
            ws.cell(row=row_idx, column=col).fill = note_fill

    ws.append(())
    ws.append(("LƯU Ý:",))
    note_rows = [
        ("", "File hỗ trợ định dạng .xlsx và .csv"),
        ("", "Mã cột có thể viết hoa hoặc thường đều được"),
        ("", "Tên cột chấp nhận: \"Mã SV\" / \"ma_sv\" và \"Mã GV\" / \"ma_gv\""),
    ]
    luu_y_font = Font(bold=True, color="92400E", size=10)
    for r in note_rows:
        row_idx = ws.max_row + 1
        ws.append(r)
        ws.cell(row=row_idx, column=1).font = luu_y_font
        ws.cell(row=row_idx, column=2).font = Font(color="78350F", size=10)
        for col in (1, 2):
            ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="FEF3C7")


def _build_data_sheet(wb, sheet_name: str, rows: list[tuple], note: str):
    ws = wb.create_sheet(sheet_name)
    _apply_header(ws, ["Mã SV", "Mã GV"])
    _set_col_widths(ws, [18, 14])
    ws.row_dimensions[1].height = 22

    for i, (ma_sv, ma_gv) in enumerate(rows, start=2):
        ws.append([ma_sv, ma_gv])
        _style_data_row(ws, i, 2, alt=(i % 2 == 0))
        ws.row_dimensions[i].height = 20

    # Note row below data
    note_row = ws.max_row + 2
    ws.cell(row=note_row, column=1, value=note)
    ws.merge_cells(f"A{note_row}:B{note_row}")
    nc = ws.cell(row=note_row, column=1)
    nc.font = NOTE_FONT
    nc.fill = NOTE_FILL
    nc.alignment = Alignment(wrap_text=True)

    return ws


# ── Tạo file đầy đủ ──────────────────────────────────────────────────────────

def generate_full():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    ws = _build_data_sheet(
        wb, "Phân công",
        FULL_ROWS,
        note="* Dòng 2–6: SV có trong hệ thống test.  Dòng 7–11: SV giả định chưa tồn tại — sẽ bị bỏ qua khi import.",
    )

    # Freeze header row
    ws.freeze_panes = "A2"

    # Build guide sheet
    _build_guide_sheet(wb)

    out = OUT_DIR / "mau_phan_cong_day_du.xlsx"
    wb.save(out)
    print(f"[OK] {out}")


# ── Tạo file tối giản ────────────────────────────────────────────────────────

def generate_simple():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = _build_data_sheet(
        wb, "Phân công",
        SIMPLE_ROWS,
        note="File mẫu tối giản — chỉ gồm header và 3 dòng ví dụ.",
    )
    ws.freeze_panes = "A2"

    out = OUT_DIR / "mau_phan_cong_toi_gian.xlsx"
    wb.save(out)
    print(f"[OK] {out}")


if __name__ == "__main__":
    generate_full()
    generate_simple()
    print("Done.")
