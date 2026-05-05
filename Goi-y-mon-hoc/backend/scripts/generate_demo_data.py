"""generate_demo_data.py — Tạo bộ demo data cho EduGuide (mô hình Lớp 2026-05-05).

Chạy:
    python -m backend.scripts.generate_demo_data --reset

Output (data/demo/):
    co_van/bulk_import.xlsx          — 10 GV (4 cột: Mã GV | Họ tên | Email | Chuyên ngành)
    lop/bulk_import.xlsx             — 6 Lớp (5 cột: Mã lớp | Tên | Khoá | CN | Mã GVCN)
    sinh_vien/bulk_import.xlsx       — 30 SV (4 cột: MSSV | Họ tên | Email | Mã lớp)
    sinh_vien/bang_diem/svXXXXX.xlsx — Bảng điểm cho SV có semesters_done > 0

Mô hình:
    - 1 Khoá K66 (cohort=66), MSSV format sv66XXX
    - 6 chuyên ngành (KHMT/MMT/CNPM/HTTT/THKT/CNTTDH), mỗi spec 1 lớp
    - Mã lớp: DCCTCT66_{spec_2digit}A (vd DCCTCT66_07A = lớp KHMT)
    - 5 SV/lớp = 30 SV tổng
    - 10 GV: 6 GVCN (1/lớp) + 4 GV "trợ giảng" không chủ nhiệm
"""
from __future__ import annotations

import argparse
import os
import random
import sys
from collections import defaultdict
from dataclasses import dataclass
from typing import List, Optional

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

_HERE = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(os.path.dirname(_HERE))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

import bcrypt  # noqa: E402
from sqlalchemy import text  # noqa: E402
from backend.db.db import SessionLocal  # noqa: E402
from backend.db import models  # noqa: E402


# ── Mapping spec ──────────────────────────────────────────────────────────────
SPEC_CODES = {
    "KHMT":   "7480201_07",
    "MMT":    "7480201_06",
    "CNPM":   "7480201_05",
    "HTTT":   "7480201_09",
    "THKT":   "7480201_04",
    "CNTTDH": "7480201_08",
}
SPEC_2DIGIT = {
    "KHMT":   "07",
    "MMT":    "06",
    "CNPM":   "05",
    "HTTT":   "09",
    "THKT":   "04",
    "CNTTDH": "08",
}
SPEC_NAMES_VI = {
    "KHMT":   "Khoa học máy tính",
    "MMT":    "Mạng máy tính & Truyền thông",
    "CNPM":   "Công nghệ phần mềm",
    "HTTT":   "Hệ thống thông tin",
    "THKT":   "Tin học kinh tế",
    "CNTTDH": "CNTT định hướng ứng dụng",
}

COHORT = "66"
DEFAULT_PASSWORD = "Test1234!"
HASH_CACHE: Optional[str] = None


def _hash_pw(pw: str) -> str:
    global HASH_CACHE
    if HASH_CACHE is None:
        HASH_CACHE = bcrypt.hashpw(pw.encode(), bcrypt.gensalt(12)).decode()
    return HASH_CACHE


# ── 6 Lớp (1/spec) ────────────────────────────────────────────────────────────
@dataclass
class ClassDef:
    code: str            # "DCCTCT66_07A"
    name: str            # "Lớp K66 KHMT A"
    spec_short: str      # "KHMT"
    advisor_tc: str      # GVCN — Mã GV chủ nhiệm

    @property
    def spec_code(self) -> str:
        return SPEC_CODES[self.spec_short]


CLASSES: List[ClassDef] = [
    ClassDef(f"DCCTCT{COHORT}_{SPEC_2DIGIT[s]}A", f"Lớp K{COHORT} {s} A", s, f"{s}001")
    for s in ["KHMT", "MMT", "CNPM", "HTTT", "THKT", "CNTTDH"]
]


# ── 10 GV: 6 GVCN (XX001 cho mỗi spec) + 4 GV trợ giảng ──────────────────────
@dataclass
class Advisor:
    teacher_code: str
    full_name: str
    spec_short: Optional[str]   # None = GV không phụ trách spec cụ thể


ADVISORS: List[Advisor] = [
    # 6 GVCN — mỗi spec 1 GV chủ nhiệm
    Advisor("KHMT001",   "Nguyễn Văn A",     "KHMT"),
    Advisor("MMT001",    "Trần Thị B",       "MMT"),
    Advisor("CNPM001",   "Lê Văn C",         "CNPM"),
    Advisor("HTTT001",   "Hoàng Văn D",      "HTTT"),
    Advisor("THKT001",   "Vũ Thị E",         "THKT"),
    Advisor("CNTTDH001", "Bùi Văn F",        "CNTTDH"),
    # 4 GV trợ giảng (không chủ nhiệm) — vẫn có managed_specialization để filter
    Advisor("KHMT002",   "Đỗ Thị G",         "KHMT"),
    Advisor("CNPM002",   "Đinh Văn H",       "CNPM"),
    Advisor("MMT002",    "Mai Thị I",        "MMT"),
    Advisor("GV001",     "Phạm Anh J",       None),  # GV chung
]


# ── 30 SV (5/lớp) ─────────────────────────────────────────────────────────────
@dataclass
class Profile:
    username: str
    full_name: str
    class_code: str          # link với CLASSES
    profile_type: str
    semesters_done: int      # 0 = chưa upload điểm
    target_gpa4: float
    fail_count: int = 0
    retake_count: int = 0
    has_internship: bool = False
    has_thesis: bool = False


def _gen_profiles() -> List[Profile]:
    """Tạo 30 SV — 5 SV cho mỗi lớp với profile đa dạng (top/typical/at_risk)."""
    profiles: List[Profile] = []
    sample_names = [
        "Nguyễn Văn", "Trần Thị", "Lê Hoàng", "Phạm Minh", "Hoàng Văn",
        "Vũ Thị", "Đỗ Anh", "Bùi Thị", "Mai Văn", "Đinh Thị",
        "Phan Hoàng", "Tô Anh", "Lý Thị", "Trịnh Văn", "Cao Thị",
        "Lưu Văn", "Hồ Anh", "Nông Thị", "Đặng Văn", "Tống Thị",
        "Chu Văn", "Văn Anh", "Trương Thị", "Quản Văn", "La Thị",
        "Ma Anh", "Sa Thị", "Tạ Văn", "Hà Thị", "Du Anh",
    ]
    suffixes = ["Sắc", "Giỏi", "Khá", "Trung", "Yếu"]

    sv_idx = 0
    for cls in CLASSES:
        for slot in range(5):
            sv_idx += 1
            username = f"sv{COHORT}{sv_idx:03d}"
            base_name = sample_names[(sv_idx - 1) % len(sample_names)]
            # 5 profiles distinct trong mỗi lớp:
            #   slot 0: top student
            #   slot 1: typical good
            #   slot 2: typical avg
            #   slot 3: needs_attention
            #   slot 4: at_risk
            if slot == 0:
                p = Profile(username, f"{base_name} {suffixes[0]}", cls.code,
                            "top_student", 6, 3.85, has_internship=True)
            elif slot == 1:
                p = Profile(username, f"{base_name} {suffixes[1]}", cls.code,
                            "typical_good", 5, 3.30)
            elif slot == 2:
                p = Profile(username, f"{base_name} {suffixes[2]}", cls.code,
                            "typical_avg", 5, 2.85)
            elif slot == 3:
                p = Profile(username, f"{base_name} {suffixes[3]}", cls.code,
                            "needs_attention", 4, 2.30, fail_count=1, retake_count=1)
            else:
                p = Profile(username, f"{base_name} {suffixes[4]}", cls.code,
                            "at_risk", 4, 1.80, fail_count=2, retake_count=1)
            profiles.append(p)
    return profiles


PROFILES: List[Profile] = _gen_profiles()


# ── XLSX writers ──────────────────────────────────────────────────────────────
def _import_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise RuntimeError("Cần cài openpyxl: pip install openpyxl")


def _write_bulk_advisors(out_path: str):
    """File template GV — 4 cột: Mã GV | Họ tên | Email | Chuyên ngành."""
    openpyxl = _import_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CoVan"
    ws.append(["Mã GV", "Họ tên", "Email", "Chuyên ngành"])
    for a in ADVISORS:
        spec_code = SPEC_CODES.get(a.spec_short) if a.spec_short else ""
        email = f"{a.teacher_code.lower()}@edu.vn"
        ws.append([a.teacher_code, a.full_name, email, spec_code])
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)


def _write_bulk_classes(out_path: str):
    """File template Lớp — 5 cột: Mã lớp | Tên lớp | Khoá | Chuyên ngành | Mã GVCN."""
    openpyxl = _import_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lop"
    ws.append(["Mã lớp", "Tên lớp", "Khoá", "Chuyên ngành", "Mã GVCN"])
    for cls in CLASSES:
        ws.append([cls.code, cls.name, COHORT, cls.spec_code, cls.advisor_tc])
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)


def _write_bulk_students(out_path: str):
    """File template SV — 4 cột: MSSV | Họ tên | Email | Mã lớp."""
    openpyxl = _import_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SinhVien"
    ws.append(["MSSV", "Họ tên", "Email", "Mã lớp"])
    for p in PROFILES:
        email = f"{p.username}@student.edu.vn"
        ws.append([p.username, p.full_name, email, p.class_code])
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)


# ── Grade XLSX generation (per SV) ────────────────────────────────────────────
def _get_course_pool(db, semester: int, spec_code: Optional[str]) -> List[dict]:
    sql = """
    SELECT course_code, course_name, credits, count_toward_credits, required_specialization
    FROM courses
    WHERE typical_semester = :sem
      AND count_toward_credits = TRUE
      AND credits IS NOT NULL
    """
    params = {"sem": semester}
    if spec_code is None:
        sql += " AND required_specialization IS NULL"
    else:
        sql += " AND (required_specialization IS NULL OR required_specialization = :spec)"
        params["spec"] = spec_code
    rows = db.execute(text(sql), params).fetchall()
    return [{"course_code": r[0], "course_name": r[1], "credits": float(r[2]),
             "count_toward": bool(r[3]), "required_spec": r[4]} for r in rows]


def _score4_to_score10(s4: float) -> float:
    if s4 >= 3.7: return min(10.0, 8.5 + (s4 - 3.7) * 5)
    if s4 >= 3.5: return 8.0 + (s4 - 3.5) * 2.5
    if s4 >= 3.0: return 7.0 + (s4 - 3.0) * 2
    if s4 >= 2.5: return 6.5 + (s4 - 2.5) * 1
    if s4 >= 2.0: return 5.5 + (s4 - 2.0) * 2
    if s4 >= 1.5: return 5.0 + (s4 - 1.5) * 1
    if s4 >= 1.0: return 4.0 + (s4 - 1.0) * 2
    return 1.0 + s4 * 3


def _score4_to_letter(s4: float) -> str:
    if s4 >= 3.7: return "A"
    if s4 >= 3.5: return "A-"
    if s4 >= 3.0: return "B+"
    if s4 >= 2.5: return "B"
    if s4 >= 2.0: return "C+"
    if s4 >= 1.5: return "C"
    if s4 >= 1.0: return "D+"
    if s4 >= 0.5: return "D"
    return "F"


def _generate_grades(db, profile: Profile, spec_code: str) -> List[dict]:
    if profile.semesters_done == 0:
        return []
    target = profile.target_gpa4
    selected = []
    for sem in range(1, profile.semesters_done + 1):
        pool = _get_course_pool(db, sem, spec_code)
        if not pool:
            continue
        random.Random(profile.username + str(sem)).shuffle(pool)
        for c in pool[:6]:
            term = f"Học kỳ {(sem - 1) % 2 + 1} - Năm học 20{20 + (sem - 1) // 2}-20{21 + (sem - 1) // 2}"
            selected.append({**c, "term": term})
    if not selected:
        return []

    grades = []
    fail_remaining = profile.fail_count
    retake_remaining = profile.retake_count
    g_rng = random.Random(sum(ord(c) for c in profile.username))
    for c in selected:
        s4 = max(0.0, min(4.0, target + g_rng.gauss(0, 0.4)))
        passed = s4 >= 1.0
        if fail_remaining > 0 and g_rng.random() < 0.3:
            s4 = max(0.0, target - 1.5 + g_rng.uniform(-0.3, 0.3))
            if s4 < 1.0:
                passed = False
                fail_remaining -= 1
        s4 = round(s4, 1)
        grades.append({
            "course_code": c["course_code"], "course_name": c["course_name"],
            "credits": c["credits"], "score10": round(_score4_to_score10(s4), 1),
            "score4": s4, "letter": _score4_to_letter(s4), "passed": passed,
            "term": c["term"],
        })
    if retake_remaining > 0:
        fails = [g for g in grades if not g["passed"]][:retake_remaining]
        for f in fails:
            f["score4"] = round(max(1.0, target - 0.5 + g_rng.uniform(-0.2, 0.2)), 1)
            f["score10"] = round(_score4_to_score10(f["score4"]), 1)
            f["letter"] = _score4_to_letter(f["score4"])
            f["passed"] = f["score4"] >= 1.0
            f["term"] = f["term"] + " (học lại pass)"
            retake_remaining -= 1

    if profile.has_internship:
        r = db.execute(text(
            "SELECT course_code, course_name, credits FROM courses "
            "WHERE course_name LIKE 'Thực tập%' AND required_specialization = :s LIMIT 1"
        ), {"s": spec_code}).fetchone()
        if r:
            grades.append({"course_code": r[0], "course_name": r[1], "credits": float(r[2]),
                           "score10": 7.5, "score4": 3.0, "letter": "B+", "passed": True,
                           "term": "Học kỳ 1 - Năm học 2025-2026"})
    if profile.has_thesis:
        r = db.execute(text(
            "SELECT course_code, course_name, credits FROM courses "
            "WHERE course_name LIKE 'Đồ án tốt nghiệp%' AND required_specialization = :s LIMIT 1"
        ), {"s": spec_code}).fetchone()
        if r:
            grades.append({"course_code": r[0], "course_name": r[1], "credits": float(r[2]),
                           "score10": 8.5, "score4": 3.7, "letter": "A", "passed": True,
                           "term": "Học kỳ 2 - Năm học 2025-2026"})
    return grades


def _write_grades_xlsx(profile: Profile, grades: List[dict], out_path: str):
    """XLSX bảng điểm theo format admin import yêu cầu (header MSSV ở dòng đầu)."""
    openpyxl = _import_openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value=f"Mã sinh viên: {profile.username}")
    ws.cell(row=2, column=1, value=f"Họ và tên: {profile.full_name}")
    headers = ["Stt", "Mã MH", "Nhóm/tổ môn học", "Tên môn học", "Số tín chỉ",
               "Điểm thi", "Điểm TK (10)", "Điểm TK (4)", "Điểm TK (C)",
               "Kết quả", "Chi tiết"]
    for col, val in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=val)

    if not grades:
        os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
        wb.save(out_path)
        return

    by_term: dict = defaultdict(list)
    for g in grades:
        by_term[g["term"]].append(g)
    cumulative_tc = 0.0
    current_row = 5
    terms = list(by_term.keys())
    for term_idx, term in enumerate(terms):
        ws.cell(row=current_row, column=1, value=term)
        current_row += 1
        passed_tc = 0.0
        for stt, g in enumerate(by_term[term], start=1):
            ws.cell(row=current_row, column=1, value=stt)
            ws.cell(row=current_row, column=2, value=g["course_code"])
            ws.cell(row=current_row, column=3, value="01")
            ws.cell(row=current_row, column=4, value=g["course_name"])
            ws.cell(row=current_row, column=5, value=g["credits"])
            ws.cell(row=current_row, column=6, value=g["score10"])
            ws.cell(row=current_row, column=7, value=g["score10"])
            ws.cell(row=current_row, column=8, value=g["score4"])
            ws.cell(row=current_row, column=9, value=g["letter"])
            ws.cell(row=current_row, column=10, value="Đạt" if g["passed"] else "Không đạt")
            current_row += 1
            if g["passed"]:
                passed_tc += g["credits"]
        cumulative_tc += passed_tc
        ws.cell(row=current_row, column=1, value="- Số tín chỉ đạt học kỳ:")
        ws.cell(row=current_row, column=2, value=passed_tc)
        current_row += 1
        if term_idx == len(terms) - 1:
            ws.cell(row=current_row, column=1, value="- Số tín chỉ tích lũy:")
            ws.cell(row=current_row, column=2, value=cumulative_tc)
            current_row += 1
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)


# ── DB seeding (GV → Lớp → SV → Grades) ──────────────────────────────────────
def _ensure_advisor(db, a: Advisor, hash_pw: str):
    user = db.query(models.User).filter_by(username=a.teacher_code).first()
    spec_code = SPEC_CODES.get(a.spec_short) if a.spec_short else None
    if not user:
        user = models.User(
            username=a.teacher_code, full_name=a.full_name,
            password_hash=hash_pw, role="advisor",
            teacher_code=a.teacher_code,
            managed_specialization=spec_code,
            is_first_login=False,
            default_password=DEFAULT_PASSWORD,
        )
        db.add(user); db.commit(); db.refresh(user)
        return user, True
    user.full_name = a.full_name
    user.managed_specialization = spec_code
    user.password_hash = hash_pw; user.is_first_login = False
    user.default_password = DEFAULT_PASSWORD
    db.commit()
    return user, False


def _ensure_class(db, cls: ClassDef, advisor_user: models.User):
    cg = db.query(models.ClassGroup).filter_by(code=cls.code).first()
    if not cg:
        cg = models.ClassGroup(
            code=cls.code, name=cls.name,
            cohort=COHORT, specialization=cls.spec_code,
            advisor_id=advisor_user.id,
        )
        db.add(cg); db.commit(); db.refresh(cg)
        return cg, True
    cg.name = cls.name
    cg.cohort = COHORT
    cg.specialization = cls.spec_code
    cg.advisor_id = advisor_user.id
    db.commit()
    return cg, False


def _ensure_student(db, p: Profile, class_group: models.ClassGroup, hash_pw: str):
    user = db.query(models.User).filter_by(username=p.username).first()
    if not user:
        user = models.User(
            username=p.username, full_name=p.full_name,
            password_hash=hash_pw, role="student", cohort=COHORT,
            specialization=class_group.specialization,
            class_group_id=class_group.id,
            is_first_login=False,
            default_password=DEFAULT_PASSWORD,
        )
        db.add(user); db.commit(); db.refresh(user)
        return user, True
    user.full_name = p.full_name
    user.cohort = COHORT
    user.specialization = class_group.specialization
    user.class_group_id = class_group.id
    user.password_hash = hash_pw; user.is_first_login = False
    user.default_password = DEFAULT_PASSWORD
    db.commit()
    return user, False


def _seed_grades(db, user, grades: List[dict]):
    db.query(models.UserGrade).filter_by(user_id=user.id).delete()
    for g in grades:
        db.add(models.UserGrade(
            user_id=user.id, course_code=g["course_code"],
            score10=g["score10"], score4=g["score4"], letter=g["letter"],
            passed=g["passed"], term=g["term"],
        ))
    db.commit()


def _sync_assignments(db):
    """Sync AdvisorAssignment cho tất cả SV → advisor = GVCN của lớp SV."""
    students = db.query(models.User).filter(
        models.User.role == "student",
        models.User.class_group_id.isnot(None),
    ).all()
    classes_map = {c.id: c for c in db.query(models.ClassGroup).all()}

    student_ids = [s.id for s in students]
    if student_ids:
        db.query(models.AdvisorAssignment).filter(
            models.AdvisorAssignment.student_id.in_(student_ids)
        ).delete(synchronize_session=False)

    for s in students:
        cg = classes_map.get(s.class_group_id)
        if cg:
            db.add(models.AdvisorAssignment(advisor_id=cg.advisor_id, student_id=s.id))
    db.commit()


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Tạo bộ dữ liệu demo cho EduGuide")
    parser.add_argument("--reset", action="store_true",
                        help="Xoá user demo cũ + ghi đè XLSX")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview, không động DB hay file")
    parser.add_argument("--no-db", action="store_true",
                        help="Chỉ ghi XLSX, không động DB")
    parser.add_argument("--output", default="data/demo",
                        help="Folder output (mặc định data/demo/)")
    args = parser.parse_args()

    out_root = os.path.join(_PROJECT_ROOT, args.output)
    sv_dir = os.path.join(out_root, "sinh_vien")
    gv_dir = os.path.join(out_root, "co_van")
    lop_dir = os.path.join(out_root, "lop")
    grades_dir = os.path.join(sv_dir, "bang_diem")

    print("=== generate_demo_data (mô hình Lớp 2026-05-05) ===")
    print(f"Output: {out_root}")
    print(f"  6 Lớp K{COHORT} (1 lớp/chuyên ngành)")
    print(f"  30 SV (5 SV/lớp)")
    print(f"  10 GV: 6 GVCN + 4 trợ giảng/chung")

    if args.dry_run:
        print("\n--- Lớp ---")
        for c in CLASSES:
            print(f"  {c.code:18} {c.spec_short:8} GVCN={c.advisor_tc}")
        print("\n--- GV ---")
        for a in ADVISORS:
            print(f"  {a.teacher_code:10} {a.spec_short or '(general)':10}")
        print("\n--- SV (mẫu 5 dòng đầu) ---")
        for p in PROFILES[:5]:
            print(f"  {p.username:10} {p.class_code:18} {p.profile_type:18} GPA={p.target_gpa4}")
        print(f"  ... ({len(PROFILES)} SV tổng)")
        return

    # Seed DB
    if not args.no_db:
        db = SessionLocal()
        try:
            if args.reset:
                print("\n[1/5] Reset: xoá user + lớp demo cũ...")
                usernames = (
                    [p.username for p in PROFILES]
                    + [a.teacher_code for a in ADVISORS]
                )
                # Xoá class_groups trước (vì có FK ON DELETE RESTRICT lên users)
                # Phải clear class_group_id của student trước
                db.execute(text(
                    "UPDATE users SET class_group_id = NULL WHERE username IN :us"
                ).bindparams(__import__('sqlalchemy').bindparam('us', expanding=True)),
                    {"us": tuple(usernames) if usernames else ("__none__",)})
                db.query(models.ClassGroup).filter(
                    models.ClassGroup.code.in_([c.code for c in CLASSES])
                ).delete(synchronize_session=False)
                # Xoá assignments + grades + users
                for u in db.query(models.User).filter(models.User.username.in_(usernames)).all():
                    db.query(models.UserGrade).filter_by(user_id=u.id).delete()
                    db.query(models.AdvisorAssignment).filter(
                        (models.AdvisorAssignment.student_id == u.id) |
                        (models.AdvisorAssignment.advisor_id == u.id)
                    ).delete()
                    db.delete(u)
                db.commit()

            hash_pw = _hash_pw(DEFAULT_PASSWORD)

            print(f"\n[2/5] Seed {len(ADVISORS)} GV...")
            advisor_map: dict[str, models.User] = {}
            for a in ADVISORS:
                user, is_new = _ensure_advisor(db, a, hash_pw)
                advisor_map[a.teacher_code] = user
                tag = "✓ NEW" if is_new else "↻ UPD"
                print(f"  {tag} {a.teacher_code} {a.spec_short or 'chung'}")

            print(f"\n[3/5] Seed {len(CLASSES)} Lớp...")
            class_map: dict[str, models.ClassGroup] = {}
            for cls in CLASSES:
                advisor_user = advisor_map.get(cls.advisor_tc)
                if not advisor_user:
                    print(f"  ⚠ Skip {cls.code}: không có GVCN {cls.advisor_tc}")
                    continue
                cg, is_new = _ensure_class(db, cls, advisor_user)
                class_map[cls.code] = cg
                tag = "✓ NEW" if is_new else "↻ UPD"
                print(f"  {tag} {cls.code} → GVCN {cls.advisor_tc}")

            print(f"\n[4/5] Seed {len(PROFILES)} SV + bảng điểm...")
            for p in PROFILES:
                cg = class_map.get(p.class_code)
                if not cg:
                    print(f"  ⚠ Skip {p.username}: lớp {p.class_code} chưa tồn tại")
                    continue
                user, is_new = _ensure_student(db, p, cg, hash_pw)
                grades = _generate_grades(db, p, cg.specialization)
                _seed_grades(db, user, grades)
                tag = "✓ NEW" if is_new else "↻ UPD"
                print(f"  {tag} {p.username} {p.class_code}: {len(grades)} grades")
                _write_grades_xlsx(p, grades, os.path.join(grades_dir, f"{p.username}.xlsx"))

            print("\n[5/5] Sync AdvisorAssignment SV ↔ GVCN của lớp...")
            _sync_assignments(db)
            ass_count = db.query(models.AdvisorAssignment).count()
            print(f"  ✓ {ass_count} assignments")
        finally:
            db.close()
    else:
        print("\n[--no-db] Bỏ qua seed DB.")

    # Bulk imports — luôn ghi (idempotent)
    print("\n[Files] Ghi XLSX bulk import...")
    sv_bulk = os.path.join(sv_dir, "bulk_import.xlsx")
    gv_bulk = os.path.join(gv_dir, "bulk_import.xlsx")
    lop_bulk = os.path.join(lop_dir, "bulk_import.xlsx")
    _write_bulk_advisors(gv_bulk); print(f"  ✓ {gv_bulk}")
    _write_bulk_classes(lop_bulk); print(f"  ✓ {lop_bulk}")
    _write_bulk_students(sv_bulk); print(f"  ✓ {sv_bulk}")

    print("\n=== Done ===")
    print(f"Default password: {DEFAULT_PASSWORD}")
    print(f"Workflow upload: GV → Lớp → SV (đúng thứ tự)")


if __name__ == "__main__":
    main()
