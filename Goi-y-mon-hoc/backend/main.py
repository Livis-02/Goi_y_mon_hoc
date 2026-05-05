from __future__ import annotations

from dotenv import load_dotenv
load_dotenv()

import hashlib
import json
import os
import re
import secrets
# import smtplib  # removed 2026-05-05 — không còn SMTP forgot password
from contextlib import asynccontextmanager
# MIMEMultipart/MIMEText imports removed 2026-05-05 — không còn SMTP
from datetime import datetime, timedelta, timezone
from pathlib import Path
import unicodedata

import bcrypt
import httpx
from fastapi import FastAPI, Depends, HTTPException, Header, UploadFile, File, Body, Query, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, RedirectResponse
from sqlalchemy.orm import Session

from backend.db.db import get_db, engine
from backend.db import models, schemas
from backend.scripts.curriculum_importer import import_curriculum_from_excel
from backend.core import academic_engine
from backend.core.academic_engine import build_progress_snapshot, build_recommendations, build_semester_roadmap, build_analytics, invalidate_difficulty_stats_cache
from backend.core.chat_assistant import chat_reply, get_chat_history, invalidate_student_context_cache
from backend.core.parser import read_rows_from_upload, extract_grades, extract_tich_luy, extract_student_code, extract_full_name, _is_missing, normalize_text

TOKEN_TTL_HOURS = 24

def _cleanup_expired_tokens(db: Session) -> None:
    db.query(models.AuthToken).filter(
        models.AuthToken.expires_at < datetime.utcnow()
    ).delete()
    db.commit()


def _migrate_teacher_codes(db):
    """Migrate teacher_code từ format GV0001 → KHMT001/GV001 (chạy một lần khi startup)."""
    import re as _re
    old_fmt = _re.compile(r"^GV\d{4}$")
    _SPEC_PREFIX_MAP = {
        "7480201_07": "KHMT", "7480201_06": "MMT",  "7480201_05": "CNPM",
        "7480201_09": "HTTT", "7480201_04": "THKT", "7480201_08": "CNTTDH",
    }
    advisors = db.query(models.User).filter(
        models.User.role == "advisor",
        models.User.teacher_code.isnot(None),
    ).all()
    changed = 0
    for adv in advisors:
        tc = (adv.teacher_code or "").upper()
        if not old_fmt.match(tc):
            continue
        spec = adv.managed_specialization
        prefix = _SPEC_PREFIX_MAP.get(spec, "GV")
        seq = int(tc[2:])
        new_tc = f"{prefix}{seq:03d}"
        # Check unique
        conflict = db.query(models.User).filter(
            models.User.teacher_code == new_tc,
            models.User.id != adv.id,
        ).first()
        if conflict:
            continue
        conflict_u = db.query(models.User).filter(
            models.User.username == new_tc,
            models.User.id != adv.id,
        ).first()
        if conflict_u:
            continue
        adv.teacher_code = new_tc
        adv.username = new_tc
        changed += 1
    if changed:
        db.commit()


def on_startup():
    # Migration & schema bootstrap chỉ chạy khi env EDU_RUN_MIGRATIONS=1 (default OFF).
    # Lý do: idempotent nhưng chậm + spam log; dev không cần chạy mỗi lần restart.
    # Cách dùng:
    #   - Lần đầu setup DB:        EDU_RUN_MIGRATIONS=1 uvicorn backend.main:app
    #   - Sau khi sửa migrate.py:  python -m backend.db.migrate    (chạy 1 lần là xong)
    #   - Bình thường:              uvicorn backend.main:app       (skip migration)
    _run_migrations = os.getenv("EDU_RUN_MIGRATIONS", "0").lower() in ("1", "true", "yes", "on")
    if _run_migrations:
        models.Base.metadata.create_all(bind=engine)
        try:
            from backend.db import migrate as _migrate_module
            _migrate_module.run()
        except Exception as exc:
            print(f"[on_startup] migrate.run() failed: {exc}", flush=True)
    else:
        print("[on_startup] skip migrations (set EDU_RUN_MIGRATIONS=1 to enable)", flush=True)

    # Maintenance tasks (nhanh, không phải migration) — vẫn chạy mỗi lần boot.
    from backend.db import SessionLocal
    db = SessionLocal()
    try:
        _cleanup_expired_tokens(db)
        if _run_migrations:
            _migrate_teacher_codes(db)
    finally:
        db.close()


@asynccontextmanager
async def lifespan(_app: FastAPI):
    on_startup()
    yield


app = FastAPI(title="CNTT GPA API", lifespan=lifespan)

_CORS_ORIGINS_ENV = os.getenv("CORS_ORIGINS", "")
_APP_ENV = (os.getenv("APP_ENV") or os.getenv("ENV") or "development").lower()
_DEV_DEFAULT_ORIGINS = [
    "http://localhost", "http://127.0.0.1",
    "http://localhost:5500", "http://127.0.0.1:5500",
    "http://localhost:5501", "http://127.0.0.1:5501",
    "http://localhost:8000", "http://127.0.0.1:8000",
]

if _CORS_ORIGINS_ENV:
    _CORS_ORIGINS = [o.strip() for o in _CORS_ORIGINS_ENV.split(",") if o.strip()]
elif _APP_ENV in ("production", "prod", "staging"):
    # Refuse to start with permissive default origins in production. Operator must set
    # CORS_ORIGINS explicitly. allow_credentials=True + bare-host wildcards is a footgun.
    raise RuntimeError(
        f"CORS_ORIGINS must be set when APP_ENV={_APP_ENV!r}. "
        "Provide a comma-separated list of allowed origins, e.g. "
        "CORS_ORIGINS=https://eduguide.example.edu,https://www.eduguide.example.edu"
    )
else:
    _CORS_ORIGINS = _DEV_DEFAULT_ORIGINS

app.add_middleware(
    CORSMiddleware,
    allow_origins=_CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type"],
)

_FRONTEND_DIR = Path(__file__).parent.parent / "frontend"

# Uploads dir cho file attachments (messages, etc.)
_UPLOADS_DIR = Path(__file__).parent.parent / "uploads"
_UPLOADS_DIR.mkdir(exist_ok=True)
(_UPLOADS_DIR / "messages").mkdir(exist_ok=True)

# Mount static để serve uploaded files
from fastapi.staticfiles import StaticFiles
app.mount("/uploads", StaticFiles(directory=str(_UPLOADS_DIR)), name="uploads")

@app.get("/")
def root():
    return RedirectResponse("/frontend/pages/landing.html")

_DOCS_TEMPLATES_DIR = Path(__file__).parent.parent / "docs" / "templates"


@app.get("/docs/templates/{filename}")
def serve_template(filename: str):
    """Phục vụ file mẫu trong docs/templates/."""
    file_path = (_DOCS_TEMPLATES_DIR / filename).resolve()
    try:
        file_path.relative_to(_DOCS_TEMPLATES_DIR.resolve())
    except ValueError:
        raise HTTPException(status_code=403, detail="Access denied")
    if file_path.is_file():
        return FileResponse(str(file_path), filename=filename)
    raise HTTPException(status_code=404, detail="File not found")


_TEST_SCENARIOS_DIR = Path(__file__).parent.parent / "data" / "test_scenarios"


@app.get("/admin/templates/assignment-full.xlsx")
def download_assignment_full_template():
    f = _TEST_SCENARIOS_DIR / "mau_phan_cong_day_du.xlsx"
    if not f.is_file():
        from backend.scripts.generate_assignment_templates import generate_full
        generate_full()
    if not f.is_file():
        raise HTTPException(status_code=404, detail="File mẫu chưa được tạo")
    return FileResponse(str(f), filename="mau_phan_cong_day_du.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/admin/templates/assignment-simple.xlsx")
def download_assignment_simple_template():
    f = _TEST_SCENARIOS_DIR / "mau_phan_cong_toi_gian.xlsx"
    if not f.is_file():
        from backend.scripts.generate_assignment_templates import generate_simple
        generate_simple()
    if not f.is_file():
        raise HTTPException(status_code=404, detail="File mẫu chưa được tạo")
    return FileResponse(str(f), filename="mau_phan_cong_toi_gian.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/public/contact")
def public_contact(db: Session = Depends(get_db)):
    """Thông tin liên hệ quản trị viên hiển thị trên trang login (không cần auth)."""
    keys = ("contact_email", "contact_phone", "contact_office")
    rows = db.query(models.SystemConfig).filter(models.SystemConfig.key.in_(keys)).all()
    data = {r.key: (r.value or "").strip() for r in rows}
    # email field on users đã DROP 2026-05-05 — chỉ lấy từ system_config
    return {
        "email": data.get("contact_email", ""),
        "phone": data.get("contact_phone", ""),
        "office": data.get("contact_office", ""),
    }


@app.get("/frontend/{path:path}")
def serve_frontend(path: str):
    file_path = (_FRONTEND_DIR / path).resolve()
    try:
        file_path.relative_to(_FRONTEND_DIR.resolve())
    except ValueError:
        raise HTTPException(status_code=403, detail="Access denied")
    if file_path.is_file():
        return FileResponse(str(file_path))
    raise HTTPException(status_code=404, detail="File not found")

PASSWORD_POLICY = {
    "min_length": 8,
    "require_uppercase": True,
    "require_lowercase": True,
    "require_digit": True,
}

# ── Login rate limiting (DB-backed sliding window) ───────────────────────────
# Stored in DB so it survives restarts and works across multiple workers.
_RATE_LIMIT_WINDOW = 300   # 5 minutes
_RATE_LIMIT_MAX    = 5     # max failed attempts per window

def _check_rate_limit(username: str, db: Session | None = None) -> None:
    if db is None:
        return
    now = datetime.utcnow()
    cutoff = now - timedelta(seconds=_RATE_LIMIT_WINDOW)
    count = db.query(models.LoginAttempt).filter(
        models.LoginAttempt.username == username,
        models.LoginAttempt.attempted_at > cutoff,
    ).count()
    if count >= _RATE_LIMIT_MAX:
        oldest = db.query(models.LoginAttempt).filter(
            models.LoginAttempt.username == username,
            models.LoginAttempt.attempted_at > cutoff,
        ).order_by(models.LoginAttempt.attempted_at.asc()).first()
        wait = max(0, int(_RATE_LIMIT_WINDOW - (now - oldest.attempted_at).total_seconds())) if oldest else _RATE_LIMIT_WINDOW
        raise HTTPException(
            status_code=429,
            detail=f"Quá nhiều lần đăng nhập sai. Thử lại sau {wait} giây.",
        )

def _record_failed_login(username: str, db: Session | None = None) -> None:
    if db is None:
        return
    db.add(models.LoginAttempt(username=username, attempted_at=datetime.utcnow()))
    # Prune old entries for this user to keep the table clean
    cutoff = datetime.utcnow() - timedelta(seconds=_RATE_LIMIT_WINDOW * 2)
    db.query(models.LoginAttempt).filter(
        models.LoginAttempt.username == username,
        models.LoginAttempt.attempted_at < cutoff,
    ).delete()
    db.commit()

def _clear_failed_login(username: str, db: Session | None = None) -> None:
    if db is None:
        return
    db.query(models.LoginAttempt).filter(
        models.LoginAttempt.username == username,
    ).delete()
    db.commit()

DEFAULT_CURRICULUM_FILE = Path("data/7480201.docx")


def _hash_password(raw: str) -> str:
    return bcrypt.hashpw(raw.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def _hash_temp_password(raw: str) -> str:
    """Dùng cho mật khẩu tạm (6 số, bắt buộc đổi sau lần đầu). rounds=4 (~1ms vs ~230ms)."""
    return bcrypt.hashpw(raw.encode("utf-8"), bcrypt.gensalt(4)).decode("utf-8")


def normalize_vietnamese_name(name: str) -> str:
    """Title-case mỗi từ, chuẩn hóa khoảng trắng. Giữ nguyên dấu tiếng Việt."""
    if not name:
        return ""
    return " ".join(p.capitalize() for p in name.strip().split())


_TC_REGEX = r"^(KHMT|MMT|CNPM|HTTT|THKT|CNTTDH|GV)\d{3}$"
_TC_PREFIXES = ("KHMT", "MMT", "CNPM", "HTTT", "THKT", "CNTTDH", "GV")

# Mapping từ managed_specialization → teacher_code prefix
_SPEC_TC_PREFIX: dict[str | None, str] = {
    "7480201_07": "KHMT",
    "7480201_06": "MMT",
    "7480201_05": "CNPM",
    "7480201_09": "HTTT",
    "7480201_04": "THKT",
    "7480201_08": "CNTTDH",
    None: "GV",
}


def validate_teacher_code(code: str) -> str:
    """Validate và chuẩn hóa mã GV. Raise HTTPException 422 nếu sai format."""
    import re as _re
    code = code.strip().upper()
    if not _re.match(_TC_REGEX, code):
        raise HTTPException(
            status_code=422,
            detail=f"Mã GV không hợp lệ. Format: PREFIX + 3 chữ số (VD: KHMT001, MMT001, GV001). Prefix hợp lệ: {', '.join(_TC_PREFIXES)}",
        )
    return code


def _validate_password(raw: str) -> None:
    import re as _re
    errors = []
    if len(raw) < PASSWORD_POLICY["min_length"]:
        errors.append(f"ít nhất {PASSWORD_POLICY['min_length']} ký tự")
    if PASSWORD_POLICY["require_uppercase"] and not _re.search(r"[A-Z]", raw):
        errors.append("ít nhất 1 chữ hoa")
    if PASSWORD_POLICY["require_lowercase"] and not _re.search(r"[a-z]", raw):
        errors.append("ít nhất 1 chữ thường")
    if PASSWORD_POLICY["require_digit"] and not _re.search(r"[0-9]", raw):
        errors.append("ít nhất 1 chữ số")
    if errors:
        raise HTTPException(status_code=400, detail="Mật khẩu cần có " + ", ".join(errors))


def _is_bcrypt_hash(hashed: str) -> bool:
    return hashed.startswith(("$2b$", "$2a$", "$2y$"))


def _verify_password(raw: str, hashed: str) -> bool:
    try:
        if _is_bcrypt_hash(hashed):
            return bcrypt.checkpw(raw.encode("utf-8"), hashed.encode("utf-8"))
        # Legacy SHA256 — still accept but will be re-hashed on login
        return hashlib.sha256(raw.encode("utf-8")).hexdigest() == hashed
    except Exception:
        return False


def _issue_token(user: models.User, db: Session) -> str:
    # Opportunistically clean up expired tokens (every login)
    db.query(models.AuthToken).filter(
        models.AuthToken.expires_at < datetime.utcnow()
    ).delete()
    raw_token = secrets.token_urlsafe(32)
    token_hash = hashlib.sha256(raw_token.encode()).hexdigest()
    expires = datetime.utcnow() + timedelta(hours=TOKEN_TTL_HOURS)
    db.add(models.AuthToken(token_hash=token_hash, user_id=user.id, expires_at=expires))
    db.commit()
    return raw_token


def _strip_accents(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text.replace("\u0111", "d").replace("\u0110", "d"))
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _normalize_text(text: str) -> str:
    return _strip_accents(str(text or "")).lower().strip()


def _resolve_user_role(user: models.User) -> str:
    return user.role if user.role else "student"


def _to_user_out(user: models.User) -> schemas.UserOut:
    # Lookup class_group code (chỉ áp dụng cho student có class_group_id)
    class_group_code = None
    try:
        if getattr(user, "class_group_id", None):
            from sqlalchemy.orm import object_session
            sess = object_session(user)
            if sess is not None:
                cg = sess.query(models.ClassGroup).filter(
                    models.ClassGroup.id == user.class_group_id
                ).first()
                class_group_code = cg.code if cg else None
    except Exception:
        class_group_code = None

    return schemas.UserOut(
        id=user.id,
        username=user.username,
        full_name=user.full_name,
        role=_resolve_user_role(user),
        specialization=user.specialization,
        cohort=user.cohort,
        managed_specialization=user.managed_specialization,
        teacher_code=user.teacher_code,
        is_first_login=bool(user.is_first_login),
        class_group_id=getattr(user, "class_group_id", None),
        class_group_code=class_group_code,
        grades_locked=False,  # Deprecated — luôn False sau refactor 2026-05-05
    )


def _authenticate(username: str, password: str, db: Session) -> models.User:
    _check_rate_limit(username, db)
    from sqlalchemy import func as _sqlfunc
    user = db.query(models.User).filter(
        _sqlfunc.lower(models.User.username) == username.lower()
    ).first()
    if not user or not _verify_password(password, user.password_hash):
        _record_failed_login(username, db)
        raise HTTPException(status_code=401, detail="Tên đăng nhập hoặc mật khẩu không đúng")
    _clear_failed_login(username, db)
    # Lazy migration: upgrade SHA256 hash to bcrypt on successful login
    if not _is_bcrypt_hash(user.password_hash):
        user.password_hash = _hash_password(password)
        db.commit()
    return user


def _get_user_by_token(authorization: str | None, db: Session) -> models.User:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing bearer token")

    raw_token = authorization.replace("Bearer ", "", 1).strip()
    token_hash = hashlib.sha256(raw_token.encode()).hexdigest()

    record = db.query(models.AuthToken).filter(
        models.AuthToken.token_hash == token_hash,
    ).first()

    if not record:
        raise HTTPException(status_code=401, detail="Token không hợp lệ")
    if record.expires_at < datetime.utcnow():
        db.delete(record)
        db.commit()
        raise HTTPException(status_code=401, detail="Phiên đăng nhập đã hết hạn, vui lòng đăng nhập lại")

    user = db.query(models.User).filter(models.User.id == record.user_id).first()
    if not user:
        raise HTTPException(status_code=401, detail="Tài khoản không tồn tại")
    return user


def _require_admin(authorization: str | None, db: Session) -> models.User:
    user = _get_user_by_token(authorization, db)
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Bạn không có quyền thực hiện thao tác này")
    return user


def _log(
    db: Session,
    admin: models.User,
    action: str,
    target_type: str | None = None,
    target_id: str | None = None,
    detail: str | None = None,
) -> None:
    """Write one row to admin_logs. Fire-and-forget — never raises."""
    try:
        db.add(models.AdminLog(
            admin_id=admin.id,
            admin_username=admin.username,
            action=action,
            target_type=target_type,
            target_id=target_id,
            detail=detail,
        ))
        db.commit()
    except Exception:
        db.rollback()



@app.get("/health")
def health():
    import os as _os
    gkey = _os.getenv("GEMINI_API_KEY", "")
    gqkey = _os.getenv("GROQ_API_KEY", "")
    okey = _os.getenv("OPENAI_API_KEY", "")
    akey = _os.getenv("ANTHROPIC_API_KEY", "") or _os.getenv("CLAUDE_API_KEY", "")
    return {
        "status": "ok",
        "gemini_key_loaded": bool(gkey),
        "groq_key_loaded": bool(gqkey),
        "openai_key_loaded": bool(okey),
        "anthropic_key_loaded": bool(akey),
        "ai_available": bool(gkey or gqkey or okey or akey),
        "gemini_key_prefix": gkey[:8] + "..." if gkey else None,
    }


@app.get("/health/ai")
def health_ai(authorization: str | None = Header(default=None), db: Session = Depends(get_db)):
    """Test Gemini API trực tiếp — admin only."""
    _require_admin(authorization, db)
    from backend.core.ai_advisor import _gemini_chat
    try:
        result = _gemini_chat("Trả lời đúng một từ: OK", max_tokens=10)
        return {"gemini_ok": bool(result), "response": result}
    except Exception as e:
        return {"gemini_ok": False, "error": str(e)}


@app.get("/auth/password-policy")
def password_policy():
    return PASSWORD_POLICY


@app.post("/auth/register", response_model=schemas.UserOut)
def register(payload: schemas.RegisterIn, db: Session = Depends(get_db)):
    existing = db.query(models.User).filter(models.User.username == payload.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    _validate_password(payload.password)

    user = models.User(
        username=payload.username,
        password_hash=_hash_password(payload.password),
        full_name=normalize_vietnamese_name(payload.full_name) or None,
        specialization=payload.specialization or None,
    )
    db.add(user)
    db.flush()  # get user.id before commit
    auto_assign_advisor(user, db)
    db.commit()
    db.refresh(user)
    return _to_user_out(user)


@app.post("/auth/login")
def login(payload: schemas.LoginIn, db: Session = Depends(get_db)):
    user = _authenticate(payload.username, payload.password, db)
    token = _issue_token(user, db)
    # require_setup REMOVED 2026-05-05 — bỏ first-login modal. SV vào thẳng app.
    return {
        "access_token": token,
        "token_type": "bearer",
        "require_setup": False,  # kept False cho backward compat — UI cũ check nhưng skip modal
        "role": user.role,
    }


@app.post("/auth/admin/login")
def admin_login(payload: schemas.LoginIn, db: Session = Depends(get_db)):
    user = _authenticate(payload.username, payload.password, db)
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Tài khoản không có quyền quản trị")
    token = _issue_token(user, db)
    return {"access_token": token, "token_type": "bearer"}


@app.post("/auth/admin/register", response_model=schemas.UserOut)
def admin_register(payload: schemas.AdminRegisterIn, db: Session = Depends(get_db)):
    """Create a new admin account. Username is auto-generated (admin001, admin002, …).
    Protected by ADMIN_SECRET — call this from Postman/curl, not from the public UI."""
    secret = os.getenv("ADMIN_SECRET", "")
    if not secret:
        raise HTTPException(status_code=503, detail="ADMIN_SECRET chưa được cấu hình trên server.")
    if payload.admin_secret != secret:
        raise HTTPException(status_code=403, detail="Mã bí mật không đúng.")
    _validate_password(payload.password)

    # Auto-generate username: admin001, admin002, …
    admin_count = db.query(models.User).filter(models.User.role == "admin").count()
    username = f"admin{admin_count + 1:03d}"
    # Ensure uniqueness in the unlikely case of a collision
    while db.query(models.User).filter(models.User.username == username).first():
        admin_count += 1
        username = f"admin{admin_count + 1:03d}"

    user = models.User(
        username=username,
        password_hash=_hash_password(payload.password),
        full_name=payload.full_name or None,
        role="admin",
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return _to_user_out(user)


@app.post("/auth/logout")
def logout(authorization: str | None = Header(default=None), db: Session = Depends(get_db)):
    if authorization and authorization.startswith("Bearer "):
        raw_token = authorization.replace("Bearer ", "", 1).strip()
        token_hash = hashlib.sha256(raw_token.encode()).hexdigest()
        db.query(models.AuthToken).filter(models.AuthToken.token_hash == token_hash).delete()
        db.commit()
    return {"message": "Đăng xuất thành công"}


@app.get("/auth/config")
def auth_config():
    """Return public auth configuration (client IDs etc.) — no auth required."""
    return {"google_client_id": os.getenv("GOOGLE_CLIENT_ID", "")}


# REMOVED 2026-05-05: POST /auth/google đã bỏ — không dùng Google OAuth.
# Tool nội bộ chỉ dùng password login (admin tạo tài khoản, SV nhận default password).


@app.get("/auth/me", response_model=schemas.UserOut)
def me(authorization: str | None = Header(default=None), db: Session = Depends(get_db)):
    user = _get_user_by_token(authorization, db)
    return _to_user_out(user)


@app.post("/curriculum/import", response_model=schemas.CurriculumImportOut)
def import_curriculum(
    file_path: str = str(DEFAULT_CURRICULUM_FILE),
    replace_existing: bool = True,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    _require_admin(authorization, db)

    # Prevent path traversal: file must stay inside project data/ directory
    _data_dir = Path("data").resolve()
    try:
        _resolved = Path(file_path).resolve()
        _resolved.relative_to(_data_dir)   # raises ValueError if outside
    except (ValueError, Exception):
        raise HTTPException(status_code=400, detail="file_path phải nằm trong thư mục data/")

    try:
        stats = import_curriculum_from_excel(
            db,
            file_path=str(_resolved),
            replace_existing=replace_existing,
        )
        db.commit()
        return schemas.CurriculumImportOut(**stats.__dict__)
    except FileNotFoundError as exc:
        db.rollback()
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Failed to import curriculum: {exc}")


# ── CTDT Template / Preview / File-upload Import ──────────────────────────────

@app.get("/admin/ctdt/template.xlsx")
def admin_ctdt_template(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Return a CTDT import template for download."""
    _require_admin(authorization, db)
    from fastapi.responses import FileResponse as _FileResponse

    _sample = Path(__file__).parent.parent / "data" / "ctdt" / "source" / "ChuongTrinhDaoTao (2).csv"
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    filename = "ChuongTrinhDaoTao_mau.xlsx"
    if not _sample.exists():
        import io
        from fastapi.responses import Response as _Response
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "CTDT"
        ws.append([
            "ma_mon",
            "ten_mon_hoc",
            "so_tin_chi",
            "hoc_ky",
            "bat_buoc",
            "nhom",
            "chuyen_nganh",
            "tien_quyet",
            "hoc_truoc",
        ])
        ws.append(["7080208", "Co so lap trinh", 3, 2, "x", "", "Chung", "", ""])
        ws.append(["7080206", "Cau truc du lieu va giai thuat", 3, 4, "x", "", "Chung", "7080208", ""])
        ws.append(["7080508", "Khai pha du lieu", 3, 7, "", "Tu chon A", "7480201_07", "7080207", ""])

        guide = wb.create_sheet("Huong dan")
        guide.append(["Cot", "Mo ta"])
        guide.append(["ma_mon", "Ma hoc phan, bat buoc va khong trung"])
        guide.append(["ten_mon_hoc", "Ten hoc phan"])
        guide.append(["so_tin_chi", "So tin chi"])
        guide.append(["hoc_ky", "Hoc ky khuyen nghi, so 1-9"])
        guide.append(["bat_buoc", "Nhap x neu la mon bat buoc"])
        guide.append(["nhom", "Ten nhom tu chon neu co, vi du Tu chon A"])
        guide.append(["chuyen_nganh", "Chung hoac ma chuyen nganh"])
        guide.append(["tien_quyet", "Ma mon tien quyet, co the ngan cach bang dau phay"])
        guide.append(["hoc_truoc", "Ma mon hoc truoc neu co"])

        buf = io.BytesIO()
        wb.save(buf)
        return _Response(
            content=buf.getvalue(),
            media_type=media_type,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    return _FileResponse(
        path=str(_sample),
        media_type=media_type,
        filename=filename,
    )


@app.post("/admin/ctdt/preview", response_model=schemas.CtdtPreviewOut)
async def admin_ctdt_preview(
    file: UploadFile = File(...),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Parse an uploaded CTDT file and return preview without touching the DB."""
    _require_admin(authorization, db)
    from backend.scripts.curriculum_importer import _parse_docx, _parse_tabular, _NON_COUNTING_CODES
    import tempfile, os as _os

    suffix = Path(file.filename or "").suffix.lower() or ".xlsx"
    if suffix not in {".xlsx", ".xls", ".xlsm", ".csv", ".docx"}:
        raise HTTPException(status_code=400, detail="Chỉ hỗ trợ .xlsx / .xls / .csv / .docx")

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="File rỗng")

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp.write(contents)
            tmp_path = tmp.name

        tmp_file = Path(tmp_path)
        if suffix == ".docx":
            parsed, elective_mappings, _ = _parse_docx(tmp_file)
        else:
            parsed, elective_mappings, _ = _parse_tabular(tmp_file)

        elective_map: dict[str, list] = {}
        for code, spec, group in elective_mappings:
            elective_map.setdefault(code, []).append((spec, group))

        valid_courses = []
        for entry in parsed:
            code, name, credits = entry[0], entry[1], entry[2]
            groups = elective_map.get(code, [])
            valid_courses.append(schemas.CtdtPreviewCourse(
                code=code,
                name=name,
                credits=float(credits) if credits is not None else None,
                elective_group=groups[0][1] if groups else None,
                specialization=groups[0][0] if groups else None,
            ))

        warnings = []
        if not elective_mappings:
            warnings.append("Không tìm thấy thông tin nhóm tự chọn. Tiến độ nhóm tự chọn sẽ không hiển thị.")

        return schemas.CtdtPreviewOut(
            valid_count=len(valid_courses),
            elective_count=len(set(c for _, c, _ in elective_mappings)),
            valid_courses=valid_courses,
            warnings=warnings,
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Lỗi phân tích file: {exc}")
    finally:
        if tmp_path:
            try:
                _os.unlink(tmp_path)
            except OSError:
                pass


@app.post("/admin/ctdt/import", response_model=schemas.CurriculumImportOut)
async def admin_ctdt_import(
    file: UploadFile = File(...),
    replace_existing: bool = True,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Import CTDT from an uploaded file (replaces legacy /curriculum/import file-path endpoint)."""
    _require_admin(authorization, db)
    import tempfile, os as _os

    suffix = Path(file.filename or "").suffix.lower() or ".xlsx"
    if suffix not in {".xlsx", ".xls", ".xlsm", ".csv", ".docx"}:
        raise HTTPException(status_code=400, detail="Chỉ hỗ trợ .xlsx / .xls / .csv / .docx")

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="File rỗng")

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp.write(contents)
            tmp_path = tmp.name

        stats = import_curriculum_from_excel(
            db,
            file_path=tmp_path,
            replace_existing=replace_existing,
        )
        db.commit()
        # Curriculum just changed → flush all reference-data caches so SV-facing
        # endpoints don't serve stale Course/Prereq/ElectiveGroup data.
        try:
            from backend.core.academic_engine import _invalidate_ref_cache
            _invalidate_ref_cache()
        except Exception:
            pass
        return schemas.CurriculumImportOut(**stats.__dict__)
    except HTTPException:
        raise
    except FileNotFoundError as exc:
        db.rollback()
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Lỗi import CTĐT: {exc}")
    finally:
        if tmp_path:
            try:
                _os.unlink(tmp_path)
            except OSError:
                pass


@app.get("/courses", response_model=list[schemas.CourseOut])
def list_courses(db: Session = Depends(get_db)):
    return db.query(models.Course).order_by(models.Course.course_code.asc()).all()


@app.get("/admin/courses", response_model=list[schemas.CourseOut])
def admin_list_courses(
    q: str | None = None,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    _require_admin(authorization, db)
    query = db.query(models.Course)
    if q:
        like = f"%{q.strip()}%"
        query = query.filter(
            models.Course.course_code.ilike(like) | models.Course.course_name.ilike(like)
        )
    return query.order_by(models.Course.course_code.asc()).all()


@app.get("/admin/courses/grouped")
def admin_courses_grouped(
    spec: str = "common",
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    _require_admin(authorization, db)
    from backend.core.academic_engine import CURRICULUM_ORDER
    from sqlalchemy import func as _sf

    all_prereqs = db.query(models.CoursePrerequisite).all()
    prereq_map: dict[str, list[str]] = {}
    for p in all_prereqs:
        prereq_map.setdefault(p.course_code, []).append(p.prerequisite_code)

    # Map course_code → list[spec] từ bảng M2M
    m2m_rows = db.query(models.CourseSpecialization).all()
    m2m_map: dict[str, list[str]] = {}
    for r in m2m_rows:
        m2m_map.setdefault(r.course_code, []).append(r.specialization)

    # Aggregate ratings per course
    rating_rows = db.query(
        models.CourseRating.course_code,
        _sf.avg(models.CourseRating.rating).label("avg_r"),
        _sf.count(models.CourseRating.id).label("count_r"),
    ).group_by(models.CourseRating.course_code).all()
    rating_map: dict[str, tuple[float, int]] = {
        r.course_code: (round(float(r.avg_r), 1), int(r.count_r)) for r in rating_rows
    }

    # Skill count per course
    skill_count_rows = db.query(
        models.CourseSkill.course_code,
        _sf.count(models.CourseSkill.id).label("n"),
    ).group_by(models.CourseSkill.course_code).all()
    skill_count_map: dict[str, int] = {r.course_code: int(r.n) for r in skill_count_rows}

    def _specs_for(c):
        s = set()
        if c.required_specialization:
            s.add(c.required_specialization)
        for x in m2m_map.get(c.course_code, []):
            s.add(x)
        return sorted(s)

    def _cd(c):
        avg_r, count_r = rating_map.get(c.course_code, (None, 0))
        return {
            "id": c.id,
            "course_code": c.course_code,
            "course_name": c.course_name,
            "credits": float(c.credits) if c.credits is not None else None,
            "count_toward_credits": c.count_toward_credits,
            "description": c.description,
            "required_specialization": c.required_specialization,
            "specializations": _specs_for(c),
            "typical_semester": c.typical_semester if c.typical_semester is not None else CURRICULUM_ORDER.get(c.course_code),
            "prereqs": prereq_map.get(c.course_code, []),
            "avg_rating": avg_r,
            "rating_count": count_r,
            "skill_count": skill_count_map.get(c.course_code, 0),
        }

    def _sort(lst):
        lst.sort(key=lambda x: (x["typical_semester"] or 99, x["course_code"]))
        return lst

    if spec == "common":
        # Môn dùng chung: required_specialization=NULL VÀ không có row M2M nào
        all_courses = db.query(models.Course).filter(
            models.Course.required_specialization == None
        ).order_by(models.Course.course_code).all()
        # Lọc bỏ những môn có row M2M (vì đó là môn nhiều CN, không phải dùng chung)
        courses = [c for c in all_courses if c.course_code not in m2m_map]
        return {
            "spec": "common",
            "compulsory": _sort([_cd(c) for c in courses]),
            "elective_a": {"min_credits": 0, "courses": []},
            "elective_b": {"min_credits": 0, "courses": []},
            "elective_c": {"min_credits": 0, "courses": []},
        }

    # Pool A uses specialization='Chung'; pool B/C use the spec code.
    elective_map: dict[str, str] = {}
    for eg in db.query(models.CourseElectiveGroup).filter(
        models.CourseElectiveGroup.specialization.in_([spec, "Chung"])
    ).all():
        elective_map[eg.course_code] = eg.group_type.upper()

    rules = db.query(models.ElectiveRule).filter(
        models.ElectiveRule.specialization.in_([spec, "Chung"])
    ).all()
    min_creds = {r.group_type.upper(): float(r.min_credits_required) for r in rules}

    from sqlalchemy import or_
    # Codes thuộc CN này qua bảng M2M
    m2m_codes_for_spec = {code for code, specs in m2m_map.items() if spec in specs}
    # Include:
    #   - dùng chung thực sự (required=NULL VÀ không có row M2M nào),
    #   - bắt buộc cho CN này (required=spec),
    #   - thuộc CN này qua bảng M2M (course_code in m2m_codes_for_spec)
    base_courses = db.query(models.Course).filter(
        or_(
            models.Course.required_specialization == None,
            models.Course.required_specialization == spec,
            models.Course.course_code.in_(m2m_codes_for_spec) if m2m_codes_for_spec else False,
        )
    ).all()
    # Lọc bỏ môn có M2M nhưng không thuộc CN này (giữ môn dùng chung & môn thuộc CN này)
    base_courses = [
        c for c in base_courses
        if c.required_specialization == spec
        or c.course_code in m2m_codes_for_spec
        or (c.required_specialization is None and c.course_code not in m2m_map)
    ]
    base_codes = {c.course_code for c in base_courses}

    # Also pull any pool B/C courses not already covered (edge case)
    extra_codes = [code for code in elective_map if code not in base_codes]
    extra_courses = db.query(models.Course).filter(
        models.Course.course_code.in_(extra_codes)
    ).all() if extra_codes else []

    comp, ea, eb, ec = [], [], [], []
    for c in list(base_courses) + list(extra_courses):
        gt = elective_map.get(c.course_code)
        cd = _cd(c)
        if gt == "A":   ea.append(cd)
        elif gt == "B": eb.append(cd)
        elif gt == "C": ec.append(cd)
        else:           comp.append(cd)

    return {
        "spec": spec,
        "compulsory": _sort(comp),
        "elective_a": {"min_credits": min_creds.get("A", 0), "courses": _sort(ea)},
        "elective_b": {"min_credits": min_creds.get("B", 0), "courses": _sort(eb)},
        "elective_c": {"min_credits": min_creds.get("C", 0), "courses": _sort(ec)},
    }


@app.get("/admin/courses/all-ctdt")
def admin_courses_all_ctdt(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Return every course in all specs, grouped by BB / TC-A / TC-B / TC-C.
    Each course carries applied_specs (list of spec codes) and group_type."""
    _require_admin(authorization, db)
    from backend.core.academic_engine import CURRICULUM_ORDER

    all_courses = db.query(models.Course).all()

    all_prereqs = db.query(models.CoursePrerequisite).all()
    prereq_map: dict[str, list[str]] = {}
    for p in all_prereqs:
        prereq_map.setdefault(p.course_code, []).append(p.prerequisite_code)

    # course_code → {group: 'A'/'B'/'C', specs: [spec1,...]}
    elective_by_course: dict[str, dict] = {}
    for eg in db.query(models.CourseElectiveGroup).all():
        entry = elective_by_course.setdefault(eg.course_code, {"group": eg.group_type.upper(), "specs": []})
        entry["specs"].append(eg.specialization)

    # M2M course_specializations
    m2m_for_course: dict[str, list[str]] = {}
    for r in db.query(models.CourseSpecialization).all():
        m2m_for_course.setdefault(r.course_code, []).append(r.specialization)

    ALL_SPECS = ["7480201_07", "7480201_06", "7480201_05", "7480201_09", "7480201_04", "7480201_08"]

    def _cd(c, group_type, applied_specs):
        return {
            "id": c.id,
            "course_code": c.course_code,
            "course_name": c.course_name,
            "credits": float(c.credits) if c.credits is not None else None,
            "count_toward_credits": c.count_toward_credits,
            "description": c.description,
            "required_specialization": c.required_specialization,
            "typical_semester": c.typical_semester if c.typical_semester is not None else CURRICULUM_ORDER.get(c.course_code),
            "prereqs": prereq_map.get(c.course_code, []),
            "group_type": group_type,
            "applied_specs": applied_specs,
        }

    def _sort(lst):
        lst.sort(key=lambda x: (x["typical_semester"] or 99, x["course_code"]))
        return lst

    comp, ea, eb, ec = [], [], [], []
    for c in all_courses:
        eg_info = elective_by_course.get(c.course_code)
        if eg_info is None:
            # Compulsory: tổng hợp specs từ legacy + M2M
            specs_set: set[str] = set()
            if c.required_specialization:
                specs_set.add(c.required_specialization)
            for s in m2m_for_course.get(c.course_code, []):
                specs_set.add(s)
            if not specs_set:
                applied = ALL_SPECS  # dùng chung mọi CN
            else:
                applied = sorted(specs_set)
            comp.append(_cd(c, "BB", applied))
        else:
            gt = eg_info["group"]
            applied = eg_info["specs"]
            cd = _cd(c, f"TC-{gt}", applied)
            if gt == "A":   ea.append(cd)
            elif gt == "B": eb.append(cd)
            elif gt == "C": ec.append(cd)
            else:           comp.append(cd)

    return {
        "spec": "all",
        "compulsory": _sort(comp),
        "elective_a": {"min_credits": 0, "courses": _sort(ea)},
        "elective_b": {"min_credits": 0, "courses": _sort(eb)},
        "elective_c": {"min_credits": 0, "courses": _sort(ec)},
    }


def _course_specs_set(db: Session, course_code: str) -> list[str]:
    """Trả về tập hợp tất cả CN môn này yêu cầu (union legacy + M2M)."""
    course = db.query(models.Course).filter(models.Course.course_code == course_code).first()
    specs: set[str] = set()
    if course and course.required_specialization:
        specs.add(course.required_specialization)
    rows = db.query(models.CourseSpecialization).filter(
        models.CourseSpecialization.course_code == course_code
    ).all()
    for r in rows:
        if r.specialization:
            specs.add(r.specialization)
    return sorted(specs)


def _course_to_out(db: Session, c: models.Course) -> dict:
    """Pack 1 Course thành dict có thêm `specializations`."""
    return {
        "id": c.id,
        "course_code": c.course_code,
        "course_name": c.course_name,
        "credits": float(c.credits) if c.credits is not None else None,
        "count_toward_credits": c.count_toward_credits,
        "description": c.description,
        "required_specialization": c.required_specialization,
        "typical_semester": c.typical_semester,
        "specializations": _course_specs_set(db, c.course_code),
    }


def _apply_specs_to_course(db: Session, course: models.Course, specs: list[str] | None):
    """Áp dụng list CN cho 1 môn:
    - None → không sửa M2M (chỉ dùng required_specialization từ payload)
    - [] → dùng chung mọi CN: required_specialization=NULL, xoá course_specs
    - [X] → 1 CN: required_specialization=X, xoá course_specs
    - [X,Y,...] → nhiều CN: required_specialization=NULL, ghi vào course_specs
    """
    if specs is None:
        return
    cleaned = [s.strip() for s in specs if s and s.strip()]
    # Xoá toàn bộ rows hiện tại
    db.query(models.CourseSpecialization).filter(
        models.CourseSpecialization.course_code == course.course_code
    ).delete()
    if not cleaned:
        course.required_specialization = None
    elif len(cleaned) == 1:
        course.required_specialization = cleaned[0]
    else:
        course.required_specialization = None
        for s in cleaned:
            db.add(models.CourseSpecialization(course_code=course.course_code, specialization=s))


@app.post("/admin/courses", response_model=schemas.CourseOut)
def admin_create_course(
    payload: schemas.CourseAdminIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    existing = db.query(models.Course).filter(models.Course.course_code == payload.course_code).first()
    if existing:
        raise HTTPException(status_code=409, detail=f"Mã môn '{payload.course_code}' đã tồn tại")
    # Auto-detect GDTC/QPAN → không tính TC tích lũy
    _GDTC_CODES = {'7010701', '7010702', '7010703'}
    code_stripped = (payload.course_code or '').strip()
    if code_stripped in _GDTC_CODES or code_stripped.startswith('73'):
        count_toward = False
    else:
        count_toward = payload.count_toward_credits
    course = models.Course(
        course_code=payload.course_code,
        course_name=payload.course_name,
        credits=payload.credits,
        description=payload.description,
        count_toward_credits=count_toward,
        required_specialization=payload.required_specialization,
        typical_semester=payload.typical_semester,
    )
    db.add(course)
    db.flush()
    _apply_specs_to_course(db, course, payload.specializations)
    db.commit()
    db.refresh(course)
    _log(db, admin, "CREATE_COURSE", "course", course.course_code,
         f"{course.course_name} | {course.credits} TC")
    return _course_to_out(db, course)


@app.put("/admin/courses/{course_id}", response_model=schemas.CourseOut)
def admin_update_course(
    course_id: int,
    payload: schemas.CourseAdminIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    import json as _json
    admin = _require_admin(authorization, db)
    course = db.query(models.Course).filter(models.Course.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Không tìm thấy môn học")
    conflict = db.query(models.Course).filter(
        models.Course.course_code == payload.course_code,
        models.Course.id != course_id,
    ).first()
    if conflict:
        raise HTTPException(status_code=409, detail=f"Mã môn '{payload.course_code}' đã tồn tại")
    before = {
        "code": course.course_code,
        "name": course.course_name,
        "credits": str(course.credits),
        "specs": _course_specs_set(db, course.course_code),
        "hk": course.typical_semester,
    }
    old_code = course.course_code
    course.course_code = payload.course_code
    course.course_name = payload.course_name
    course.credits = payload.credits
    course.description = payload.description
    course.count_toward_credits = payload.count_toward_credits
    course.typical_semester = payload.typical_semester
    # Nếu specializations được truyền: áp dụng (có thể override required_specialization)
    if payload.specializations is not None:
        # Course code có thể đã đổi → cập nhật course_specs theo code mới
        if old_code != payload.course_code:
            db.query(models.CourseSpecialization).filter(
                models.CourseSpecialization.course_code == old_code
            ).update({"course_code": payload.course_code})
        _apply_specs_to_course(db, course, payload.specializations)
    elif payload.required_specialization is not None:
        # Chỉ set required_specialization khi client truyền giá trị cụ thể.
        # Nếu cả 2 fields None → là partial update (vd. _savePanelHk), giữ nguyên spec.
        course.required_specialization = payload.required_specialization
    db.commit()
    db.refresh(course)
    after = {
        "code": course.course_code,
        "name": course.course_name,
        "credits": str(course.credits),
        "specs": _course_specs_set(db, course.course_code),
        "hk": course.typical_semester,
    }
    _log(db, admin, "UPDATE_COURSE", "course", course.course_code,
         _json.dumps({"before": before, "after": after}, ensure_ascii=False))
    return _course_to_out(db, course)


@app.patch("/admin/courses/{course_id}/count-toward-credits", response_model=schemas.CourseOut)
def admin_toggle_count_toward_credits(
    course_id: int,
    count_toward_credits: bool,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    course = db.query(models.Course).filter(models.Course.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Không tìm thấy môn học")
    old_val = course.count_toward_credits
    course.count_toward_credits = count_toward_credits
    db.commit()
    db.refresh(course)
    _log(db, admin, "TOGGLE_COUNT_CREDITS", "course", course.course_code,
         f"{course.course_name}: {old_val} → {count_toward_credits}")
    return course


@app.patch("/admin/courses/{course_code}/semester")
def admin_update_course_semester(
    course_code: str,
    payload: dict,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Update typical_semester for plan view drag-drop."""
    _require_admin(authorization, db)
    course = db.query(models.Course).filter(models.Course.course_code == course_code).first()
    if not course:
        raise HTTPException(status_code=404, detail="Không tìm thấy môn học")
    sem = payload.get("semester")
    if sem is not None and not (1 <= int(sem) <= 9):
        raise HTTPException(status_code=422, detail="semester phải từ 1 đến 9")
    course.typical_semester = int(sem) if sem is not None else None
    db.commit()
    return {"ok": True, "course_code": course_code, "semester": course.typical_semester}


@app.get("/admin/courses/{course_code}/relations")
def admin_course_relations(
    course_code: str,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Return prerequisite graph + elective groups for a course."""
    _require_admin(authorization, db)
    course = db.query(models.Course).filter(models.Course.course_code == course_code).first()
    if not course:
        raise HTTPException(status_code=404, detail="Không tìm thấy môn học")

    prereq_codes = [r.prerequisite_code for r in
                    db.query(models.CoursePrerequisite)
                    .filter(models.CoursePrerequisite.course_code == course_code).all()]
    prereq_courses = db.query(models.Course).filter(models.Course.course_code.in_(prereq_codes)).all() if prereq_codes else []

    successor_codes = [r.course_code for r in
                       db.query(models.CoursePrerequisite)
                       .filter(models.CoursePrerequisite.prerequisite_code == course_code).all()]
    successor_courses = db.query(models.Course).filter(models.Course.course_code.in_(successor_codes)).all() if successor_codes else []

    eg_rows = db.query(models.CourseElectiveGroup).filter(
        models.CourseElectiveGroup.course_code == course_code
    ).all()

    return {
        "prereqs": [{"code": c.course_code, "name": c.course_name, "credits": float(c.credits or 0)} for c in prereq_courses],
        "successors": [{"code": c.course_code, "name": c.course_name, "credits": float(c.credits or 0)} for c in successor_courses],
        "elective_groups": [{"program_code": eg.program_code, "specialization": eg.specialization, "group_type": eg.group_type} for eg in eg_rows],
    }


@app.delete("/admin/courses/{course_id}", response_model=schemas.MessageOut)
def admin_delete_course(
    course_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    course = db.query(models.Course).filter(models.Course.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Không tìm thấy môn học")
    name, code = course.course_name, course.course_code
    db.delete(course)
    db.commit()
    _log(db, admin, "DELETE_COURSE", "course", code, name)
    return schemas.MessageOut(message=f"Đã xóa môn '{name}'")


@app.post("/admin/courses/bootstrap-default", response_model=schemas.CourseBootstrapOut)
def admin_bootstrap_default_courses(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    _require_admin(authorization, db)
    try:
        stats = import_curriculum_from_excel(db, file_path=str(DEFAULT_CURRICULUM_FILE), replace_existing=True)
        db.commit()
        return schemas.CourseBootstrapOut(
            inserted=stats.inserted_courses,
            skipped=stats.protected_courses,
            total_rows=stats.total_rows,
        )
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Lỗi import: {exc}")


@app.post("/admin/ml/retrain", response_model=schemas.MessageOut)
def admin_ml_retrain(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Retrain the ML recommendation model using synthetic data generated from current DB."""
    _require_admin(authorization, db)
    try:
        from backend.core.ml_trainer import train as ml_train
        admin = _require_admin(authorization, db)
        artifact = ml_train(db=db)
        if artifact is None:
            raise HTTPException(status_code=500, detail="Không đủ dữ liệu để huấn luyện model")
        roc = artifact.get("cv_roc_auc", 0)
        _log(db, admin, "ml_retrain", detail=f"ROC-AUC={roc:.3f}")
        return schemas.MessageOut(message=f"Model đã được huấn luyện lại. ROC-AUC = {roc:.3f}")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Lỗi huấn luyện model: {exc}")


@app.get("/admin/prerequisites", response_model=list[schemas.CoursePrerequisiteOut])
def list_prerequisites(
    course_code: str | None = None,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    _require_admin(authorization, db)
    q = db.query(models.CoursePrerequisite)
    if course_code:
        q = q.filter(models.CoursePrerequisite.course_code == course_code.strip())
    return q.order_by(models.CoursePrerequisite.course_code.asc()).all()


@app.post("/admin/prerequisites", response_model=schemas.CoursePrerequisiteOut)
def add_prerequisite(
    payload: schemas.CoursePrerequisiteIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    if payload.course_code == payload.prerequisite_code:
        raise HTTPException(status_code=400, detail="Môn học không thể là tiên quyết của chính nó")
    for code in (payload.course_code, payload.prerequisite_code):
        if not db.query(models.Course).filter(models.Course.course_code == code).first():
            raise HTTPException(status_code=404, detail=f"Không tìm thấy môn '{code}'")
    existing = db.query(models.CoursePrerequisite).filter(
        models.CoursePrerequisite.course_code == payload.course_code,
        models.CoursePrerequisite.prerequisite_code == payload.prerequisite_code,
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail="Quan hệ tiên quyết này đã tồn tại")

    # DFS cycle check: adding prereq → course would cycle if course already reaches prereq
    _all_p = db.query(models.CoursePrerequisite).all()
    _graph: dict[str, set[str]] = {}
    for _p in _all_p:
        _graph.setdefault(_p.course_code, set()).add(_p.prerequisite_code)

    def _reachable(start: str, target: str) -> bool:
        visited: set[str] = set()
        stack = [start]
        while stack:
            node = stack.pop()
            if node in visited:
                continue
            visited.add(node)
            for nb in _graph.get(node, set()):
                if nb == target:
                    return True
                stack.append(nb)
        return False

    if _reachable(payload.prerequisite_code, payload.course_code):
        raise HTTPException(status_code=400, detail="Thêm tiên quyết này sẽ tạo vòng lặp trong đồ thị tiên quyết")

    p = models.CoursePrerequisite(
        course_code=payload.course_code,
        prerequisite_code=payload.prerequisite_code,
    )
    db.add(p)
    db.commit()
    db.refresh(p)
    _log(db, admin, "ADD_PREREQ", "prerequisite",
         f"{payload.prerequisite_code}→{payload.course_code}",
         f"{payload.prerequisite_code} là tiên quyết của {payload.course_code}")
    return p


@app.delete("/admin/prerequisites/{prereq_id}", response_model=schemas.MessageOut)
def delete_prerequisite(
    prereq_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    p = db.query(models.CoursePrerequisite).filter(models.CoursePrerequisite.id == prereq_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Không tìm thấy quan hệ tiên quyết")
    detail = f"{p.prerequisite_code} → {p.course_code}"
    db.delete(p)
    db.commit()
    _log(db, admin, "DELETE_PREREQ", "prerequisite", detail, detail)
    return schemas.MessageOut(message=f"Đã xóa tiên quyết {detail}")


@app.delete("/admin/prerequisites/{course_code}/{prereq_code}", response_model=schemas.MessageOut)
def delete_prerequisite_by_codes(
    course_code: str,
    prereq_code: str,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    p = db.query(models.CoursePrerequisite).filter(
        models.CoursePrerequisite.course_code == course_code,
        models.CoursePrerequisite.prerequisite_code == prereq_code,
    ).first()
    if not p:
        raise HTTPException(status_code=404, detail="Không tìm thấy quan hệ tiên quyết")
    detail = f"{prereq_code} → {course_code}"
    db.delete(p)
    db.commit()
    _log(db, admin, "DELETE_PREREQ", "prerequisite", detail, detail)
    return schemas.MessageOut(message=f"Đã xóa tiên quyết {detail}")


@app.get("/elective-rules", response_model=list[schemas.ElectiveRuleOut])
def list_elective_rules(
    specialization: str | None = None,
    program_code: str | None = None,
    db: Session = Depends(get_db),
):
    q = db.query(models.ElectiveRule)
    if specialization:
        q = q.filter(models.ElectiveRule.specialization == specialization)
    if program_code:
        q = q.filter(models.ElectiveRule.program_code == program_code)
    return q.order_by(models.ElectiveRule.specialization.asc(), models.ElectiveRule.group_type.asc()).all()


_MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB

# Minimum number of matching specialization-specific courses to be confident
_SPEC_DETECT_MIN_MATCHES = 2


def _detect_specialization(grade_codes: set[str], db) -> str | None:
    """
    Infer a student's specialization from the course codes in their transcript.

    Strategy (two passes):
    1. required_specialization courses — each is unique to one spec, strongest signal
    2. course_elective_groups — courses in elective pools, weighted by group
    Returns the best-match specialization, or None if confidence is too low.
    """
    from collections import Counter
    scores: Counter = Counter()

    # Pass 1: mandatory specialization courses (strong signal — weight 3)
    spec_required = db.query(models.Course).filter(
        models.Course.required_specialization.isnot(None),
        models.Course.required_specialization != "Chung",
    ).all()
    for c in spec_required:
        if c.course_code in grade_codes:
            scores[c.required_specialization] += 3

    # Pass 2: elective group memberships (softer signal — weight 1)
    elective_mappings = db.query(models.CourseElectiveGroup).filter(
        models.CourseElectiveGroup.specialization != "Chung"
    ).all()
    for m in elective_mappings:
        if m.course_code in grade_codes:
            scores[m.specialization] += 1

    if not scores:
        return None

    best_spec, best_score = scores.most_common(1)[0]
    # Require at least a minimum confidence score before committing
    if best_score < _SPEC_DETECT_MIN_MATCHES:
        return None
    return best_spec


@app.post("/grades/upload", response_model=schemas.GradeUploadOut)
def upload_grades(
    file: UploadFile,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """SV upload bảng điểm cá nhân.

    Mục đích: SV dùng cho lập kế hoạch cá nhân (lộ trình, AI advisor, gợi ý).
    Đây là DỮ LIỆU TỰ KHAI — EduGuide là tool cá nhân, không thay thế cổng
    SIS chính thức. Refactor 2026-05-05: bỏ logic merge admin/self, bỏ
    grades_locked, bỏ source field — mỗi lần upload là full replace.
    """
    user = _get_user_by_token(authorization, db)
    if user.role != "student":
        raise HTTPException(
            status_code=403,
            detail="Endpoint này chỉ dành cho sinh viên.",
        )

    data = file.file.read(_MAX_UPLOAD_BYTES + 1)
    if len(data) > _MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File quá lớn. Giới hạn tối đa 10 MB.")
    filename = file.filename or "upload.xlsx"

    try:
        _parsed = read_rows_from_upload(filename, data)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception:
        raise HTTPException(status_code=400, detail="Khong the doc file. Kiem tra lai dinh dang (xlsx/csv).")
    rows = _parsed["rows"]
    parsed_name = _parsed.get("full_name") or None
    if parsed_name and not user.full_name:
        user.full_name = normalize_vietnamese_name(parsed_name)

    grade_records = extract_grades(rows)
    if not grade_records:
        raise HTTPException(status_code=400, detail="Khong tim thay du lieu diem trong file.")

    all_courses = {c.course_code: c for c in db.query(models.Course).all()}

    # ── Pre-compute valid courses cho chuyên ngành của SV (CTĐT) ─────────────
    # Để cảnh báo "môn ngoài CTDT": môn tồn tại trong system nhưng không thuộc
    # required_specialization của SV. SV vẫn được upload (insert), chỉ thêm
    # warning vào issues.
    user_spec = user.specialization
    courses_in_user_ctdt: set[str] = set()
    if user_spec:
        # Single-spec qua required_specialization
        for c in all_courses.values():
            if c.required_specialization == user_spec or c.required_specialization is None:
                courses_in_user_ctdt.add(c.course_code)
        # M2M qua course_specializations
        m2m_codes = db.query(models.CourseSpecialization.course_code).filter(
            models.CourseSpecialization.specialization == user_spec
        ).all()
        for (code,) in m2m_codes:
            courses_in_user_ctdt.add(code)

    # Snapshot GPA before replacing
    old_grades = db.query(models.UserGrade).filter(
        models.UserGrade.user_id == user.id, models.UserGrade.passed == True
    ).all()
    def _calc_gpa4(grades):
        num, den = 0.0, 0.0
        for g in grades:
            tc = float(all_courses[g.course_code].credits) if g.course_code in all_courses else 0.0
            s4 = float(g.score4) if g.score4 is not None else None
            if s4 is not None and tc > 0:
                num += s4 * tc; den += tc
        return round(num / den, 2) if den > 0 else None
    gpa4_before = _calc_gpa4(old_grades)
    credits_before = int(sum(float(all_courses[g.course_code].credits) for g in old_grades if g.course_code in all_courses))

    # Full replace: xoá hết grades cũ, insert grades mới.
    db.query(models.UserGrade).filter(
        models.UserGrade.user_id == user.id,
    ).delete(synchronize_session=False)

    # Dedup: CSV may contain retakes for the same course — keep best score per course_code.
    best_per_course: dict[str, dict] = {}
    for rec in grade_records:
        code = rec["course_code"]
        existing = best_per_course.get(code)
        if existing is None:
            best_per_course[code] = rec
        else:
            s_new = rec.get("score10") or -1
            s_old = existing.get("score10") or -1
            if s_new > s_old:
                best_per_course[code] = rec
    grade_records = list(best_per_course.values())

    inserted = 0
    updated = 0  # kept for API compatibility; always 0 in replace mode
    skipped_unknown = 0
    skipped_admin_verified = 0  # always 0 sau refactor — kept for API compat
    issues: list[schemas.GradeUploadIssue] = []

    for rec in grade_records:
        code = rec["course_code"]
        if code not in all_courses:
            issues.append(schemas.GradeUploadIssue(
                course_code_in_file=code,
                course_name_in_file=rec.get("course_name"),
                reason="Mã môn không có trong CTĐT",
            ))
            skipped_unknown += 1
            continue

        # Cảnh báo môn ngoài CTDT của SV (nhưng vẫn insert — SV tự khai)
        if user_spec and courses_in_user_ctdt and code not in courses_in_user_ctdt:
            course_obj = all_courses.get(code)
            issues.append(schemas.GradeUploadIssue(
                course_code_in_file=code,
                course_name_in_file=rec.get("course_name") or (course_obj.course_name if course_obj else None),
                reason=f"Môn không thuộc CTĐT chuyên ngành của bạn (vẫn được lưu, nhưng không tính vào tốt nghiệp)",
            ))

        db.add(models.UserGrade(
            user_id=user.id,
            course_code=code,
            score10=rec.get("score10"),
            score4=rec.get("score4"),
            letter=rec.get("letter"),
            passed=rec.get("passed", False),
            term=rec.get("term"),
        ))
        inserted += 1

    # Audit log: SV upload điểm
    try:
        _log(db, user, "STUDENT_UPLOAD_GRADES", "user_grades", str(user.id),
             f"{inserted} môn (filename={filename})")
    except Exception:
        pass

    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Loi luu du lieu: {exc}")

    # Auto-update registration outcomes for matched grades
    for rec in grade_records:
        code = rec["course_code"]
        if code not in all_courses:
            continue
        reg = db.query(models.CourseRegistration).filter(
            models.CourseRegistration.user_id == user.id,
            models.CourseRegistration.course_code == code,
            models.CourseRegistration.outcome == None,
        ).first()
        if reg:
            reg.outcome = "passed" if rec.get("passed") else "failed"
    try:
        db.commit()
    except Exception:
        db.rollback()

    invalidate_student_context_cache(user.id)
    invalidate_difficulty_stats_cache()  # refresh aggregate difficulty stats

    # Snapshot GPA after
    new_grades = db.query(models.UserGrade).filter(
        models.UserGrade.user_id == user.id, models.UserGrade.passed == True
    ).all()
    gpa4_after = _calc_gpa4(new_grades)
    credits_after = int(sum(float(all_courses[g.course_code].credits) for g in new_grades if g.course_code in all_courses))

    # Approach C: KHÔNG auto-detect specialization từ self-upload.
    # CN do trường phân (qua admin import roster) — app không tự suy luận.
    # Cố vấn cũng KHÔNG được auto-assign từ self-upload (advisor view chỉ
    # dùng source='admin', nên việc gán CV trên data tự khai là vô nghĩa).
    return schemas.GradeUploadOut(
        inserted=inserted, updated=updated,
        skipped_unknown=skipped_unknown,
        skipped_admin_verified=skipped_admin_verified,
        issues=issues,
        gpa4_before=gpa4_before, gpa4_after=gpa4_after,
        credits_before=credits_before, credits_after=credits_after,
        official_tich_luy=tich_luy,
        detected_specialization=None,
        spec_auto_set=False,
        advisor_assigned=None,
        advisor_warning=None,
    )


@app.get("/grades/detected-spec")
def grades_detected_spec(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    grade_codes = {
        g.course_code
        for g in db.query(models.UserGrade).filter(models.UserGrade.user_id == user.id).all()
    }
    detected = _detect_specialization(grade_codes, db) if grade_codes else None
    return {"detected_specialization": detected}


@app.get("/grades/me", response_model=list[schemas.UserGradeOut])
def list_my_grades(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    return (
        db.query(models.UserGrade)
        .filter(models.UserGrade.user_id == user.id)
        .order_by(models.UserGrade.term.asc().nulls_last(), models.UserGrade.course_code.asc())
        .all()
    )


# REMOVED 2026-05-05: PATCH /admin/users/{id}/grades-lock và GET /grades/me/status
# đã bị bỏ. Mô hình mới: SV luôn có thể upload điểm cá nhân, không có khái niệm
# admin lock hoặc verified source. UI luôn hiển thị "data tự khai".


@app.get("/grades/me/status")
def get_my_grades_status(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Trả trạng thái upload cho frontend SV. Sau refactor 2026-05-05: luôn cho upload."""
    user = _get_user_by_token(authorization, db)
    if user.role != "student":
        return {"locked": False, "can_upload": False}
    grade_count = db.query(models.UserGrade).filter(
        models.UserGrade.user_id == user.id
    ).count()
    return {
        "locked": False,
        "can_upload": True,
        "lock_reason": None,
        "grade_count": grade_count,
    }


@app.patch("/auth/me/specialization", response_model=schemas.UserOut)
def update_specialization(
    payload: schemas.SpecializationIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    old_spec = user.specialization
    new_spec = payload.specialization
    user.specialization = new_spec
    db.commit()
    db.refresh(user)

    # Triết lý: đổi CN = bắt đầu lại việc chọn TC.
    #  1. Pool TC-B/C khác giữa CN, môn cũ có thể không còn pool mới
    #  2. Ngay cả cùng pool, độ ưu tiên/recommendation khác theo định hướng CN
    #  3. Buộc SV ý thức chọn lại TC phù hợp CN mới (UX rõ ràng > implicit filter)
    # → Wipe TẤT CẢ TC đã plan (PlannedElective + StudyPlanItem TC).
    # BB cơ sở ngành sẽ tự update qua CTDT mới.
    if old_spec != new_spec and new_spec:
        # Lấy set tất cả TC course codes (xuất hiện trong course_elective_groups)
        tc_codes = {
            row[0] for row in db.query(models.CourseElectiveGroup.course_code).distinct().all()
        }
        if tc_codes:
            # 1. Xoá PlannedElective TC
            (
                db.query(models.PlannedElective)
                .filter(
                    models.PlannedElective.user_id == user.id,
                    models.PlannedElective.course_code.in_(tc_codes),
                )
                .delete(synchronize_session=False)
            )
            # 2. Xoá StudyPlanItem TC (môn TC drag-drop trong roadmap_custom)
            custom_plan = db.query(models.StudyPlan).filter(
                models.StudyPlan.user_id == user.id,
                models.StudyPlan.plan_name == "roadmap_custom",
            ).first()
            if custom_plan:
                (
                    db.query(models.StudyPlanItem)
                    .filter(
                        models.StudyPlanItem.plan_id == custom_plan.id,
                        models.StudyPlanItem.course_code.in_(tc_codes),
                    )
                    .delete(synchronize_session=False)
                )
            db.commit()

    invalidate_student_context_cache(user.id)
    return _to_user_out(user)


@app.patch("/auth/me/full-name", response_model=schemas.UserOut)
def update_full_name(
    payload: schemas.UpdateFullNameIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    user.full_name = normalize_vietnamese_name(payload.full_name) or None
    db.commit()
    db.refresh(user)
    return _to_user_out(user)


@app.post("/auth/me/change-password", response_model=schemas.MessageOut)
def change_password(
    payload: schemas.ChangePasswordIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    if not _verify_password(payload.current_password, user.password_hash):
        raise HTTPException(status_code=400, detail="Mật khẩu hiện tại không đúng")
    _validate_password(payload.new_password)
    user.password_hash = _hash_password(payload.new_password)
    user.default_password = None
    user.is_first_login = False
    # Revoke all existing sessions so other devices must re-login
    db.query(models.AuthToken).filter(models.AuthToken.user_id == user.id).delete()
    db.commit()
    return schemas.MessageOut(message="Đổi mật khẩu thành công. Vui lòng đăng nhập lại.")


# REMOVED 2026-05-05 (Phase 6):
#   • PATCH /auth/me/setup-account     — first-login modal đã bỏ
#   • POST  /auth/forgot-password-edu  — không còn email cá nhân
#   • POST  /auth/forgot-password      — SMTP flow không dùng
#   • POST  /auth/reset-password       — token reset SMTP
#   • _send_reset_email() helper
# Mô hình mới: Admin tạo SV với default_password (6 chữ số). SV login →
# vào thẳng app. SV đổi password tự do qua Settings (POST /auth/change-password).
# Nếu SV quên pw → admin reset (POST /admin/users/{id}/reset-password) → admin
# nhận pw mới + giao SV. Không có email cá nhân, không có self-service forgot.


@app.get("/study-goal/me", response_model=schemas.StudyGoalOut)
def get_study_goal(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    plan = db.query(models.StudyPlan).filter(models.StudyPlan.user_id == user.id).order_by(models.StudyPlan.id.desc()).first()
    target_gpa = float(plan.target_gpa) if plan and plan.target_gpa is not None else None
    return schemas.StudyGoalOut(target_gpa=target_gpa)


@app.put("/study-goal/me", response_model=schemas.StudyGoalOut)
def set_study_goal(
    payload: schemas.StudyGoalIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    plan = db.query(models.StudyPlan).filter(models.StudyPlan.user_id == user.id).order_by(models.StudyPlan.id.desc()).first()
    if plan:
        plan.target_gpa = payload.target_gpa
    else:
        plan = models.StudyPlan(user_id=user.id, plan_name="default", target_gpa=payload.target_gpa)
        db.add(plan)
    db.commit()
    db.refresh(plan)
    return schemas.StudyGoalOut(target_gpa=float(plan.target_gpa) if plan.target_gpa is not None else None)


@app.get("/progress/me", response_model=schemas.ProgressOut)
def progress_me(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    plan = db.query(models.StudyPlan).filter(models.StudyPlan.user_id == user.id).order_by(models.StudyPlan.id.desc()).first()
    target_gpa = float(plan.target_gpa) if plan and plan.target_gpa is not None else None
    return schemas.ProgressOut(**build_progress_snapshot(db, user.id, target_gpa=target_gpa))


@app.get("/recommendations/me", response_model=schemas.RecommendationOut)
def recommendations_me(
    limit: int = 5,
    career_goal: str | None = None,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    payload = build_recommendations(
        db, user.id, limit=limit,
        available_course_codes=None,
        career_goal=career_goal,
    )
    # Get active registrations (outcome=NULL = currently studying) and filter from recommendations
    active_reg_codes = {
        r.course_code for r in db.query(models.CourseRegistration).filter(
            models.CourseRegistration.user_id == user.id,
            models.CourseRegistration.outcome == None,
        ).all()
    }
    if active_reg_codes:
        payload["recommendations"] = [
            r for r in payload["recommendations"]
            if r["course_code"] not in active_reg_codes
        ]
    return schemas.RecommendationOut(**payload)


@app.get("/recommendations/why-not/{course_code}")
def recommendations_why_not(
    course_code: str,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Explain why a course is/isn't recommended — structured + multi-reason.

    Response shape:
    {
      course_code, course_name, credits,
      verdict: "recommended" | "eligible_low_priority" | "blocked" | "completed" | "studying",
      reasons: [{kind, category, icon, title, detail, data?}, ...],
      // Legacy single-reason fields (backwards compat):
      reason_code, explanation, missing_prerequisites?
    }
    """
    from collections import defaultdict as _dd
    from sqlalchemy import or_ as _or
    from backend.core.academic_engine import (
        _build_snapshot, _find_internship_thesis, _elective_group_remaining,
        _course_credits, INTERNSHIP_REMAINING_BUFFER, _get_academic_thresholds,
        _get_difficulty_stats, _compute_unlock_scores, _counts_toward_credits,
    )

    user = _get_user_by_token(authorization, db)
    specialization = user.specialization

    course = db.query(models.Course).filter(
        models.Course.course_code == course_code
    ).first()
    if not course:
        raise HTTPException(status_code=404, detail="Không tìm thấy môn học")

    snapshot = _build_snapshot(db, user.id, specialization=specialization)
    # Note: official_earned_credits đã được drop ở refactor 2026-05-05.
    # earned_credits giờ tính trực tiếp từ user_grades trong _build_snapshot.

    code = course.course_code
    course_name = course.course_name
    credits_val = float(course.credits) if course.credits else 0.0

    def _build(verdict: str, reasons: list, legacy_code: str, legacy_explanation: str, **extra) -> dict:
        out = {
            "course_code": code,
            "course_name": course_name,
            "credits": credits_val,
            "verdict": verdict,
            "reasons": reasons,
            "reason_code": legacy_code,
            "explanation": legacy_explanation,
        }
        out.update(extra)
        return out

    # ── EARLY EXITS (status of the course in user's life) ────────────────────
    if code not in snapshot.course_by_code:
        r = {
            "kind": "blocker", "category": "spec", "icon": "block",
            "title": "Không thuộc chương trình",
            "detail": "Môn học này không thuộc chương trình đào tạo của bạn.",
        }
        return _build("blocked", [r], "NOT_IN_CURRICULUM", r["detail"])

    if code in snapshot.completed_codes:
        grade = snapshot.best_grades.get(code)
        score_str = ""
        if grade:
            if grade.score10 is not None:
                score_str = f" (điểm {grade.score10:.1f}/10)"
            elif grade.letter:
                score_str = f" (điểm {grade.letter})"
        r = {
            "kind": "info", "category": "completed", "icon": "task_alt",
            "title": "Đã hoàn thành",
            "detail": f"Bạn đã pass môn này{score_str}.",
        }
        return _build("completed", [r], "ALREADY_COMPLETED", r["detail"])

    active_reg = db.query(models.CourseRegistration).filter(
        models.CourseRegistration.user_id == user.id,
        models.CourseRegistration.course_code == code,
        models.CourseRegistration.outcome == None,  # noqa: E711
    ).first()
    if active_reg:
        r = {
            "kind": "info", "category": "schedule", "icon": "schedule",
            "title": "Đang học",
            "detail": "Bạn đang học môn này trong học kỳ hiện tại.",
        }
        return _build("studying", [r], "CURRENTLY_STUDYING", r["detail"])

    # ── BLOCKERS (collect ALL applicable, not just first) ────────────────────
    blockers = []

    if course.required_specialization and course.required_specialization != specialization:
        spec_label = course.required_specialization
        blockers.append({
            "kind": "blocker", "category": "spec", "icon": "school",
            "title": "Sai chuyên ngành",
            "detail": f"Môn này dành cho chuyên ngành {spec_label}, không phải CN của bạn.",
            "data": {"required_spec": course.required_specialization, "user_spec": specialization},
        })

    elective_group_remaining = _elective_group_remaining(
        db, snapshot.course_by_code, snapshot.completed_codes, specialization
    )
    elective_mappings = db.query(models.CourseElectiveGroup).filter(
        _or(
            models.CourseElectiveGroup.specialization == specialization,
            models.CourseElectiveGroup.specialization == "Chung",
        )
    ).all() if specialization else []
    course_to_group = {m.course_code: (m.program_code, m.specialization, m.group_type) for m in elective_mappings}

    group_key = course_to_group.get(code)
    if group_key:
        group_remaining = elective_group_remaining.get(group_key, 0.0)
        if group_remaining <= 0.0:
            blockers.append({
                "kind": "warning", "category": "elective_quota", "icon": "playlist_add_check",
                "title": f"Đã đủ TC tự chọn {group_key[2]}",
                "detail": f"Nhóm tự chọn {group_key[2]} của bạn đã đủ tín chỉ yêu cầu — học thêm sẽ không tính vào chuẩn TN.",
                "data": {"group_type": group_key[2]},
            })

    prereq_map: dict[str, list[str]] = _dd(list)
    for p in db.query(models.CoursePrerequisite).all():
        prereq_map[p.course_code].append(p.prerequisite_code)
    missing_prereqs = [p for p in prereq_map.get(code, []) if p not in snapshot.completed_codes]
    if missing_prereqs:
        prereq_details = []
        for pc in missing_prereqs:
            pc_course = snapshot.course_by_code.get(pc)
            prereq_details.append({
                "course_code": pc,
                "course_name": pc_course.course_name if pc_course else pc,
            })
        names = [p["course_name"] for p in prereq_details]
        blockers.append({
            "kind": "blocker", "category": "prereq", "icon": "lock",
            "title": "Thiếu môn tiên quyết",
            "detail": f"Cần hoàn thành: {', '.join(names)}.",
            "data": {"missing": prereq_details},
        })

    # Internship / thesis gating
    internship_course, thesis_course = _find_internship_thesis(snapshot.courses, specialization)
    internship_done = internship_course and internship_course.course_code in snapshot.completed_codes
    internship_outstanding = _course_credits(internship_course) if internship_course and not internship_done else 0.0
    thesis_outstanding = _course_credits(thesis_course) if thesis_course and thesis_course.course_code not in snapshot.completed_codes else 0.0
    remaining_non_special = snapshot.total_credits - snapshot.earned_credits - thesis_outstanding - internship_outstanding
    _at = _get_academic_thresholds(db)
    _gpa = snapshot.avg_score4 if snapshot.avg_score4 is not None else 0.0

    if thesis_course and code == thesis_course.course_code:
        if not internship_done:
            blockers.append({
                "kind": "blocker", "category": "schedule", "icon": "work_history",
                "title": "Chưa pass thực tập",
                "detail": "Bạn cần hoàn thành thực tập doanh nghiệp trước khi làm đồ án TN.",
            })
        if snapshot.earned_credits < _at["thesis_min_credits"]:
            blockers.append({
                "kind": "blocker", "category": "threshold", "icon": "scoreboard",
                "title": "Chưa đủ TC làm ĐATN",
                "detail": f"Cần ≥ {_at['thesis_min_credits']:.0f} TC, hiện có {snapshot.earned_credits:.1f} TC.",
                "data": {"threshold": _at["thesis_min_credits"], "current": snapshot.earned_credits},
            })
        if _gpa < _at["thesis_min_gpa4"]:
            blockers.append({
                "kind": "blocker", "category": "threshold", "icon": "trending_up",
                "title": "Chưa đủ GPA làm ĐATN",
                "detail": f"Cần GPA hệ 4 ≥ {_at['thesis_min_gpa4']:.2f}, hiện {_gpa:.2f}.",
                "data": {"threshold": _at["thesis_min_gpa4"], "current": _gpa},
            })

    if internship_course and code == internship_course.course_code:
        if snapshot.earned_credits < _at["internship_min_credits"]:
            blockers.append({
                "kind": "blocker", "category": "threshold", "icon": "scoreboard",
                "title": "Chưa đủ TC đi thực tập",
                "detail": f"Cần ≥ {_at['internship_min_credits']:.0f} TC, hiện {snapshot.earned_credits:.1f} TC.",
                "data": {"threshold": _at["internship_min_credits"], "current": snapshot.earned_credits},
            })
        if remaining_non_special > INTERNSHIP_REMAINING_BUFFER:
            blockers.append({
                "kind": "blocker", "category": "threshold", "icon": "pending_actions",
                "title": "Còn quá nhiều TC chưa hoàn thành",
                "detail": f"Còn {remaining_non_special:.1f} TC cần học, ngưỡng cho phép TT: ≤ {INTERNSHIP_REMAINING_BUFFER} TC còn lại.",
                "data": {"remaining": remaining_non_special, "buffer": INTERNSHIP_REMAINING_BUFFER},
            })

    # Có blocker → blocked
    if blockers:
        first = blockers[0]
        legacy_map = {
            "spec": "WRONG_SPECIALIZATION", "elective_quota": "ELECTIVE_QUOTA_FILLED",
            "prereq": "MISSING_PREREQUISITES", "schedule": "THESIS_NEEDS_INTERNSHIP",
            "threshold": "NOT_ELIGIBLE_INTERNSHIP",
        }
        legacy_code = legacy_map.get(first["category"], "BLOCKED")
        extra = {}
        if missing_prereqs:
            extra["missing_prerequisites"] = missing_prereqs
        return _build("blocked", blockers, legacy_code, first["detail"], **extra)

    # ── ĐỦ ĐIỀU KIỆN — phân biệt BB (bắt buộc CTĐT) vs TC (cá nhân hóa được) ──
    # CLAUDE.md §5.9: BB không cá nhân hóa — message phải khác TC.
    is_compulsory_bb_ctx = (
        group_key is None
        and not (thesis_course and code == thesis_course.course_code)
        and not (internship_course and code == internship_course.course_code)
    )
    positives = []

    # BB: hiển thị thông điệp chính rằng đây là môn bắt buộc trong CTĐT
    if is_compulsory_bb_ctx:
        hk = course.typical_semester
        hk_note = f" (HK chuẩn: {hk})" if hk else ""
        positives.append({
            "kind": "info", "category": "required", "icon": "school",
            "title": f"Môn bắt buộc{hk_note}",
            "detail": "Đây là môn bắt buộc trong CTĐT của trường — bạn phải học để đủ TC tốt nghiệp, không phụ thuộc định hướng nghề.",
        })

    try:
        rec_payload = build_recommendations(db, user.id, limit=10)
        rec_list = rec_payload.get("recommendations", [])
        rec_codes = [r["course_code"] for r in rec_list]
        if code in rec_codes:
            rank = rec_codes.index(code) + 1
            positives.append({
                "kind": "positive", "category": "ranking", "icon": "star",
                "title": f"Top #{rank} kỳ tới",
                "detail": f"Hệ thống xếp môn này hạng #{rank}/{len(rec_codes)} cho bạn.",
            })
    except Exception:
        pass

    # Unlock score
    remaining_codes = {c.course_code for c in snapshot.courses if c.course_code not in snapshot.completed_codes}
    unlock = _compute_unlock_scores(remaining_codes, prereq_map).get(code, 0)
    if unlock > 0:
        positives.append({
            "kind": "positive", "category": "unlock", "icon": "lock_open",
            "title": f"Mở khóa {unlock} môn sau",
            "detail": f"Học môn này giúp bạn đủ tiên quyết để học {unlock} môn tiếp theo.",
            "data": {"unlock_count": unlock},
        })

    # Pass rate
    diff_stats = _get_difficulty_stats(db).get(code, {})
    if diff_stats.get("pass_rate") is not None:
        pr = diff_stats["pass_rate"] * 100
        if pr >= 80:
            positives.append({
                "kind": "positive", "category": "performance", "icon": "verified",
                "title": f"Tỉ lệ pass cao ({pr:.0f}%)",
                "detail": f"{pr:.0f}% sinh viên đã học pass môn này — môn được đánh giá vừa sức.",
                "data": {"pass_rate": pr / 100},
            })
        elif pr < 60:
            positives.append({
                "kind": "warning", "category": "performance", "icon": "warning",
                "title": f"Tỉ lệ pass thấp ({pr:.0f}%)",
                "detail": f"Chỉ {pr:.0f}% sinh viên pass — môn khó, cần đầu tư thời gian.",
                "data": {"pass_rate": pr / 100},
            })

    # Prereq performance — nếu SV đã pass tiên quyết với điểm cao
    if prereq_map.get(code):
        prereq_scores = []
        for pc in prereq_map[code]:
            grade = snapshot.best_grades.get(pc)
            if grade and grade.score10 is not None:
                prereq_scores.append(float(grade.score10))
        if prereq_scores:
            avg_prereq = sum(prereq_scores) / len(prereq_scores)
            if avg_prereq >= 7.5:
                positives.append({
                    "kind": "positive", "category": "performance", "icon": "trending_up",
                    "title": f"Tiên quyết điểm cao ({avg_prereq:.1f}/10)",
                    "detail": f"Bạn pass các tiên quyết với điểm TB {avg_prereq:.1f}, sẵn sàng học môn này.",
                })

    # CLAUDE.md §5.9: Track / Skill / Peer-rating chỉ áp dụng cho TC (môn tự chọn).
    # BB là môn bắt buộc — không bị/được "đề xuất" theo định hướng cá nhân.
    if not is_compulsory_bb_ctx:
        # Track match
        if specialization:
            from backend.core.academic_engine import _get_spec_tracks, _build_spec_track_strengths
            spec_tracks = _get_spec_tracks(specialization)
            track_strengths = _build_spec_track_strengths(snapshot, spec_tracks)
            if track_strengths:
                top_track = max(track_strengths.items(), key=lambda x: x[1])[0]
                top_track_courses = spec_tracks.get(top_track, set())
                if code in top_track_courses:
                    positives.append({
                        "kind": "positive", "category": "track", "icon": "psychology",
                        "title": f"Phù hợp định hướng {top_track}",
                        "detail": f"Môn này thuộc định hướng nghề nghiệp mạnh nhất của bạn ({top_track}).",
                        "data": {"track": top_track},
                    })

        # Skill match — feature đã bỏ (cột users.career_skills DROP 2026-05-05)
        user_skills: set[str] = set()
        if user_skills:
            course_skills = db.query(
                models.CourseSkill.skill_code, models.CourseSkill.weight, models.Skill.name,
            ).join(
                models.Skill, models.CourseSkill.skill_code == models.Skill.code
            ).filter(models.CourseSkill.course_code == code).all()
            matched = [(sc, float(w), nm) for sc, w, nm in course_skills if sc in user_skills]
            if matched:
                names = [m[2] for m in matched[:4]]
                positives.append({
                    "kind": "positive", "category": "skill", "icon": "label",
                    "title": f"Dạy {len(matched)} kỹ năng bạn quan tâm",
                    "detail": f"Môn này phát triển: {', '.join(names)}{('...' if len(matched) > 4 else '')}.",
                    "data": {"matched_skills": [m[0] for m in matched]},
                })

        # Peer rating — SV trước đánh giá cao (chỉ TC, không áp dụng cho BB)
        from sqlalchemy import func as _sf2
        rating_agg = db.query(
            _sf2.avg(models.CourseRating.rating).label("avg_r"),
            _sf2.count(models.CourseRating.id).label("count_r"),
        ).filter(models.CourseRating.course_code == code).first()
        if rating_agg and rating_agg.count_r and rating_agg.count_r >= 3:
            avg_r = float(rating_agg.avg_r or 0)
            cnt_r = int(rating_agg.count_r)
            if avg_r >= 4.0:
                positives.append({
                    "kind": "positive", "category": "peer_rating", "icon": "star",
                    "title": f"Đánh giá cao ({avg_r:.1f}/5★)",
                    "detail": f"{cnt_r} sinh viên trước đánh giá môn này — môn được đánh giá tốt.",
                    "data": {"avg_rating": avg_r, "count": cnt_r},
                })
            elif avg_r < 3.0:
                positives.append({
                    "kind": "warning", "category": "peer_rating", "icon": "warning",
                    "title": f"Đánh giá thấp ({avg_r:.1f}/5★)",
                    "detail": f"{cnt_r} sinh viên đánh giá môn dưới TB — cần cân nhắc.",
                    "data": {"avg_rating": avg_r, "count": cnt_r},
                })

    in_top = any(p["category"] == "ranking" for p in positives)
    if is_compulsory_bb_ctx:
        # CLAUDE.md §5.9: BB không có khái niệm "low priority" — luôn phải học.
        verdict = "required"
        legacy_code = "IS_REQUIRED"
        legacy_exp = "Môn bắt buộc trong CTĐT — bạn cần học để đủ TC tốt nghiệp."
    elif in_top:
        verdict = "recommended"
        legacy_code = "IS_RECOMMENDED"
        legacy_exp = positives[0]["detail"]
    else:
        verdict = "eligible_low_priority"
        legacy_code = "LOW_PRIORITY"
        legacy_exp = (
            "Môn này đủ điều kiện nhưng chưa nằm top gợi ý. "
            "Hệ thống ưu tiên môn mở khóa nhiều tiên quyết hơn, phù hợp định hướng, hoặc có tỉ lệ pass cao hơn."
        )
        positives.append({
            "kind": "info", "category": "ranking", "icon": "low_priority",
            "title": "Ngoài top gợi ý",
            "detail": legacy_exp,
        })

    return _build(verdict, positives, legacy_code, legacy_exp)


@app.get("/recommendations/evaluate")
def recommendations_evaluate(
    top_k: int = 5,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """
    Retrospective temporal evaluation (leave-one-term-out):
    For each term t (starting from term 2), simulate what the rule-based engine
    would have recommended using ONLY grades from terms < t, then measure how
    many of the top-K recommendations the student actually enrolled in at term t.

    Returns Precision@K per term and mean Precision@K.
    """
    from collections import defaultdict as _dd
    import re as _re
    from backend.core.academic_engine import (
        _build_snapshot, _course_credits,
        _get_difficulty_stats, _compute_unlock_scores,
        _build_track_strengths, _prereq_performance, _course_tracks,
        _clamp, TRACKS, INTERNSHIP_REMAINING_BUFFER,
    )
    from math import log as _log

    user = _get_user_by_token(authorization, db)
    top_k = max(1, min(top_k, 10))

    all_grades = db.query(models.UserGrade).filter(
        models.UserGrade.user_id == user.id
    ).all()

    # Group grades by term
    term_courses: dict[str, list[str]] = _dd(list)   # term → passed course codes
    for g in all_grades:
        if g.term:
            term_courses[g.term].append(g.course_code)

    def _sort_term_key(t: str):
        nums = _re.findall(r"\d+", t)
        year = next((int(n) for n in nums if 2000 < int(n) < 2100), 9999)
        hk  = next((int(n) for n in nums if 1 <= int(n) <= 4), 0)
        return (year, hk)

    sorted_terms = sorted(term_courses.keys(), key=_sort_term_key)
    if len(sorted_terms) < 2:
        return {
            "message": "Cần ít nhất 2 học kỳ dữ liệu để đánh giá.",
            "terms_evaluated": 0,
            "mean_precision": None,
        }

    prereqs = db.query(models.CoursePrerequisite).all()
    prereq_map: dict[str, list[str]] = _dd(list)
    for p in prereqs:
        prereq_map[p.course_code].append(p.prerequisite_code)

    all_difficulty_stats = _get_difficulty_stats(db)
    term_results = []
    all_precisions = []

    for i, term in enumerate(sorted_terms[1:], 1):
        prior_terms = set(sorted_terms[:i])

        # Build past state: passed codes from all prior terms
        prior_passed = {
            g.course_code for g in all_grades
            if g.term in prior_terms and g.passed
        }
        # Build past best_grades map
        prior_grades: dict[str, models.UserGrade] = {}
        for g in all_grades:
            if g.term in prior_terms:
                existing = prior_grades.get(g.course_code)
                if existing is None or (g.score10 or 0) > (existing.score10 or 0):
                    prior_grades[g.course_code] = g

        try:
            # Full snapshot (current), then override to past state
            snapshot = _build_snapshot(db, user.id, specialization=user.specialization)
            snapshot.completed_codes = prior_passed
            snapshot.best_grades = prior_grades
            snapshot.earned_credits = sum(
                _course_credits(snapshot.course_by_code[c])
                for c in prior_passed if c in snapshot.course_by_code
            )
            # Estimate past GPA from prior grades
            past_scores = [
                float(g.score10) for g in prior_grades.values()
                if g.score10 is not None and g.passed
            ]
            snapshot.avg_score10 = (sum(past_scores) / len(past_scores)) if past_scores else None

            remaining = [
                c for c in snapshot.courses
                if c.course_code not in prior_passed
            ]
            remaining_codes_set = {c.course_code for c in remaining}
            unlock_scores = _compute_unlock_scores(remaining_codes_set, prereq_map)
            track_strengths = _build_track_strengths(snapshot)
            preferred_track = max(track_strengths.items(), key=lambda x: x[1])[0] if track_strengths else None
            baseline = snapshot.avg_score10 if snapshot.avg_score10 else 6.5

            scored: list[tuple[float, str]] = []
            for c in remaining:
                code = c.course_code
                # Skip if prereqs not met
                if any(p not in prior_passed for p in prereq_map.get(code, [])):
                    continue

                score = 50.0
                # Required bonus
                score += 20.0

                # Difficulty fit
                dstat = all_difficulty_stats.get(code, {})
                avg10 = dstat.get("avg_score10")
                pr    = dstat.get("pass_rate")
                if avg10 is not None and pr is not None:
                    gap = baseline - avg10
                    if pr < 0.6:
                        score += _clamp(8.0 + gap * 2.0, 0.0, 20.0)
                    elif pr >= 0.85:
                        score += 14.0 if baseline < 6.5 else 7.0
                    else:
                        score += _clamp(10.0 + gap * 1.0, 0.0, 20.0)

                # Prereq performance penalty
                _, warn = _prereq_performance(code, prereq_map, prior_grades)
                if warn:
                    score -= 5.0

                # Track alignment
                candidate_tracks = _course_tracks(c.course_name)
                if preferred_track and preferred_track in candidate_tracks:
                    score += 14.0
                elif candidate_tracks:
                    score += 6.0

                # Unlock bonus
                ul = unlock_scores.get(code, 0)
                if ul > 0:
                    score += min(16.0, 4.0 + 4.0 * _log(1 + ul))

                scored.append((score, code))

            scored.sort(key=lambda x: -x[0])
            top_recommended = [code for _, code in scored[:top_k]]

            # Actual courses taken in this term (enrolled, regardless of pass/fail)
            actual = set(term_courses[term])
            # Only count courses that were actually eligible (prereqs met)
            eligible_actual = {
                c for c in actual
                if not any(p not in prior_passed for p in prereq_map.get(c, []))
                and c in {x.course_code for x in snapshot.courses}
                and c not in prior_passed
            }

            hits = len(eligible_actual & set(top_recommended))
            denom = min(top_k, len(eligible_actual)) if eligible_actual else top_k
            precision = hits / denom if denom > 0 else 0.0
            all_precisions.append(precision)

            term_results.append({
                "term": term,
                "recommended_top_k": top_recommended,
                "actual_eligible": sorted(eligible_actual),
                "hits": hits,
                f"precision@{top_k}": round(precision, 3),
            })
        except Exception:
            continue

    mean_p = round(sum(all_precisions) / len(all_precisions), 3) if all_precisions else None
    return {
        "top_k": top_k,
        "terms_evaluated": len(term_results),
        f"mean_precision@{top_k}": mean_p,
        "per_term": term_results,
    }


@app.post("/grades/preview", response_model=schemas.GradePreviewOut)
def preview_grades(
    file: UploadFile,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Parse grades file and return preview — does NOT write to the database."""
    _get_user_by_token(authorization, db)

    data = file.file.read(_MAX_UPLOAD_BYTES + 1)
    if len(data) > _MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File quá lớn. Giới hạn tối đa 10 MB.")
    filename = file.filename or "upload.xlsx"

    try:
        _parsed = read_rows_from_upload(filename, data)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception:
        raise HTTPException(status_code=400, detail="Không thể đọc file. Kiểm tra lại định dạng (xlsx/csv/pdf).")
    rows = _parsed["rows"]

    grade_records = extract_grades(rows)
    tich_luy = extract_tich_luy(rows)
    all_courses = {c.course_code: c for c in db.query(models.Course).all()}

    matched: list[schemas.GradePreviewItem] = []
    unmatched: list[schemas.GradeUploadIssue] = []

    for rec in grade_records:
        code = rec["course_code"]
        course = all_courses.get(code)
        if course:
            matched.append(schemas.GradePreviewItem(
                course_code=code,
                course_name_in_file=rec.get("course_name"),
                course_name_in_db=course.course_name,
                credits=float(course.credits) if course.credits is not None else None,
                score10=rec.get("score10"),
                score4=rec.get("score4"),
                letter=rec.get("letter"),
                passed=rec.get("passed", False),
                term=rec.get("term"),
                matched=True,
            ))
        else:
            unmatched.append(schemas.GradeUploadIssue(
                course_code_in_file=code,
                course_name_in_file=rec.get("course_name"),
                reason="Mã môn không có trong CTĐT",
            ))

    student_code_out = _parsed.get("student_code") or None

    # For admin callers: enrich with account existence + spec detection
    account_created: bool | None = None
    old_spec: str | None = None
    new_spec: str | None = None
    spec_changed: bool | None = None
    calling_user = _get_user_by_token(authorization, db)
    if calling_user.role in ("admin", "advisor") and student_code_out:
        existing = db.query(models.User).filter(models.User.username == student_code_out).first()
        account_created = existing is None
        if existing:
            old_spec = existing.specialization
        passed_codes = {
            r["course_code"] for r in grade_records
            if r.get("passed") and r["course_code"] in all_courses
        }
        if passed_codes:
            new_spec = _detect_specialization(passed_codes, db)
        if new_spec and new_spec != old_spec:
            spec_changed = True
        elif new_spec is not None:
            spec_changed = False

    return schemas.GradePreviewOut(
        matched=matched,
        unmatched=unmatched,
        tich_luy=tich_luy,
        total_rows_in_file=len(grade_records),
        student_code=student_code_out,
        account_created=account_created,
        old_specialization=old_spec,
        new_specialization=new_spec,
        specialization_changed=spec_changed,
    )


# ── Profile ────────────────────────────────────────────────────────────────────

@app.get("/profile/me", response_model=schemas.UserProfileOut)
def get_profile(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Lấy profile preferences SV — sau refactor 2026-05-05 chỉ còn target_gpa."""
    user = _get_user_by_token(authorization, db)
    plan = db.query(models.StudyPlan).filter(models.StudyPlan.user_id == user.id).order_by(models.StudyPlan.id.desc()).first()
    return schemas.UserProfileOut(
        target_gpa=float(plan.target_gpa) if plan and plan.target_gpa is not None else None,
    )


@app.patch("/profile/me", response_model=schemas.UserProfileOut)
def update_profile(
    payload: schemas.UserProfileIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    if payload.target_gpa is not None:
        plan = db.query(models.StudyPlan).filter(models.StudyPlan.user_id == user.id).order_by(models.StudyPlan.id.desc()).first()
        if plan:
            plan.target_gpa = payload.target_gpa
        else:
            plan = models.StudyPlan(user_id=user.id, plan_name="default", target_gpa=payload.target_gpa)
            db.add(plan)
    db.commit()
    db.refresh(user)
    invalidate_student_context_cache(user.id)
    plan = db.query(models.StudyPlan).filter(models.StudyPlan.user_id == user.id).order_by(models.StudyPlan.id.desc()).first()
    return schemas.UserProfileOut(
        target_gpa=float(plan.target_gpa) if plan and plan.target_gpa is not None else None,
    )


def _normalize_term(raw: str) -> str:
    """Convert any term string to canonical 'HK{n}/{startYear}' format.

    Examples:
      "HK1 2025-2026"     → "HK1/2025"
      "Học kỳ 2 2024-2025"→ "HK2/2024"
      "HK1/2025"          → "HK1/2025"  (already canonical)
      "hè 2024"           → "HKhe/2024"
    """
    import re as _re
    if not raw:
        return raw
    raw = raw.strip()
    # Already canonical
    if _re.match(r'^HK\w+/\d{4}$', raw):
        return raw
    nums = [int(n) for n in _re.findall(r'\d+', raw)]
    years = [n for n in nums if 2000 < n < 2100]
    if not years:
        return raw  # cannot normalise — keep as-is
    # Two consecutive years (e.g. 2025 and 2026) → start year is the lower
    if len(years) == 2 and abs(years[1] - years[0]) == 1:
        start_year = min(years)
    else:
        start_year = years[0]
    raw_lower = raw.lower()
    if any(w in raw_lower for w in ('hè', 'he', 'summer', 'hk3', 'hk 3')):
        sem = 'he'
    else:
        sem_nums = [n for n in nums if 1 <= n <= 2 and n not in years]
        sem = str(sem_nums[0]) if sem_nums else '1'
    return f"HK{sem}/{start_year}"


def _get_active_semester(db: Session) -> str | None:
    cfg = db.query(models.SystemConfig).filter(models.SystemConfig.key == "active_semester").first()
    return cfg.value if cfg else None



def _compute_overdue_student_ids(db: Session, graduation_threshold: float) -> set[int]:
    """SV quá hạn: span năm học từ term điểm > 5 năm AND TC tích lũy < graduation_threshold.
    Đo theo dữ liệu điểm thực tế (KHÔNG dùng cohort của mã SV) — mã SV có thể là test data
    với prefix cũ mà SV mới chỉ học 1 kỳ. Helper này được dùng chung bởi /admin/users
    (gắn flag is_overdue) và /admin/dashboard/stats (đếm số lượng cảnh báo)."""
    import re as _re
    _year_pattern = _re.compile(r"Năm học (\d{4})")
    sv_year_spans: dict[int, set[int]] = {}
    for user_id, term in db.query(models.UserGrade.user_id, models.UserGrade.term).filter(
        models.UserGrade.term != None  # noqa: E711
    ).distinct():
        if not term:
            continue
        m = _year_pattern.search(term)
        if not m:
            continue
        sv_year_spans.setdefault(user_id, set()).add(int(m.group(1)))

    candidate_ids = [uid for uid, years in sv_year_spans.items()
                     if years and (max(years) - min(years) + 1) > 5]
    if not candidate_ids:
        return set()

    # Tính credits từ user_grades (passed) thay vì official_earned_credits đã drop
    overdue: set[int] = set()
    course_credits = {c.course_code: float(c.credits or 0) for c in db.query(models.Course).all()}
    for uid in candidate_ids:
        passed = db.query(models.UserGrade.course_code).filter(
            models.UserGrade.user_id == uid,
            models.UserGrade.passed == True,
        ).all()
        tc = sum(course_credits.get(code, 0) for (code,) in passed)
        if tc < graduation_threshold:
            overdue.add(uid)
    return overdue


@app.get("/admin/users", response_model=schemas.AdminUserListOut)
def admin_list_users(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    _require_admin(authorization, db)
    from sqlalchemy import func as sqlfunc
    users = db.query(models.User).order_by(models.User.id).all()
    graduation_threshold = academic_engine._get_graduation_threshold(db)
    # Refactor 2026-05-05: bỏ tính grade_count, gpa, overdue (admin không quản lý điểm)

    # Batch: advisor assignment per student → advisor user
    assignments = db.query(models.AdvisorAssignment).all()
    assn_map = {a.student_id: a.advisor_id for a in assignments}
    advisor_ids = set(assn_map.values())
    advisor_map = {
        u.id: u
        for u in db.query(models.User).filter(models.User.id.in_(advisor_ids)).all()
    } if advisor_ids else {}

    # Batch: class_group_id → code (1 query thay vì N)
    class_group_ids = {u.class_group_id for u in users if getattr(u, "class_group_id", None)}
    class_code_map: dict[int, str] = {}
    if class_group_ids:
        for cg in db.query(models.ClassGroup).filter(models.ClassGroup.id.in_(class_group_ids)).all():
            class_code_map[cg.id] = cg.code

    items = []
    for u in users:
        adv_id = assn_map.get(u.id) if u.role == "student" else None
        adv = advisor_map.get(adv_id) if adv_id else None
        cg_id = getattr(u, "class_group_id", None)
        items.append(schemas.AdminUserItem(
            id=u.id,
            username=u.username,
            full_name=u.full_name,
            role=u.role,
            specialization=u.specialization,
            cohort=u.cohort,
            class_group_id=cg_id,
            class_group_code=class_code_map.get(cg_id) if cg_id else None,
            career_goal=None,  # Deprecated 2026-05-05
            official_earned_credits=None,  # Deprecated 2026-05-05
            grade_count=0,                  # Deprecated 2026-05-05 — admin không xem điểm
            avg_score4=None,                # Deprecated 2026-05-05
            default_password=u.default_password,
            is_overdue=False,               # Deprecated 2026-05-05
            advisor_id=adv_id,
            advisor_teacher_code=adv.teacher_code if adv else None,
            advisor_full_name=adv.full_name if adv else None,
        ))
    return schemas.AdminUserListOut(total=len(items), users=items)


# REMOVED 2026-05-05: GET /admin/users/{id}/grades đã bị bỏ.
# Mô hình mới: EduGuide là tool nội bộ — admin chỉ quản lý tài khoản + thông báo.
# Bảng điểm là dữ liệu cá nhân của SV, admin không cần xem.


@app.patch("/admin/users/{user_id}/role", response_model=schemas.MessageOut)
def admin_set_user_role(
    user_id: int,
    payload: schemas.AdminSetRoleIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    u = db.query(models.User).filter(models.User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Không tìm thấy người dùng")
    old_role = u.role
    u.role = payload.role
    db.commit()
    _log(db, admin, "CHANGE_ROLE", "user", u.username, f"{old_role} → {payload.role}")
    return schemas.MessageOut(message=f"Đã cập nhật vai trò {u.username} → {payload.role}")


@app.patch("/admin/users/{user_id}", response_model=schemas.UserOut)
def admin_update_user(
    user_id: int,
    payload: schemas.AdminUpdateUserIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Cập nhật thông tin SV/Advisor: full_name, cohort, specialization, email.

    Partial update — field None bỏ qua. specialization=='' để clear (chưa CN).
    Khi spec thay đổi cho student → reassign advisor.
    """
    admin = _require_admin(authorization, db)
    u = db.query(models.User).filter(models.User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Không tìm thấy người dùng")

    changes: list[str] = []
    if payload.full_name is not None:
        new_name = normalize_vietnamese_name(payload.full_name) or None
        if not new_name:
            raise HTTPException(status_code=422, detail="Họ tên không được rỗng")
        if new_name != u.full_name:
            changes.append(f"full_name: '{u.full_name}' → '{new_name}'")
            u.full_name = new_name

    if payload.cohort is not None:
        new_cohort = (payload.cohort or "").strip() or None
        if new_cohort != u.cohort:
            changes.append(f"cohort: '{u.cohort}' → '{new_cohort}'")
            u.cohort = new_cohort

    spec_changed = False
    if payload.specialization is not None:
        # "" hoặc None → clear; mã spec hợp lệ → set
        new_spec = (payload.specialization or "").strip() or None
        if new_spec is not None and new_spec not in _IMPORT_VALID_SPECIALIZATIONS:
            raise HTTPException(
                status_code=422,
                detail=f"Mã CN '{new_spec}' không hợp lệ. Phải là 1 trong: {', '.join(sorted(_IMPORT_VALID_SPECIALIZATIONS))}"
            )
        if new_spec != u.specialization:
            changes.append(f"specialization: '{u.specialization}' → '{new_spec}'")
            u.specialization = new_spec
            spec_changed = True

    # Email column REMOVED 2026-05-05 — không còn email cá nhân. SV đổi pw qua Settings.

    # Đổi lớp (chỉ áp dụng cho student) → trigger reassign advisor
    class_changed = False
    if getattr(payload, "class_code", None) is not None:
        new_class_code = (payload.class_code or "").strip().upper() or None
        old_class_id = getattr(u, "class_group_id", None)
        if u.role != "student":
            raise HTTPException(status_code=400, detail="Chỉ sinh viên mới có lớp")
        if new_class_code:
            cg = db.query(models.ClassGroup).filter(
                models.ClassGroup.code == new_class_code
            ).first()
            if not cg:
                raise HTTPException(
                    status_code=400,
                    detail=f"Lớp '{new_class_code}' chưa tồn tại"
                )
            if cg.id != old_class_id:
                changes.append(f"class: {old_class_id} → {cg.id} ({new_class_code})")
                u.class_group_id = cg.id
                u.specialization = cg.specialization  # tự cập nhật spec
                u.cohort = cg.cohort
                spec_changed = True
                class_changed = True
        else:
            # Clear class
            if old_class_id is not None:
                changes.append(f"class: {old_class_id} → null")
                u.class_group_id = None
                class_changed = True

    if not changes:
        return _to_user_out(u)  # nothing to update

    db.commit()
    db.refresh(u)

    # Nếu là student và spec/class thay đổi → reassign advisor
    if u.role == "student" and (spec_changed or class_changed):
        try:
            assign_advisor_for_student(db, u.id, u.specialization)
            db.commit()
        except Exception:
            pass

    _log(db, admin, "UPDATE_USER", "user", u.username, "; ".join(changes))
    return _to_user_out(u)


@app.post("/admin/users/{user_id}/recover-account", response_model=schemas.AdminResetPasswordOut)
def admin_recover_account(
    user_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Khôi phục tài khoản SV bị mất Gmail + quên mật khẩu.

    Flow offline: SV xác minh identity với phòng đào tạo (CCCD, gặp mặt) → admin
    gọi endpoint này. Reset 3 thứ atomic:
      1. Mật khẩu mới (6 số ngẫu nhiên) — admin đưa cho SV trực tiếp
      2. Xoá email cũ + google_sub → SV phải link Gmail mới khi đăng nhập lại
      3. Set is_first_login=TRUE → buộc SV setup lại tài khoản

    Trả về password mới để admin copy → đưa cho SV.
    """
    import random as _rnd
    admin = _require_admin(authorization, db)
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Không tìm thấy người dùng")
    if user.role != "student":
        raise HTTPException(
            status_code=400,
            detail="Khôi phục tài khoản chỉ áp dụng cho sinh viên (advisor/admin dùng reset-password thường)."
        )

    new_pw = str(_rnd.randint(100000, 999999))
    user.password_hash = _hash_temp_password(new_pw)
    user.default_password = new_pw  # admin xem được trong UI
    user.is_first_login = True
    db.commit()

    _log(db, admin, "RECOVER_ACCOUNT", "user", user.username,
         "Reset password + force first-login")

    return schemas.AdminResetPasswordOut(
        password=new_pw,
        message=f"Đã khôi phục {user.username}. Đưa mật khẩu mới cho SV."
    )


@app.post("/admin/users/{user_id}/reset-password", response_model=schemas.AdminResetPasswordOut)
def admin_reset_user_password(
    user_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    import random as _rnd
    admin = _require_admin(authorization, db)
    u = db.query(models.User).filter(models.User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Không tìm thấy người dùng")
    new_pw = str(_rnd.randint(100000, 999999))
    u.password_hash = _hash_temp_password(new_pw)
    u.default_password = new_pw
    u.is_first_login = True
    db.commit()
    _clear_failed_login(u.username, db)  # xóa rate limit để user login ngay được
    _log(db, admin, "RESET_PASSWORD", "user", u.username, "admin reset")
    return schemas.AdminResetPasswordOut(
        message=f"Đã reset mật khẩu cho {u.username}",
        password=new_pw,
    )


@app.get("/admin/users/{user_id}/default-password", response_model=schemas.AdminDefaultPasswordOut)
def admin_get_default_password(
    user_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    _require_admin(authorization, db)
    u = db.query(models.User).filter(models.User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Không tìm thấy người dùng")
    return schemas.AdminDefaultPasswordOut(
        has_default=bool(u.default_password),
        password=u.default_password,
    )


@app.get("/admin/passwords-pending")
def admin_passwords_pending(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Users whose default_password is still set (admin-created / reset but not yet changed)."""
    _require_admin(authorization, db)
    users = (
        db.query(models.User)
        .filter(models.User.default_password != None)
        .order_by(models.User.id.desc())
        .all()
    )
    return [
        {
            "id": u.id,
            "username": u.username,
            "full_name": u.full_name,
            "role": u.role,
        }
        for u in users
    ]


@app.post("/admin/users/{user_id}/view-password")
def admin_view_user_password(
    user_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Return + log admin viewing a user's default password."""
    admin = _require_admin(authorization, db)
    u = db.query(models.User).filter(models.User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Không tìm thấy người dùng")
    if not u.default_password:
        raise HTTPException(status_code=404, detail="User đã đổi mật khẩu, không còn mật khẩu mặc định")
    _log(db, admin, "VIEW_PASSWORD", "user", u.username, f"admin xem mật khẩu mặc định của {u.username}")
    return {"password": u.default_password}


@app.delete("/admin/users/{user_id}", response_model=schemas.MessageOut)
def admin_delete_user(
    user_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    if admin.id == user_id:
        raise HTTPException(status_code=400, detail="Không thể xóa chính mình")
    u = db.query(models.User).filter(models.User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404, detail="Không tìm thấy người dùng")
    username = u.username
    db.delete(u)
    db.commit()
    _log(db, admin, "DELETE_USER", "user", username, f"role={u.role}")
    return schemas.MessageOut(message=f"Đã xóa tài khoản {username}")


# ── Admin bulk import users ───────────────────────────────────────────────────

_IMPORT_VALID_SPECIALIZATIONS = {
    # Canonical spec codes (sync với DB users.specialization + frontend dropdown values)
    "7480201_07",  # KHMT
    "7480201_06",  # MMT
    "7480201_05",  # CNPM
    "7480201_09",  # HTTT
    "7480201_04",  # THKT
    "7480201_08",  # CNTTDH
}


def _parse_users_csv(data: bytes, filename: str) -> list[dict]:
    """Parse CSV/Excel file thành list of dicts.

    File template MỚI 4 cột bắt buộc + 2 cột optional:
      | MSSV | Họ tên | Email | Mã lớp | (Mật khẩu, Khoá) |

    Mã lớp BẮT BUỘC — backend tự derive specialization + cohort từ class_groups.code.
    Chuyên ngành/Khoá là legacy fields giữ lại để backward-compat (sẽ override
    bằng giá trị từ class_groups khi gán SV vào lớp).
    """
    import io
    rows = read_rows_from_upload(filename, data)["rows"]
    if not rows:
        raise ValueError("File rỗng hoặc không đọc được")

    # Tìm header row: dòng đầu tiên có ít nhất cột 'username' hoặc 'mssv'
    header_idx = None
    headers: list[str] = []
    for i, row in enumerate(rows[:10]):
        normalized = [str(c or "").strip().lower() for c in row]
        if any(k in normalized for k in ("username", "mssv", "mã sv", "masv")):
            header_idx = i
            headers = normalized
            break

    if header_idx is None:
        # Fallback: assume first row is header
        header_idx = 0
        headers = [str(c or "").strip().lower() for c in rows[0]]

    def _col(*keys):
        for k in keys:
            for i, h in enumerate(headers):
                if k in h:
                    return i
        return None

    idx_user   = _col("username", "mssv", "mã sv", "masv", "tài khoản", "tai khoan")
    idx_pass   = _col("password", "mật khẩu", "mat khau")
    idx_name   = _col("full_name", "họ tên", "ho ten", "tên", "ten", "họ và tên")
    idx_email  = _col("email", "email")
    idx_class  = _col("mã lớp", "ma lop", "class_code", "lớp", "lop")
    idx_spec   = _col("specialization", "chuyên ngành", "chuyen nganh")
    idx_cohort = _col("cohort", "khoá", "khoa hoc", "khóa")

    if idx_user is None:
        raise ValueError("Không tìm thấy cột username/MSSV trong file")

    result = []
    for row in rows[header_idx + 1:]:
        if not row or all(_is_missing(c) for c in row):
            continue
        username = str(row[idx_user] or "").strip() if idx_user < len(row) else ""
        if not username:
            continue
        result.append({
            "username": username,
            "password": str(row[idx_pass] or "").strip() if idx_pass is not None and idx_pass < len(row) else "",
            "full_name": str(row[idx_name] or "").strip() if idx_name is not None and idx_name < len(row) else "",
            "email": str(row[idx_email] or "").strip().lower() if idx_email is not None and idx_email < len(row) else "",
            "class_code": str(row[idx_class] or "").strip().upper() if idx_class is not None and idx_class < len(row) else "",
            "specialization": str(row[idx_spec] or "").strip() if idx_spec is not None and idx_spec < len(row) else "",
            "cohort": str(row[idx_cohort] or "").strip() if idx_cohort is not None and idx_cohort < len(row) else "",
        })
    return result


_MSSV_REGEX = r"^sv\d{5}$"  # 1 chuẩn duy nhất: sv + 2 số khoá + 3 số STT (vd sv14001)


def _validate_mssv_format(username: str) -> str:
    """Chuẩn hoá + validate MSSV. Format chuẩn: sv14001 (lowercase).

    Raise HTTPException 422 nếu sai format.
    """
    import re as _re
    u = (username or "").strip().lower()
    if not _re.match(_MSSV_REGEX, u):
        raise HTTPException(
            status_code=422,
            detail=(
                f"MSSV '{username}' không đúng format chuẩn 'sv14001' "
                "(sv + 2 số khoá + 3 số STT). VD: sv14001, sv14002, sv25001."
            ),
        )
    return u


def _extract_cohort_from_username(username: str | None) -> str | None:
    """Đoán cohort từ MSSV. Format chuẩn: sv14001 → '14'.

    Vẫn nhận diện format legacy (SV140001, 1400000010) cho backfill data cũ —
    nhưng không tạo SV mới với format đó nữa (validate ở `_validate_mssv_format`).
    """
    import re as _re
    if not username:
        return None
    u = str(username).strip()
    m = _re.match(r"^sv(\d{2})\d{3}$", u, _re.IGNORECASE)
    if m:
        return m.group(1)
    m = _re.match(r"^SV(\d{2})\d{4}$", u, _re.IGNORECASE)
    if m:
        return m.group(1)
    m = _re.match(r"^(\d{2})\d{8}$", u)
    if m:
        return m.group(1)
    return None


@app.post("/admin/users", response_model=schemas.AdminCreateUserOut)
def admin_create_user(
    payload: schemas.AdminCreateUserIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Tạo một tài khoản sinh viên thủ công."""
    admin = _require_admin(authorization, db)
    role = payload.role or "student"
    # Validate MSSV format chuẩn (sv14001) cho student. Advisor/admin format khác.
    username = payload.username
    if role == "student":
        username = _validate_mssv_format(username)
    existing = db.query(models.User).filter(models.User.username == username).first()
    if existing:
        raise HTTPException(status_code=409, detail=f"Tài khoản '{username}' đã tồn tại.")

    # ── Resolve class_group nếu có (chỉ áp dụng cho student) ──────────────
    class_group: models.ClassGroup | None = None
    derived_spec: str | None = None
    derived_cohort: str | None = None
    class_code = (getattr(payload, "class_code", None) or "").strip().upper()
    if class_code and role == "student":
        class_group = db.query(models.ClassGroup).filter(
            models.ClassGroup.code == class_code
        ).first()
        if not class_group:
            raise HTTPException(
                status_code=400,
                detail=f"Lớp '{class_code}' chưa tồn tại — tạo lớp trước khi gán SV"
            )
        derived_spec = class_group.specialization
        derived_cohort = class_group.cohort

    # Fallback: derive cohort từ MSSV nếu không có class_group
    if role == "student" and not derived_cohort:
        derived_cohort = _extract_cohort_from_username(username)

    import random as _random
    password_plain = str(_random.randint(100000, 999999))
    hashed = _hash_temp_password(password_plain)

    user = models.User(
        username=username,
        password_hash=hashed,
        full_name=normalize_vietnamese_name(payload.full_name) or payload.username,
        role=role,
        cohort=derived_cohort,
        specialization=derived_spec,
        email=getattr(payload, "email", None),
        class_group_id=class_group.id if class_group else None,
        is_first_login=True,
        default_password=password_plain,
    )
    db.add(user)
    db.flush()
    # Auto-assign advisor (qua class_group nếu có, fallback round-robin)
    if role == "student":
        try:
            assign_advisor_for_student(db, user.id, derived_spec)
        except Exception:
            pass
    db.commit()
    db.refresh(user)
    _log(db, admin, "CREATE_USER", "user", payload.username,
         f"role={user.role} cohort={derived_cohort} class={class_code or '—'}")
    return schemas.AdminCreateUserOut(
        id=user.id,
        username=user.username,
        full_name=user.full_name,
        role=user.role,
        email=getattr(user, "email", None),
        class_group_id=getattr(user, "class_group_id", None),
        class_group_code=class_group.code if class_group else None,
        password_plain=password_plain,
    )


@app.post("/admin/users/import", response_model=schemas.AdminUsersImportOut)
def admin_import_users(
    file: UploadFile,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    import traceback as _tb
    admin = _require_admin(authorization, db)

    data = file.file.read(_MAX_UPLOAD_BYTES + 1)
    if len(data) > _MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File quá lớn. Giới hạn 10 MB.")

    try:
        records = _parse_users_csv(data, file.filename or "import.xlsx")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        # Log full trace để debug — không nuốt
        print(f"[admin_import_users] Parse error: {_tb.format_exc()}", flush=True)
        raise HTTPException(status_code=400, detail=f"Không đọc được file: {exc}")

    if not records:
        raise HTTPException(status_code=400, detail="Không tìm thấy dữ liệu sinh viên trong file")

    # Approach C: UPSERT — tạo mới hoặc cập nhật. Re-upload roster sau khi đổi
    # lớp sẽ tự derive lại specialization + advisor.
    existing_users = {u.username: u for u in db.query(models.User).all()}
    # Pre-load class_groups (code → ClassGroup) để tránh N queries
    classes_by_code = {c.code: c for c in db.query(models.ClassGroup).all()}

    created_count = 0
    updated_count = 0
    spec_changed_count = 0
    errors: list[schemas.UserImportError] = []
    generated_passwords: dict[str, str] = {}

    for row_idx, rec in enumerate(records, start=2):  # row 1 = header
        username = rec["username"]
        try:
            # ── Resolve class_group (priority: file column "Mã lớp")
            class_code = (rec.get("class_code") or "").strip().upper()
            class_group: models.ClassGroup | None = None
            if class_code:
                class_group = classes_by_code.get(class_code)
                if not class_group:
                    errors.append(schemas.UserImportError(
                        row=row_idx, username=username,
                        reason=f"Lớp '{class_code}' chưa tồn tại — tạo lớp trước khi import SV"
                    ))
                    continue

            # Derive spec + cohort: ưu tiên class_group, fallback CSV cohort/spec
            if class_group:
                new_spec = class_group.specialization
                new_cohort = class_group.cohort
            else:
                new_spec = rec["specialization"] if rec["specialization"] in _IMPORT_VALID_SPECIALIZATIONS else None
                new_cohort = (rec.get("cohort") or "").strip() or None
                if not new_cohort:
                    new_cohort = _extract_cohort_from_username(username)

            new_name = normalize_vietnamese_name(rec["full_name"]) or None
            # email field removed 2026-05-05 (Phase 6)

            existing = existing_users.get(username)
            if existing:
                # UPDATE — không đụng password. Cập nhật fields nếu CSV có giá trị.
                old_spec = existing.specialization
                if new_spec is not None and new_spec != existing.specialization:
                    existing.specialization = new_spec
                if new_cohort and new_cohort != existing.cohort:
                    existing.cohort = new_cohort
                if new_name and not existing.full_name:
                    existing.full_name = new_name
                # email field removed 2026-05-05
                # Đổi lớp → trigger reassign advisor qua sync
                old_class_id = getattr(existing, "class_group_id", None)
                if class_group and old_class_id != class_group.id:
                    existing.class_group_id = class_group.id
                    db.flush()
                    try:
                        assign_advisor_for_student(db, existing.id, existing.specialization)
                    except Exception:
                        pass
                # Reassign advisor nếu CN thay đổi (vd không qua lớp)
                elif existing.specialization != old_spec:
                    spec_changed_count += 1
                    try:
                        assign_advisor_for_student(db, existing.id, existing.specialization)
                    except Exception:
                        pass
                updated_count += 1
                continue

            # Validate MSSV format chuẩn: sv14001 (sv + 2 số khoá + 3 số STT)
            import re as _re
            normalized_username = username.strip().lower()
            if not _re.match(_MSSV_REGEX, normalized_username):
                errors.append(schemas.UserImportError(
                    row=row_idx, username=username,
                    reason="Sai format. Phải là sv14001 (sv + 2 số khoá + 3 số STT). VD: sv14001, sv25010."
                ))
                continue
            username = normalized_username  # dùng lowercase cho insert
            # Re-derive cohort từ username chuẩn hoá (đảm bảo nhất quán)
            if not new_cohort:
                new_cohort = _extract_cohort_from_username(username)

            import random as _rnd
            plain_pw = rec["password"] or str(_rnd.randint(100000, 999999))
            if not rec["password"]:
                generated_passwords[username] = plain_pw

            pw_hash = _hash_temp_password(plain_pw) if not rec["password"] else _hash_password(plain_pw)

            new_user = models.User(
                username=username,
                password_hash=pw_hash,
                full_name=new_name,
                role="student",
                specialization=new_spec,
                cohort=new_cohort,
                class_group_id=class_group.id if class_group else None,
                default_password=plain_pw if not rec["password"] else None,
            )
            db.add(new_user)
            db.flush()  # get new_user.id
            # Auto-assign advisor cho SV mới có Lớp HOẶC CN
            if class_group or new_user.specialization:
                try:
                    assign_advisor_for_student(db, new_user.id, new_user.specialization)
                except Exception:
                    pass
            existing_users[username] = new_user
            created_count += 1
        except Exception as exc:
            # Lỗi từng dòng — rollback partial state, log lỗi, tiếp tục row khác.
            print(f"[admin_import_users] Row {row_idx} ({username}) failed: {_tb.format_exc()}", flush=True)
            db.rollback()
            errors.append(schemas.UserImportError(
                row=row_idx, username=username,
                reason=f"Lỗi DB: {type(exc).__name__}: {str(exc)[:200]}"
            ))
            # Reload existing_users + classes sau rollback để row sau dùng state đúng
            try:
                existing_users = {u.username: u for u in db.query(models.User).all()}
                classes_by_code = {c.code: c for c in db.query(models.ClassGroup).all()}
            except Exception:
                existing_users = {}
                classes_by_code = {}
            continue

    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Lỗi lưu dữ liệu: {exc}")

    _log(db, admin, "BULK_IMPORT_USERS", "users", None,
         f"created={created_count} updated={updated_count} spec_changed={spec_changed_count} errors={len(errors)}")

    return schemas.AdminUsersImportOut(
        created_count=created_count,
        skipped_count=updated_count,  # giữ tên field cũ; nay dùng cho "updated"
        errors=errors,
        generated_passwords=generated_passwords,
    )


# ── Admin bulk import grades ──────────────────────────────────────────────────
# REMOVED 2026-05-05: endpoint POST /admin/grades/import đã bị bỏ.
# Mô hình mới: EduGuide là tool cá nhân — SV tự upload bảng điểm qua POST
# /grades/upload. Admin không quản lý điểm chính thức (đã có SIS riêng).
# → Bỏ luôn logic merge admin/self, drop column user_grades.source.


# REMOVED 2026-05-05: GET /admin/reports/graduation đã bị bỏ.
# Mô hình mới: EduGuide là tool nội bộ — admin không quản lý điểm/tốt nghiệp.
# Báo cáo tốt nghiệp chính thức xem ở SIS của trường.


# ── Admin Audit Log ───────────────────────────────────────────────────────────

@app.get("/admin/logs", response_model=list[schemas.AdminLogOut])
def admin_get_logs(
    limit: int = 100,
    offset: int = 0,
    action: str | None = None,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    _require_admin(authorization, db)
    q = db.query(models.AdminLog)
    if action:
        q = q.filter(models.AdminLog.action == action.upper())
    return (
        q.order_by(models.AdminLog.created_at.desc())
        .offset(offset)
        .limit(min(limit, 2000))
        .all()
    )


# ── Admin Dashboard ───────────────────────────────────────────────────────────

@app.get("/admin/dashboard", response_model=schemas.AdminDashboardOut)
def admin_dashboard(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    _require_admin(authorization, db)
    from sqlalchemy import func as sqlfunc
    from datetime import timezone

    total_students = db.query(sqlfunc.count(models.User.id)).filter(models.User.role == "student").scalar() or 0
    total_advisors = db.query(sqlfunc.count(models.User.id)).filter(models.User.role == "advisor").scalar() or 0
    total_courses  = db.query(sqlfunc.count(models.Course.id)).scalar() or 0

    # Refactor 2026-05-05: bỏ stats GPA-related (at_risk_students). Tool nội bộ
    # không quản lý điểm. at_risk = 0 (giữ field cho API compat).

    # active_users_this_week: số SV unique có UserGrade.uploaded_at trong 7 ngày qua
    week_ago = datetime.now(timezone.utc) - timedelta(days=7)
    week_ago_naive = week_ago.replace(tzinfo=None)
    active_this_week = (
        db.query(sqlfunc.count(sqlfunc.distinct(models.UserGrade.user_id)))
        .filter(models.UserGrade.uploaded_at >= week_ago_naive)
        .scalar() or 0
    )

    graduation_threshold = academic_engine._get_graduation_threshold(db)

    # SV chưa upload điểm cá nhân
    students_with_grades_ids = {
        row[0] for row in db.query(models.UserGrade.user_id).distinct().all()
    }
    all_student_ids = {
        row[0] for row in db.query(models.User.id).filter(models.User.role == "student").all()
    }
    students_no_grades = len(all_student_ids - students_with_grades_ids)

    students_first_login = (
        db.query(sqlfunc.count(models.User.id))
        .filter(models.User.role == "student", models.User.is_first_login == True)  # noqa: E712
        .scalar() or 0
    )

    # Phân bố theo khóa + chuyên ngành (cho overview)
    cohort_distribution: dict = {}
    spec_distribution: dict = {}
    students = db.query(models.User.username, models.User.specialization).filter(
        models.User.role == "student"
    ).all()
    for username, spec in students:
        if username and len(username) >= 3:
            cohort_key = username[1:3]
            cohort_distribution[cohort_key] = cohort_distribution.get(cohort_key, 0) + 1
        spec_key = spec or "Chưa chọn"
        spec_distribution[spec_key] = spec_distribution.get(spec_key, 0) + 1

    return schemas.AdminDashboardOut(
        total_students=total_students,
        total_advisors=total_advisors,
        total_courses=total_courses,
        at_risk_students=0,  # Deprecated — không quản lý điểm nữa
        active_users_this_week=active_this_week,
        graduation_threshold=graduation_threshold,
        students_no_grades=students_no_grades,
        students_first_login=students_first_login,
        cohort_distribution=cohort_distribution,
        spec_distribution=spec_distribution,
    )


# ── Admin Dashboard Stats (extended — 8 cards + Chart.js + warnings) ─────────

@app.get("/admin/dashboard/stats", response_model=schemas.AdminDashboardStatsOut)
def admin_dashboard_stats(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    _require_admin(authorization, db)
    from sqlalchemy import func as sqlfunc
    from datetime import timezone

    total_students = db.query(sqlfunc.count(models.User.id)).filter(models.User.role == "student").scalar() or 0
    total_advisors = db.query(sqlfunc.count(models.User.id)).filter(models.User.role == "advisor").scalar() or 0
    total_courses  = db.query(sqlfunc.count(models.Course.id)).scalar() or 0

    # at_risk: GPA tích lũy weighted < 2.0 (đồng bộ với /admin/users.avg_score4)
    # Quy tắc CLAUDE.md §5.4: chỉ tính môn passed + count_toward_credits=True, weighted by credits.
    # Refactor 2026-05-05: bỏ tính at_risk, thesis_eligible (admin không quản lý điểm).
    at_risk = 0
    thesis_eligible = 0
    graduation_threshold = academic_engine._get_graduation_threshold(db)

    week_ago_naive = (datetime.now(timezone.utc) - timedelta(days=7)).replace(tzinfo=None)
    active_this_week = (
        db.query(sqlfunc.count(sqlfunc.distinct(models.UserGrade.user_id)))
        .filter(models.UserGrade.uploaded_at >= week_ago_naive)
        .scalar() or 0
    )

    all_student_ids = {
        row[0] for row in db.query(models.User.id).filter(models.User.role == "student").all()
    }
    students_with_grades_ids = {
        row[0] for row in db.query(models.UserGrade.user_id).distinct().all()
    }
    students_no_grades = len(all_student_ids - students_with_grades_ids)

    students_with_advisor_ids = {
        row[0] for row in db.query(models.AdvisorAssignment.student_id).distinct().all()
    }
    students_no_advisor = len(all_student_ids - students_with_advisor_ids)

    students_first_login = (
        db.query(sqlfunc.count(models.User.id))
        .filter(models.User.role == "student", models.User.is_first_login == True)  # noqa: E712
        .scalar() or 0
    )

    # SV chưa đủ điều kiện làm đồ án TN
    total_students_count = len(all_student_ids)
    not_thesis_eligible = total_students_count - thesis_eligible

    notifications_total = db.query(sqlfunc.count(models.SystemNotification.id)).scalar() or 0
    total_reads = db.query(sqlfunc.count(models.NotificationRead.id)).scalar() or 0
    notifications_avg_read_rate = 0.0
    if notifications_total > 0 and total_students > 0:
        notifications_avg_read_rate = round(total_reads / (notifications_total * total_students), 3)

    _SPEC_SHORT = {
        "7480201_07": "KHMT", "7480201_06": "MMT", "7480201_05": "CNPM",
        "7480201_09": "HTTT", "7480201_04": "THKT", "7480201_08": "CNTTDH",
    }
    cohort_distribution: dict = {}
    spec_distribution: dict = {}
    for username, spec in db.query(models.User.username, models.User.specialization).filter(
        models.User.role == "student"
    ).all():
        if username:
            raw = username[2:4] if username.upper().startswith("SV") and len(username) >= 4 else username[:2]
            try:
                yr = int(raw)
                cohort_distribution[yr] = cohort_distribution.get(yr, 0) + 1
            except ValueError:
                pass
        spec_key = _SPEC_SHORT.get(spec, spec) if spec else "Chưa xác định"
        spec_distribution[spec_key] = spec_distribution.get(spec_key, 0) + 1

    # Chỉ lấy 5 khoá gần nhất
    all_cohorts = sorted(cohort_distribution.items())
    recent_cohorts = all_cohorts[-5:]

    sorted_spec = sorted(spec_distribution.items(), key=lambda x: -x[1])

    # Bỏ overdue tính theo TC — không quản lý điểm.
    overdue_count = 0

    # Warnings: chỉ những vấn đề liên quan tới quản lý tài khoản (không liên quan điểm)
    warnings_list = []
    if students_no_grades > 0:
        warnings_list.append({"level": "info", "message": f"{students_no_grades} sinh viên chưa upload bảng điểm cá nhân", "count": students_no_grades, "tab": "users", "filter": "no_grades"})
    if students_first_login > 0:
        warnings_list.append({"level": "warning", "message": f"{students_first_login} sinh viên chưa đăng nhập lần đầu", "count": students_first_login, "tab": "users", "filter": "default_pw"})
    if students_no_advisor > 0:
        warnings_list.append({"level": "info", "message": f"{students_no_advisor} sinh viên chưa có cố vấn phụ trách", "count": students_no_advisor, "tab": "users", "filter": "no_advisor"})

    return schemas.AdminDashboardStatsOut(
        total_students=total_students,
        total_advisors=total_advisors,
        total_courses=total_courses,
        at_risk_students=at_risk,
        thesis_eligible=thesis_eligible,
        students_no_grades=students_no_grades,
        students_no_advisor=students_no_advisor,
        not_thesis_eligible=not_thesis_eligible,
        students_first_login=students_first_login,
        active_users_this_week=active_this_week,
        graduation_threshold=graduation_threshold,
        notifications_total=notifications_total,
        notifications_avg_read_rate=notifications_avg_read_rate,
        cohort_labels=[f"K{yr:02d}" for yr, _ in recent_cohorts],
        cohort_values=[v for _, v in recent_cohorts],
        spec_labels=[k for k, _ in sorted_spec],
        spec_values=[v for _, v in sorted_spec],
        warnings=warnings_list,
    )


# ── Admin Config — graduation threshold ───────────────────────────────────────

@app.get("/admin/config/graduation-threshold", response_model=schemas.GraduationThresholdOut)
def admin_get_graduation_threshold(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    _require_admin(authorization, db)
    cfg = db.query(models.SystemConfig).filter(
        models.SystemConfig.key == "graduation_credit_threshold"
    ).first()
    if cfg:
        try:
            return schemas.GraduationThresholdOut(threshold=float(cfg.value), source="db")
        except (ValueError, TypeError):
            pass
    return schemas.GraduationThresholdOut(
        threshold=academic_engine._DEFAULT_GRADUATION_THRESHOLD, source="default"
    )


@app.put("/admin/config/graduation-threshold", response_model=schemas.GraduationThresholdOut)
def admin_set_graduation_threshold(
    body: schemas.GraduationThresholdIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    cfg = db.query(models.SystemConfig).filter(
        models.SystemConfig.key == "graduation_credit_threshold"
    ).first()
    if cfg:
        cfg.value = str(body.threshold)
    else:
        db.add(models.SystemConfig(key="graduation_credit_threshold", value=str(body.threshold)))
    db.commit()
    academic_engine._invalidate_graduation_threshold_cache()
    _log(db, admin, "SET_GRADUATION_THRESHOLD", "system_config",
         "graduation_credit_threshold", f"new={body.threshold}")
    return schemas.GraduationThresholdOut(threshold=body.threshold, source="db")


# ── Cấu hình ngưỡng học vụ (TT DN, ĐATN) ─────────────────────────────────────
_ACADEMIC_THRESHOLD_DEFAULTS = {
    "internship_min_credits": 90.0,
    "thesis_min_credits": 130.0,
    "thesis_min_gpa4": 2.0,
}


def _read_threshold(db: Session, key: str) -> tuple[float, str]:
    """Đọc 1 threshold từ SystemConfig → (value, source)."""
    cfg = db.query(models.SystemConfig).filter(models.SystemConfig.key == key).first()
    if cfg:
        try:
            return float(cfg.value), "db"
        except (ValueError, TypeError):
            pass
    return _ACADEMIC_THRESHOLD_DEFAULTS[key], "default"


@app.get("/admin/config/academic-thresholds", response_model=schemas.AcademicThresholdsOut)
def admin_get_academic_thresholds(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    _require_admin(authorization, db)
    intern_v, intern_src = _read_threshold(db, "internship_min_credits")
    thesis_v, thesis_src = _read_threshold(db, "thesis_min_credits")
    gpa_v, gpa_src = _read_threshold(db, "thesis_min_gpa4")
    sources = {intern_src, thesis_src, gpa_src}
    src = sources.pop() if len(sources) == 1 else "mixed"
    return schemas.AcademicThresholdsOut(
        internship_min_credits=intern_v,
        thesis_min_credits=thesis_v,
        thesis_min_gpa4=gpa_v,
        source=src,
    )


@app.put("/admin/config/academic-thresholds", response_model=schemas.AcademicThresholdsOut)
def admin_set_academic_thresholds(
    body: schemas.AcademicThresholdsIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    payload = {
        "internship_min_credits": body.internship_min_credits,
        "thesis_min_credits": body.thesis_min_credits,
        "thesis_min_gpa4": body.thesis_min_gpa4,
    }
    for key, val in payload.items():
        cfg = db.query(models.SystemConfig).filter(models.SystemConfig.key == key).first()
        if cfg:
            cfg.value = str(val)
        else:
            db.add(models.SystemConfig(key=key, value=str(val)))
    db.commit()
    from backend.core.academic_engine import _invalidate_academic_thresholds_cache
    _invalidate_academic_thresholds_cache()
    _log(db, admin, "SET_ACADEMIC_THRESHOLDS", "system_config",
         "academic_thresholds",
         f"internship={body.internship_min_credits}, thesis_tc={body.thesis_min_credits}, thesis_gpa={body.thesis_min_gpa4}")
    return schemas.AcademicThresholdsOut(
        internship_min_credits=body.internship_min_credits,
        thesis_min_credits=body.thesis_min_credits,
        thesis_min_gpa4=body.thesis_min_gpa4,
        source="db",
    )


# ── Skills (admin manage + public read) ──────────────────────────────────────

@app.get("/skills", response_model=list[schemas.SkillOut])
def list_skills(
    category: str | None = None,
    db: Session = Depends(get_db),
):
    """Public — list tất cả skills, optionally filter by category."""
    q = db.query(models.Skill)
    if category:
        q = q.filter(models.Skill.category == category)
    return q.order_by(models.Skill.category, models.Skill.name).all()


@app.post("/admin/skills", response_model=schemas.SkillOut)
def admin_create_skill(
    payload: schemas.SkillIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    if db.query(models.Skill).filter(models.Skill.code == payload.code).first():
        raise HTTPException(status_code=409, detail=f"Skill '{payload.code}' đã tồn tại")
    s = models.Skill(
        code=payload.code, name=payload.name,
        category=payload.category, description=payload.description,
    )
    db.add(s); db.commit(); db.refresh(s)
    _log(db, admin, "CREATE_SKILL", "skill", s.code, f"{s.name} ({s.category})")
    return s


@app.put("/admin/skills/{skill_code}", response_model=schemas.SkillOut)
def admin_update_skill(
    skill_code: str,
    payload: schemas.SkillIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    s = db.query(models.Skill).filter(models.Skill.code == skill_code).first()
    if not s:
        raise HTTPException(status_code=404, detail="Skill không tồn tại")
    if payload.code != skill_code:
        # Đổi code → cập nhật cả course_skills
        if db.query(models.Skill).filter(models.Skill.code == payload.code).first():
            raise HTTPException(status_code=409, detail=f"Skill code '{payload.code}' đã tồn tại")
        db.query(models.CourseSkill).filter(
            models.CourseSkill.skill_code == skill_code
        ).update({"skill_code": payload.code})
        s.code = payload.code
    s.name = payload.name
    s.category = payload.category
    s.description = payload.description
    db.commit(); db.refresh(s)
    _log(db, admin, "UPDATE_SKILL", "skill", s.code, s.name)
    return s


@app.delete("/admin/skills/{skill_code}", response_model=schemas.MessageOut)
def admin_delete_skill(
    skill_code: str,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    s = db.query(models.Skill).filter(models.Skill.code == skill_code).first()
    if not s:
        raise HTTPException(status_code=404, detail="Skill không tồn tại")
    # Cascade tự xóa CourseSkill rows nhờ ON DELETE CASCADE FK
    db.delete(s); db.commit()
    _log(db, admin, "DELETE_SKILL", "skill", skill_code, s.name)
    return {"message": f"Đã xóa skill '{skill_code}'"}


@app.get("/courses/{course_code}/details")
def get_course_details(
    course_code: str,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Trả tổng hợp thông tin 1 môn học cho trang chi tiết — phục vụ SV.

    Bao gồm: course info, skills, prereqs với status SV, downstream (môn nào
    yêu cầu môn này), rating summary, my_rating + my_review, recent reviews.
    """
    user = _get_user_by_token(authorization, db)

    course = db.query(models.Course).filter(
        models.Course.course_code == course_code
    ).first()
    if not course:
        raise HTTPException(status_code=404, detail="Không tìm thấy môn học")

    # Status của SV với môn này
    student_status: dict = {"status": "not_taken"}
    if user.role == "student":
        # Best grade
        best = db.query(models.UserGrade).filter(
            models.UserGrade.user_id == user.id,
            models.UserGrade.course_code == course_code,
        ).order_by(models.UserGrade.score10.desc().nullslast()).first()
        if best:
            student_status = {
                "status": "passed" if best.passed else "failed",
                "score10": float(best.score10) if best.score10 is not None else None,
                "score4": float(best.score4) if best.score4 is not None else None,
                "letter": best.letter,
                "term": best.term,
            }

    # Skills
    skill_rows = db.query(
        models.CourseSkill.skill_code, models.CourseSkill.weight,
        models.Skill.name, models.Skill.category, models.Skill.description,
    ).join(
        models.Skill, models.CourseSkill.skill_code == models.Skill.code
    ).filter(models.CourseSkill.course_code == course_code).all()
    skills = [
        {
            "skill_code": r[0], "weight": float(r[1]),
            "skill_name": r[2], "category": r[3], "description": r[4],
        }
        for r in sorted(skill_rows, key=lambda x: -float(x[1]))
    ]

    # Prereqs với status SV
    prereq_rows = db.query(models.CoursePrerequisite.prerequisite_code).filter(
        models.CoursePrerequisite.course_code == course_code
    ).all()
    prereq_codes = [r[0] for r in prereq_rows]
    prereqs = []
    if prereq_codes:
        pcourses = {c.course_code: c for c in db.query(models.Course).filter(
            models.Course.course_code.in_(prereq_codes)
        ).all()}
        # Best grade per prereq for this user
        my_prereq_grades = {}
        if user.role == "student":
            for g in db.query(models.UserGrade).filter(
                models.UserGrade.user_id == user.id,
                models.UserGrade.course_code.in_(prereq_codes),
            ).order_by(models.UserGrade.score10.desc().nullslast()).all():
                if g.course_code not in my_prereq_grades:
                    my_prereq_grades[g.course_code] = g
        for pc in prereq_codes:
            pcourse = pcourses.get(pc)
            grade = my_prereq_grades.get(pc)
            prereqs.append({
                "course_code": pc,
                "course_name": pcourse.course_name if pcourse else pc,
                "credits": float(pcourse.credits) if pcourse and pcourse.credits else None,
                "status": (
                    "passed" if grade and grade.passed
                    else "failed" if grade else "not_taken"
                ),
                "score10": float(grade.score10) if grade and grade.score10 is not None else None,
            })

    # Downstream — môn nào tiên quyết môn này
    downstream_rows = db.query(
        models.CoursePrerequisite.course_code,
        models.Course.course_name,
        models.Course.credits,
    ).join(
        models.Course, models.CoursePrerequisite.course_code == models.Course.course_code
    ).filter(models.CoursePrerequisite.prerequisite_code == course_code).all()
    downstream = [
        {
            "course_code": r[0],
            "course_name": r[1],
            "credits": float(r[2]) if r[2] is not None else None,
        }
        for r in downstream_rows
    ]

    # Rating summary
    from sqlalchemy import func as _sf
    ratings = db.query(models.CourseRating).filter(
        models.CourseRating.course_code == course_code
    ).all()
    count = len(ratings)
    avg = round(sum(r.rating for r in ratings) / count, 1) if count > 0 else None
    distribution = {str(i): sum(1 for r in ratings if r.rating == i) for i in range(1, 6)}
    my_r = next((r for r in ratings if r.user_id == user.id), None)

    # Có thể rate? — chỉ SV đã pass mới rate được
    can_rate = bool(
        user.role == "student" and student_status.get("status") == "passed"
    )

    # Specialization label
    _SPEC_LABELS_LOCAL = {
        "7480201_07": "Khoa học máy tính",
        "7480201_06": "Mạng máy tính",
        "7480201_05": "Công nghệ phần mềm",
        "7480201_09": "Hệ thống thông tin",
        "7480201_04": "Tin học kinh tế",
        "7480201_08": "CNTT Địa học",
    }
    spec_label = None
    if course.required_specialization:
        spec_label = _SPEC_LABELS_LOCAL.get(course.required_specialization, course.required_specialization)

    return {
        "course_code": course.course_code,
        "course_name": course.course_name,
        "credits": float(course.credits) if course.credits else None,
        "count_toward_credits": bool(course.count_toward_credits),
        "description": course.description or "",
        "typical_semester": course.typical_semester,
        "required_specialization": course.required_specialization,
        "specialization_label": spec_label,
        "skills": skills,
        "prereqs": prereqs,
        "downstream": downstream,
        "rating": {
            "avg": avg,
            "count": count,
            "distribution": distribution,
            "my_rating": my_r.rating if my_r else None,
            "my_review": my_r.review if my_r else None,
        },
        "can_rate": can_rate,
        "student_status": student_status,
    }


@app.get("/courses/{course_code}/reviews")
def get_course_reviews(
    course_code: str,
    limit: int = 20,
    offset: int = 0,
    sort: str = "recent",  # 'recent' | 'helpful' | 'rating_desc' | 'rating_asc'
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """List reviews chi tiết (có text comment) cho 1 môn.

    Anonymized — chỉ lộ initial chữ + cohort, không lộ full mã SV.
    """
    user = _get_user_by_token(authorization, db)
    safe_limit = max(1, min(int(limit), 50))
    safe_offset = max(0, int(offset))

    # Chỉ lấy review có text (loại stars-only) + chưa bị admin ẩn (hidden=False)
    base_q = db.query(models.CourseRating).filter(
        models.CourseRating.course_code == course_code,
        models.CourseRating.review.isnot(None),
        models.CourseRating.review != "",
        models.CourseRating.hidden == False,  # noqa: E712
    )
    total = base_q.count()

    if sort == "rating_desc":
        base_q = base_q.order_by(models.CourseRating.rating.desc(), models.CourseRating.id.desc())
    elif sort == "rating_asc":
        base_q = base_q.order_by(models.CourseRating.rating.asc(), models.CourseRating.id.desc())
    else:  # recent (default)
        base_q = base_q.order_by(models.CourseRating.id.desc())

    rows = base_q.offset(safe_offset).limit(safe_limit).all()

    # Map user_id → username để lấy cohort
    user_ids = {r.user_id for r in rows}
    users = {u.id: u for u in db.query(models.User).filter(
        models.User.id.in_(user_ids)
    ).all()} if user_ids else {}

    def _anonymize(u):
        if not u:
            return {"display_name": "Sinh viên ẩn danh", "cohort": None}
        # Cohort từ username
        cohort = None
        un = u.username or ""
        if un.upper().startswith("SV") and len(un) >= 4 and un[2:4].isdigit():
            cohort = "K" + un[2:4]
        elif len(un) >= 2 and un[:2].isdigit():
            cohort = "K" + un[:2]
        # Tên chỉ hiện chữ cái đầu
        first = (u.full_name or u.username or "?").strip()
        initial = first[0].upper() if first else "?"
        return {
            "display_name": f"{initial}***",
            "cohort": cohort,
        }

    items = []
    for r in rows:
        # Nếu SV chọn is_anonymous → ẩn hoàn toàn (kể cả cohort).
        # Else → partial anonymize (initial + cohort) như default.
        if r.is_anonymous and r.user_id != user.id:
            display_name = "Sinh viên ẩn danh"
            cohort = None
        else:
            anon = _anonymize(users.get(r.user_id))
            display_name = "Bạn" if r.user_id == user.id else anon["display_name"]
            cohort = anon["cohort"]
        items.append({
            "id": r.id,
            "rating": r.rating,
            "review": r.review,
            "is_mine": r.user_id == user.id,
            "is_anonymous": bool(r.is_anonymous),
            "display_name": display_name,
            "cohort": cohort,
        })

    return {
        "course_code": course_code,
        "total": total,
        "limit": safe_limit,
        "offset": safe_offset,
        "items": items,
    }


@app.get("/courses/{course_code}/skills", response_model=list[schemas.CourseSkillOut])
def get_course_skills(course_code: str, db: Session = Depends(get_db)):
    """Public — list skills + weight của 1 môn."""
    rows = db.query(
        models.CourseSkill.skill_code, models.CourseSkill.weight,
        models.Skill.name, models.Skill.category,
    ).join(
        models.Skill, models.CourseSkill.skill_code == models.Skill.code
    ).filter(models.CourseSkill.course_code == course_code).all()
    return [
        {"skill_code": r[0], "weight": float(r[1]), "skill_name": r[2], "category": r[3]}
        for r in rows
    ]


# REMOVED 2026-05-05: PUT /admin/courses/{code}/skills đã bỏ.
# Admin không quản lý skill mapping qua UI nữa — course_skills được seed bởi
# scripts/seed_careers.py và dùng trực tiếp cho recommendation rerank +
# career fit %. GET /courses/{code}/skills (public — SV xem) vẫn còn.


# REMOVED 2026-05-05: GET/PUT /me/career-skills đã bỏ.
# Cột users.career_skills đã DROP — feature SV chọn skills quan tâm không dùng nữa.


# ── Notifications v2 — flexible target groups ─────────────────────────────────

_VALID_TARGET_TYPES = {"all", "all_students", "all_advisors", "cohort",
                       "specialization", "students", "advisors", "department"}
_VALID_SEVERITIES = {"info", "warning", "urgent"}


def _notif_visible_to(notif: models.SystemNotification, user: models.User) -> bool:
    tt = notif.target_type or "all"
    tv = notif.target_value or ""
    values = [v.strip() for v in tv.split(",") if v.strip()]
    if tt == "all":
        return True
    if tt == "all_students":
        return user.role == "student"
    if tt == "all_advisors":
        return user.role == "advisor"
    if tt == "cohort":
        cohort_nums = [v.lstrip("K") for v in values]
        uname = user.username or ""
        cohort_of_user = uname[2:4] if uname.upper().startswith("SV") and len(uname) >= 4 else uname[:2]
        return user.role == "student" and cohort_of_user in cohort_nums
    if tt == "specialization":
        return user.role == "student" and user.specialization in values
    if tt == "students":
        return user.username in values
    if tt == "advisors":
        return user.role == "advisor" and (user.teacher_code or "") in values
    if tt == "department":
        return user.role == "advisor" and user.managed_specialization == tv.strip()
    return False


def _count_reach(db: Session, tt: str, tv: str) -> int:
    from sqlalchemy import func as sqlfunc, or_
    values = [v.strip() for v in (tv or "").split(",") if v.strip()]
    if tt == "all":
        return db.query(sqlfunc.count(models.User.id)).filter(models.User.role != "admin").scalar() or 0
    if tt == "all_students":
        return db.query(sqlfunc.count(models.User.id)).filter(models.User.role == "student").scalar() or 0
    if tt == "all_advisors":
        return db.query(sqlfunc.count(models.User.id)).filter(models.User.role == "advisor").scalar() or 0
    if tt == "cohort":
        cohort_nums = [v.lstrip("K") for v in values]
        if not cohort_nums:
            return 0
        patterns = []
        for c in cohort_nums:
            patterns.append(models.User.username.ilike(f"SV{c}%"))
            patterns.append(models.User.username.ilike(f"{c}%"))
        return db.query(sqlfunc.count(models.User.id)).filter(
            models.User.role == "student",
            or_(*patterns),
        ).scalar() or 0
    if tt == "specialization":
        return db.query(sqlfunc.count(models.User.id)).filter(
            models.User.role == "student", models.User.specialization.in_(values)
        ).scalar() or 0
    if tt == "students":
        return db.query(sqlfunc.count(models.User.id)).filter(
            models.User.username.in_(values)
        ).scalar() or 0
    if tt == "advisors":
        return db.query(sqlfunc.count(models.User.id)).filter(
            models.User.role == "advisor", models.User.teacher_code.in_(values)
        ).scalar() or 0
    if tt == "department":
        return db.query(sqlfunc.count(models.User.id)).filter(
            models.User.role == "advisor", models.User.managed_specialization == tv.strip()
        ).scalar() or 0
    return 0


@app.post("/admin/notifications")
def admin_create_notification_v2(
    payload: schemas.NotificationCreateIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    title = payload.title.strip()
    content = payload.content.strip()
    if len(title) < 3 or len(title) > 200:
        raise HTTPException(400, "title phải từ 3–200 ký tự")
    if not content:
        raise HTTPException(400, "content không được rỗng")
    if payload.severity not in _VALID_SEVERITIES:
        raise HTTPException(400, f"severity phải là: {', '.join(sorted(_VALID_SEVERITIES))}")
    tt = (payload.target_type or "all").strip()
    if tt not in _VALID_TARGET_TYPES:
        raise HTTPException(400, f"target_type không hợp lệ")
    tv = (payload.target_value or "").strip()
    values = [v.strip() for v in tv.split(",") if v.strip()]
    if tt == "students" and values:
        existing = {u.username for u in db.query(models.User).filter(
            models.User.username.in_(values), models.User.role == "student"
        ).all()}
        missing = [v for v in values if v not in existing]
        if missing:
            raise HTTPException(400, f"Không tìm thấy sinh viên: {', '.join(missing)}")
    elif tt == "advisors" and values:
        existing = {u.teacher_code for u in db.query(models.User).filter(
            models.User.teacher_code.in_(values)
        ).all()}
        missing = [v for v in values if v not in existing]
        if missing:
            raise HTTPException(400, f"Không tìm thấy cố vấn: {', '.join(missing)}")
    notif = models.SystemNotification(
        title=title, body=content,
        type=payload.severity, severity=payload.severity,
        target_type=tt, target_value=tv or None,
        admin_id=admin.id, admin_username=admin.username,
        is_active=True,
    )
    db.add(notif)
    db.commit()
    db.refresh(notif)
    _log(db, admin, "SEND_NOTIF", "notification", str(notif.id), title)
    return {
        "id": notif.id, "title": notif.title, "body": notif.body,
        "severity": notif.severity, "target_type": notif.target_type,
        "target_value": notif.target_value, "admin_username": notif.admin_username,
        "is_active": notif.is_active, "created_at": notif.created_at,
        "read_count": 0, "total_reach": _count_reach(db, tt, tv),
    }


@app.get("/admin/notifications")
def admin_list_notifications_v2(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    from sqlalchemy import func as sqlfunc
    _require_admin(authorization, db)
    notifs = (
        db.query(models.SystemNotification)
        .order_by(models.SystemNotification.created_at.desc())
        .limit(200).all()
    )
    result = []
    for n in notifs:
        read_count = db.query(sqlfunc.count(models.NotificationRead.id)).filter(
            models.NotificationRead.notification_id == n.id
        ).scalar() or 0
        result.append({
            "id": n.id, "title": n.title, "body": n.body,
            "severity": n.severity or n.type or "info",
            "target_type": n.target_type or "all",
            "target_value": n.target_value,
            "admin_username": n.admin_username,
            "is_active": n.is_active, "created_at": n.created_at,
            "read_count": read_count,
            "total_reach": _count_reach(db, n.target_type or "all", n.target_value or ""),
        })
    return result


@app.delete("/admin/notifications/{notif_id}")
def admin_delete_notification(
    notif_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    notif = db.query(models.SystemNotification).filter(
        models.SystemNotification.id == notif_id
    ).first()
    if not notif:
        raise HTTPException(404, "Không tìm thấy thông báo")
    title = notif.title
    db.delete(notif)
    db.commit()
    _log(db, admin, "DELETE_NOTIF", "notification", str(notif_id), title)
    return schemas.MessageOut(message="Đã xóa thông báo")


@app.patch("/admin/notifications/{notif_id}/deactivate")
def admin_deactivate_notification_v2(
    notif_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    notif = db.query(models.SystemNotification).filter(
        models.SystemNotification.id == notif_id
    ).first()
    if not notif:
        raise HTTPException(404, "Không tìm thấy thông báo")
    notif.is_active = False
    db.commit()
    _log(db, admin, "DEACTIVATE_NOTIF", "notification", str(notif_id), notif.title)
    return schemas.MessageOut(message="Đã thu hồi thông báo")


@app.get("/admin/notifications/estimate-reach")
def estimate_notification_reach(
    target_type: str = "all",
    target_value: str = "",
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    _require_admin(authorization, db)
    if target_type not in _VALID_TARGET_TYPES:
        raise HTTPException(400, "target_type không hợp lệ")
    return {"count": _count_reach(db, target_type, target_value)}


@app.get("/notifications/me")
def get_my_notifications(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    all_notifs = (
        db.query(models.SystemNotification)
        .filter(models.SystemNotification.is_active == True)
        .order_by(models.SystemNotification.created_at.desc())
        .limit(100).all()
    )
    read_ids = {
        r.notification_id for r in db.query(models.NotificationRead).filter(
            models.NotificationRead.user_id == user.id
        ).all()
    }
    return [
        {
            "id": n.id, "title": n.title, "body": n.body,
            "severity": n.severity or n.type or "info",
            "created_at": n.created_at,
            "is_read": n.id in read_ids,
        }
        for n in all_notifs if _notif_visible_to(n, user)
    ]


@app.post("/notifications/{notif_id}/mark-read")
def mark_notification_read(
    notif_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    existing = db.query(models.NotificationRead).filter(
        models.NotificationRead.notification_id == notif_id,
        models.NotificationRead.user_id == user.id,
    ).first()
    if not existing:
        db.add(models.NotificationRead(notification_id=notif_id, user_id=user.id))
        db.commit()
    return schemas.MessageOut(message="ok")


@app.post("/notifications/mark-all-read")
def mark_all_notifications_read(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    all_notifs = (
        db.query(models.SystemNotification)
        .filter(models.SystemNotification.is_active == True).all()
    )
    read_ids = {
        r.notification_id for r in db.query(models.NotificationRead).filter(
            models.NotificationRead.user_id == user.id
        ).all()
    }
    for n in all_notifs:
        if _notif_visible_to(n, user) and n.id not in read_ids:
            db.add(models.NotificationRead(notification_id=n.id, user_id=user.id))
    db.commit()
    return schemas.MessageOut(message="ok")


@app.get("/notifications/me/unread-count")
def get_unread_notification_count(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    all_notifs = (
        db.query(models.SystemNotification)
        .filter(models.SystemNotification.is_active == True).all()
    )
    read_ids = {
        r.notification_id for r in db.query(models.NotificationRead).filter(
            models.NotificationRead.user_id == user.id
        ).all()
    }
    count = sum(1 for n in all_notifs if _notif_visible_to(n, user) and n.id not in read_ids)
    return {"count": count}


# ── Specializations comparison (Phase 11) ─────────────────────────────────────

# Static metadata for 6 CN — used by /specs/compare and SV năm 1-2 lúc chọn CN
_SPEC_META = {
    "7480201_07": {
        "name": "Khoa học máy tính (ứng dụng)", "short": "KHMT", "icon": "psychology", "color": "indigo",
        "tagline": "AI/ML, khai phá DL, thuật toán",
        "description": "Nghiên cứu nền tảng lý thuyết tính toán, thuật toán và trí tuệ nhân tạo. Phù hợp sinh viên yêu thích nghiên cứu, AI và khoa học dữ liệu.",
        "careers": ["Kỹ sư AI", "Nhà khoa học dữ liệu", "Kỹ sư ML", "Nghiên cứu viên"],
        "salary_range": {"min": 15, "max": 30, "unit": "triệu VND/tháng"},
        "fit_keywords": ["AI Engineer", "Data Scientist", "ML Engineer", "Research"],
    },
    "7480201_06": {
        "name": "Mạng máy tính", "short": "MMT", "icon": "lan", "color": "sky",
        "tagline": "An ninh mạng, hạ tầng, DevOps",
        "description": "Thiết kế, quản trị hạ tầng mạng và bảo mật hệ thống. Phù hợp sinh viên yêu thích an ninh mạng, hệ thống phân tán và điện toán đám mây.",
        "careers": ["Kỹ sư mạng", "Chuyên gia an ninh mạng", "DevOps Engineer", "Cloud Architect"],
        "salary_range": {"min": 13, "max": 25, "unit": "triệu VND/tháng"},
        "fit_keywords": ["Network Engineer", "DevOps", "SecOps", "Cloud"],
    },
    "7480201_05": {
        "name": "Công nghệ phần mềm", "short": "CNPM", "icon": "code", "color": "emerald",
        "tagline": ".NET, mobile, OOP, kỹ nghệ PM",
        "description": "Quy trình phát triển, kiểm thử và bảo trì phần mềm quy mô lớn. Phù hợp sinh viên muốn làm lập trình viên chuyên nghiệp theo chuẩn công nghiệp.",
        "careers": ["Lập trình viên Backend", "Fullstack Developer", "Mobile Developer", "QA Engineer"],
        "salary_range": {"min": 14, "max": 28, "unit": "triệu VND/tháng"},
        "fit_keywords": ["Backend Dev", "Fullstack", "Mobile Dev", "QA"],
    },
    "7480201_09": {
        "name": "Hệ thống thông tin", "short": "HTTT", "icon": "schema", "color": "violet",
        "tagline": "ERP, BA, BI, hệ quản trị",
        "description": "Phân tích, thiết kế và quản trị hệ thống thông tin doanh nghiệp. Phù hợp sinh viên muốn làm cầu nối giữa công nghệ và nghiệp vụ kinh doanh.",
        "careers": ["Business Analyst", "Nhà phân tích dữ liệu", "BI Developer", "ERP Consultant"],
        "salary_range": {"min": 13, "max": 24, "unit": "triệu VND/tháng"},
        "fit_keywords": ["Business Analyst", "Data Analyst", "BI Dev"],
    },
    "7480201_04": {
        "name": "Tin học kinh tế", "short": "THKT", "icon": "trending_up", "color": "amber",
        "tagline": "Tài chính - kế toán + IT, chuyển đổi số",
        "description": "Ứng dụng toán học và công nghệ vào phân tích tài chính, kế toán và chuyển đổi số doanh nghiệp. Phù hợp sinh viên muốn kết hợp IT và kinh tế.",
        "careers": ["FinTech Developer", "Kế toán + IT", "BA tài chính", "Chuyên viên chuyển đổi số"],
        "salary_range": {"min": 12, "max": 22, "unit": "triệu VND/tháng"},
        "fit_keywords": ["FinTech", "Accountant + IT", "BA tài chính"],
    },
    "7480201_08": {
        "name": "CNTT Địa học", "short": "CNTTDH", "icon": "public", "color": "teal",
        "tagline": "GIS, viễn thám, mỏ - khoáng sản",
        "description": "Phát triển ứng dụng GIS, viễn thám và hệ thống thông tin địa lý phục vụ ngành mỏ - khoáng sản. Phù hợp sinh viên yêu thích bản đồ và dữ liệu không gian.",
        "careers": ["Kỹ sư GIS", "Chuyên viên viễn thám", "Lập trình viên bản đồ số"],
        "salary_range": {"min": 12, "max": 20, "unit": "triệu VND/tháng"},
        "fit_keywords": ["GIS Engineer", "Geo-IT", "Remote Sensing"],
    },
}
_SPEC_ORDER = ["7480201_07", "7480201_06", "7480201_05", "7480201_09", "7480201_04", "7480201_08"]


@app.get("/specs/compare")
def compare_specializations(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Trả về bảng so sánh 6 chuyên ngành cho SV chuẩn bị chọn CN.
    Mỗi spec gồm: BB courses (12 TC), pool TC-B/TC-C, internship+thesis codes.
    Endpoint công khai cho mọi role authenticated.
    """
    user = _get_user_by_token(authorization, db)
    from collections import defaultdict

    # Pre-load all data
    all_courses = {c.course_code: c for c in db.query(models.Course).all()}
    all_eg = db.query(models.CourseElectiveGroup).all()  # course_code, specialization, group_type
    all_m2m = db.query(models.CourseSpecialization).all()  # course_code → spec
    all_prereqs = db.query(models.CoursePrerequisite).all()
    all_course_skills = db.query(models.CourseSkill).all()  # course_code, skill_code, weight
    all_skills = {s.code: s for s in db.query(models.Skill).all()}
    prereq_map: dict[str, list[str]] = defaultdict(list)
    for p in all_prereqs:
        prereq_map[p.course_code].append(p.prerequisite_code)

    # Index course → list of (skill_code, weight, skill_name, skill_category)
    course_to_skills: dict[str, list] = defaultdict(list)
    for cs in all_course_skills:
        s = all_skills.get(cs.skill_code)
        course_to_skills[cs.course_code].append({
            "code": cs.skill_code,
            "weight": float(cs.weight),
            "name": s.name if s else cs.skill_code,
            "category": s.category if s else None,
        })

    # Build M2M index: course → set(specs)
    m2m_specs: dict[str, set] = defaultdict(set)
    for r in all_m2m:
        m2m_specs[r.course_code].add(r.specialization)

    # Build elective pool index: (spec, group_type) → list[course_code]
    pool_index: dict[tuple, list[str]] = defaultdict(list)
    for eg in all_eg:
        if eg.specialization == "Chung":
            continue
        pool_index[(eg.specialization, eg.group_type.upper())].append(eg.course_code)

    def _course_brief(code: str) -> dict | None:
        c = all_courses.get(code)
        if not c:
            return None
        return {
            "course_code": c.course_code,
            "course_name": c.course_name,
            "credits": float(c.credits) if c.credits else 0.0,
            "typical_semester": c.typical_semester,
        }

    result_specs = []
    for spec_code in _SPEC_ORDER:
        meta = _SPEC_META[spec_code]
        # BB courses for this spec — required_specialization == spec OR M2M tied to ONLY this spec
        bb_courses = []
        for c in all_courses.values():
            if c.required_specialization == spec_code:
                bb_courses.append(c)
            elif spec_code in m2m_specs.get(c.course_code, set()) and len(m2m_specs[c.course_code]) == 1:
                # M2M tied exclusively to this spec
                if c.course_code not in {eg.course_code for eg in all_eg}:
                    bb_courses.append(c)
        bb_briefs = [_course_brief(c.course_code) for c in bb_courses if _course_brief(c.course_code)]
        bb_briefs.sort(key=lambda x: (x.get("typical_semester") or 99, x["course_code"]))
        bb_credits = round(sum(b["credits"] for b in bb_briefs), 1)

        # Internship/thesis identification by name pattern
        internship_code = None
        thesis_code = None
        for b in bb_briefs:
            n = (b.get("course_name") or "").lower()
            if "thực tập" in n and "tốt nghiệp" in n:
                internship_code = b["course_code"]
            elif "đồ án" in n and "tốt nghiệp" in n:
                thesis_code = b["course_code"]

        # Filter out internship/thesis from BB list (show separately)
        core_bb = [b for b in bb_briefs if b["course_code"] not in (internship_code, thesis_code)]

        # TC-B / TC-C pools
        b_pool = [_course_brief(code) for code in pool_index.get((spec_code, "B"), [])]
        c_pool = [_course_brief(code) for code in pool_index.get((spec_code, "C"), [])]
        b_pool = [x for x in b_pool if x]
        c_pool = [x for x in c_pool if x]

        # Top skills emphasis for this spec — aggregate from BB + B + C course→skill mappings
        # Weight a skill by sum of (course_skill.weight) across spec-specific courses
        skill_score: dict[str, float] = defaultdict(float)
        skill_meta: dict[str, dict] = {}
        spec_courses = (
            [b["course_code"] for b in bb_briefs]
            + [c["course_code"] for c in b_pool]
            + [c["course_code"] for c in c_pool]
        )
        for code in spec_courses:
            for sk in course_to_skills.get(code, []):
                skill_score[sk["code"]] += sk["weight"]
                if sk["code"] not in skill_meta:
                    skill_meta[sk["code"]] = {"name": sk["name"], "category": sk["category"]}

        # Sort by score desc, take top 6
        top_skills_sorted = sorted(skill_score.items(), key=lambda kv: -kv[1])[:6]
        max_score = top_skills_sorted[0][1] if top_skills_sorted else 1.0
        top_skills = [
            {
                "code": code,
                "name": skill_meta[code]["name"],
                "category": skill_meta[code]["category"],
                "score": round(score, 2),
                "percent": round(score / max_score * 100, 1) if max_score > 0 else 0,
            }
            for code, score in top_skills_sorted
        ]

        # Core topics — short labels for first 3 BB courses (Vietnamese-friendly)
        # These appear in matrix view to differentiate at a glance
        core_topics = [b["course_name"] for b in core_bb[:4]]

        result_specs.append({
            "code": spec_code,
            "name": meta["name"],
            "short_name": meta["short"],
            "icon": meta["icon"],
            "color": meta["color"],
            "tagline": meta["tagline"],
            "description": meta.get("description", ""),
            "careers": meta.get("careers", []),
            "salary_range": meta.get("salary_range"),
            "fit_keywords": meta["fit_keywords"],
            "compulsory_credits": bb_credits,
            "compulsory_courses": core_bb,
            "core_topics": core_topics,
            "top_skills": top_skills,
            "internship": _course_brief(internship_code) if internship_code else None,
            "thesis": _course_brief(thesis_code) if thesis_code else None,
            "elective_b": {"min_credits": 9, "courses": b_pool},
            "elective_c": {"min_credits": 9, "courses": c_pool},
        })

    return {
        "specs": result_specs,
        "common_credits": 117,  # đại cương 65 + cơ sở ngành 52
        "spec_credits": 40,     # 12 BB + 9 TC-B + 9 TC-C + 10 TT + 10 ĐATN
        "current_spec": user.specialization,
    }


# ── Roadmap ────────────────────────────────────────────────────────────────────

@app.get("/roadmap/me", response_model=schemas.RoadmapOut)
def roadmap_me(
    max_credits: float | None = None,
    explore_spec: str | None = None,
    target_terms: int = 9,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Get roadmap. Khi SV chưa chốt CN có thể truyền explore_spec=<code> để xem
    lộ trình giả lập với CN đó (kho TC sẽ chứa elective slots theo CN này).

    target_terms: tổng HK dự định TN (mặc định 9 — CTĐT chuẩn 4.5 năm).
    Cho phép SV chỉnh nhịp học (8 = TN sớm, 10-11 = kéo dài).
    """
    user = _get_user_by_token(authorization, db)
    # explore_spec chỉ có tác dụng khi user chưa có CN thật (tránh override người đã chốt)
    override = explore_spec if (explore_spec and not user.specialization) else None
    # Clamp target_terms an toàn
    target_terms = max(6, min(12, int(target_terms or 9)))
    result = build_semester_roadmap(
        db, user.id,
        max_credits_per_term=max_credits,
        available_course_codes=None,
        override_spec=override,
        target_terms=target_terms,
    )
    return schemas.RoadmapOut(**result)


@app.get("/roadmap/custom/me")
def get_custom_roadmap(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    plan = db.query(models.StudyPlan).filter(
        models.StudyPlan.user_id == user.id,
        models.StudyPlan.plan_name == "roadmap_custom",
    ).first()
    if not plan:
        return {"saved": False, "semesters": []}
    items = db.query(models.StudyPlanItem).filter(
        models.StudyPlanItem.plan_id == plan.id
    ).all()
    from collections import defaultdict
    sem_map = defaultdict(list)
    for item in items:
        sem_map[item.term_label].append(item.course_code)
    semesters = [{"semester_label": k, "course_codes": v} for k, v in sem_map.items()]
    return {"saved": True, "semesters": semesters}


@app.post("/roadmap/whatif")
def roadmap_whatif(
    payload: dict = Body(...),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """What-if simulator: returns impact of simulating one or more completed courses as failed."""
    user = _get_user_by_token(authorization, db)
    simulate_failed: list[str] = payload.get("simulate_failed", [])
    if not simulate_failed:
        raise HTTPException(status_code=422, detail="simulate_failed list required")

    original = build_semester_roadmap(db, user.id)
    simulated = build_semester_roadmap(db, user.id, simulate_failed=simulate_failed)

    orig_terms = original.get("estimated_terms", 0) or 0
    sim_terms = simulated.get("estimated_terms", 0) or 0

    # Find courses that are now blocked because a prereq is in simulate_failed
    prereq_map: dict = simulated.get("prereq_map", {})
    failed_set = set(simulate_failed)
    blocked = []
    for course_code, prereqs in prereq_map.items():
        if any(p in failed_set for p in prereqs):
            blocked.append(course_code)

    # Collect course names for blocked courses
    all_courses_map = {c.course_code: c.course_name for c in db.query(models.Course).all()}
    blocked_details = [
        {"course_code": code, "course_name": all_courses_map.get(code, code)}
        for code in blocked
    ]

    # Include first 4 upcoming simulated semesters
    upcoming = simulated.get("semesters", [])[:4]

    return {
        "original_terms": orig_terms,
        "simulated_terms": sim_terms,
        "delta_terms": sim_terms - orig_terms,
        "blocked_courses": blocked_details,
        "upcoming_semesters": upcoming,
    }


@app.get("/me/upcoming-deadlines")
def get_upcoming_deadlines(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Trả về list các deadline / nhắc nhở quan trọng cho SV đang đăng nhập.

    Item shape:
      { kind: 'registration'|'risk'|'reminder'|'milestone',
        severity: 'critical'|'warning'|'info',
        title, description, due_at (ISO|null), days_left (int|null),
        action_label, action_href }

    Source:
    - SystemConfig key 'next_registration_open_at' → registration deadline
    - Active risk cases (case_state='open') → cảnh báo học vụ
    - User state heuristics (chưa upload BĐ / chưa quiz / sắp TN) → milestone
    """
    from datetime import datetime as _dt
    user = _get_user_by_token(authorization, db)
    items: list[dict] = []
    now = _dt.utcnow()

    # 1. Registration deadline
    cfg = db.query(models.SystemConfig).filter(
        models.SystemConfig.key == "next_registration_open_at"
    ).first()
    if cfg and cfg.value:
        try:
            due = _dt.fromisoformat(cfg.value.replace("Z", "+00:00").rstrip("Z"))
            if due.tzinfo:
                due = due.replace(tzinfo=None)
            days_left = (due - now).days
            if -1 <= days_left <= 30:  # show within 30 days before, +1 day after
                severity = "critical" if days_left <= 3 else ("warning" if days_left <= 7 else "info")
                items.append({
                    "kind": "registration",
                    "severity": severity,
                    "title": f"Đăng ký HK tới {'mở vào' if days_left > 0 else 'đã mở'}",
                    "description": f"{days_left} ngày nữa" if days_left > 0 else "Hôm nay/đã quá hạn — đăng ký ngay",
                    "due_at": due.isoformat(),
                    "days_left": days_left,
                    "action_label": "Xem kế hoạch HK tới",
                    "action_href": "integrated-roadmap.html",
                })
        except Exception:
            pass

    # Risk cases removed 2026-05-05 — bảng + workflow đã drop, tool nội bộ
    # không cần tracking risk (advisor xem trực tiếp progress SV).

    # 3. State milestones (passive nudges)
    has_grades = db.query(models.UserGrade).filter(models.UserGrade.user_id == user.id).count() > 0
    if not has_grades:
        items.append({
            "kind": "milestone",
            "severity": "info",
            "title": "Upload bảng điểm",
            "description": "Hệ thống chưa có bảng điểm — upload để mở khoá tính năng",
            "due_at": None, "days_left": None,
            "action_label": "Upload",
            "action_href": "grades.html",
        })

    # Sort by severity then days_left ascending
    SEV_ORDER = {"critical": 0, "warning": 1, "info": 2}
    items.sort(key=lambda x: (SEV_ORDER.get(x["severity"], 9),
                              x["days_left"] if x["days_left"] is not None else 999))

    return {"items": items, "fetched_at": now.isoformat()}


@app.get("/me/next-term-plan")
def get_next_term_plan(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Lấy kế hoạch SV đã lưu cho HK kế tiếp (tab 'Kỳ tới' trong roadmap).

    Khác với /roadmap/custom/me (full 9-HK roadmap):
    - Đây là plan ngắn hạn (1 HK), tách riêng plan_name='next_term'
    - Term + course list được lưu đơn giản, không có ràng buộc CTĐT
    """
    user = _get_user_by_token(authorization, db)
    plan = db.query(models.StudyPlan).filter(
        models.StudyPlan.user_id == user.id,
        models.StudyPlan.plan_name == "next_term",
    ).first()
    if not plan:
        return {"saved": False, "term": None, "courses": [], "saved_at": None}
    items = db.query(models.StudyPlanItem).filter(
        models.StudyPlanItem.plan_id == plan.id
    ).all()
    course_lookup = {
        c.course_code: c for c in db.query(models.Course).filter(
            models.Course.course_code.in_([i.course_code for i in items])
        ).all()
    } if items else {}
    courses = []
    for it in items:
        c = course_lookup.get(it.course_code)
        courses.append({
            "course_code": it.course_code,
            "course_name": c.course_name if c else it.course_code,
            "credits": float(c.credits) if c and c.credits else 0,
        })
    return {
        "saved": True,
        "term": items[0].term_label if items else None,
        "courses": courses,
        "saved_at": plan.updated_at.isoformat() if plan.updated_at else None,
    }


@app.put("/me/next-term-plan", response_model=schemas.MessageOut)
def save_next_term_plan(
    payload: dict = Body(...),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Lưu/Cập nhật kế hoạch HK kế tiếp.

    Body: {term: 'HK7', courses: [{course_code: '...'}]}
    Replace all existing items for this plan.
    """
    user = _get_user_by_token(authorization, db)
    term = payload.get("term") or "HK"
    course_codes = [c.get("course_code") for c in (payload.get("courses") or []) if c.get("course_code")]
    if not course_codes:
        raise HTTPException(status_code=422, detail="Cần ít nhất 1 môn để lưu kế hoạch")

    # Validate course_codes exist
    valid_codes = {c.course_code for c in db.query(models.Course).filter(
        models.Course.course_code.in_(course_codes)
    ).all()}
    invalid = [c for c in course_codes if c not in valid_codes]
    if invalid:
        raise HTTPException(status_code=400, detail=f"Mã môn không hợp lệ: {', '.join(invalid[:3])}")

    plan = db.query(models.StudyPlan).filter(
        models.StudyPlan.user_id == user.id,
        models.StudyPlan.plan_name == "next_term",
    ).first()
    if not plan:
        plan = models.StudyPlan(user_id=user.id, plan_name="next_term")
        db.add(plan)
        db.flush()
    else:
        db.query(models.StudyPlanItem).filter(
            models.StudyPlanItem.plan_id == plan.id
        ).delete()

    from datetime import datetime as _dtnow
    plan.updated_at = _dtnow.utcnow()
    for code in course_codes:
        db.add(models.StudyPlanItem(plan_id=plan.id, course_code=code, term_label=term))
    db.commit()
    return {"message": f"Đã lưu kế hoạch {len(course_codes)} môn cho {term}"}


@app.put("/roadmap/custom/me")
def save_custom_roadmap(
    payload: dict = Body(...),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Save user's custom term placement.
    Accepts two formats:
      New: {"items": [{"course_code": "...", "term_label": "HK1"}]}
      Old: {"semesters": [{"semester_label": "HK1", "course_codes": [...]}]}
    Validates prereq ordering and rejects 422 on violation."""
    user = _get_user_by_token(authorization, db)

    # Normalise both body formats into flat list of (course_code, term_label)
    raw_items: list[tuple[str, str]] = []
    if "items" in payload:
        items_in = payload["items"]
        if not isinstance(items_in, list):
            raise HTTPException(status_code=422, detail="'items' phải là mảng")
        if len(items_in) > 200:
            raise HTTPException(status_code=422, detail="Tối đa 200 môn")
        for itm in items_in:
            code = (itm.get("course_code") or "").strip()
            label = (itm.get("term_label") or "").strip()
            if code:
                raw_items.append((code, label))
    else:
        for sem in (payload.get("semesters") or []):
            label = sem.get("semester_label", "")
            for code in (sem.get("course_codes") or []):
                if code:
                    raw_items.append((str(code).strip(), str(label).strip()))

    # Validate course codes exist
    codes = {c for c, _ in raw_items}
    valid_codes = {
        c.course_code
        for c in db.query(models.Course).filter(models.Course.course_code.in_(codes)).all()
    }
    for code, _ in raw_items:
        if code not in valid_codes:
            raise HTTPException(status_code=422, detail=f"Môn {code} không tồn tại")

    # Validate term_label format. Accept: HK1..HK99 (HK chính, không có HK hè).
    _TERM_RE = re.compile(r'^HK\d+$')
    for code, label in raw_items:
        if not _TERM_RE.match(label):
            raise HTTPException(
                status_code=422,
                detail=f"HK '{label}' không hợp lệ (phải dạng HK1, HK10, ...)"
            )

    # Validate Quy chế: tối đa 16 HK chính (8 năm)
    unique_terms = {l for _, l in raw_items}
    if len(unique_terms) > 16:
        raise HTTPException(
            status_code=422,
            detail=f"Vượt Quy chế: tối đa 16 HK chính (đang có {len(unique_terms)})"
        )

    # Build placement map: course_code -> term index (HK1=1, HK10=10, ...)
    def _term_idx(label: str) -> int:
        try:
            return int(label.replace("HK", "").strip())
        except Exception:
            return 99

    custom_placement: dict[str, int] = {code: _term_idx(label) for code, label in raw_items}
    label_of: dict[str, str] = {code: label for code, label in raw_items}

    # Build EFFECTIVE placement = CTDT default + customOverride.
    # Lý do: client chỉ gửi items đã di chuyển khỏi default. Backend cần biết
    # vị trí default của các môn KHÁC để validate prereq cho cả những môn user
    # chưa move (mà có thể bị ảnh hưởng bởi move hiện tại).
    # Default = Course.typical_semester (HK theo CTĐT 1-9).
    effective_placement: dict[str, int] = {}
    try:
        spec_code = (user.specialization or "").strip() or None
        # Lấy tất cả courses áp dụng cho spec hoặc shared (required_spec NULL)
        q = db.query(models.Course).filter(models.Course.typical_semester.isnot(None))
        if spec_code:
            q = q.filter(
                (models.Course.required_specialization.is_(None))
                | (models.Course.required_specialization == spec_code)
            )
        for c in q.all():
            if c.course_code and c.typical_semester:
                effective_placement[c.course_code] = int(c.typical_semester)
    except Exception:
        # Fallback: nếu không lấy được standard plan, validate chỉ trên payload (như cũ)
        pass

    # Override: customItems chiếm ưu tiên hơn default
    for code, idx in custom_placement.items():
        effective_placement[code] = idx
        label_of.setdefault(code, f"HK{idx}")

    # Validate prereq ordering trên effective_placement (full picture)
    prereq_rows = db.query(models.CoursePrerequisite).all()
    # Filter chỉ những prereq liên quan đến course có trong effective placement
    effective_codes = set(effective_placement.keys())
    for p in prereq_rows:
        if p.course_code not in effective_codes:
            continue
        if p.prerequisite_code not in effective_placement:
            continue  # prereq không trong CTDT/customItems → skip (đã pass hoặc không liên quan)
        course_idx = effective_placement[p.course_code]
        prereq_idx = effective_placement[p.prerequisite_code]
        if prereq_idx >= course_idx:
            # Chỉ raise lỗi nếu vi phạm liên quan đến items user vừa gửi
            # (tránh raise lỗi cho default config cũ user chưa đụng vào)
            related = (p.course_code in custom_placement) or (p.prerequisite_code in custom_placement)
            if not related:
                continue
            raise HTTPException(
                status_code=422,
                detail=(
                    f"Môn {p.course_code} (HK{course_idx}) yêu cầu tiên quyết "
                    f"{p.prerequisite_code} phải học trước "
                    f"(hiện ở HK{prereq_idx})"
                ),
            )

    # Replace existing plan
    plan = db.query(models.StudyPlan).filter(
        models.StudyPlan.user_id == user.id,
        models.StudyPlan.plan_name == "roadmap_custom",
    ).first()
    if not plan:
        plan = models.StudyPlan(user_id=user.id, plan_name="roadmap_custom")
        db.add(plan)
        db.flush()
    else:
        db.query(models.StudyPlanItem).filter(
            models.StudyPlanItem.plan_id == plan.id
        ).delete()

    from datetime import datetime as _dtnow
    plan.updated_at = _dtnow.utcnow()
    for code, label in raw_items:
        db.add(models.StudyPlanItem(plan_id=plan.id, course_code=code, term_label=label))
    db.commit()
    return {"ok": True, "message": "Đã lưu lộ trình.", "saved_count": len(raw_items)}


# ── Roadmap Optimizer (delta-based, theo mục tiêu cá nhân) ─────────────────
@app.post("/v2/roadmap/optimize")
def optimize_roadmap(
    body: dict = Body(...),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Suggest delta moves to optimize roadmap toward goal.
    Body: {
      "goal": "early" | "balance" | "extend",
      "current_placement": {course_code: "HKn", ...}
    }
    Returns: { suggested_moves: [{course_code, course_name, from_term, to_term, reason}], stats }

    Algorithm:
    - early: try to move courses from last term earlier
    - balance: redistribute from overloaded HK (>22 TC) to lighter
    - extend: spread to more HKs, lower max load

    Constraints: locked (graded) courses don't move; prereqs must be in earlier HK; min 12 / max 22 TC per HK.
    """
    from collections import defaultdict
    user = _get_user_by_token(authorization, db)
    goal = body.get("goal", "balance")
    current = dict(body.get("current_placement", {}))  # mutable copy
    if goal not in ("early", "balance", "extend"):
        raise HTTPException(422, "goal phải là 'early' | 'balance' | 'extend'")

    # Locked: courses with grades
    grades = db.query(models.UserGrade).filter(models.UserGrade.user_id == user.id).all()
    locked_codes = {g.course_code for g in grades}

    # Prereqs map
    prereq_rows = db.query(models.CoursePrerequisite).all()
    prereq_map = defaultdict(list)
    for p in prereq_rows:
        prereq_map[p.course_code].append(p.prerequisite_code)

    # Course info
    courses_data = {c.course_code: c for c in db.query(models.Course).all()}

    def t_idx(t):
        try: return int(str(t).replace('HK', ''))
        except: return 99

    def credits(code):
        c = courses_data.get(code)
        return float(c.credits or 0) if c else 0.0

    def name(code):
        c = courses_data.get(code)
        return c.course_name if c else code

    # Compute TC per HK
    tc_per_hk = defaultdict(float)
    for code, term in current.items():
        if not str(term).startswith('HK'): continue
        tc_per_hk[t_idx(term)] += credits(code)

    def can_move(code, target_idx):
        """Check if moving 'code' to HK target_idx respects prereqs + dependents."""
        if code in locked_codes:
            return False, "đã có điểm"
        if target_idx < 1 or target_idx > 9:
            return False, "HK không hợp lệ"
        # All prereqs in earlier HK
        for p in prereq_map.get(code, []):
            p_term = current.get(p)
            if not p_term: continue
            if t_idx(p_term) >= target_idx:
                return False, f"vi phạm tiên quyết {p}"
        # All dependents (môn cần code) trong HK sau
        for dep_code, prereqs in prereq_map.items():
            if code not in prereqs: continue
            d_term = current.get(dep_code)
            if not d_term: continue
            if t_idx(d_term) <= target_idx:
                return False, f"môn {dep_code} cần môn này trước"
        return True, None

    def move_capacity(target_idx, code, max_tc=22):
        """Will moving code into target_idx exceed max_tc?"""
        return tc_per_hk[target_idx] + credits(code) <= max_tc

    moves = []

    if goal == "early":
        # Move courses from highest HK earlier to reduce total HK count
        max_hk = max((t_idx(t) for t in current.values() if str(t).startswith('HK')), default=9)
        # Try to empty HK 9 first, then HK 8 if possible
        for source_hk in [max_hk, max_hk - 1]:
            if source_hk < 1: break
            source_codes = [code for code, t in current.items() if t_idx(t) == source_hk and code not in locked_codes]
            for code in list(source_codes):
                # Try earliest possible target
                for try_idx in range(1, source_hk):
                    ok, reason = can_move(code, try_idx)
                    if ok and move_capacity(try_idx, code):
                        moves.append({
                            "course_code": code,
                            "course_name": name(code),
                            "credits": credits(code),
                            "from_term": f"HK{source_hk}",
                            "to_term": f"HK{try_idx}",
                            "reason": f"Tốt nghiệp sớm: chuyển lên HK{try_idx} ({tc_per_hk[try_idx]:.0f}+{credits(code):.0f} TC ≤ 22)",
                        })
                        tc_per_hk[try_idx] += credits(code)
                        tc_per_hk[source_hk] -= credits(code)
                        current[code] = f"HK{try_idx}"
                        break
                if len(moves) >= 8: break
            if len(moves) >= 8: break

    elif goal == "balance":
        # Find overloaded HK (>22 TC), move course out
        overloaded = [hk for hk in range(1, 10) if tc_per_hk.get(hk, 0) > 22]
        for hk in overloaded:
            attempts = 0
            while tc_per_hk.get(hk, 0) > 22 and attempts < 5:
                attempts += 1
                # Movable courses in this HK (sort by credits desc — move biggest first)
                movables = sorted(
                    [code for code, t in current.items() if t_idx(t) == hk and code not in locked_codes],
                    key=lambda c: -credits(c)
                )
                moved = False
                for code in movables:
                    # Find HK to move to (under 18 TC, valid)
                    for try_idx in range(1, 10):
                        if try_idx == hk: continue
                        ok, _ = can_move(code, try_idx)
                        if ok and tc_per_hk[try_idx] + credits(code) <= 22 and tc_per_hk[try_idx] < 18:
                            moves.append({
                                "course_code": code,
                                "course_name": name(code),
                                "credits": credits(code),
                                "from_term": f"HK{hk}",
                                "to_term": f"HK{try_idx}",
                                "reason": f"Cân bằng tải: HK{hk} đang {tc_per_hk[hk]:.0f} TC > 22. Chuyển sang HK{try_idx} ({tc_per_hk[try_idx]:.0f} TC).",
                            })
                            tc_per_hk[try_idx] += credits(code)
                            tc_per_hk[hk] -= credits(code)
                            current[code] = f"HK{try_idx}"
                            moved = True
                            break
                    if moved: break
                if not moved: break

    elif goal == "extend":
        # Spread courses to lower max load. Move from heaviest HK to lightest (if room).
        for _iter in range(8):
            heaviest_hk = max(range(1, 10), key=lambda h: tc_per_hk.get(h, 0))
            lightest_hk = min(range(1, 10), key=lambda h: tc_per_hk.get(h, 0))
            if tc_per_hk[heaviest_hk] - tc_per_hk[lightest_hk] < 4: break  # already balanced
            if tc_per_hk[heaviest_hk] <= 16: break  # already low
            # Try to move 1 course from heaviest to a non-heaviest
            movables = sorted(
                [code for code, t in current.items() if t_idx(t) == heaviest_hk and code not in locked_codes],
                key=lambda c: credits(c)  # smallest first (gentle move)
            )
            moved = False
            for code in movables:
                for try_idx in sorted(range(1, 10), key=lambda h: tc_per_hk.get(h, 0)):  # try lightest first
                    if try_idx == heaviest_hk: continue
                    ok, _ = can_move(code, try_idx)
                    if ok and tc_per_hk[try_idx] + credits(code) <= 18:
                        moves.append({
                            "course_code": code,
                            "course_name": name(code),
                            "credits": credits(code),
                            "from_term": f"HK{heaviest_hk}",
                            "to_term": f"HK{try_idx}",
                            "reason": f"Giãn tải: HK{heaviest_hk} ({tc_per_hk[heaviest_hk]:.0f} TC) → HK{try_idx} ({tc_per_hk[try_idx]:.0f} TC). Tổng tải đều hơn.",
                        })
                        tc_per_hk[try_idx] += credits(code)
                        tc_per_hk[heaviest_hk] -= credits(code)
                        current[code] = f"HK{try_idx}"
                        moved = True
                        break
                if moved: break
            if not moved: break

    # Compute stats
    final_max_hk = max((t_idx(t) for t in current.values() if str(t).startswith('HK')), default=9)
    final_max_tc = max(tc_per_hk.values()) if tc_per_hk else 0
    final_min_tc = min((v for k, v in tc_per_hk.items() if v > 0), default=0)

    return {
        "goal": goal,
        "suggested_moves": moves,
        "stats": {
            "max_hk_used": final_max_hk,
            "max_tc_per_hk": round(final_max_tc, 1),
            "min_tc_per_hk": round(final_min_tc, 1),
            "total_moves": len(moves),
        },
    }


# ── Analytics ─────────────────────────────────────────────────────────────────

@app.get("/analytics/me", response_model=schemas.AnalyticsOut)
def analytics_me(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    return schemas.AnalyticsOut(**build_analytics(db, user.id))


@app.get("/track-switch-cost/me")
def track_switch_cost(
    target: str,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    current_spec = user.specialization or ""
    target_spec = target.strip()

    if current_spec == target_spec:
        raise HTTPException(status_code=400, detail="Chuyên ngành mục tiêu giống chuyên ngành hiện tại")

    # Passed courses
    grades = db.query(models.UserGrade).filter(
        models.UserGrade.user_id == user.id,
        models.UserGrade.score10 >= 5.0,
    ).all()
    score_map: dict[str, float] = {}
    for g in grades:
        score_map[g.course_code] = max(score_map.get(g.course_code, 0.0), float(g.score10 or 0))

    all_courses = {c.course_code: c for c in db.query(models.Course).all()}

    # Current spec elective credits already earned
    cur_elec_mappings = db.query(models.CourseElectiveGroup).filter(
        models.CourseElectiveGroup.specialization == current_spec
    ).all()
    cur_elec_codes = {m.course_code for m in cur_elec_mappings}

    # Target spec elective rules and course mappings
    target_rules = db.query(models.ElectiveRule).filter(
        models.ElectiveRule.specialization == target_spec
    ).order_by(models.ElectiveRule.group_type).all()

    target_mappings = db.query(models.CourseElectiveGroup).filter(
        models.CourseElectiveGroup.specialization == target_spec
    ).all()
    # group_type -> set of course_codes
    target_group_courses: dict[str, set] = {}
    for m in target_mappings:
        target_group_courses.setdefault(m.group_type, set()).add(m.course_code)

    # Spec-specific required courses (non-elective) for target
    target_required = db.query(models.Course).filter(
        models.Course.required_specialization == target_spec
    ).all()
    cur_required = db.query(models.Course).filter(
        models.Course.required_specialization == current_spec
    ).all()

    # Elective group analysis
    elective_analysis = []
    total_new_elective_needed = 0.0
    for rule in target_rules:
        group = rule.group_type
        group_codes = target_group_courses.get(group, set())
        # Credits already earned that count toward this group
        earned = sum(
            float(all_courses[c].credits) for c in group_codes
            if c in score_map and c in all_courses
        )
        earned = min(earned, float(rule.min_credits_required))
        needed = max(0.0, float(rule.min_credits_required) - earned)
        total_new_elective_needed += needed
        transferable = [
            {"code": c, "name": all_courses[c].course_name, "credits": float(all_courses[c].credits)}
            for c in group_codes if c in score_map and c in all_courses
        ]
        elective_analysis.append({
            "group_type": group,
            "min_required": float(rule.min_credits_required),
            "already_earned": round(earned, 1),
            "still_needed": round(needed, 1),
            "done": needed == 0,
            "transferable_courses": transferable[:5],
        })

    # Required courses unique to target spec (not in current)
    cur_req_codes = {c.course_code for c in cur_required}
    new_required = [
        {"code": c.course_code, "name": c.course_name, "credits": float(c.credits), "already_passed": c.course_code in score_map}
        for c in target_required if c.course_code not in cur_req_codes
    ]
    new_required_credits = sum(r["credits"] for r in new_required if not r["already_passed"])

    # Estimate extra semesters
    total_extra = total_new_elective_needed + new_required_credits
    p = db.query(models.StudyPlan).filter(models.StudyPlan.user_id == user.id).first()
    avg_tc = 15.0  # fallback
    extra_semesters = round(total_extra / avg_tc, 1) if total_extra > 0 else 0

    return {
        "current_specialization": current_spec,
        "target_specialization": target_spec,
        "elective_groups": elective_analysis,
        "new_required_courses": new_required,
        "total_new_elective_credits": round(total_new_elective_needed, 1),
        "total_new_required_credits": round(new_required_credits, 1),
        "total_extra_credits": round(total_extra, 1),
        "estimated_extra_semesters": extra_semesters,
    }


@app.get("/career-fit/me")
def get_career_fit(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    from backend.core.academic_engine import SPECIALIZATION_TRACKS, _normalize_text

    user = _get_user_by_token(authorization, db)

    grades = db.query(models.UserGrade).filter(
        models.UserGrade.user_id == user.id,
        models.UserGrade.score10 >= 5.0,
    ).all()
    if not grades:
        return []

    all_courses = {c.course_code: c for c in db.query(models.Course).all()}

    # Best score per course (cast to float to avoid Decimal*float TypeError)
    score_map: dict[str, float] = {}
    for g in grades:
        s10 = float(g.score10 or 0.0)
        score_map[g.course_code] = max(score_map.get(g.course_code, 0.0), s10)

    overall_avg = sum(score_map.values()) / len(score_map) if score_map else 5.0

    def _career_match(keywords: list[str]) -> tuple[float, int, list[str]]:
        matched_scores, matched_names = [], []
        for code, score in score_map.items():
            course = all_courses.get(code)
            if not course:
                continue
            norm = _normalize_text(course.course_name)
            if any(kw in norm for kw in keywords):
                matched_scores.append(score)
                matched_names.append(course.course_name)
        if not matched_scores:
            return overall_avg * 0.6, 0, []
        return sum(matched_scores) / len(matched_scores), len(matched_scores), matched_names[:4]

    results = []
    for spec_name, tracks in SPECIALIZATION_TRACKS.items():
        for track_key, info in tracks.items():
            avg_in_field, matched_count, matched_names = _career_match(info["keywords"])
            blended = avg_in_field * 0.65 + overall_avg * 0.35
            fit_pct = min(99, max(10, round((blended / 10) * 100)))
            results.append({
                "specialization": spec_name,
                "career_key": track_key,
                "career_label": info["label"],
                "icon": info["icon"],
                "description": info["desc"],
                "fit_percent": fit_pct,
                "matched_courses_count": matched_count,
                "matched_course_names": matched_names,
                "avg_score_in_field": round(avg_in_field, 1),
                "is_current_spec": spec_name == (user.specialization or ""),
            })

    results.sort(key=lambda x: (-x["fit_percent"], not x["is_current_spec"]))
    return results


# ── Career Learning Path (new feature) ────────────────────────────────────────

@app.get("/career-path/tracks")
def get_career_path_tracks():
    """All career tracks with skills and resources — no auth needed."""
    from backend.career_data import CAREER_TRACKS
    out = []
    for track_id, t in CAREER_TRACKS.items():
        skills = t["skills"]
        out.append({
            "id": track_id,
            "name": t["name"],
            "icon": t["icon"],
            "color": t["color"],
            "desc": t["desc"],
            "total_skills": len(skills),
            "school_covered": sum(1 for s in skills if s.get("school_covered")),
            "skills": skills,
        })
    return out


@app.get("/career-path/me")
def get_my_career_path(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """ML-scored career fit + per-skill progress for the current user."""
    from backend.career_data import score_career_fit, CAREER_TRACKS

    user = _get_user_by_token(authorization, db)

    grades = db.query(models.UserGrade).filter(models.UserGrade.user_id == user.id).all()
    grade_list = [{"course_code": g.course_code, "score10": float(g.score10 or 0), "passed": g.passed} for g in grades]

    fit_scores = score_career_fit(grade_list)

    progress_rows = db.query(models.UserSkillProgress).filter(
        models.UserSkillProgress.user_id == user.id
    ).all()
    skill_status: dict[str, str] = {row.skill_id: row.status for row in progress_rows}

    tracks_out = []
    for track_id, t in CAREER_TRACKS.items():
        skills_flat = t["skills"]
        skills_with_status = [
            {**s, "status": skill_status.get(f"{track_id}.{s['id']}", "not_started")}
            for s in skills_flat
        ]
        school_covered = sum(1 for s in skills_flat if s.get("school_covered"))
        completed = sum(1 for s in skills_flat if skill_status.get(f"{track_id}.{s['id']}") == "completed")
        in_progress = sum(1 for s in skills_flat if skill_status.get(f"{track_id}.{s['id']}") == "in_progress")

        tracks_out.append({
            "id": track_id,
            "name": t["name"],
            "icon": t["icon"],
            "color": t["color"],
            "desc": t["desc"],
            "fit_score": fit_scores.get(track_id, 0),
            "total_skills": len(skills_flat),
            "school_covered": school_covered,
            "completed": completed,
            "in_progress": in_progress,
            "skills": skills_with_status,
        })

    tracks_out.sort(key=lambda x: -x["fit_score"])
    recommended = tracks_out[0]["id"] if tracks_out else None
    return {"tracks": tracks_out, "recommended_track": recommended}


@app.patch("/career-path/me/skills/{skill_key:path}")
def update_career_skill(
    skill_key: str,
    payload: dict,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Update skill progress. skill_key = 'track_id.skill_id'"""
    user = _get_user_by_token(authorization, db)
    status = payload.get("status", "not_started")
    if status not in ("not_started", "in_progress", "completed"):
        raise HTTPException(status_code=400, detail="Invalid status value")

    row = db.query(models.UserSkillProgress).filter(
        models.UserSkillProgress.user_id == user.id,
        models.UserSkillProgress.skill_id == skill_key,
    ).first()
    if row:
        row.status = status
    else:
        db.add(models.UserSkillProgress(user_id=user.id, skill_id=skill_key, status=status))
    db.commit()
    return {"skill_id": skill_key, "status": status}


@app.get("/courses/bottleneck/me")
def get_bottleneck_courses(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)

    grades = db.query(models.UserGrade).filter(
        models.UserGrade.user_id == user.id,
        models.UserGrade.score10 >= 5.0,
    ).all()
    passed_codes = {g.course_code for g in grades}

    prereqs = db.query(models.CoursePrerequisite).all()

    # prerequisite_code → set of courses that require it
    dependents: dict[str, set] = {}
    for p in prereqs:
        dependents.setdefault(p.prerequisite_code, set()).add(p.course_code)

    all_courses = {c.course_code: c for c in db.query(models.Course).all()}

    def count_blocked(code: str) -> int:
        visited: set[str] = set()
        queue = list(dependents.get(code, set()))
        while queue:
            cur = queue.pop()
            if cur in visited:
                continue
            visited.add(cur)
            queue.extend(dependents.get(cur, set()))
        return len(visited)

    bottlenecks = []
    for prereq_code, deps in dependents.items():
        if prereq_code in passed_codes:
            continue
        course = all_courses.get(prereq_code)
        if not course:
            continue
        blocks = count_blocked(prereq_code)
        if blocks == 0:
            continue
        direct_blocked = [c for c in deps if c not in passed_codes]
        bottlenecks.append({
            "course_code": prereq_code,
            "course_name": course.course_name,
            "credits": course.credits,
            "blocks_count": blocks,
            "direct_blocked": direct_blocked[:5],
        })

    bottlenecks.sort(key=lambda x: x["blocks_count"], reverse=True)
    return bottlenecks[:8]


# ── Spec overview (public — no auth needed) ────────────────────────────────────

@app.get("/courses/spec-overview")
def get_spec_overview(specialization: str, db: Session = Depends(get_db)):
    """Return required + elective courses for a given specialization."""
    from sqlalchemy import or_
    required = db.query(models.Course).filter(
        models.Course.required_specialization == specialization
    ).order_by(models.Course.course_code).all()

    elective_mappings = db.query(models.CourseElectiveGroup).filter(
        or_(
            models.CourseElectiveGroup.specialization == specialization,
            models.CourseElectiveGroup.specialization == "Chung",
        )
    ).order_by(models.CourseElectiveGroup.group_type).all()

    rules = db.query(models.ElectiveRule).filter(
        or_(
            models.ElectiveRule.specialization == specialization,
            models.ElectiveRule.specialization == "Chung",
        )
    ).all()
    rule_map = {r.group_type: float(r.min_credits_required) for r in rules}

    course_lookup = {c.course_code: c for c in db.query(models.Course).all()}
    groups: dict[str, list] = {}
    for m in elective_mappings:
        groups.setdefault(m.group_type, [])
        c = course_lookup.get(m.course_code)
        if c:
            groups[m.group_type].append({
                "course_code": c.course_code,
                "course_name": c.course_name,
                "credits": float(c.credits) if c.credits else 0,
            })

    return {
        "specialization": specialization,
        "required_courses": [
            {"course_code": c.course_code, "course_name": c.course_name, "credits": float(c.credits) if c.credits else 0}
            for c in required
        ],
        "elective_groups": [
            {"group_type": k, "min_credits": rule_map.get(k, 0), "courses": v}
            for k, v in groups.items()
        ],
    }


# ── Course Standard Plan ──────────────────────────────────────────────────────

# HK7-9 course codes grouped by specialization (derived from CURRICULUM_ORDER comments).
# IMPORTANT: must match academic_engine.CURRICULUM_ORDER exactly for HK7-9, otherwise
# the standard-plan filter in get_standard_plan() leaks courses from other CNs (a code not
# in _ALL_SPEC_CODES short-circuits the filter check `code in _ALL_SPEC_CODES`).
_SPEC_HK7_PLUS: dict[str, set[str]] = {
    # CHỈ chứa môn ĐẶC THÙ của CN (không có 7080504 — đó là cơ sở ngành CHUNG HK7).
    # 7080504 nằm trong CURRICULUM_ORDER nhưng KHÔNG vào _ALL_SPEC_CODES → filter
    # `code in _ALL_SPEC_CODES` False → môn này luôn xuất hiện trong mọi spec view.
    "khmt":   {"7080508","7080515","7080510","7080506","7080513","7080519"},
    "cnttdh": {"7080313","7050303","7080309","7080403","7080311","7080314"},
    "thkt":   {"7080633","7080616","7080638","7080603","7080617","7080604"},
    "mmt":    {"7080721","7080728","7080720","7080729","7080715","7080723"},
    "httt":   {"7080212","7080213","7080204","7080210","7080224","7080218"},
    "cnpm":   {"7080104","7080108","7080114","7080102","7080106","7080119","7080110"},
}
_ALL_SPEC_CODES: set[str] = {c for s in _SPEC_HK7_PLUS.values() for c in s}

# Map spec code (7480201_07) → spec key (khmt) used in _SPEC_HK7_PLUS
_SPEC_CODE_TO_KEY: dict[str, str] = {
    "7480201_07": "khmt",
    "7480201_06": "mmt",
    "7480201_05": "cnpm",
    "7480201_09": "httt",
    "7480201_04": "thkt",
    "7480201_08": "cnttdh",
}

# Fixed placeholder slot distribution per pool (matches CTDT chuẩn KHMT layout)
# Format: (group_type, [(semester, count), ...], credits_per_slot)
_POOL_PLACEHOLDER_LAYOUT = {
    "A": ([(4, 1), (5, 1)], 3.0),  # Pool A: 2 slots × 3 TC = 6 TC (min)
    "B": ([(7, 3)], 3.0),           # Pool B: 3 slots × 3 TC = 9 TC (min)
    "C": ([(8, 3)], 3.0),           # Pool C: 3 slots × 3 TC = 9 TC (min)
}


@app.get("/courses/prerequisites")
def list_all_prerequisites(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Return all course prereqs as {course_code: [prereq_codes]}.
    Used by frontend roadmap to validate drag-drop ordering."""
    _get_user_by_token(authorization, db)
    from collections import defaultdict as _dd
    rows = db.query(models.CoursePrerequisite).all()
    result: dict[str, list[str]] = _dd(list)
    for r in rows:
        result[r.course_code].append(r.prerequisite_code)
    return dict(result)


@app.get("/courses/standard-plan")
def get_standard_plan(
    spec: str = Query(default=""),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Kế hoạch đào tạo chuẩn theo CURRICULUM_ORDER + placeholder Tự chọn A/B/C.

    Hiển thị môn bắt buộc + slot "Tự chọn A/B/C" theo CTDT chuẩn (KHÔNG liệt kê môn cụ thể của pool).
    Spec accepted both: code "7480201_07" hoặc key "khmt".
    """
    _get_user_by_token(authorization, db)

    all_courses: dict[str, models.Course] = {
        c.course_code: c for c in db.query(models.Course).all()
    }
    raw_spec = (spec or "").strip().lower()
    spec_key = _SPEC_CODE_TO_KEY.get(raw_spec.upper().replace(raw_spec, raw_spec), raw_spec)
    if raw_spec in _SPEC_CODE_TO_KEY:
        spec_key = _SPEC_CODE_TO_KEY[raw_spec]
    spec_code = next((c for c, k in _SPEC_CODE_TO_KEY.items() if k == spec_key), None)
    if not spec_code and raw_spec.startswith("7480201_"):
        spec_code = raw_spec  # fallback: dùng raw spec làm code

    allowed_spec_codes: set[str] = (
        _SPEC_HK7_PLUS.get(spec_key, set()) if spec_key else _ALL_SPEC_CODES
    )

    # Build prereqs map
    prereqs_map: dict[str, list[str]] = {}
    for p in db.query(models.CoursePrerequisite).all():
        prereqs_map.setdefault(p.course_code, []).append(p.prerequisite_code)

    # Tập course_code thuộc elective pool (cần loại khỏi std plan vì là alternatives)
    pool_member_codes: set[str] = set()
    if spec_code:
        for eg in db.query(models.CourseElectiveGroup).filter(
            models.CourseElectiveGroup.specialization.in_([spec_code, "Chung"])
        ).all():
            pool_member_codes.add(eg.course_code)

    # Đọc min_credits per pool để biết pool nào có placeholder
    pool_min_credits: dict[str, float] = {}
    if spec_code:
        for r in db.query(models.ElectiveRule).filter(
            models.ElectiveRule.specialization.in_([spec_code, "Chung"])
        ).all():
            g = (r.group_type or "").upper()
            if g in {"A", "B", "C"}:
                pool_min_credits[g] = max(pool_min_credits.get(g, 0), float(r.min_credits_required))

    semesters: dict[int, list] = {i: [] for i in range(1, 10)}

    # 1. Compulsory courses (filter pool members + spec mismatch)
    for code, sem_num in academic_engine.CURRICULUM_ORDER.items():
        if sem_num >= 7 and code in _ALL_SPEC_CODES and code not in allowed_spec_codes:
            continue
        if code in pool_member_codes:
            continue  # Pool members là alternatives, không phải std plan
        course = all_courses.get(code)
        if not course:
            continue
        semesters[sem_num].append({
            "course_code": code,
            "course_name": course.course_name,
            "credits": float(course.credits) if course.credits else 0,
            "count_toward_credits": bool(course.count_toward_credits),
            "required_specialization": course.required_specialization,
            "prereqs": prereqs_map.get(code, []),
            "group": "compulsory",
        })

    # 2. Placeholder Tự chọn A/B/C (chỉ khi spec có pool tương ứng với min_credits > 0)
    for pool_group, (slot_layout, credits_per_slot) in _POOL_PLACEHOLDER_LAYOUT.items():
        if pool_min_credits.get(pool_group, 0) <= 0:
            continue
        for hk, count in slot_layout:
            for i in range(1, count + 1):
                semesters[hk].append({
                    "course_code": f"__POOL_{pool_group}_{hk}_{i}__",
                    "course_name": f"Tự chọn {pool_group}",
                    "credits": credits_per_slot,
                    "count_toward_credits": True,
                    "required_specialization": spec_code,
                    "prereqs": [],
                    "group": pool_group.lower(),  # "a"/"b"/"c" cho frontend color-code
                    "is_placeholder": True,
                })

    result_sems = [
        {
            "semester_number": k,
            "total_credits": round(sum(c["credits"] for c in v if c["count_toward_credits"]), 1),
            "courses": sorted(v, key=lambda c: (
                # Compulsory đứng trước, placeholder đứng sau theo group
                {"compulsory": 0, "a": 1, "b": 2, "c": 3}.get(c.get("group", "compulsory"), 9),
                c["course_code"],
            )),
        }
        for k, v in sorted(semesters.items()) if v
    ]
    total = round(sum(s["total_credits"] for s in result_sems), 1)

    return {
        "semesters": result_sems,
        "total_credits": total,
        "specialization": spec_key or "all",
        "pool_min_credits": pool_min_credits,
    }


# ── Course Catalog ─────────────────────────────────────────────────────────────

@app.get("/courses/catalog/me")
def get_course_catalog(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Enriched course list for the student catalog: completion status, score, prereqs, curriculum semester."""
    user = _get_user_by_token(authorization, db)

    all_courses = db.query(models.Course).order_by(models.Course.course_code.asc()).all()

    grades = db.query(models.UserGrade).filter(models.UserGrade.user_id == user.id).all()
    best: dict[str, models.UserGrade] = {}
    for g in grades:
        prev = best.get(g.course_code)
        if prev is None or (g.score10 or 0) > (prev.score10 or 0):
            best[g.course_code] = g

    prereqs_all = db.query(models.CoursePrerequisite).all()
    prereq_map: dict[str, list[str]] = {}
    for p in prereqs_all:
        prereq_map.setdefault(p.course_code, []).append(p.prerequisite_code)

    # courses that unlock after a given course (direct dependents)
    unlocks_map: dict[str, list[str]] = {}
    for p in prereqs_all:
        unlocks_map.setdefault(p.prerequisite_code, []).append(p.course_code)

    completed_codes = {code for code, g in best.items() if g.passed}

    # Build elective group map: course_code → group_type (first match for user's spec or Chung)
    spec = user.specialization or ""
    elective_group_map: dict[str, str] = {}
    elective_rows = db.query(models.CourseElectiveGroup).filter(
        (models.CourseElectiveGroup.specialization == spec) |
        (models.CourseElectiveGroup.specialization == "Chung"),
    ).all()
    for em in elective_rows:
        # Prefer user's specific spec over Chung
        if em.course_code not in elective_group_map or em.specialization == spec:
            elective_group_map[em.course_code] = em.group_type

    result = []
    for c in all_courses:
        code = c.course_code
        g = best.get(code)
        if g:
            if g.passed:
                status = "completed"
            else:
                status = "failed"
        else:
            status = "not_started"

        prereq_codes = prereq_map.get(code, [])
        prereq_info = []
        for pcode in prereq_codes:
            pc = db.query(models.Course).filter(models.Course.course_code == pcode).first()
            prereq_info.append({
                "course_code": pcode,
                "course_name": pc.course_name if pc else pcode,
                "completed": pcode in completed_codes,
            })

        unlock_codes = unlocks_map.get(code, [])
        unlock_names = []
        for ucode in unlock_codes[:5]:
            uc = db.query(models.Course).filter(models.Course.course_code == ucode).first()
            unlock_names.append({"course_code": ucode, "course_name": uc.course_name if uc else ucode})

        result.append({
            "course_code": code,
            "course_name": c.course_name,
            "credits": float(c.credits) if c.credits else None,
            "description": c.description,
            "required_specialization": c.required_specialization,
            "count_toward_credits": c.count_toward_credits,
            "curriculum_semester": academic_engine.CURRICULUM_ORDER.get(code, 0),
            "status": status,
            "score10": float(g.score10) if g and g.score10 else None,
            "letter": g.letter if g else None,
            "prerequisites": prereq_info,
            "unlocks": unlock_names,
            "prereqs_met": all(p in completed_codes for p in prereq_codes),
            "elective_group": elective_group_map.get(code),  # 'A','B','C' or None
        })

    return result


# ── ML Ranking: xếp hạng môn tự chọn theo profile + career ────────────────────

@app.get("/courses/rank-for-me")
def rank_courses_for_me(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """ML-based ranking for elective selection.

    Combines 3 signals per course:
    - career_alignment: cosine similarity giữa course domains và primary career's domain_profile
    - history_fit: user's strength in course's domains (weighted avg score10 của môn cùng domain đã học)
    - peer_rating: avg rating từ course_ratings (peer signal)

    Final score = 0.5*career + 0.3*history + 0.2*rating (normalized to [0,1]).
    """
    user = _get_user_by_token(authorization, db)

    # User's domain strengths from grades
    user_profile = _user_domain_profile(db, user.id)

    # Primary career's domain_profile (if chosen)
    choice = db.query(models.UserCareerChoice).filter(
        models.UserCareerChoice.user_id == user.id
    ).first()
    career_profile: dict = {}
    career_code: str | None = None
    career_name: str | None = None
    if choice and choice.primary_path_id:
        path = db.query(models.CareerPath).filter(
            models.CareerPath.id == choice.primary_path_id
        ).first()
        if path:
            career_profile = path.domain_profile or {}
            career_code = path.code
            career_name = path.name

    # Peer ratings per course
    from sqlalchemy import func as _sf
    ratings = db.query(
        models.CourseRating.course_code,
        _sf.avg(models.CourseRating.rating).label("avg_r"),
        _sf.count(models.CourseRating.id).label("count_r"),
    ).group_by(models.CourseRating.course_code).all()
    rating_map = {r.course_code: (float(r.avg_r or 0), int(r.count_r or 0)) for r in ratings}

    # All courses
    all_courses = db.query(models.Course).all()
    results: list[dict] = []

    def _cos(a: dict, b: dict) -> float:
        import math
        keys = set(a.keys()) | set(b.keys())
        dot = sum(a.get(k, 0.0) * float(b.get(k, 0.0)) for k in keys)
        na = math.sqrt(sum(v * v for v in a.values()))
        nb = math.sqrt(sum(float(v) * float(v) for v in b.values()))
        return dot / (na * nb) if na and nb else 0.0

    for c in all_courses:
        domains = _infer_course_domains(c.course_name or "")

        # Career alignment
        career_align = _cos(domains, career_profile) if career_profile and domains else 0.0

        # History fit — averaged user's strength across this course's domains
        if domains and user_profile:
            strengths = [user_profile.get(d, 0.0) for d in domains]
            history_fit = sum(strengths) / len(strengths) if strengths else 0.0
        else:
            history_fit = 0.0

        # Peer rating (normalize 1-5 → 0-1)
        avg_r, count_r = rating_map.get(c.course_code, (0.0, 0))
        rating_norm = ((avg_r - 1) / 4) if avg_r > 0 else 0.0

        # Weighted composite
        final = 0.5 * career_align + 0.3 * history_fit + 0.2 * rating_norm

        # Build reason tags
        tags: list[str] = []
        if career_align >= 0.6: tags.append("career")
        if history_fit >= 0.7: tags.append("strong_fit")
        elif history_fit <= 0.3 and user_profile: tags.append("weak_area")
        if count_r >= 5 and avg_r >= 4.0: tags.append("peer_loved")
        if not domains: tags.append("no_domain_match")

        results.append({
            "course_code": c.course_code,
            "ml_score": round(final, 3),
            "ml_percent": round(final * 100, 1),
            "breakdown": {
                "career_alignment": round(career_align, 3),
                "history_fit": round(history_fit, 3),
                "peer_rating": round(rating_norm, 3),
                "avg_rating_stars": round(avg_r, 1) if avg_r > 0 else None,
                "rating_count": count_r,
            },
            "tags": tags,
            "reason": _build_ml_reason(career_name, tags, career_align, history_fit, domains),
        })

    results.sort(key=lambda x: x["ml_score"], reverse=True)
    return {
        "career_code": career_code,
        "career_name": career_name,
        "user_has_grades": bool(user_profile),
        "courses": results,
    }


def _build_ml_reason(career_name: str | None, tags: list[str], career_align: float,
                     history_fit: float, domains: dict) -> str:
    """Build a short Vietnamese reason for a course's ranking."""
    if not domains:
        return "Chưa xếp hạng — không nhận diện được domain"
    parts: list[str] = []
    if "career" in tags and career_name:
        parts.append(f"phù hợp định hướng {career_name}")
    if "strong_fit" in tags:
        parts.append("tận dụng điểm mạnh của bạn")
    elif "weak_area" in tags:
        parts.append("thử thách lĩnh vực bạn chưa mạnh")
    if "peer_loved" in tags:
        parts.append("sinh viên khác đánh giá cao")
    if not parts:
        if career_align > 0.3:
            parts.append("liên quan định hướng nghề")
        else:
            parts.append("môn tự chọn chung")
    return " · ".join(parts).capitalize()


# ── Grade Simulator (SV đặt mục tiêu + mô phỏng điểm kỳ tới) ──────────────────

@app.get("/simulator/snapshot/me")
def simulator_current_snapshot(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Trả về trạng thái hiện tại + danh sách môn kỳ tới (lấy từ roadmap custom, fallback engine)."""
    user = _get_user_by_token(authorization, db)
    snap = academic_engine._build_snapshot(db, user.id, specialization=user.specialization)
    # Note: official_earned_credits đã drop. earned_credits tính từ grades.

    # Pull planned courses for next term from custom roadmap first
    custom_plan = db.query(models.StudyPlan).filter(
        models.StudyPlan.user_id == user.id,
        models.StudyPlan.plan_name == "roadmap_custom",
    ).first()
    next_term_courses: list[dict] = []
    source = "auto_engine"
    if custom_plan:
        items = db.query(models.StudyPlanItem).filter(
            models.StudyPlanItem.plan_id == custom_plan.id
        ).order_by(models.StudyPlanItem.term_label, models.StudyPlanItem.id).all()
        # Find smallest (earliest) term label — skip kho sentinel
        terms = sorted({it.term_label or "" for it in items if it.term_label and it.term_label != "__KHO__"})
        if terms:
            first_term = terms[0]
            first_items = [it for it in items if (it.term_label or "") == first_term]
            codes = [it.course_code for it in first_items]
            course_map = {
                c.course_code: c for c in db.query(models.Course).filter(
                    models.Course.course_code.in_(codes)
                ).all()
            }
            for it in first_items:
                c = course_map.get(it.course_code)
                if not c:
                    continue
                next_term_courses.append({
                    "course_code": it.course_code,
                    "course_name": c.course_name,
                    "credits": float(c.credits) if c.credits else 0.0,
                    "term_label": first_term,
                })
            source = "custom_roadmap"

    # Fallback: use auto-generated roadmap
    if not next_term_courses:
        rm = academic_engine.build_semester_roadmap(db, user.id)
        sems = rm.get("semesters", [])
        if sems:
            first = sems[0]
            for c in first.get("courses", []):
                if str(c.get("course_code", "")).startswith("__"):
                    continue
                next_term_courses.append({
                    "course_code": c["course_code"],
                    "course_name": c.get("course_name", c["course_code"]),
                    "credits": float(c.get("credits") or 0),
                    "term_label": first.get("semester_label", "HK tới"),
                })

    grad_threshold = academic_engine._get_graduation_threshold(db)

    # Build list of remaining_terms (multi-semester simulator)
    remaining_terms: list[dict] = []
    if custom_plan:
        items_all = db.query(models.StudyPlanItem).filter(
            models.StudyPlanItem.plan_id == custom_plan.id
        ).order_by(models.StudyPlanItem.term_label, models.StudyPlanItem.id).all()
        from collections import defaultdict as _dd
        _term_buckets: dict[str, list] = _dd(list)
        for it in items_all:
            _term_buckets[it.term_label or ""].append(it)
        all_codes = {it.course_code for it in items_all}
        course_map_all = {
            c.course_code: c for c in db.query(models.Course).filter(
                models.Course.course_code.in_(all_codes)
            ).all()
        }
        for term_label in sorted(_term_buckets.keys()):
            # Skip kho sentinel — môn ở kho chưa xếp HK, không vào simulator
            if term_label == "__KHO__":
                continue
            tcourses = []
            for it in _term_buckets[term_label]:
                c = course_map_all.get(it.course_code)
                if not c:
                    continue
                tcourses.append({
                    "course_code": it.course_code,
                    "course_name": c.course_name,
                    "credits": float(c.credits) if c.credits else 0.0,
                })
            if tcourses:
                remaining_terms.append({
                    "term_label": term_label,
                    "courses": tcourses,
                    "total_credits": round(sum(c["credits"] for c in tcourses), 1),
                })
    else:
        # Fallback: engine roadmap (skip __ELECTIVE placeholders)
        rm_full = academic_engine.build_semester_roadmap(db, user.id)
        for sem in rm_full.get("semesters", []):
            tcourses = []
            for c in sem.get("courses", []):
                if str(c.get("course_code", "")).startswith("__"):
                    continue
                tcourses.append({
                    "course_code": c["course_code"],
                    "course_name": c.get("course_name", c["course_code"]),
                    "credits": float(c.get("credits") or 0),
                })
            if tcourses:
                remaining_terms.append({
                    "term_label": sem.get("semester_label", ""),
                    "courses": tcourses,
                    "total_credits": round(sum(c["credits"] for c in tcourses), 1),
                })

    return {
        "current": {
            "avg_score4": round(snap.avg_score4, 2) if snap.avg_score4 is not None else None,
            "avg_score10": round(snap.avg_score10, 2) if snap.avg_score10 is not None else None,
            "earned_credits": snap.earned_credits,
            "graduation_threshold": grad_threshold,
            "remaining_credits": max(0.0, grad_threshold - snap.earned_credits),
            "classification": academic_engine.classify_gpa4(snap.avg_score4),
        },
        "next_term": {
            "term_label": next_term_courses[0]["term_label"] if next_term_courses else None,
            "source": source,
            "courses": next_term_courses,
            "total_credits": round(sum(c["credits"] for c in next_term_courses), 1),
        },
        "remaining_terms": remaining_terms,
        "source": source,
    }


@app.post("/simulator/target-gpa")
def simulator_target_gpa(
    payload: dict = Body(...),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Tính điểm trung bình cần đạt ở các kỳ còn lại để đạt target GPA tổng kết.

    Input: { target_gpa4: float, remaining_credits: float (optional) }
    Output: required avg score4 + score10 + difficulty label + per-letter hint.
    """
    user = _get_user_by_token(authorization, db)
    try:
        target_gpa4 = float(payload.get("target_gpa4"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=422, detail="target_gpa4 không hợp lệ")
    if target_gpa4 < 0 or target_gpa4 > 4.0:
        raise HTTPException(status_code=422, detail="target_gpa4 phải trong [0, 4.0]")

    snap = academic_engine._build_snapshot(db, user.id, specialization=user.specialization)
    # Note: official_earned_credits đã drop. earned_credits tính từ grades.

    grad_threshold = academic_engine._get_graduation_threshold(db)
    current_avg4 = snap.avg_score4 or 0.0
    earned = snap.earned_credits or 0.0
    # Default remaining: threshold - earned (cap at 0)
    remaining = payload.get("remaining_credits")
    if remaining is None:
        remaining = max(0.0, grad_threshold - earned)
    else:
        try: remaining = float(remaining)
        except (TypeError, ValueError): remaining = max(0.0, grad_threshold - earned)

    if remaining <= 0:
        return {
            "achievable": True,
            "already_met": current_avg4 >= target_gpa4,
            "current_gpa4": round(current_avg4, 2),
            "target_gpa4": target_gpa4,
            "remaining_credits": 0,
            "required_avg_score4": None,
            "required_avg_score10": None,
            "difficulty": "not_applicable",
            "message": "Bạn đã đủ tín chỉ tốt nghiệp. Không còn kỳ nào để cải thiện GPA.",
        }

    # target_gpa = (current_avg × earned + required_avg × remaining) / (earned + remaining)
    # required_avg = (target_gpa × (earned+remaining) - current_avg × earned) / remaining
    total_credits = earned + remaining
    required_avg4 = (target_gpa4 * total_credits - current_avg4 * earned) / remaining

    achievable = required_avg4 <= 4.0
    required_avg4_capped = max(0.0, min(4.0, required_avg4))

    # Inverse mapping score4 → score10 (approx midpoint of each band)
    def _s4_to_s10(s4: float) -> float:
        if s4 >= 3.85: return 9.25
        if s4 >= 3.5:  return 8.5
        if s4 >= 3.25: return 8.25
        if s4 >= 3.0:  return 7.5
        if s4 >= 2.5:  return 6.75
        if s4 >= 2.0:  return 5.75
        if s4 >= 1.5:  return 5.25
        if s4 >= 1.0:  return 4.5
        return 3.0
    required_avg10 = _s4_to_s10(required_avg4_capped)

    if not achievable:
        difficulty = "impossible"
        message = f"Không thể đạt GPA {target_gpa4} với {remaining} TC còn lại (cần TB {round(required_avg4,2)}/4 > 4.0)."
    elif required_avg4_capped >= 3.5:
        difficulty = "very_hard"
        message = f"Cần trung bình {academic_engine.letter_from_score10(required_avg10)} ({required_avg10:.1f}/10) — rất khó."
    elif required_avg4_capped >= 3.0:
        difficulty = "hard"
        message = f"Cần trung bình {academic_engine.letter_from_score10(required_avg10)} ({required_avg10:.1f}/10) — cần nỗ lực cao."
    elif required_avg4_capped >= 2.0:
        difficulty = "moderate"
        message = f"Cần trung bình {academic_engine.letter_from_score10(required_avg10)} ({required_avg10:.1f}/10) — vừa sức."
    else:
        difficulty = "easy"
        message = f"Mục tiêu dễ đạt — chỉ cần trung bình {required_avg10:.1f}/10."

    return {
        "achievable": achievable,
        "already_met": current_avg4 >= target_gpa4,
        "current_gpa4": round(current_avg4, 2),
        "target_gpa4": target_gpa4,
        "remaining_credits": remaining,
        "required_avg_score4": round(required_avg4_capped, 2),
        "required_avg_score10": round(required_avg10, 2),
        "difficulty": difficulty,
        "message": message,
    }


@app.post("/simulator/what-if")
def simulator_what_if(
    payload: dict = Body(...),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Mô phỏng GPA kỳ + GPA tích lũy mới dựa trên điểm dự kiến SV nhập cho kỳ tới.

    Input: { grades: [{ course_code: str, score10: float }] }
    Output: per-course s4 + term_gpa4 + new_cumulative_gpa4 + classification change.
    """
    user = _get_user_by_token(authorization, db)
    raw = payload.get("grades") or []
    if not isinstance(raw, list) or not raw:
        raise HTTPException(status_code=422, detail="grades phải là list không rỗng")

    # Fetch snapshot for current stats
    snap = academic_engine._build_snapshot(db, user.id, specialization=user.specialization)
    # Note: official_earned_credits đã drop. earned_credits tính từ grades.

    current_avg4 = snap.avg_score4 or 0.0
    current_earned = snap.earned_credits or 0.0

    # Look up course credits
    codes = [str(r.get("course_code", "")).strip() for r in raw if r.get("course_code")]
    course_map = {
        c.course_code: c for c in db.query(models.Course).filter(
            models.Course.course_code.in_(codes)
        ).all()
    }

    per_course = []
    term_weighted_score4 = 0.0
    term_credits_passed = 0.0
    term_credits_total = 0.0
    failed_courses: list[str] = []

    for row in raw:
        code = str(row.get("course_code", "")).strip()
        try: s10 = float(row.get("score10"))
        except (TypeError, ValueError): continue
        s10 = max(0.0, min(10.0, s10))
        c = course_map.get(code)
        if not c: continue
        credits = float(c.credits or 0.0)
        s4 = academic_engine.score10_to_score4(s10) or 0.0
        passed = s10 >= 4.0
        if passed:
            term_weighted_score4 += s4 * credits
            term_credits_passed += credits
        else:
            failed_courses.append(c.course_name or code)
        term_credits_total += credits
        per_course.append({
            "course_code": code,
            "course_name": c.course_name,
            "credits": credits,
            "score10": s10,
            "score4": round(s4, 2),
            "letter": academic_engine.letter_from_score10(s10),
            "passed": passed,
        })

    term_gpa4 = round(term_weighted_score4 / term_credits_passed, 2) if term_credits_passed > 0 else 0.0

    # New cumulative GPA
    new_total_credits = current_earned + term_credits_passed
    if new_total_credits > 0:
        new_cum_gpa4 = round((current_avg4 * current_earned + term_weighted_score4) / new_total_credits, 2)
    else:
        new_cum_gpa4 = 0.0

    delta = round(new_cum_gpa4 - current_avg4, 2) if current_avg4 > 0 else 0.0

    return {
        "per_course": per_course,
        "term": {
            "credits_total": round(term_credits_total, 1),
            "credits_passed": round(term_credits_passed, 1),
            "failed_count": len(failed_courses),
            "failed_course_names": failed_courses,
            "gpa4": term_gpa4,
            "classification": academic_engine.classify_gpa4(term_gpa4) if term_credits_passed else "Không có môn qua",
        },
        "cumulative": {
            "before_gpa4": round(current_avg4, 2),
            "before_credits": round(current_earned, 1),
            "after_gpa4": new_cum_gpa4,
            "after_credits": round(new_total_credits, 1),
            "delta": delta,
            "direction": "up" if delta > 0.01 else "down" if delta < -0.01 else "same",
            "classification_before": academic_engine.classify_gpa4(current_avg4) if current_avg4 > 0 else "Chưa có",
            "classification_after": academic_engine.classify_gpa4(new_cum_gpa4) if new_cum_gpa4 > 0 else "Chưa có",
        },
    }


# ── Function #8: Switch-spec what-if ──────────────────────────────────────────
@app.post("/simulator/switch-spec")
def simulator_switch_spec(
    payload: dict = Body(...),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Mô phỏng đổi chuyên ngành: trả về TC hiện vẫn dùng được, TC bị mất, TC còn thiếu, kỳ TN dự kiến mới.

    Input: { new_spec: "7480201_05" }
    Output: {
      current_spec, new_spec, current_spec_label, new_spec_label,
      kept_credits,       // TC vẫn count cho CN mới
      lost_credits,       // TC từ môn riêng CN cũ — KHÔNG count cho CN mới
      remaining_credits,  // TC còn thiếu để TN
      total_threshold,
      reusable_courses,   // môn đã pass + vẫn count cho CN mới
      lost_courses,       // môn đã pass + bị bỏ vì spec cũ
      missing_required_count,  // số môn bắt buộc CN mới chưa pass
      predicted_grad_term_current,
      predicted_grad_term_new,
      extra_terms_needed,
    }
    """
    user = _get_user_by_token(authorization, db)
    if user.role != "student":
        raise HTTPException(status_code=403, detail="Chỉ sinh viên mới dùng simulator")

    new_spec = (payload or {}).get("new_spec", "").strip()
    if not new_spec:
        raise HTTPException(status_code=400, detail="new_spec không được rỗng")
    if new_spec == user.specialization:
        raise HTTPException(status_code=400, detail="new_spec trùng với CN hiện tại")

    # Resolve spec labels
    spec_label_map = {
        "7480201_07": "Khoa học máy tính",
        "7480201_06": "Mạng máy tính",
        "7480201_05": "Công nghệ phần mềm",
        "7480201_09": "Hệ thống thông tin",
        "7480201_04": "Tin học kinh tế",
        "7480201_08": "CNTT Địa học",
    }
    if new_spec not in spec_label_map:
        raise HTTPException(status_code=400, detail=f"Spec không hợp lệ: {new_spec}")

    # 1. Build snapshot với CN HIỆN TẠI (để biết TC hiện tại)
    snap_current = academic_engine._build_snapshot(db, user.id, specialization=user.specialization)
    current_earned = float(snap_current.earned_credits or 0.0)

    # 2. Lấy tất cả grades + courses đã pass
    grades = db.query(models.UserGrade).filter(
        models.UserGrade.user_id == user.id,
        models.UserGrade.passed == True,
    ).all()
    passed_codes = list({g.course_code for g in grades})
    passed_courses_objs = db.query(models.Course).filter(
        models.Course.course_code.in_(passed_codes)
    ).all() if passed_codes else []
    course_by_code = {c.course_code: c for c in passed_courses_objs}

    # 3. Phân loại từng môn pass:
    #    - kept: đại cương (no spec) HOẶC required_specialization == new_spec
    #    - lost: required_specialization == old_spec (và != new_spec)
    kept_courses = []
    lost_courses = []
    for code in passed_codes:
        c = course_by_code.get(code)
        if not c:
            continue
        if not academic_engine._counts_toward_credits(c):
            continue  # GDTC, QPAN không count vào TC tích lũy
        if not c.required_specialization or c.required_specialization == new_spec:
            kept_courses.append(c)
        else:
            lost_courses.append(c)

    kept_credits = sum(float(c.credits or 0) for c in kept_courses)
    lost_credits = sum(float(c.credits or 0) for c in lost_courses)

    # 4. Threshold tốt nghiệp
    thresholds = academic_engine._get_academic_thresholds(db)
    total_threshold = float(thresholds.get("graduation_credit_threshold", 153.0))
    remaining_credits = max(0.0, total_threshold - kept_credits)

    # 5. Đếm môn bắt buộc CN mới chưa pass
    new_spec_required = db.query(models.Course).filter(
        models.Course.required_specialization == new_spec
    ).all()
    new_spec_required_codes = {c.course_code for c in new_spec_required}
    missing_required_count = len(new_spec_required_codes - set(passed_codes))

    # 6. Dự đoán kỳ TN — dùng risk_engine.predict_graduation_term
    from backend.core import risk_engine as _risk
    # Pace dùng main_terms studied của SV
    grades_all = db.query(models.UserGrade).filter(
        models.UserGrade.user_id == user.id
    ).all()
    main_terms = sum(1 for g in grades_all if g.term and _risk._is_main_term(g.term))

    # Current scenario: dùng current_earned vs threshold
    f_current = _risk.RiskFeatures(
        earned_credits=current_earned, expected_credits=main_terms * 19.0,
        credits_pct=current_earned / total_threshold,
        main_terms_studied=main_terms, summer_terms_studied=0,
        gpa4=snap_current.avg_score4, gpa4_recent_2=None, gpa4_trend=None,
        failed_count=0, failed_unrecovered=0, retake_count=0,
        pace_ratio=(current_earned / (main_terms * 19.0)) if main_terms else 1.0,
        on_internship_eligible=False, on_thesis_eligible=False,
    )
    # New scenario: kept_credits làm starting point
    f_new = _risk.RiskFeatures(
        earned_credits=kept_credits, expected_credits=main_terms * 19.0,
        credits_pct=kept_credits / total_threshold,
        main_terms_studied=main_terms, summer_terms_studied=0,
        gpa4=snap_current.avg_score4, gpa4_recent_2=None, gpa4_trend=None,
        failed_count=0, failed_unrecovered=0, retake_count=0,
        pace_ratio=(kept_credits / (main_terms * 19.0)) if main_terms else 1.0,
        on_internship_eligible=False, on_thesis_eligible=False,
    )
    latest = _risk._latest_term_of(db, user.id)
    grad_current = _risk.predict_graduation_term(f_current, latest)
    grad_new = _risk.predict_graduation_term(f_new, latest)

    # Tính extra terms needed (kỳ chính)
    extra_terms = 0
    if lost_credits > 0:
        # Mỗi kỳ kỳ vọng ~19 TC → chậm thêm ceil(lost_credits / 19)
        import math
        extra_terms = math.ceil(lost_credits / 19.0)

    return {
        "current_spec": user.specialization,
        "new_spec": new_spec,
        "current_spec_label": spec_label_map.get(user.specialization or "", user.specialization or "—"),
        "new_spec_label": spec_label_map[new_spec],
        "kept_credits": round(kept_credits, 1),
        "lost_credits": round(lost_credits, 1),
        "current_credits": round(current_earned, 1),
        "remaining_credits": round(remaining_credits, 1),
        "total_threshold": total_threshold,
        "kept_courses": [
            {"course_code": c.course_code, "course_name": c.course_name, "credits": float(c.credits or 0)}
            for c in kept_courses
        ],
        "lost_courses": [
            {"course_code": c.course_code, "course_name": c.course_name, "credits": float(c.credits or 0)}
            for c in lost_courses
        ],
        "missing_required_count": missing_required_count,
        "predicted_grad_term_current": grad_current,
        "predicted_grad_term_new": grad_new,
        "extra_terms_needed": extra_terms,
        "verdict": (
            "easy" if lost_credits <= 6 else
            "moderate" if lost_credits <= 18 else
            "hard"
        ),
    }


# ── Career Paths ──────────────────────────────────────────────────────────────

_CAREER_DOMAIN_KEYWORDS = {
    "programming":         ["lập trình", "lap trinh", "python", "java", "c++", "c #", "c#", "javascript", "nhập môn", "kỹ thuật lập trình"],
    "algorithms":          ["giải thuật", "thuật toán", "cấu trúc dữ liệu", "ctdl", "algorithm"],
    "database":            ["cơ sở dữ liệu", "csdl", "database", "sql", "nosql", "data warehouse"],
    "networking":          ["mạng", "mạng máy tính", "network", "truyền thông"],
    "web":                 ["web", "html", "css", "front-end", "backend"],
    "mobile":              ["di động", "mobile", "android", "ios"],
    "ai_ml":               ["trí tuệ nhân tạo", "học máy", "machine learning", "deep", "ai", "ttnt", "nlp"],
    "security":            ["an toàn", "an ninh", "bảo mật", "security"],
    "devops":              ["hệ điều hành", "triển khai", "cloud", "docker", "kub"],
    "data_analytics":      ["khai phá dữ liệu", "dữ liệu", "phân tích", "data mining", "analytics", "bi"],
    "math":                ["toán", "giải tích", "đại số", "xác suất", "thống kê", "rời rạc", "discrete", "calculus", "algebra"],
    "software_engineering":["phần mềm", "công nghệ phần mềm", "quản trị dự án", "đồ án", "phân tích thiết kế"],
}


def _infer_course_domains(course_name: str) -> dict[str, float]:
    n = (course_name or "").lower()
    out: dict[str, float] = {}
    for dom, kws in _CAREER_DOMAIN_KEYWORDS.items():
        for kw in kws:
            if kw in n:
                out[dom] = 1.0
                break
    return out


def _user_domain_profile(db: Session, user_id: int) -> dict[str, float]:
    """Aggregate student's domain strengths from passed courses weighted by score10.

    Returns {domain: normalized score in [0,1]}.
    """
    grades = db.query(models.UserGrade).filter(
        models.UserGrade.user_id == user_id,
        models.UserGrade.passed == True,  # noqa: E712
    ).all()
    if not grades:
        return {}

    course_codes = {g.course_code for g in grades if g.course_code}
    course_map = {
        c.course_code: c
        for c in db.query(models.Course).filter(models.Course.course_code.in_(course_codes)).all()
    }

    accum: dict[str, float] = {}
    weights: dict[str, float] = {}
    for g in grades:
        course = course_map.get(g.course_code)
        if not course:
            continue
        domains = _infer_course_domains(course.course_name or "")
        if not domains:
            continue
        weight = float(g.score10 or 5.0) / 10.0
        for d in domains:
            accum[d] = accum.get(d, 0.0) + weight
            weights[d] = weights.get(d, 0.0) + 1.0

    normalized: dict[str, float] = {}
    for d, total in accum.items():
        count = weights[d]
        normalized[d] = min(1.0, (total / count))
    return normalized


def _career_fit_score(user_profile: dict[str, float], career_profile: dict) -> float:
    """Weighted cosine similarity with bias toward careers that require domains the student excels at."""
    import math
    if not user_profile or not career_profile:
        return 0.0
    keys = set(user_profile.keys()) | set(career_profile.keys())
    dot = 0.0
    un, cn = 0.0, 0.0
    for k in keys:
        u = user_profile.get(k, 0.0)
        c = float(career_profile.get(k, 0.0))
        dot += u * c
        un += u * u
        cn += c * c
    return dot / (math.sqrt(un) * math.sqrt(cn)) if un and cn else 0.0


def _format_career(
    path: models.CareerPath,
    skills: list[models.CareerSkill] | None = None,
    course_maps: list[models.CareerCourseMap] | None = None,
    course_lookup: dict[str, models.Course] | None = None,
) -> dict:
    out = {
        "id": path.id,
        "code": path.code,
        "name": path.name,
        "icon": path.icon,
        "color": path.color,
        "short_description": path.short_description,
        "long_description": path.long_description,
        "domain_profile": path.domain_profile,
    }
    if skills is not None:
        out["skills"] = [
            {
                "id": s.id,
                "skill_name": s.skill_name,
                "skill_type": s.skill_type,
                "level": s.level,
                "priority": s.priority,
                "source_type": s.source_type,
                "source_name": s.source_name,
                "source_url": s.source_url,
                "description": s.description,
                "estimated_hours": s.estimated_hours,
            }
            for s in skills
        ]
    if course_maps is not None:
        course_lookup = course_lookup or {}
        out["related_courses"] = [
            {
                "course_code": cm.course_code,
                "course_name": (course_lookup.get(cm.course_code).course_name if course_lookup.get(cm.course_code) else cm.course_code),
                "relevance": float(cm.relevance or 1.0),
            }
            for cm in course_maps
        ]
    return out


@app.get("/me/readiness")
def get_my_readiness(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Tổng hợp readiness 3 mốc lớn: tốt nghiệp / thực tập DN / đồ án TN.

    Mỗi mốc trả về:
    - status: 'ready' | 'almost' | 'not_ready'
    - met: list checklist đã đạt
    - blockers: list checklist chưa đạt + remaining number
    - estimate: tham chiếu thời gian dự kiến (cho graduation only)
    """
    from datetime import datetime as _dt
    from backend.core.academic_engine import (
        _build_snapshot, _find_internship_thesis, _course_credits,
        _get_graduation_threshold, _get_academic_thresholds,
        INTERNSHIP_REMAINING_BUFFER, _graduation_estimate,
    )

    user = _get_user_by_token(authorization, db)
    if user.role != "student":
        raise HTTPException(status_code=403, detail="Chỉ SV mới xem readiness")

    specialization = user.specialization
    snapshot = _build_snapshot(db, user.id, specialization=specialization)
    # Note: official_earned_credits đã được drop ở refactor 2026-05-05.
    # earned_credits giờ tính trực tiếp từ user_grades trong _build_snapshot.

    grad_threshold = _get_graduation_threshold(db)
    at = _get_academic_thresholds(db)
    internship_min = at["internship_min_credits"]
    thesis_min = at["thesis_min_credits"]
    thesis_min_gpa = at["thesis_min_gpa4"]

    earned = snapshot.earned_credits or 0.0
    gpa4 = snapshot.avg_score4 if snapshot.avg_score4 is not None else 0.0

    # ── GDTC & QPAN completion ───────────────────────────────────────────────
    _GDTC_CODES = {"7010701", "7010702", "7010703"}
    _GDTC_TOTAL = 3
    all_grades = db.query(models.UserGrade).filter(models.UserGrade.user_id == user.id).all()
    gdtc_passed = len({
        g.course_code for g in all_grades
        if g.course_code.strip() in _GDTC_CODES and g.passed
    })
    qpan_passed = len({
        g.course_code for g in all_grades
        if g.course_code.strip().startswith("73") and g.passed
    })
    # Xác định tổng QPAN theo khóa: K21 có 4 môn (7300103/104/202/203), ver mới 3 môn
    cohort_str = user.username[:2] if user.username and len(user.username) >= 2 else "21"
    try:
        cohort_year = int(cohort_str)
    except ValueError:
        cohort_year = 21
    _QPAN_TOTAL = 4 if cohort_year <= 21 else 3
    gdtc_done = gdtc_passed >= _GDTC_TOTAL
    qpan_done = qpan_passed >= _QPAN_TOTAL

    # ── Graduation readiness ─────────────────────────────────────────────────
    grad_remaining = max(0.0, grad_threshold - earned)
    grad_pct = min(100, round(earned / grad_threshold * 100, 1)) if grad_threshold > 0 else 0
    est = _graduation_estimate(db, user.id, snapshot)
    estimated_terms = est.get("estimated_terms_remaining")
    avg_per_term = est.get("avg_credits_per_term")
    terms_studied = est.get("terms_studied", 0)

    # Project to actual semester label
    projected_label = None
    if earned >= grad_threshold:
        projected_label = "Đã đủ điều kiện"
    elif estimated_terms is not None and estimated_terms > 0:
        # Roughly: assume now is mid-academic year; round up
        terms_to_go = max(1, int(estimated_terms + 0.5))
        # Find next academic year start (academic year = current calendar year if month >= 8 else year - 1)
        now = _dt.now()
        cur_year_start = now.year if now.month >= 8 else now.year - 1
        # Each "term" = 1 HK; roughly HK1 (Sep-Jan), HK2 (Feb-Jun) — alternate
        # Simple: assume current term + N more terms
        # Output format: HKx/yyyy-yyyy
        cur_hk = 1 if now.month >= 8 or now.month <= 1 else 2
        future_hk = cur_hk
        future_year = cur_year_start
        for _ in range(terms_to_go):
            future_hk += 1
            if future_hk > 2:
                future_hk = 1
                future_year += 1
        projected_label = f"HK{future_hk}/{future_year}-{future_year + 1}"

    grad_met = []
    grad_blockers = []
    if earned >= grad_threshold:
        grad_met.append({"label": f"Đạt {earned:.1f}/{grad_threshold:.0f} TC", "icon": "check_circle"})
    else:
        grad_blockers.append({
            "label": f"Còn thiếu {grad_remaining:.1f} TC tích lũy",
            "icon": "credit_score", "value": grad_remaining,
        })

    if earned >= grad_threshold:
        grad_status = "ready"
    elif grad_remaining <= 20:
        grad_status = "almost"
    else:
        grad_status = "not_ready"

    # ── Internship readiness ─────────────────────────────────────────────────
    internship_course, thesis_course = _find_internship_thesis(snapshot.courses, specialization)
    internship_done = bool(
        internship_course and internship_course.course_code in snapshot.completed_codes
    )
    thesis_done = bool(
        thesis_course and thesis_course.course_code in snapshot.completed_codes
    )
    internship_outstanding = (
        _course_credits(internship_course) if internship_course and not internship_done else 0.0
    )
    thesis_outstanding = (
        _course_credits(thesis_course) if thesis_course and not thesis_done else 0.0
    )
    remaining_non_special = (
        snapshot.total_credits - snapshot.earned_credits - thesis_outstanding - internship_outstanding
    )

    int_met = []
    int_blockers = []
    if earned >= internship_min:
        int_met.append({
            "label": f"TC tích lũy ≥ {internship_min:.0f} (hiện {earned:.1f})",
            "icon": "credit_score",
        })
    else:
        int_blockers.append({
            "label": f"Cần thêm {internship_min - earned:.1f} TC tích lũy (≥ {internship_min:.0f})",
            "icon": "credit_score",
            "value": round(internship_min - earned, 1),
        })

    if remaining_non_special <= INTERNSHIP_REMAINING_BUFFER:
        int_met.append({
            "label": f"Còn ≤ {INTERNSHIP_REMAINING_BUFFER:.0f} TC chưa hoàn thành (hiện {remaining_non_special:.1f})",
            "icon": "task_alt",
        })
    else:
        int_blockers.append({
            "label": f"Còn {remaining_non_special:.1f} TC chưa học, cần ≤ {INTERNSHIP_REMAINING_BUFFER:.0f} mới được đi TT",
            "icon": "pending_actions",
            "value": round(remaining_non_special - INTERNSHIP_REMAINING_BUFFER, 1),
        })

    if internship_done:
        int_status = "ready"
        int_met.append({"label": "Đã hoàn thành thực tập", "icon": "check_circle"})
    elif not int_blockers:
        int_status = "ready"
    elif len(int_blockers) == 1 and (int_blockers[0].get("value") or 0) < 10:
        int_status = "almost"
    else:
        int_status = "not_ready"

    # ── Thesis readiness ─────────────────────────────────────────────────────
    th_met = []
    th_blockers = []
    if internship_done:
        th_met.append({"label": "Đã hoàn thành thực tập DN", "icon": "check_circle"})
    else:
        th_blockers.append({"label": "Phải pass thực tập DN trước", "icon": "work_history"})

    if earned >= thesis_min:
        th_met.append({
            "label": f"TC tích lũy ≥ {thesis_min:.0f} (hiện {earned:.1f})",
            "icon": "credit_score",
        })
    else:
        th_blockers.append({
            "label": f"Cần thêm {thesis_min - earned:.1f} TC (≥ {thesis_min:.0f})",
            "icon": "credit_score",
            "value": round(thesis_min - earned, 1),
        })

    if gpa4 >= thesis_min_gpa:
        th_met.append({
            "label": f"GPA hệ 4 ≥ {thesis_min_gpa:.2f} (hiện {gpa4:.2f})",
            "icon": "grade",
        })
    else:
        th_blockers.append({
            "label": f"GPA chưa đủ — cần ≥ {thesis_min_gpa:.2f} (hiện {gpa4:.2f})",
            "icon": "trending_up",
            "value": round(thesis_min_gpa - gpa4, 2),
        })

    if thesis_done:
        th_status = "ready"
        th_met.append({"label": "Đã hoàn thành đồ án TN", "icon": "check_circle"})
    elif not th_blockers:
        th_status = "ready"
    elif len(th_blockers) <= 1:
        th_status = "almost"
    else:
        th_status = "not_ready"

    # ── Tỷ lệ % completion từng mốc ─────────────────────────────────────────
    def _pct(numerator, denominator):
        if denominator <= 0:
            return 100
        return min(100, round(numerator / denominator * 100, 1))

    # ── Suggested courses for each milestone — actionable list of next courses ──
    # For each blocked milestone, suggest courses SV chưa hoàn thành cần ưu tiên học
    completed = set(snapshot.completed_codes)
    remaining_courses = [
        c for c in snapshot.courses
        if c.course_code not in completed
        and c.course_code not in (
            internship_course.course_code if internship_course else "",
            thesis_course.course_code if thesis_course else "",
        )
    ]

    def _course_dict(c, with_role: str = ""):
        return {
            "course_code": c.course_code,
            "course_name": c.course_name,
            "credits": float(c.credits) if c.credits else 0.0,
            "typical_semester": c.typical_semester,
            "role": with_role,
        }

    # Sort: BB (no elective_group) first by typical_semester, then by credits desc
    def _sort_key(c):
        is_bb = c.required_specialization is None or c.required_specialization == specialization
        sem = c.typical_semester if c.typical_semester is not None else 99
        creds = float(c.credits) if c.credits else 0.0
        return (0 if is_bb else 1, sem, -creds)

    remaining_courses.sort(key=_sort_key)

    # Graduation: top 6 remaining required-side courses
    grad_suggested = [_course_dict(c, "Bắt buộc còn lại") for c in remaining_courses[:6]]

    # Internship: courses needed to either reach internship_min or reduce remaining_non_special
    # Strategy: same priority list since SV cần học các môn còn lại
    int_suggested = []
    if not internship_done and (earned < internship_min or remaining_non_special > INTERNSHIP_REMAINING_BUFFER):
        int_suggested = [_course_dict(c, "Cần pass trước khi TT") for c in remaining_courses[:5]]
    if internship_course and not internship_done:
        int_suggested.append(_course_dict(internship_course, "Môn thực tập"))

    # Thesis: must complete internship + reach thesis_min + GPA threshold
    th_suggested = []
    if not thesis_done:
        if not internship_done and internship_course:
            th_suggested.append(_course_dict(internship_course, "Phải qua trước"))
        if earned < thesis_min:
            # Need more credits → list top remaining
            th_suggested.extend([_course_dict(c, "Tích lũy TC") for c in remaining_courses[:4]])
        if thesis_course:
            th_suggested.append(_course_dict(thesis_course, "Môn ĐATN"))

    return {
        "earned_credits": earned,
        "graduation_threshold": grad_threshold,
        "gpa4": round(gpa4, 2),
        "specialization": specialization,
        "gdtc": {
            "done": gdtc_done,
            "passed": gdtc_passed,
            "total": _GDTC_TOTAL,
            "label": f"GDTC ({gdtc_passed}/{_GDTC_TOTAL} HP)",
        },
        "qpan": {
            "done": qpan_done,
            "passed": qpan_passed,
            "total": _QPAN_TOTAL,
            "label": f"QPAN ({qpan_passed}/{_QPAN_TOTAL} môn)",
        },
        "english_cert": {
            "note": "Liên hệ Phòng Đào tạo (P.102) để xác nhận chứng chỉ ngoại ngữ đầu ra.",
        },
        "graduation": {
            "status": grad_status,
            "percent": grad_pct,
            "remaining_credits": round(grad_remaining, 1),
            "estimated_terms_remaining": estimated_terms,
            "avg_credits_per_term": avg_per_term,
            "terms_studied": terms_studied,
            "projected_term_label": projected_label,
            "met": grad_met,
            "blockers": grad_blockers,
            "suggested_courses": grad_suggested,
        },
        "internship": {
            "status": int_status,
            "percent": _pct(earned, internship_min),
            "min_credits_required": internship_min,
            "remaining_to_min_credits": max(0, round(internship_min - earned, 1)),
            "remaining_non_special": round(remaining_non_special, 1),
            "buffer": INTERNSHIP_REMAINING_BUFFER,
            "is_done": internship_done,
            "course_code": internship_course.course_code if internship_course else None,
            "course_name": internship_course.course_name if internship_course else None,
            "met": int_met,
            "blockers": int_blockers,
            "suggested_courses": int_suggested,
        },
        "thesis": {
            "status": th_status,
            "percent": _pct(earned, thesis_min),
            "min_credits_required": thesis_min,
            "min_gpa_required": thesis_min_gpa,
            "remaining_to_min_credits": max(0, round(thesis_min - earned, 1)),
            "remaining_to_min_gpa": max(0, round(thesis_min_gpa - gpa4, 2)),
            "is_done": thesis_done,
            "course_code": thesis_course.course_code if thesis_course else None,
            "course_name": thesis_course.course_name if thesis_course else None,
            "met": th_met,
            "blockers": th_blockers,
            "suggested_courses": th_suggested,
        },
    }


@app.get("/me/skill-tree")
def get_my_skill_tree(
    career_code: str | None = None,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Evidence-based skill tree cho SV.

    Replace cũ "% mastery": mỗi skill có 2 nguồn bằng chứng:
    - evidence_courses: môn đã pass có cover skill (auto, verifiable)
    - evidence_personal: link cá nhân SV upload (GitHub, certificate, project)

    Status: evidenced (có ít nhất 1 nguồn) | no_evidence (chưa có gì)
    Tag: core (importance >= 0.7) | bonus (< 0.7 nhưng required) | extra (không trong career)
    """
    user = _get_user_by_token(authorization, db)
    if user.role != "student":
        raise HTTPException(status_code=403, detail="Chỉ SV mới có skill tree")

    # 1. Resolve career path: param → user choice → null
    target_path = None
    if career_code:
        target_path = db.query(models.CareerPath).filter(
            models.CareerPath.code == career_code
        ).first()
        if not target_path:
            raise HTTPException(status_code=404, detail="Career path không tồn tại")
    else:
        choice = db.query(models.UserCareerChoice).filter(
            models.UserCareerChoice.user_id == user.id
        ).first()
        if choice and choice.primary_path_id:
            target_path = db.query(models.CareerPath).filter(
                models.CareerPath.id == choice.primary_path_id
            ).first()

    # 2. Lấy passed courses + grades chi tiết (cho evidence_courses)
    passed_grades = db.query(models.UserGrade).filter(
        models.UserGrade.user_id == user.id,
        models.UserGrade.passed == True,  # noqa: E712
    ).all()
    passed_codes = {g.course_code for g in passed_grades}

    # Course details (name) for evidence display
    pass_course_lookup = {
        c.course_code: c for c in db.query(models.Course).filter(
            models.Course.course_code.in_(passed_codes)
        ).all()
    } if passed_codes else {}

    # Best-grade per course (handle retake — dùng điểm cao nhất)
    best_grade_per_course: dict[str, models.UserGrade] = {}
    for g in passed_grades:
        existing = best_grade_per_course.get(g.course_code)
        if not existing or (g.score4 or 0) > (existing.score4 or 0):
            best_grade_per_course[g.course_code] = g

    # 3. Build evidence_courses_per_skill: skill_code → [course evidence]
    evidence_courses_map: dict[str, list[dict]] = {}
    if passed_codes:
        for cs in db.query(models.CourseSkill).filter(
            models.CourseSkill.course_code.in_(passed_codes)
        ).all():
            c = pass_course_lookup.get(cs.course_code)
            g = best_grade_per_course.get(cs.course_code)
            if not c:
                continue
            evidence_courses_map.setdefault(cs.skill_code, []).append({
                "course_code": cs.course_code,
                "course_name": c.course_name,
                "score10": float(g.score10) if g and g.score10 is not None else None,
                "letter": g.letter if g else None,
                "weight": round(float(cs.weight), 2),
            })

    # Sort each list by weight desc (best evidence first)
    for sk in evidence_courses_map:
        evidence_courses_map[sk].sort(key=lambda x: -x["weight"])

    # 4. Build evidence_personal_per_skill: skill_code → [user-uploaded evidence]
    evidence_personal_map: dict[str, list[dict]] = {}
    personal_evs = db.query(models.SkillEvidence).filter(
        models.SkillEvidence.user_id == user.id
    ).all()
    for ev in personal_evs:
        evidence_personal_map.setdefault(ev.skill_code, []).append({
            "id": ev.id,
            "type": ev.evidence_type,
            "url": ev.url,
            "label": ev.label,
            "description": ev.description,
            "verified": bool(ev.verified),
            "created_at": ev.created_at.isoformat() if ev.created_at else None,
        })

    # 5. Skill catalog
    all_skills_q = db.query(models.Skill).all()
    skill_info_by_code = {
        s.code: {
            "code": s.code, "name": s.name, "category": s.category,
            "description": s.description,
        }
        for s in all_skills_q
    }

    # 7. Required skills from path
    required_map: dict[str, float] = {}  # skill_code → importance
    if target_path:
        for cps in db.query(models.CareerPathSkill).filter(
            models.CareerPathSkill.path_id == target_path.id
        ).all():
            required_map[cps.skill_code] = float(cps.importance)

    # 8. Build skill list — evidence-based status
    def _evidence_status(code: str) -> str:
        """evidenced if has any course OR personal evidence, else no_evidence."""
        if evidence_courses_map.get(code) or evidence_personal_map.get(code):
            return "evidenced"
        return "no_evidence"

    def _tag(required: bool, importance: float) -> str:
        """core (>=0.7 importance), bonus (<0.7 in path), extra (not in path)."""
        if not required:
            return "extra"
        return "core" if importance >= 0.7 else "bonus"

    skills_out = []
    seen = set()
    # Iterate required first (priority order)
    for code, imp in sorted(required_map.items(), key=lambda x: -x[1]):
        seen.add(code)
        info = skill_info_by_code.get(code)
        if not info:
            continue
        skills_out.append({
            **info,
            "required": True,
            "importance": imp,
            "tag": _tag(True, imp),
            "status": _evidence_status(code),
            "evidence_courses": evidence_courses_map.get(code, []),
            "evidence_personal": evidence_personal_map.get(code, []),
        })
    # Then extra (SV có evidence nhưng không required cho career)
    extra_codes = set(evidence_courses_map.keys()) | set(evidence_personal_map.keys())
    for code in sorted(extra_codes):
        if code in seen:
            continue
        info = skill_info_by_code.get(code)
        if not info:
            continue
        skills_out.append({
            **info,
            "required": False,
            "importance": 0.0,
            "tag": "extra",
            "status": "evidenced",  # extra always has evidence by definition
            "evidence_courses": evidence_courses_map.get(code, []),
            "evidence_personal": evidence_personal_map.get(code, []),
        })

    # 9. Compute summary — counts, NO percentages
    required_skills = [s for s in skills_out if s["required"]]
    evidenced = sum(1 for s in required_skills if s["status"] == "evidenced")
    no_evidence = sum(1 for s in required_skills if s["status"] == "no_evidence")
    total_required = len(required_skills)

    # 10. Recommend courses to fill gap (skills no_evidence)
    recommended_courses: list[dict] = []
    if target_path and no_evidence > 0:
        missing_codes = {s["code"] for s in required_skills if s["status"] == "no_evidence"}
        candidate_rows = db.query(
            models.CourseSkill.course_code,
            models.CourseSkill.skill_code,
            models.CourseSkill.weight,
        ).filter(
            models.CourseSkill.skill_code.in_(missing_codes),
            ~models.CourseSkill.course_code.in_(passed_codes),
        ).all() if missing_codes else []

        course_score: dict[str, dict] = {}
        for code, sk, w in candidate_rows:
            entry = course_score.setdefault(code, {
                "course_code": code, "matched_skills": [], "total_weight": 0.0,
            })
            entry["matched_skills"].append(sk)
            entry["total_weight"] += float(w)
        course_codes_list = list(course_score.keys())
        if course_codes_list:
            course_lookup = {
                c.course_code: c for c in db.query(models.Course).filter(
                    models.Course.course_code.in_(course_codes_list)
                ).all()
            }
            for code, info in course_score.items():
                c = course_lookup.get(code)
                if not c:
                    continue
                if c.required_specialization and user.specialization and \
                   c.required_specialization != user.specialization:
                    continue
                matched_unique = list(set(info["matched_skills"]))
                # skills_unlocked: số skill sẽ chuyển từ no_evidence → evidenced
                skills_unlocked = len(matched_unique)
                recommended_courses.append({
                    "course_code": c.course_code,
                    "course_name": c.course_name,
                    "credits": float(c.credits) if c.credits is not None else None,
                    "typical_semester": c.typical_semester,
                    "matched_skill_codes": matched_unique,
                    "match_score": round(info["total_weight"], 2),
                    "skills_unlocked": skills_unlocked,
                })
            recommended_courses.sort(key=lambda x: (-x["skills_unlocked"], -x["match_score"]))
            recommended_courses = recommended_courses[:8]

    return {
        "career_path": {
            "code": target_path.code, "name": target_path.name,
            "icon": target_path.icon, "color": target_path.color,
            "short_description": target_path.short_description,
        } if target_path else None,
        "summary": {
            "total_required": total_required,
            "evidenced": evidenced,
            "no_evidence": no_evidence,
            "extra": sum(1 for s in skills_out if s["tag"] == "extra"),
            "passed_courses": len(passed_codes),
            "personal_evidence_total": sum(len(v) for v in evidence_personal_map.values()),
        },
        "skills": skills_out,
        "recommended_courses": recommended_courses,
    }


# ── CRUD endpoints for skill evidence (SV) + resources (admin) ──────────────

@app.post("/skills/{skill_code}/evidence", status_code=201)
def add_skill_evidence(
    skill_code: str,
    payload: dict = Body(...),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """SV thêm bằng chứng cá nhân cho 1 skill.

    Input: {type: "github"|"certificate"|"project"|"other", url?, label, description?}
    """
    user = _get_user_by_token(authorization, db)
    if user.role != "student":
        raise HTTPException(status_code=403, detail="Chỉ sinh viên")

    ev_type = (payload or {}).get("type", "").strip()
    if ev_type not in ("github", "certificate", "project", "other"):
        raise HTTPException(status_code=400, detail="type phải: github | certificate | project | other")
    label = ((payload or {}).get("label") or "").strip()
    if not label:
        raise HTTPException(status_code=400, detail="label không được rỗng")
    if len(label) > 200:
        raise HTTPException(status_code=400, detail="label tối đa 200 ký tự")

    url = (payload or {}).get("url")
    if url:
        url = url.strip()
        if not (url.startswith("http://") or url.startswith("https://")):
            raise HTTPException(status_code=400, detail="url phải bắt đầu bằng http:// hoặc https://")
        if len(url) > 1000:
            raise HTTPException(status_code=400, detail="url tối đa 1000 ký tự")
    description = (payload or {}).get("description")
    if description and len(description) > 500:
        raise HTTPException(status_code=400, detail="description tối đa 500 ký tự")

    # Verify skill exists in catalog
    skill = db.query(models.Skill).filter(models.Skill.code == skill_code).first()
    if not skill:
        raise HTTPException(status_code=404, detail=f"Skill {skill_code} không có trong catalog")

    # Limit: 10 evidence/skill/user
    existing_count = db.query(models.SkillEvidence).filter(
        models.SkillEvidence.user_id == user.id,
        models.SkillEvidence.skill_code == skill_code,
    ).count()
    if existing_count >= 10:
        raise HTTPException(status_code=409, detail="Đã đạt giới hạn 10 bằng chứng/skill")

    ev = models.SkillEvidence(
        user_id=user.id,
        skill_code=skill_code,
        evidence_type=ev_type,
        url=url,
        label=label,
        description=description,
    )
    db.add(ev)
    db.commit()
    db.refresh(ev)
    return {
        "id": ev.id,
        "skill_code": ev.skill_code,
        "type": ev.evidence_type,
        "url": ev.url,
        "label": ev.label,
        "description": ev.description,
        "verified": bool(ev.verified),
        "created_at": ev.created_at.isoformat() if ev.created_at else None,
    }


@app.delete("/skills/evidence/{evidence_id}")
def delete_skill_evidence(
    evidence_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """SV xoá bằng chứng cá nhân của mình."""
    user = _get_user_by_token(authorization, db)
    ev = db.query(models.SkillEvidence).filter(models.SkillEvidence.id == evidence_id).first()
    if not ev:
        raise HTTPException(status_code=404, detail="Evidence không tồn tại")
    if ev.user_id != user.id and user.role != "admin":
        raise HTTPException(status_code=403, detail="Không có quyền")
    db.delete(ev)
    db.commit()
    return {"deleted": True, "id": evidence_id}


@app.get("/careers")
def list_careers(db: Session = Depends(get_db)):
    """Danh sách tất cả định hướng nghề nghiệp (public — không cần auth)."""
    paths = db.query(models.CareerPath).order_by(models.CareerPath.id.asc()).all()
    return [_format_career(p) for p in paths]


@app.get("/careers/recommend/me")
def recommend_careers_for_me(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """ML-based career recommendation dựa trên profile học tập của SV."""
    user = _get_user_by_token(authorization, db)
    user_profile = _user_domain_profile(db, user.id)

    paths = db.query(models.CareerPath).all()
    scored = []
    for p in paths:
        fit = _career_fit_score(user_profile, p.domain_profile or {})
        scored.append({
            "path": p,
            "fit_score": round(fit, 3),
            "fit_percent": round(fit * 100, 1),
        })
    scored.sort(key=lambda x: x["fit_score"], reverse=True)

    # Top drivers: pick top-3 shared domains between user and top career
    top_shared: list[tuple[str, float]] = []
    if scored:
        top = scored[0]
        cp = top["path"].domain_profile or {}
        joint = [
            (k, user_profile.get(k, 0.0) * float(cp.get(k, 0.0)))
            for k in set(user_profile.keys()) | set(cp.keys())
        ]
        joint.sort(key=lambda x: x[1], reverse=True)
        top_shared = [(k, v) for k, v in joint[:3] if v > 0]

    return {
        "user_profile": user_profile,
        "total_courses_analyzed": sum(1 for _ in db.query(models.UserGrade).filter(
            models.UserGrade.user_id == user.id,
            models.UserGrade.passed == True,  # noqa: E712
        ).all()),
        "recommendations": [
            {
                **_format_career(x["path"]),
                "fit_score": x["fit_score"],
                "fit_percent": x["fit_percent"],
            }
            for x in scored
        ],
        "top_drivers": [{"domain": k, "contribution": round(v, 3)} for k, v in top_shared],
    }


@app.get("/careers/me")
def get_my_career_choice(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    choice = db.query(models.UserCareerChoice).filter(
        models.UserCareerChoice.user_id == user.id
    ).first()
    if not choice:
        return {"chosen": False}

    def _path_brief(pid: int | None) -> dict | None:
        if not pid:
            return None
        p = db.query(models.CareerPath).filter(models.CareerPath.id == pid).first()
        return _format_career(p) if p else None

    progress_rows = db.query(models.UserCareerSkillProgress).filter(
        models.UserCareerSkillProgress.user_id == user.id
    ).all()
    progress_map = {r.skill_id: r for r in progress_rows}

    # Build detail for primary path including skills + user progress
    primary_detail = None
    if choice.primary_path_id:
        p = db.query(models.CareerPath).filter(models.CareerPath.id == choice.primary_path_id).first()
        if p:
            skills = db.query(models.CareerSkill).filter(
                models.CareerSkill.path_id == p.id
            ).order_by(models.CareerSkill.priority.asc(), models.CareerSkill.id.asc()).all()
            primary_detail = _format_career(p, skills=skills)
            for s in primary_detail["skills"]:
                pr = progress_map.get(s["id"])
                s["user_status"] = pr.status if pr else "planned"
                s["completed_at"] = pr.completed_at.isoformat() if (pr and pr.completed_at) else None
            completed = sum(1 for s in primary_detail["skills"] if s.get("user_status") == "completed")
            primary_detail["progress"] = {
                "completed": completed,
                "total": len(primary_detail["skills"]),
                "percent": round(100.0 * completed / len(primary_detail["skills"]), 1) if primary_detail["skills"] else 0.0,
            }

    return {
        "chosen": True,
        "primary": primary_detail,
        "secondary": _path_brief(choice.secondary_path_id),
        "chosen_at": choice.chosen_at.isoformat() if choice.chosen_at else None,
        "updated_at": choice.updated_at.isoformat() if choice.updated_at else None,
    }


@app.post("/careers/me/choose")
def set_my_career_choice(
    payload: dict = Body(...),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Set career choice — supports 2 modes:
    1. Full: {primary_code, secondary_code?} — set cả 2 slot
    2. Slot: {code, slot: 'primary'|'secondary'} — chỉ update 1 slot
    """
    user = _get_user_by_token(authorization, db)
    code_arg = (payload.get("code") or "").strip()
    slot_arg = (payload.get("slot") or "").strip()
    primary_code = (payload.get("primary_code") or "").strip()
    secondary_code = (payload.get("secondary_code") or "").strip() or None

    from datetime import datetime as _dt
    choice = db.query(models.UserCareerChoice).filter(
        models.UserCareerChoice.user_id == user.id
    ).first()

    if code_arg and slot_arg:
        # Slot mode
        if slot_arg not in ("primary", "secondary"):
            raise HTTPException(status_code=400, detail="slot phải là 'primary' hoặc 'secondary'")
        target_path = db.query(models.CareerPath).filter(models.CareerPath.code == code_arg).first()
        if not target_path:
            raise HTTPException(status_code=404, detail="career không tồn tại")
        if not choice:
            choice = models.UserCareerChoice(user_id=user.id)
            db.add(choice)
        if slot_arg == "primary":
            # Nếu code này đang là secondary → swap (move secondary lên primary)
            if choice.secondary_path_id == target_path.id:
                choice.secondary_path_id = choice.primary_path_id
            choice.primary_path_id = target_path.id
        else:
            # secondary slot — không cho trùng primary
            if choice.primary_path_id == target_path.id:
                raise HTTPException(status_code=400, detail="Không thể đặt primary làm secondary")
            choice.secondary_path_id = target_path.id
        choice.updated_at = _dt.utcnow()
        db.commit()
        return {"message": f"Đã đặt {slot_arg}", "code": code_arg, "slot": slot_arg}

    # Full mode (legacy)
    if not primary_code:
        raise HTTPException(status_code=422, detail="primary_code (hoặc code+slot) bắt buộc")

    primary = db.query(models.CareerPath).filter(models.CareerPath.code == primary_code).first()
    if not primary:
        raise HTTPException(status_code=404, detail="primary không tồn tại")
    secondary = None
    if secondary_code:
        secondary = db.query(models.CareerPath).filter(models.CareerPath.code == secondary_code).first()
        if not secondary:
            raise HTTPException(status_code=404, detail="secondary không tồn tại")

    if not choice:
        choice = models.UserCareerChoice(
            user_id=user.id,
            primary_path_id=primary.id,
            secondary_path_id=secondary.id if secondary else None,
        )
        db.add(choice)
    else:
        choice.primary_path_id = primary.id
        choice.secondary_path_id = secondary.id if secondary else None
        choice.updated_at = _dt.utcnow()
    db.commit()
    return {"message": "Đã lưu định hướng", "primary_code": primary_code, "secondary_code": secondary_code}


@app.post("/careers/me/skills/{skill_id}/status")
def update_skill_status(
    skill_id: int,
    payload: dict = Body(...),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    status = (payload.get("status") or "").strip()
    if status not in ("planned", "in_progress", "completed"):
        raise HTTPException(status_code=422, detail="status không hợp lệ")

    skill = db.query(models.CareerSkill).filter(models.CareerSkill.id == skill_id).first()
    if not skill:
        raise HTTPException(status_code=404, detail="Skill không tồn tại")

    from datetime import datetime as _dt
    row = db.query(models.UserCareerSkillProgress).filter(
        models.UserCareerSkillProgress.user_id == user.id,
        models.UserCareerSkillProgress.skill_id == skill_id,
    ).first()
    if not row:
        row = models.UserCareerSkillProgress(
            user_id=user.id, skill_id=skill_id, status=status,
            completed_at=_dt.utcnow() if status == "completed" else None,
        )
        db.add(row)
    else:
        row.status = status
        row.completed_at = _dt.utcnow() if status == "completed" else None
    db.commit()
    return {"message": "Đã cập nhật", "skill_id": skill_id, "status": status}


@app.get("/careers/explain/me/{career_code}")
def explain_career_fit(
    career_code: str,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """LLM-generated personalized explanation for why a career fits (or not).

    Uses fallback chain Gemini → Groq → Claude via chat_assistant.
    """
    user = _get_user_by_token(authorization, db)
    path = db.query(models.CareerPath).filter(models.CareerPath.code == career_code).first()
    if not path:
        raise HTTPException(status_code=404, detail="Định hướng không tồn tại")

    user_profile = _user_domain_profile(db, user.id)
    fit = _career_fit_score(user_profile, path.domain_profile or {})

    top_courses = (
        db.query(models.UserGrade)
        .filter(models.UserGrade.user_id == user.id, models.UserGrade.passed == True)  # noqa: E712
        .order_by(models.UserGrade.score10.desc().nullslast())
        .limit(10).all()
    )
    course_lookup = {
        c.course_code: c
        for c in db.query(models.Course).filter(
            models.Course.course_code.in_([g.course_code for g in top_courses if g.course_code])
        ).all()
    }
    top_list = [
        f"- {course_lookup.get(g.course_code).course_name if course_lookup.get(g.course_code) else g.course_code}: {g.score10 or '?'}/10"
        for g in top_courses[:6]
    ]

    strengths = sorted(user_profile.items(), key=lambda x: x[1], reverse=True)[:3]
    career_needs = sorted((path.domain_profile or {}).items(), key=lambda x: float(x[1] or 0), reverse=True)[:4]

    prompt = f"""Bạn là cố vấn học tập CNTT. Hãy viết 1 đoạn ngắn (80-120 chữ) bằng tiếng Việt tư vấn cho sinh viên có thể theo hướng "{path.name}" hay không.

Dữ liệu sinh viên:
- Thế mạnh (top domains): {", ".join(f"{k}={v:.2f}" for k, v in strengths) or "chưa đủ dữ liệu"}
- Môn điểm cao nhất:
{chr(10).join(top_list) if top_list else "- (chưa có dữ liệu điểm)"}

Đặc thù hướng "{path.name}":
- Mô tả: {path.short_description}
- Domain cần: {", ".join(f"{k}={float(v):.2f}" for k, v in career_needs)}

Độ phù hợp tính bằng cosine similarity: {fit*100:.1f}%.

Viết tự nhiên, khuyên cụ thể (nên tập trung domain nào, học thêm gì). KHÔNG dùng bullet, không đề mục, chỉ 1 đoạn văn súc tích."""

    explanation = None
    try:
        from backend.core.chat_assistant import _gemini_chat, _groq_chat, _claude_chat_with_history
        msgs = [
            {"role": "system", "content": "Bạn là cố vấn học tập ngành CNTT tư vấn định hướng nghề nghiệp cho sinh viên Việt Nam."},
            {"role": "user", "content": prompt},
        ]
        explanation = (
            _gemini_chat(msgs, temperature=0.4, max_tokens=300)
            or _groq_chat(msgs, temperature=0.4, max_tokens=300)
            or _claude_chat_with_history(msgs)
        )
    except Exception:
        explanation = None

    if not explanation:
        # Rule-based fallback
        if fit >= 0.75:
            explanation = f"Với thế mạnh ở {', '.join(k for k, _ in strengths)}, bạn rất phù hợp với hướng {path.name}. Tiếp tục học sâu các môn liên quan, bổ sung skill ngoài trường theo mô tả: {path.short_description.lower()}"
        elif fit >= 0.5:
            explanation = f"Bạn có nền tảng ổn để theo {path.name}, nhưng cần cải thiện thêm ở {', '.join(k for k, _ in career_needs[:2])}."
        else:
            explanation = f"Hồ sơ hiện tại của bạn chưa thật sự phù hợp với {path.name}. Cân nhắc hướng khác hoặc bổ sung mạnh các môn về {', '.join(k for k, _ in career_needs[:2])}."

    return {
        "career_code": career_code,
        "career_name": path.name,
        "fit_percent": round(fit * 100, 1),
        "top_strengths": [{"domain": k, "score": round(v, 2)} for k, v in strengths],
        "career_needs": [{"domain": k, "weight": round(float(v), 2)} for k, v in career_needs],
        "explanation": explanation.strip() if explanation else "",
    }


@app.get("/careers/{code}")
def get_career_detail(
    code: str,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Chi tiết 1 định hướng — bổ sung peer stats + my_choice cho UI."""
    path = db.query(models.CareerPath).filter(models.CareerPath.code == code).first()
    if not path:
        raise HTTPException(status_code=404, detail="Không tìm thấy định hướng")
    skills = db.query(models.CareerSkill).filter(
        models.CareerSkill.path_id == path.id
    ).order_by(models.CareerSkill.priority.asc(), models.CareerSkill.id.asc()).all()
    cms = db.query(models.CareerCourseMap).filter(
        models.CareerCourseMap.path_id == path.id
    ).order_by(models.CareerCourseMap.relevance.desc()).all()
    codes = [cm.course_code for cm in cms]
    course_lookup = {
        c.course_code: c
        for c in db.query(models.Course).filter(models.Course.course_code.in_(codes)).all()
    }
    out = _format_career(path, skills=skills, course_maps=cms, course_lookup=course_lookup)

    # Enrich related_courses với credits + typical_semester + is_compulsory
    if "related_courses" in out:
        for rc in out["related_courses"]:
            c = course_lookup.get(rc["course_code"])
            if c:
                rc["credits"] = float(c.credits) if c.credits is not None else None
                rc["typical_semester"] = c.typical_semester
                rc["count_toward_credits"] = bool(c.count_toward_credits)

    # Peer count: bao nhiêu SV đã chọn path này làm primary hoặc secondary
    student_count = db.query(models.UserCareerChoice).filter(
        (models.UserCareerChoice.primary_path_id == path.id) |
        (models.UserCareerChoice.secondary_path_id == path.id)
    ).count()
    out["student_count"] = student_count

    # is_my_choice + my_progress (nếu user đã login)
    out["is_my_primary"] = False
    out["is_my_secondary"] = False
    out["my_skill_progress"] = None
    out["passed_related_courses"] = []
    try:
        user = _get_user_by_token(authorization, db)
        choice = db.query(models.UserCareerChoice).filter(
            models.UserCareerChoice.user_id == user.id
        ).first()
        if choice:
            out["is_my_primary"] = (choice.primary_path_id == path.id)
            out["is_my_secondary"] = (choice.secondary_path_id == path.id)
        # Skill progress
        if user.role == "student":
            total_skills = len(skills)
            if total_skills > 0:
                done_count = db.query(models.UserCareerSkillProgress).filter(
                    models.UserCareerSkillProgress.user_id == user.id,
                    models.UserCareerSkillProgress.skill_id.in_([s.id for s in skills]),
                    models.UserCareerSkillProgress.status == "completed",
                ).count()
                out["my_skill_progress"] = {
                    "completed": done_count,
                    "total": total_skills,
                    "percent": round(done_count / total_skills * 100) if total_skills else 0,
                }
            # Số môn related đã pass
            if codes:
                passed_codes = {
                    g.course_code for g in db.query(models.UserGrade).filter(
                        models.UserGrade.user_id == user.id,
                        models.UserGrade.course_code.in_(codes),
                        models.UserGrade.passed == True,  # noqa: E712
                    ).all()
                }
                out["passed_related_courses"] = sorted(passed_codes)
    except Exception:
        pass  # public access — no user context

    return out


# ── Course Ratings ────────────────────────────────────────────────────────────

@app.post("/courses/{course_code}/rate")
def rate_course(
    course_code: str,
    rating: int,
    review: str | None = None,
    is_anonymous: bool = False,
    admin_feedback: str | None = None,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Student rates a course they have completed (1-5 stars).

    Args:
        rating: 1-5 sao (bắt buộc)
        review: nhận xét công khai (tùy chọn) — hiển thị cho SV khác
        is_anonymous: nếu True → ẩn tên SV trong public reviews
        admin_feedback: góp ý riêng cho admin về chất lượng giảng dạy / GV
                        (không công khai, chỉ admin xem được)
    """
    if rating < 1 or rating > 5:
        raise HTTPException(status_code=400, detail="Điểm đánh giá phải từ 1 đến 5 sao")

    user = _get_user_by_token(authorization, db)
    if user.role != "student":
        raise HTTPException(status_code=403, detail="Chỉ sinh viên mới có thể đánh giá môn học")

    course = db.query(models.Course).filter(models.Course.course_code == course_code).first()
    if not course:
        raise HTTPException(status_code=404, detail="Không tìm thấy môn học")

    # Must have passed the course
    grade = db.query(models.UserGrade).filter(
        models.UserGrade.user_id == user.id,
        models.UserGrade.course_code == course_code,
        models.UserGrade.passed == True,
    ).first()
    if not grade:
        raise HTTPException(status_code=403, detail="Bạn chỉ có thể đánh giá môn học đã hoàn thành")

    # Sanitize review/admin_feedback (strip whitespace, collapse empty → None)
    review = (review or "").strip() or None
    admin_feedback = (admin_feedback or "").strip() or None

    existing = db.query(models.CourseRating).filter(
        models.CourseRating.user_id == user.id,
        models.CourseRating.course_code == course_code,
    ).first()

    if existing:
        existing.rating = rating
        existing.review = review
        existing.is_anonymous = is_anonymous
        existing.admin_feedback = admin_feedback
        db.commit()
        return {"message": "Đã cập nhật đánh giá", "rating": rating}

    new_rating = models.CourseRating(
        user_id=user.id,
        course_code=course_code,
        rating=rating,
        review=review,
        is_anonymous=is_anonymous,
        admin_feedback=admin_feedback,
    )
    db.add(new_rating)
    db.commit()
    return {"message": "Đã ghi nhận đánh giá", "rating": rating}


@app.delete("/courses/{course_code}/rate", response_model=schemas.MessageOut)
def delete_my_course_rating(
    course_code: str,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """SV xóa đánh giá của chính mình (rating + review) cho 1 môn."""
    user = _get_user_by_token(authorization, db)
    if user.role != "student":
        raise HTTPException(status_code=403, detail="Chỉ sinh viên có quyền xóa đánh giá")
    existing = db.query(models.CourseRating).filter(
        models.CourseRating.user_id == user.id,
        models.CourseRating.course_code == course_code,
    ).first()
    if not existing:
        raise HTTPException(status_code=404, detail="Bạn chưa đánh giá môn này")
    db.delete(existing)
    db.commit()
    return {"message": f"Đã xóa đánh giá môn {course_code}"}


@app.get("/me/reviews")
def list_my_reviews(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """List tất cả đánh giá SV đã viết — kèm course info để link sang chi tiết."""
    user = _get_user_by_token(authorization, db)
    if user.role != "student":
        raise HTTPException(status_code=403, detail="Chỉ sinh viên có /me/reviews")
    rows = db.query(models.CourseRating).filter(
        models.CourseRating.user_id == user.id
    ).order_by(models.CourseRating.id.desc()).all()
    if not rows:
        return {"total": 0, "items": []}
    course_codes = [r.course_code for r in rows]
    courses = {c.course_code: c for c in db.query(models.Course).filter(
        models.Course.course_code.in_(course_codes)
    ).all()}
    items = []
    for r in rows:
        c = courses.get(r.course_code)
        items.append({
            "course_code": r.course_code,
            "course_name": c.course_name if c else r.course_code,
            "credits": float(c.credits) if c and c.credits else None,
            "rating": r.rating,
            "review": r.review or "",
            "has_review": bool(r.review and r.review.strip()),
        })
    return {"total": len(items), "items": items}


@app.get("/courses/{course_code}/rating")
def get_course_rating(
    course_code: str,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Get aggregate rating stats for a course + current user's own rating."""
    user = _get_user_by_token(authorization, db)

    ratings = db.query(models.CourseRating).filter(
        models.CourseRating.course_code == course_code
    ).all()

    count = len(ratings)
    avg = round(sum(r.rating for r in ratings) / count, 1) if count > 0 else None
    distribution = {str(i): sum(1 for r in ratings if r.rating == i) for i in range(1, 6)}

    my_row = next((r for r in ratings if r.user_id == user.id), None)
    return {
        "course_code": course_code,
        "avg_rating": avg,
        "count": count,
        "distribution": distribution,
        "my_rating": my_row.rating if my_row else None,
        "my_review": my_row.review if my_row else None,
        "my_is_anonymous": bool(my_row.is_anonymous) if my_row else False,
        "my_admin_feedback": my_row.admin_feedback if my_row else None,
    }


# ── Course Registrations ───────────────────────────────────────────────────────

@app.post("/registrations/me", response_model=schemas.MessageOut)
def register_courses(
    payload: schemas.CourseRegistrationIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    from collections import defaultdict as _dd
    user = _get_user_by_token(authorization, db)

    # 1. Resolve semester label — normalise to canonical "HK{n}/{year}" format
    sem_label = _normalize_term(payload.semester_label or "")
    # "Hiện tại" or empty → fall back to admin-set active semester
    if sem_label in ("Hiện tại", "hien tai", ""):
        active = _get_active_semester(db)
        if active:
            sem_label = _normalize_term(active)

    # 2. Load course catalog
    all_courses = {c.course_code: c for c in db.query(models.Course).all()}

    # 3. Student's completed courses
    completed_codes = {
        g.course_code for g in db.query(models.UserGrade).filter(
            models.UserGrade.user_id == user.id,
            models.UserGrade.passed == True,
        ).all()
    }

    # 4. Prerequisite map
    prereq_map: dict[str, list[str]] = _dd(list)
    for p in db.query(models.CoursePrerequisite).all():
        prereq_map[p.course_code].append(p.prerequisite_code)

    # 5. Validate prerequisites for every requested course
    violations: list[str] = []
    for item in payload.courses:
        code = item.course_code
        if code not in all_courses:
            continue
        missing = [p for p in prereq_map.get(code, []) if p not in completed_codes]
        if missing:
            cname = all_courses[code].course_name
            violations.append(f"• {cname} ({code}): chưa hoàn thành {', '.join(missing)}")

    if violations:
        raise HTTPException(
            status_code=422,
            detail="Chưa đủ điều kiện tiên quyết:\n" + "\n".join(violations),
        )

    # 6. Validate credit load for this semester
    max_credits = 21.0  # Default — user.max_credits_per_term đã drop ở 2026-05-05
    existing_regs = db.query(models.CourseRegistration).filter(
        models.CourseRegistration.user_id == user.id,
        models.CourseRegistration.semester_label == sem_label,
        models.CourseRegistration.outcome == None,
    ).all()
    existing_reg_codes = {r.course_code for r in existing_regs}
    existing_credits = sum(
        float(all_courses[r.course_code].credits or 0)
        for r in existing_regs if r.course_code in all_courses
    )
    new_credits = sum(
        float(all_courses[item.course_code].credits or 0)
        for item in payload.courses
        if item.course_code in all_courses and item.course_code not in existing_reg_codes
    )
    total_credits = existing_credits + new_credits
    if total_credits > max_credits:
        raise HTTPException(
            status_code=422,
            detail=(
                f"Vượt giới hạn tín chỉ: kế hoạch hiện có {existing_credits:.0f} TC, "
                f"thêm {new_credits:.0f} TC → tổng {total_credits:.0f} TC "
                f"(tối đa {max_credits:.0f} TC/kỳ). "
                f"Có thể điều chỉnh giới hạn trong trang Hồ sơ."
            ),
        )

    # 7. Save valid new registrations
    inserted = 0
    for item in payload.courses:
        if item.course_code not in all_courses or item.course_code in existing_reg_codes:
            continue
        db.add(models.CourseRegistration(
            user_id=user.id,
            course_code=item.course_code,
            semester_label=sem_label,
            was_recommended=item.was_recommended,
            recommendation_score_at_time=item.recommendation_score,
        ))
        inserted += 1
    db.commit()
    return schemas.MessageOut(message=f"Đã lưu {inserted} môn vào kế hoạch dự kiến cho {sem_label}.")


@app.delete("/registrations/me/{course_code}", response_model=schemas.MessageOut)
def unregister_course(
    course_code: str,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    # Delete active (outcome=NULL) registrations for this course.
    # We no longer track "dropped" as a separate outcome — removing = deleting.
    from sqlalchemy import or_
    deleted = db.query(models.CourseRegistration).filter(
        models.CourseRegistration.user_id == user.id,
        models.CourseRegistration.course_code == course_code,
        or_(
            models.CourseRegistration.outcome.is_(None),
            models.CourseRegistration.outcome == "dropped",
        ),
    ).delete(synchronize_session=False)
    db.commit()
    if not deleted:
        raise HTTPException(status_code=404, detail="Không tìm thấy môn trong kế hoạch dự kiến")
    return schemas.MessageOut(message=f"Đã bỏ môn {course_code} khỏi kế hoạch dự kiến")


@app.get("/registrations/me", response_model=list[schemas.CourseRegistrationOut])
def list_registrations(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    return db.query(models.CourseRegistration).filter(
        models.CourseRegistration.user_id == user.id
    ).order_by(models.CourseRegistration.registered_at.desc()).all()


@app.patch("/registrations/{reg_id}/outcome", response_model=schemas.CourseRegistrationOut)
def update_registration_outcome(
    reg_id: int,
    outcome: str,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    if outcome not in ("passed", "failed", "dropped"):
        raise HTTPException(status_code=400, detail="outcome phải là 'passed', 'failed' hoặc 'dropped'")
    reg = db.query(models.CourseRegistration).filter(
        models.CourseRegistration.id == reg_id,
        models.CourseRegistration.user_id == user.id,
    ).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi đăng ký")
    reg.outcome = outcome
    db.commit()
    db.refresh(reg)
    return reg


# ── Schedule ──────────────────────────────────────────────────────────────────

def _times_overlap(s1: str, e1: str, s2: str, e2: str) -> bool:
    """Return True if time range [s1,e1) overlaps [s2,e2)."""
    return s1 < e2 and s2 < e1


def _dates_overlap(sd1, ed1, sd2, ed2) -> bool:
    """Return True if date ranges overlap (None = open-ended)."""
    if sd1 and ed2 and sd1 > ed2:
        return False
    if sd2 and ed1 and sd2 > ed1:
        return False
    return True


@app.get("/schedule/me", response_model=list[schemas.ScheduleEntryOut])
def get_schedule(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    return db.query(models.ScheduleEntry).filter(models.ScheduleEntry.user_id == user.id).all()


@app.post("/schedule/me", response_model=schemas.ScheduleEntryOut)
def add_schedule_entry(
    payload: schemas.ScheduleEntryIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    entry = models.ScheduleEntry(
        user_id=user.id,
        course_code=payload.course_code,
        course_name=payload.course_name,
        day_of_week=payload.day_of_week,
        start_time=payload.start_time,
        end_time=payload.end_time,
        start_date=payload.start_date,
        end_date=payload.end_date,
        room=payload.room,
        color=payload.color,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return entry


@app.put("/schedule/me/{entry_id}", response_model=schemas.ScheduleEntryOut)
def update_schedule_entry(
    entry_id: int,
    payload: schemas.ScheduleEntryIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    entry = db.query(models.ScheduleEntry).filter(
        models.ScheduleEntry.id == entry_id,
        models.ScheduleEntry.user_id == user.id,
    ).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Không tìm thấy lịch học")
    entry.course_code = payload.course_code
    entry.course_name = payload.course_name
    entry.day_of_week = payload.day_of_week
    entry.start_time = payload.start_time
    entry.end_time = payload.end_time
    entry.start_date = payload.start_date
    entry.end_date = payload.end_date
    entry.room = payload.room
    entry.color = payload.color
    db.commit()
    db.refresh(entry)
    return entry


@app.delete("/schedule/me/{entry_id}", response_model=schemas.MessageOut)
def delete_schedule_entry(
    entry_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    entry = db.query(models.ScheduleEntry).filter(
        models.ScheduleEntry.id == entry_id,
        models.ScheduleEntry.user_id == user.id,
    ).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Không tìm thấy lịch học")
    db.delete(entry)
    db.commit()
    return schemas.MessageOut(message="Đã xóa lịch học")


@app.get("/schedule/me/conflicts", response_model=list[schemas.ScheduleConflictOut])
def get_schedule_conflicts(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    entries = db.query(models.ScheduleEntry).filter(models.ScheduleEntry.user_id == user.id).all()
    conflicts = []
    for i, a in enumerate(entries):
        for b in entries[i + 1:]:
            if a.day_of_week != b.day_of_week:
                continue
            if not _times_overlap(a.start_time, a.end_time, b.start_time, b.end_time):
                continue
            if not _dates_overlap(a.start_date, a.end_date, b.start_date, b.end_date):
                continue
            conflicts.append({"entry_a": a, "entry_b": b, "day_of_week": a.day_of_week})
    return conflicts


@app.get("/courses/lookup")
def course_lookup(
    q: str = "",
    db: Session = Depends(get_db),
):
    """Autocomplete: search courses by code prefix or name substring (no auth needed)."""
    q = q.strip()
    if not q:
        return []
    from sqlalchemy import or_
    results = db.query(models.Course).filter(
        or_(
            models.Course.course_code.ilike(f"{q}%"),
            models.Course.course_name.ilike(f"%{q}%"),
        )
    ).limit(10).all()
    return [{"course_code": c.course_code, "course_name": c.course_name, "credits": float(c.credits or 0)} for c in results]


@app.get("/chat/suggestions/me")
def chat_suggestions_me(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Gợi ý câu hỏi mặc định theo role. Dùng cho floating chat + AI-chat page."""
    user = _get_user_by_token(authorization, db)
    role = user.role or "student"
    if role == "advisor":
        greeting = f"Chào {user.full_name.split()[0] if user.full_name else 'thầy/cô'}! Tôi là AI hỗ trợ cố vấn học tập. Bạn muốn hỏi gì?"
        chips = [
            "Những SV nào đang gặp khó khăn nhất?",
            "Tóm tắt tình hình nhóm SV tôi phụ trách",
            "Gợi ý hành động cho SV có GPA thấp",
            "SV nào có nguy cơ không tốt nghiệp đúng hạn?",
            "So sánh tiến độ nhóm tôi với chuẩn CTĐT",
        ]
    elif role == "admin":
        greeting = "Chào quản trị viên! Tôi có thể phân tích dữ liệu toàn hệ thống. Hỏi gì đây?"
        chips = [
            "Tổng quan chất lượng học tập toàn hệ thống",
            "Cohort nào có GPA trung bình thấp nhất?",
            "Chuyên ngành nào có nhiều SV nguy cơ nhất?",
            "Có cố vấn nào chưa được phân công SV không?",
            "Đề xuất hành động quản trị cho tuần này",
        ]
    else:
        first = user.full_name.split()[0] if user.full_name else "bạn"
        greeting = f"Chào {first}! Tôi tư vấn môn học, tiến độ TC, định hướng nghề. Bạn cần hỗ trợ gì?"
        chips = [
            "Tôi nên học môn gì kỳ tới?",
            "Phân tích điểm yếu của tôi",
            "Khi nào tôi có thể tốt nghiệp?",
            "Định hướng nghề nào phù hợp với tôi?",
            "Giải thích một môn học cụ thể",
        ]
    return {"role": role, "greeting": greeting, "suggestions": chips}


@app.post("/chat/me", response_model=schemas.ChatOut)
def chat_me(
    payload: schemas.ChatIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    # If thread provided, ensure ChatThread record exists (upsert title on first message)
    if payload.thread_id:
        thread = db.query(models.ChatThread).filter(
            models.ChatThread.id == payload.thread_id,
            models.ChatThread.user_id == user.id,
        ).first()
        if not thread:
            title = payload.message[:40] + ("…" if len(payload.message) > 40 else "")
            db.add(models.ChatThread(id=payload.thread_id, user_id=user.id, title=title))
            db.commit()
    result = chat_reply(
        message=payload.message,
        db=db,
        user_id=user.id,
        limit=payload.limit,
        prefer_llm=payload.prefer_llm,
        thread_id=payload.thread_id,
    )
    return schemas.ChatOut(intent=result.intent, answer=result.answer, suggestions=result.suggestions)


@app.get("/chat/history/me", response_model=list[schemas.ChatHistoryItemOut])
def chat_history_me(
    limit: int = 50,
    thread_id: str | None = None,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    return get_chat_history(db, user.id, limit=limit, thread_id=thread_id)


@app.get("/chat/threads/me", response_model=list[schemas.ChatThreadOut])
def chat_threads_me(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    return (
        db.query(models.ChatThread)
        .filter(models.ChatThread.user_id == user.id)
        .order_by(models.ChatThread.updated_at.desc())
        .all()
    )


@app.delete("/chat/threads/{thread_id}/me", response_model=schemas.MessageOut)
def delete_chat_thread(
    thread_id: str,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    user = _get_user_by_token(authorization, db)
    thread = db.query(models.ChatThread).filter(
        models.ChatThread.id == thread_id,
        models.ChatThread.user_id == user.id,
    ).first()
    if not thread:
        raise HTTPException(status_code=404, detail="Không tìm thấy đoạn chat")
    db.delete(thread)
    db.commit()
    return schemas.MessageOut(message="Đã xóa đoạn chat")


# ══════════════════════════════════════════════════════════════════════════════
# ADVISOR ROUTES
# ══════════════════════════════════════════════════════════════════════════════

def _require_advisor(authorization: str | None, db: Session) -> models.User:
    """Cho phép role advisor hoặc admin (admin có thể dùng mọi endpoint cố vấn)."""
    user = _get_user_by_token(authorization, db)
    if user.role not in ("advisor", "admin"):
        raise HTTPException(status_code=403, detail="Bạn không có quyền truy cập trang cố vấn")
    return user


def _check_advisor_has_student(advisor: models.User, student_id: int, db: Session) -> None:
    """Admin bỏ qua kiểm tra; advisor phải có assignment với student này."""
    if advisor.role == "admin":
        return
    exists = db.query(models.AdvisorAssignment).filter(
        models.AdvisorAssignment.advisor_id == advisor.id,
        models.AdvisorAssignment.student_id == student_id,
    ).first()
    if not exists:
        raise HTTPException(status_code=403, detail="Bạn không phụ trách sinh viên này")


_SPEC_DISPLAY: dict[str | None, str] = {
    "7480201_07": "KHMT",
    "7480201_06": "MMT",
    "7480201_05": "CNPM",
    "7480201_09": "HTTT",
    "7480201_04": "THKT",
    "7480201_08": "CNTTDH",
    None: "Năm đại cương",
}


def assign_advisor_for_student(
    db: Session,
    student_id: int,
    specialization: str | None,
) -> tuple[models.User | None, str | None]:
    """Gán cố vấn cho SV theo thứ tự ưu tiên:

    1. Nếu SV thuộc Lớp (class_group_id): assign GVCN của lớp đó (chuẩn nhất)
    2. Fallback: round-robin trong cùng managed_specialization → pick GV ít SV nhất
    3. Nếu không có GV nào cùng spec: trả warning

    - Xóa assignment cũ (nếu có) trước khi gán mới.
    - KHÔNG commit — caller phải tự commit.
    """
    db.query(models.AdvisorAssignment).filter(
        models.AdvisorAssignment.student_id == student_id
    ).delete(synchronize_session=False)

    # Path 1: SV thuộc Lớp → GVCN
    student = db.query(models.User).get(student_id)
    if student and getattr(student, "class_group_id", None):
        cg = db.query(models.ClassGroup).filter(
            models.ClassGroup.id == student.class_group_id
        ).first()
        if cg:
            advisor = db.query(models.User).get(cg.advisor_id)
            if advisor:
                db.add(models.AdvisorAssignment(advisor_id=advisor.id, student_id=student_id))
                return advisor, None

    # Path 2: Fallback round-robin trong cùng spec (cho SV chưa có lớp)
    advisors = db.query(models.User).filter(
        models.User.role == "advisor",
        models.User.managed_specialization == specialization,
    ).all()

    if advisors:
        # Pick GV có ít SV nhất; tie-break: alphabet teacher_code
        counts = {
            adv.id: db.query(models.AdvisorAssignment).filter(
                models.AdvisorAssignment.advisor_id == adv.id
            ).count()
            for adv in advisors
        }
        chosen = sorted(advisors, key=lambda a: (counts[a.id], a.teacher_code or a.username))[0]
        db.add(models.AdvisorAssignment(advisor_id=chosen.id, student_id=student_id))
        return chosen, None

    label = _SPEC_DISPLAY.get(specialization, specialization or "Năm đại cương")
    return None, f"Chuyên ngành {label} chưa có cố vấn"


# Kept for backwards compat with any existing assignment at startup/import flows;
# now a no-op wrapper that falls back gracefully.
def auto_assign_advisor(student: models.User, db: Session) -> None:
    """Deprecated — use assign_advisor_for_student instead. Kept to avoid call-site churn."""
    pass


def _student_status(progress: dict) -> tuple[str, str | None]:
    """
    Tính trạng thái nguy cơ học tập theo SPEC.md §7.9:
      high_risk     — GPA < 2.0 hoặc dự kiến trễ TN > 2 HK
      needs_attention — GPA 2.0–2.5 hoặc trễ 1–2 HK
      normal
    """
    gpa4 = progress.get("avg_score4")
    terms_studied = progress.get("terms_studied") or 0
    est_remaining = progress.get("estimated_terms_remaining")

    # Số HK trễ so với chuẩn 9 HK chính
    late_hk = 0
    if est_remaining is not None:
        late_hk = max(0.0, (terms_studied + float(est_remaining)) - 9.0)

    reasons: list[str] = []
    if gpa4 is not None and gpa4 < 2.0:
        reasons.append(f"GPA {round(gpa4, 2)} < 2.0")
    if late_hk > 2:
        reasons.append(f"dự kiến trễ {round(late_hk, 1)} HK")
    if reasons:
        return "high_risk", " · ".join(reasons)

    reasons2: list[str] = []
    if gpa4 is not None and 2.0 <= gpa4 < 2.5:
        reasons2.append(f"GPA {round(gpa4, 2)} < 2.5")
    if 1 <= late_hk <= 2:
        reasons2.append(f"dự kiến trễ {round(late_hk, 1)} HK")
    if reasons2:
        return "needs_attention", " · ".join(reasons2)

    return "normal", None


def _student_flags(progress: dict) -> list[dict]:
    """Extra warning badges (separate from primary status).

    - extended_duration: kế hoạch SV vượt 10 HK chính (~5 năm) → có thể trì hoãn TN
    - low_pace: TC trung bình/HK < 12 (học thưa, có thể kéo dài bất thường)
    """
    flags: list[dict] = []
    terms_studied = progress.get("terms_studied") or 0
    est_remaining = progress.get("estimated_terms_remaining")
    earned = progress.get("earned_credits") or 0.0

    # extended_duration
    if est_remaining is not None:
        total_terms = float(terms_studied) + float(est_remaining)
        if total_terms > 10:
            flags.append({
                "code": "extended_duration",
                "label": f"Kéo dài {round(total_terms,1)} HK",
                "severity": "warning",
            })

    # low_pace: chỉ cảnh báo khi đã học >=2 kỳ chính
    if terms_studied >= 2:
        avg_per_term = float(earned) / float(terms_studied)
        if avg_per_term < 12.0:
            flags.append({
                "code": "low_pace",
                "label": f"Trung bình {round(avg_per_term,1)} TC/HK",
                "severity": "warning",
            })

    return flags


# ── 1. GET /advisor/students — danh sách SV phụ trách ────────────────────────

@app.get("/advisor/students", response_model=list[schemas.AdvisorStudentItem])
def advisor_list_students(
    cohort: str | None = Query(default=None, description="Lọc theo khoá tuyển (vd 'K20'). Bỏ trống = tất cả."),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    advisor = _require_advisor(authorization, db)

    assignments = (
        db.query(models.AdvisorAssignment)
        .filter(models.AdvisorAssignment.advisor_id == advisor.id)
        .order_by(models.AdvisorAssignment.assigned_at.asc())
        .all()
    )

    cohort_filter = (cohort or "").strip() or None

    # Pre-query: SV nào có grade admin → batch để tránh N queries
    student_ids = [asgn.student_id for asgn in assignments]
    # Sau refactor 2026-05-05: không còn distinguish admin/self source.
    # Field has_admin_grades giờ luôn False — UI advisor xem "data tự khai".
    admin_grade_user_ids: set[int] = set()

    items: list[schemas.AdvisorStudentItem] = []
    for asgn in assignments:
        student = db.query(models.User).filter(models.User.id == asgn.student_id).first()
        if not student:
            continue
        # Server-side filter theo khoá (nếu client gửi)
        if cohort_filter and (student.cohort or "") != cohort_filter:
            continue
        try:
            prog = build_progress_snapshot(db, student.id)
        except Exception:
            prog = {}

        status, reason = _student_status(prog)
        items.append(schemas.AdvisorStudentItem(
            id=student.id,
            username=student.username,
            full_name=student.full_name,
            specialization=student.specialization,
            cohort=student.cohort,
            earned_credits=prog.get("earned_credits"),
            avg_score4=prog.get("avg_score4"),
            status=status,
            status_reason=reason,
            flags=_student_flags(prog),
            has_admin_grades=student.id in admin_grade_user_ids,
            assigned_at=asgn.assigned_at,
        ))

    # high_risk lên đầu, rồi needs_attention, rồi normal
    _rank = {"high_risk": 0, "needs_attention": 1, "normal": 2}
    items.sort(key=lambda x: _rank.get(x.status, 2))
    return items


# ── 2. GET /advisor/students/{student_id} — hồ sơ đầy đủ SV ─────────────────

@app.get("/advisor/students/{student_id}", response_model=schemas.ProgressOut)
def advisor_student_detail(
    student_id: int,
    target_gpa: float | None = None,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    advisor = _require_advisor(authorization, db)
    _check_advisor_has_student(advisor, student_id, db)

    student = db.query(models.User).filter(models.User.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Không tìm thấy sinh viên")

    return schemas.ProgressOut(**build_progress_snapshot(db, student_id, target_gpa=target_gpa))


# ── 3. GET /advisor/students/{student_id}/recommendations — gợi ý của SV ─────

@app.get("/advisor/students/{student_id}/recommendations")
def advisor_student_recommendations(
    student_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    advisor = _require_advisor(authorization, db)
    _check_advisor_has_student(advisor, student_id, db)

    student = db.query(models.User).filter(models.User.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Không tìm thấy sinh viên")

    return build_recommendations(db, student_id)


# ── 3.5. GET /advisor/students/{student_id}/risk-analysis — AI phân tích rủi ro ─

@app.get("/advisor/students/{student_id}/risk-analysis")
def advisor_student_risk_analysis(
    student_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """A6: AI-assisted at-risk analysis cho SV cụ thể.

    Khác với /advisor/stats (rule-based threshold GPA<2.0):
    Endpoint này gửi rich context (GPA trend, retake count, missing TC, notes
    history, recent term performance) vào LLM để có phân tích chi tiết +
    recommendation cụ thể cho từng SV.

    Token cost: gọi mỗi lần advisor click "Phân tích AI" trong UI — không tự
    động chạy nightly để tiết kiệm. Cố vấn quyết định khi cần đào sâu.

    Returns:
        {
          "risk_level": "high" | "medium" | "low",
          "summary": "1-2 câu tóm tắt",
          "factors": ["yếu tố 1", "yếu tố 2", ...],
          "recommendations": ["hành động 1", "hành động 2", ...],
          "confidence": "high" | "medium" | "low"
        }
    """
    import json as _json_lib
    advisor = _require_advisor(authorization, db)
    _check_advisor_has_student(advisor, student_id, db)

    student = db.query(models.User).filter(models.User.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Không tìm thấy sinh viên")

    # Build rich context — grades + progress + notes
    progress = build_progress_snapshot(db, student_id)
    grades = db.query(models.UserGrade).filter(
        models.UserGrade.user_id == student_id
    ).all()

    # Compute GPA trend: chia grades thành nửa cũ/nửa mới (sort by term), so sánh
    def _term_sort_key(t: str) -> int:
        # "Học kỳ 2 - Năm học 2021-2022" → 20212
        import re
        if not t:
            return 0
        yr = re.search(r"(\d{4})\s*[-–]\s*\d{4}", t)
        sm = re.search(r"(?:H[Kk]|Học\s*k[ỳy])\s*(\d+)", t, re.IGNORECASE)
        return (int(yr.group(1)) if yr else 0) * 10 + (int(sm.group(1)) if sm else 0)

    sorted_grades = sorted([g for g in grades if g.term and g.score4 is not None],
                           key=lambda g: _term_sort_key(g.term))
    gpa_trend = "không đủ dữ liệu"
    if len(sorted_grades) >= 4:
        mid = len(sorted_grades) // 2
        old_avg = sum(float(g.score4) for g in sorted_grades[:mid]) / mid
        new_avg = sum(float(g.score4) for g in sorted_grades[mid:]) / (len(sorted_grades) - mid)
        diff = new_avg - old_avg
        if diff < -0.3:
            gpa_trend = f"đang giảm ({old_avg:.2f} → {new_avg:.2f})"
        elif diff > 0.3:
            gpa_trend = f"đang tăng ({old_avg:.2f} → {new_avg:.2f})"
        else:
            gpa_trend = f"ổn định (~{new_avg:.2f})"

    # Retake / fail count
    fail_count = sum(1 for g in grades if g.score4 is not None and float(g.score4) < 1.0)
    retake_count = 0
    seen = set()
    for g in sorted_grades:
        if g.course_code in seen:
            retake_count += 1
        seen.add(g.course_code)

    # Notes về SV
    notes = db.query(models.AdvisorNote).filter(
        models.AdvisorNote.student_id == student_id
    ).order_by(models.AdvisorNote.created_at.desc()).limit(5).all()
    notes_summary = "\n".join(
        f"- {n.created_at.strftime('%Y-%m-%d')}: {(n.content or '')[:150]}"
        for n in notes
    ) or "(chưa có note)"

    # Build context for LLM
    earned = progress.get("earned_credits", 0)
    threshold = academic_engine._get_graduation_threshold(db)
    remaining = max(0, threshold - earned)
    avg4 = progress.get("avg_score4")
    # Format avg4 separately — f-string không support format-spec với conditional inline
    avg4_str = f"{avg4:.2f}" if avg4 is not None else "—"

    context = f"""Sinh viên: {student.full_name or student.username} (MSSV: {student.username})
Khoá: {student.cohort or '?'}, Chuyên ngành: {student.specialization or '?'}

Tiến độ hiện tại:
- Tín chỉ tích luỹ: {earned}/{threshold} TC (còn {remaining} TC)
- GPA hệ 4: {avg4_str}
- Tổng số môn đã học: {len(grades)}
- Số môn trượt (score4 < 1.0): {fail_count}
- Số môn học lại: {retake_count}
- Xu hướng GPA: {gpa_trend}
- Đủ điều kiện thực tập: {'có' if progress.get('internship_eligible') else 'không'}
- Đủ điều kiện tốt nghiệp: {'có' if progress.get('graduation_ready') else 'không'}

5 ghi chú gần nhất của cố vấn:
{notes_summary}
"""

    # LLM prompt — yêu cầu output JSON strict
    prompt = f"""Bạn là chuyên gia tư vấn học vụ phân tích rủi ro của sinh viên CNTT.

DỮ LIỆU SINH VIÊN:
{context}

Phân tích rủi ro học tập của SV và trả về JSON HỢP LỆ với cấu trúc:
{{
  "risk_level": "high" | "medium" | "low",
  "summary": "1-2 câu tóm tắt tình trạng",
  "factors": ["yếu tố rủi ro 1", "yếu tố 2", "yếu tố 3"],
  "recommendations": ["hành động cố vấn nên làm 1", "hành động 2", "hành động 3"],
  "confidence": "high" | "medium" | "low"
}}

YÊU CẦU:
- factors: 2-4 yếu tố cụ thể (vd: "GPA giảm từ 2.8 xuống 2.3", "trượt 3 môn cốt lõi", "thiếu 25 TC năm cuối")
- recommendations: 2-4 hành động cụ thể, có thể thực hiện (vd: "Hẹn SV gặp tuần sau để đánh giá lại lộ trình", "Khuyến cáo đăng ký lại Toán rời rạc kỳ tới")
- KHÔNG dùng markdown, KHÔNG thêm text ngoài JSON
- KHÔNG dùng code fence ```
- Bắt đầu bằng {{ kết thúc bằng }}

JSON:"""

    # Call LLM
    from backend.core.chat_assistant import _gemini_chat, _groq_chat
    raw = None
    try:
        raw = (_gemini_chat([{"role": "user", "content": prompt}], temperature=0.3, max_tokens=600)
            or _groq_chat([{"role": "user", "content": prompt}], temperature=0.3, max_tokens=600))
    except Exception:
        pass

    if not raw:
        # Fallback: rule-based mini-analysis
        risk = "high" if (avg4 is not None and avg4 < 2.0) else \
               ("medium" if (avg4 is not None and avg4 < 2.5) else "low")
        return {
            "risk_level": risk,
            "summary": "AI tạm không khả dụng. Phân tích nhanh dựa trên ngưỡng GPA cứng.",
            "factors": [
                f"GPA hệ 4: {avg4:.2f}" if avg4 is not None else "Chưa có GPA",
                f"Còn {remaining} TC để tốt nghiệp",
                f"{fail_count} môn trượt, {retake_count} môn học lại",
            ],
            "recommendations": ["Bật AI provider (Gemini/Groq) để có phân tích chi tiết."],
            "confidence": "low",
        }

    # Parse JSON từ LLM output
    raw = raw.strip()
    if raw.startswith("```"):
        raw = "\n".join(line for line in raw.split("\n") if not line.startswith("```"))
    start = raw.find("{")
    end = raw.rfind("}")
    if start >= 0 and end > start:
        try:
            parsed = _json_lib.loads(raw[start:end + 1])
            # Validate structure
            if not isinstance(parsed, dict):
                raise ValueError("not dict")
            return {
                "risk_level": parsed.get("risk_level", "medium"),
                "summary": str(parsed.get("summary", ""))[:500],
                "factors": [str(f)[:200] for f in (parsed.get("factors") or [])][:5],
                "recommendations": [str(r)[:300] for r in (parsed.get("recommendations") or [])][:5],
                "confidence": parsed.get("confidence", "medium"),
            }
        except Exception:
            pass

    # Last fallback
    return {
        "risk_level": "medium",
        "summary": raw[:300],
        "factors": [],
        "recommendations": [],
        "confidence": "low",
    }


# ── 4. GET /advisor/stats — thống kê nhóm SV ─────────────────────────────────

@app.get("/advisor/stats", response_model=schemas.AdvisorStatsOut)
def advisor_stats(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    advisor = _require_advisor(authorization, db)

    assignments = (
        db.query(models.AdvisorAssignment)
        .filter(models.AdvisorAssignment.advisor_id == advisor.id)
        .all()
    )

    if not assignments:
        return schemas.AdvisorStatsOut(
            total_students=0, high_risk_count=0,
            needs_attention_count=0, normal_count=0,
        )

    high_risk = needs_attention = normal = 0
    gpa4_vals: list[float] = []
    pct_vals: list[float] = []

    for asgn in assignments:
        try:
            prog = build_progress_snapshot(db, asgn.student_id)
        except Exception:
            prog = {}

        status, _ = _student_status(prog)
        if status == "high_risk":
            high_risk += 1
        elif status == "needs_attention":
            needs_attention += 1
        else:
            normal += 1

        if prog.get("avg_score4") is not None:
            gpa4_vals.append(float(prog["avg_score4"]))
        if prog.get("completion_percent") is not None:
            pct_vals.append(float(prog["completion_percent"]))

    return schemas.AdvisorStatsOut(
        total_students=len(assignments),
        high_risk_count=high_risk,
        needs_attention_count=needs_attention,
        normal_count=normal,
        avg_gpa4=round(sum(gpa4_vals) / len(gpa4_vals), 2) if gpa4_vals else None,
        avg_completion_percent=round(sum(pct_vals) / len(pct_vals), 1) if pct_vals else None,
    )


# ── 5. GET /advisor/notes/{student_id} — lịch sử ghi chú ─────────────────────

@app.get("/advisor/notes/{student_id}", response_model=list[schemas.AdvisorNoteOut])
def advisor_get_notes(
    student_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    advisor = _require_advisor(authorization, db)
    _check_advisor_has_student(advisor, student_id, db)

    query = db.query(models.AdvisorNote).filter(
        models.AdvisorNote.student_id == student_id,
    )
    # Cố vấn chỉ thấy ghi chú của mình; admin thấy tất cả
    if advisor.role != "admin":
        query = query.filter(models.AdvisorNote.advisor_id == advisor.id)

    return query.order_by(models.AdvisorNote.created_at.desc(), models.AdvisorNote.id.desc()).all()


# ── 6. POST /advisor/notes — tạo ghi chú tư vấn ──────────────────────────────

@app.post("/advisor/notes", response_model=schemas.AdvisorNoteOut)
def advisor_create_note(
    body: schemas.AdvisorNoteCreateIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    advisor = _require_advisor(authorization, db)
    _check_advisor_has_student(advisor, body.student_id, db)

    student = db.query(models.User).filter(models.User.id == body.student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Không tìm thấy sinh viên")

    note = models.AdvisorNote(
        advisor_id=advisor.id,
        student_id=body.student_id,
        content=body.content,
        course_code=(body.course_code or None),
    )
    db.add(note)
    db.commit()
    db.refresh(note)
    return note


# ── 6b. GET /advisor/students/{id}/course-notes — gom notes theo mã môn (F4) ──

@app.get("/advisor/students/{student_id}/course-notes")
def advisor_get_course_notes(
    student_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    advisor = _require_advisor(authorization, db)
    _check_advisor_has_student(advisor, student_id, db)

    query = db.query(models.AdvisorNote).filter(
        models.AdvisorNote.student_id == student_id,
        models.AdvisorNote.course_code.isnot(None),
    )
    if advisor.role != "admin":
        query = query.filter(models.AdvisorNote.advisor_id == advisor.id)

    notes = query.order_by(models.AdvisorNote.created_at.desc()).all()
    grouped: dict[str, list] = {}
    for n in notes:
        grouped.setdefault(n.course_code, []).append({
            "id": n.id,
            "advisor_id": n.advisor_id,
            "content": n.content,
            "created_at": n.created_at.isoformat() if n.created_at else None,
            "updated_at": n.updated_at.isoformat() if n.updated_at else None,
        })
    return grouped


# ── /me/course-notes — SV xem note advisor đã ghi vào môn của mình (read-only)
@app.get("/me/course-notes")
def my_course_notes(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """SV xem note theo môn từ cố vấn của mình. Read-only — SV không sửa được."""
    user = _get_user_by_token(authorization, db)
    if user.role != "student":
        raise HTTPException(status_code=403, detail="Chỉ sinh viên dùng được endpoint này")

    notes = db.query(models.AdvisorNote).filter(
        models.AdvisorNote.student_id == user.id,
        models.AdvisorNote.course_code.isnot(None),
    ).order_by(models.AdvisorNote.created_at.desc()).all()

    advisor_ids = {n.advisor_id for n in notes}
    advisor_map = {
        u.id: u for u in db.query(models.User).filter(models.User.id.in_(advisor_ids)).all()
    } if advisor_ids else {}

    grouped: dict[str, list] = {}
    for n in notes:
        adv = advisor_map.get(n.advisor_id)
        grouped.setdefault(n.course_code, []).append({
            "id": n.id,
            "content": n.content,
            "advisor_name": (adv.full_name if adv else "Cố vấn"),
            "created_at": n.created_at.isoformat() if n.created_at else None,
            "updated_at": n.updated_at.isoformat() if n.updated_at else None,
        })
    return grouped


# ── 7. PUT /advisor/notes/{note_id} — sửa ghi chú ───────────────────────────

@app.put("/advisor/notes/{note_id}", response_model=schemas.AdvisorNoteOut)
def advisor_update_note(
    note_id: int,
    body: schemas.AdvisorNoteIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    advisor = _require_advisor(authorization, db)

    note = db.query(models.AdvisorNote).filter(models.AdvisorNote.id == note_id).first()
    if not note:
        raise HTTPException(status_code=404, detail="Không tìm thấy ghi chú")

    # Chỉ tác giả hoặc admin được sửa
    if advisor.role != "admin" and note.advisor_id != advisor.id:
        raise HTTPException(status_code=403, detail="Bạn không có quyền sửa ghi chú này")

    note.content = body.content
    note.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(note)
    return note


@app.delete("/advisor/notes/{note_id}")
def advisor_delete_note(
    note_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    advisor = _require_advisor(authorization, db)

    note = db.query(models.AdvisorNote).filter(models.AdvisorNote.id == note_id).first()
    if not note:
        raise HTTPException(status_code=404, detail="Không tìm thấy ghi chú")

    if advisor.role != "admin" and note.advisor_id != advisor.id:
        raise HTTPException(status_code=403, detail="Bạn không có quyền xoá ghi chú này")

    db.delete(note)
    db.commit()
    return {"ok": True, "deleted_id": note_id}


# ── Admin: Quản lý cố vấn ─────────────────────────────────────────────────────

@app.get("/admin/advisors", response_model=list[schemas.AdminAdvisorItem])
def admin_list_advisors(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    _require_admin(authorization, db)
    advisors = (
        db.query(models.User)
        .filter(models.User.role == "advisor")
        .order_by(models.User.id)
        .all()
    )
    result = []
    for adv in advisors:
        count = (
            db.query(models.AdvisorAssignment)
            .filter(models.AdvisorAssignment.advisor_id == adv.id)
            .count()
        )
        # Đếm số lớp GV chủ nhiệm
        class_count = db.query(models.ClassGroup).filter(
            models.ClassGroup.advisor_id == adv.id
        ).count()
        result.append(schemas.AdminAdvisorItem(
            id=adv.id,
            username=adv.username,
            teacher_code=adv.teacher_code,
            full_name=adv.full_name,
            managed_specialization=adv.managed_specialization,
            email=getattr(adv, "email", None),
            student_count=count,
            class_count=class_count,
            default_password=adv.default_password,
        ))
    return result


# Note: _check_head_conflict + _demote_existing_heads đã bị remove cùng với
# concept "Trưởng bộ môn" — model mới quản lý SV ↔ GV qua Lớp (ClassGroup).


def _parse_advisors_csv(data: bytes, filename: str) -> list[dict]:
    """Parse CSV/Excel → list of dicts: teacher_code, full_name, managed_specialization, email.

    File template 4 cột:
      | Mã GV | Họ tên | Email | Chuyên ngành |

    Lưu ý: cột "Trưởng BM" cũ đã bị BỎ — model mới quản lý qua Lớp (ClassGroup),
    không còn concept Trưởng bộ môn.
    """
    rows = read_rows_from_upload(filename, data)["rows"]
    if not rows:
        raise ValueError("File rỗng hoặc không đọc được")

    header_idx = 0
    headers = [str(c or "").strip().lower() for c in rows[0]]

    def _col(*keys):
        for k in keys:
            for i, h in enumerate(headers):
                if k in h:
                    return i
        return None

    idx_tc    = _col("teacher_code", "mã gv", "ma gv", "magv", "mã giảng viên", "tc")
    idx_name  = _col("full_name", "họ tên", "ho ten", "tên", "ten", "họ và tên")
    idx_email = _col("email", "email")
    idx_spec  = _col("managed_specialization", "specialization", "chuyên ngành", "chuyen nganh", "bộ môn", "bo mon")

    if idx_tc is None:
        raise ValueError("Không tìm thấy cột teacher_code / Mã GV trong file")

    result = []
    for row in rows[header_idx + 1:]:
        if not row or all(_is_missing(c) for c in row):
            continue
        tc = str(row[idx_tc] or "").strip().upper() if idx_tc < len(row) else ""
        if not tc:
            continue
        result.append({
            "teacher_code": tc,
            "full_name": str(row[idx_name] or "").strip() if idx_name is not None and idx_name < len(row) else "",
            "managed_specialization": str(row[idx_spec] or "").strip() if idx_spec is not None and idx_spec < len(row) else "",
            "email": str(row[idx_email] or "").strip().lower() if idx_email is not None and idx_email < len(row) else "",
        })
    return result


_IMPORT_VALID_SPECS = {"7480201_07", "7480201_06", "7480201_05", "7480201_09", "7480201_04", "7480201_08"}


@app.post("/admin/advisors/import", response_model=schemas.AdminAdvisorsImportOut)
def admin_import_advisors(
    file: UploadFile,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Import cố vấn hàng loạt từ file CSV/Excel."""
    admin = _require_admin(authorization, db)

    data = file.file.read(_MAX_UPLOAD_BYTES + 1)
    if len(data) > _MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File quá lớn. Giới hạn 10 MB.")

    try:
        records = _parse_advisors_csv(data, file.filename or "import.xlsx")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Không đọc được file: {exc}")

    if not records:
        raise HTTPException(status_code=400, detail="Không tìm thấy dữ liệu trong file")

    existing_usernames = {u.username for u in db.query(models.User.username).all()}

    import re as _re_imp, random as _rnd_imp
    _tc_pat = _re_imp.compile(r"^(KHMT|MMT|CNPM|HTTT|THKT|CNTTDH|GV)\d{3}$")

    created_count = 0
    skipped_count = 0
    errors: list[schemas.AdvisorImportError] = []
    generated_passwords: dict[str, str] = {}

    for row_idx, rec in enumerate(records, start=2):
        tc = rec["teacher_code"]
        full_name = rec["full_name"]
        spec = rec["managed_specialization"] or None
        if spec not in _IMPORT_VALID_SPECS:
            spec = None
        email = (rec.get("email") or "").strip().lower() or None
        if email and ("@" not in email or "." not in email.split("@")[-1]):
            email = None  # silently drop invalid email

        if not tc:
            errors.append(schemas.AdvisorImportError(row=row_idx, teacher_code=tc, reason="Mã GV trống"))
            continue
        if not _tc_pat.match(tc):
            errors.append(schemas.AdvisorImportError(row=row_idx, teacher_code=tc, reason=f"Mã GV không hợp lệ: {tc}"))
            continue
        if tc in existing_usernames:
            skipped_count += 1
            continue
        if not full_name:
            errors.append(schemas.AdvisorImportError(row=row_idx, teacher_code=tc, reason="Thiếu họ tên"))
            continue

        plain_pw = str(_rnd_imp.randint(100000, 999999))
        generated_passwords[tc] = plain_pw

        user = models.User(
            username=tc,
            password_hash=_hash_temp_password(plain_pw),
            full_name=normalize_vietnamese_name(full_name) or full_name,
            role="advisor",
            teacher_code=tc,
            managed_specialization=spec,
            email=email,
            default_password=plain_pw,
        )
        db.add(user)
        existing_usernames.add(tc)
        created_count += 1

    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Lỗi lưu dữ liệu: {exc}")

    _log(db, admin, "BULK_IMPORT_ADVISORS", "users", None,
         f"created={created_count} skipped={skipped_count} errors={len(errors)}")

    return schemas.AdminAdvisorsImportOut(
        created_count=created_count,
        skipped_count=skipped_count,
        errors=errors,
        generated_passwords=generated_passwords,
    )


@app.post("/admin/advisors", response_model=schemas.AdminCreateAdvisorOut)
def admin_create_advisor(
    payload: schemas.AdminCreateAdvisorIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    teacher_code = validate_teacher_code(payload.teacher_code)
    if db.query(models.User).filter(models.User.username == teacher_code).first():
        raise HTTPException(status_code=400, detail=f"Mã GV '{teacher_code}' đã tồn tại")
    if db.query(models.User).filter(models.User.teacher_code == teacher_code).first():
        raise HTTPException(status_code=400, detail=f"Mã GV '{teacher_code}' đã tồn tại")
    spec = payload.managed_specialization or None
    import random as _random
    raw_pw = str(_random.randint(100000, 999999))
    hashed = _hash_temp_password(raw_pw)
    user = models.User(
        username=teacher_code,
        teacher_code=teacher_code,
        password_hash=hashed,
        full_name=normalize_vietnamese_name(payload.full_name),
        role="advisor",
        managed_specialization=spec,
        email=getattr(payload, "email", None),
        default_password=raw_pw,
        is_first_login=True,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    _log(db, admin, "create_advisor", "user", str(user.id), f"Tạo cố vấn {user.username}")
    return schemas.AdminCreateAdvisorOut(
        id=user.id,
        username=user.username,
        teacher_code=user.teacher_code,
        full_name=user.full_name,
        managed_specialization=user.managed_specialization,
        email=getattr(user, "email", None),
        role=user.role,
        password_plain=raw_pw,
    )


@app.patch("/admin/advisors/{advisor_id}", response_model=schemas.AdminAdvisorItem)
def admin_update_advisor(
    advisor_id: int,
    payload: schemas.AdminUpdateAdvisorIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    adv = db.query(models.User).filter(
        models.User.id == advisor_id,
        models.User.role == "advisor",
    ).first()
    if not adv:
        raise HTTPException(status_code=404, detail="Không tìm thấy cố vấn")

    if payload.full_name is not None:
        adv.full_name = normalize_vietnamese_name(payload.full_name) or adv.full_name
    if payload.managed_specialization is not None:
        adv.managed_specialization = payload.managed_specialization or None
    # email field removed 2026-05-05

    db.commit()
    db.refresh(adv)
    _log(db, admin, "update_advisor", "user", str(advisor_id), f"Cập nhật cố vấn {adv.username}")

    count = db.query(models.AdvisorAssignment).filter(
        models.AdvisorAssignment.advisor_id == adv.id
    ).count()
    class_count = db.query(models.ClassGroup).filter(
        models.ClassGroup.advisor_id == adv.id
    ).count()
    return schemas.AdminAdvisorItem(
        id=adv.id,
        username=adv.username,
        teacher_code=adv.teacher_code,
        full_name=adv.full_name,
        managed_specialization=adv.managed_specialization,
        email=getattr(adv, "email", None),
        student_count=count,
        class_count=class_count,
    )


@app.delete("/admin/advisors/{advisor_id}", response_model=schemas.MessageOut)
def admin_delete_advisor(
    advisor_id: int,
    transfer_to: int | None = None,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Xoá cố vấn — block nếu còn lớp chủ nhiệm.

    Quy tắc:
    - Nếu GV còn chủ nhiệm Lớp nào: 409 Conflict + danh sách lớp.
      Admin phải PATCH các lớp đó (đổi GVCN) trước khi xoá.
    - Nếu còn SV phụ trách qua AdvisorAssignment (không qua lớp) và bộ môn còn
      GV khác: bắt buộc `transfer_to` để chuyển SV sang người khác.
    - Nếu không còn ràng buộc: cho xoá ngay.
    """
    admin = _require_admin(authorization, db)
    adv = db.query(models.User).filter(
        models.User.id == advisor_id,
        models.User.role == "advisor",
    ).first()
    if not adv:
        raise HTTPException(status_code=404, detail="Không tìm thấy cố vấn")

    # Block: GV còn chủ nhiệm lớp nào không?
    classes_owned = db.query(models.ClassGroup).filter(
        models.ClassGroup.advisor_id == advisor_id
    ).all()
    if classes_owned:
        raise HTTPException(
            status_code=409,
            detail={
                "message": (
                    f"GV {adv.username} còn chủ nhiệm {len(classes_owned)} lớp. "
                    "Hãy phân GVCN mới cho các lớp này trước khi xoá."
                ),
                "classes": [
                    {"id": c.id, "code": c.code, "name": c.name}
                    for c in classes_owned
                ],
            },
        )

    student_count = db.query(models.AdvisorAssignment).filter(
        models.AdvisorAssignment.advisor_id == advisor_id
    ).count()
    other_in_spec = db.query(models.User).filter(
        models.User.role == "advisor",
        models.User.managed_specialization == adv.managed_specialization,
        models.User.id != adv.id,
    ).all()

    needs_transfer = student_count > 0 and len(other_in_spec) > 0

    if needs_transfer and transfer_to is None:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Không thể xóa: cần chọn cố vấn kế nhiệm để chuyển giao "
                f"({student_count} SV phụ trách)."
            ),
        )

    new_owner = None
    if transfer_to is not None:
        new_owner = db.query(models.User).filter(
            models.User.id == transfer_to,
            models.User.role == "advisor",
        ).first()
        if not new_owner:
            raise HTTPException(status_code=404, detail="Không tìm thấy cố vấn kế nhiệm")
        if new_owner.id == adv.id:
            raise HTTPException(status_code=400, detail="Cố vấn kế nhiệm phải khác người đang xóa")
        if (new_owner.managed_specialization or "") != (adv.managed_specialization or ""):
            raise HTTPException(status_code=400, detail="Cố vấn kế nhiệm phải cùng chuyên ngành")

    transferred_students = 0

    if new_owner is not None:
        result = db.query(models.AdvisorAssignment).filter(
            models.AdvisorAssignment.advisor_id == advisor_id
        ).update({"advisor_id": new_owner.id})
        transferred_students = result or 0
        db.flush()
    else:
        db.query(models.AdvisorAssignment).filter(
            models.AdvisorAssignment.advisor_id == advisor_id
        ).delete()

    db.query(models.AdvisorNote).filter(
        models.AdvisorNote.advisor_id == advisor_id
    ).delete()

    db.delete(adv)
    db.commit()

    detail_parts = [f"Xóa cố vấn {adv.username}"]
    if transferred_students:
        detail_parts.append(f"chuyển {transferred_students} SV sang {new_owner.username}")
    _log(db, admin, "delete_advisor", "user", str(advisor_id), " · ".join(detail_parts))

    msg = f"Đã xóa cố vấn {adv.username}"
    if transferred_students:
        msg += f" · Đã chuyển {transferred_students} SV sang {new_owner.full_name or new_owner.username}"
    return {"message": msg}


# Note: endpoint POST /admin/advisors/{id}/transfer-head đã bị remove cùng concept
# "Trưởng bộ môn". Để chuyển GVCN của lớp, dùng PATCH /admin/classes/{id}.


@app.get("/admin/advisors/{advisor_id}/students", response_model=list[schemas.AdminAdvisorStudentItem])
def admin_list_advisor_students(
    advisor_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    _require_admin(authorization, db)
    adv = db.query(models.User).filter(
        models.User.id == advisor_id, models.User.role == "advisor"
    ).first()
    if not adv:
        raise HTTPException(status_code=404, detail="Không tìm thấy cố vấn")
    assignments = (
        db.query(models.AdvisorAssignment)
        .filter(models.AdvisorAssignment.advisor_id == advisor_id)
        .order_by(models.AdvisorAssignment.assigned_at.desc())
        .all()
    )
    result = []
    for asgn in assignments:
        sv = db.query(models.User).filter(models.User.id == asgn.student_id).first()
        if sv:
            result.append(schemas.AdminAdvisorStudentItem(
                assignment_id=asgn.id,
                student_id=sv.id,
                username=sv.username,
                full_name=sv.full_name,
                assigned_at=asgn.assigned_at,
            ))
    return result


@app.post("/admin/advisors/{advisor_id}/assign", response_model=schemas.MessageOut)
def admin_assign_student(
    advisor_id: int,
    payload: schemas.AdminAssignStudentIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    adv = db.query(models.User).filter(
        models.User.id == advisor_id, models.User.role == "advisor"
    ).first()
    if not adv:
        raise HTTPException(status_code=404, detail="Không tìm thấy cố vấn")
    sv = db.query(models.User).filter(
        models.User.id == payload.student_id, models.User.role == "student"
    ).first()
    if not sv:
        raise HTTPException(status_code=404, detail="Không tìm thấy sinh viên")
    # Quy tắc: 1 SV chỉ có 1 cố vấn. Nếu SV đã có assignment với advisor khác → REPLACE.
    existing = db.query(models.AdvisorAssignment).filter(
        models.AdvisorAssignment.student_id == payload.student_id,
    ).first()
    if existing and existing.advisor_id == advisor_id:
        raise HTTPException(status_code=400, detail="Sinh viên đã được phân công cho cố vấn này")
    if existing:
        # Replace: drop old assignment + connection (nếu có)
        old_adv_id = existing.advisor_id
        db.query(models.AdvisorAssignment).filter(
            models.AdvisorAssignment.id == existing.id
        ).delete()
        db.flush()
        # Optional: drop old connection (giữ lại lịch sử chat thì giữ luôn)
    asgn = models.AdvisorAssignment(
        advisor_id=advisor_id,
        student_id=payload.student_id,
    )
    db.add(asgn)
    # Auto-create accepted connection advisor ↔ student để chat ngay không cần lời mời
    existing_conn = db.query(models.UserConnection).filter(
        ((models.UserConnection.from_id == advisor_id) & (models.UserConnection.to_id == payload.student_id)) |
        ((models.UserConnection.from_id == payload.student_id) & (models.UserConnection.to_id == advisor_id))
    ).first()
    if not existing_conn:
        db.add(models.UserConnection(from_id=advisor_id, to_id=payload.student_id, status="accepted"))
    elif existing_conn.status != "accepted":
        existing_conn.status = "accepted"
    db.commit()
    _log(db, admin, "assign_student", "advisor_assignment", str(asgn.id),
         f"Phân công SV {sv.username} → cố vấn {adv.username}")
    return {"message": f"Đã phân công {sv.username} cho cố vấn {adv.username}"}


@app.post("/admin/advisors/{advisor_id}/assign/bulk", response_model=schemas.AdminAdvisorBulkAssignOut)
def admin_bulk_assign_students(
    advisor_id: int,
    payload: schemas.AdminAdvisorBulkAssignIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    adv = db.query(models.User).filter(
        models.User.id == advisor_id, models.User.role == "advisor"
    ).first()
    if not adv:
        raise HTTPException(status_code=404, detail="Không tìm thấy cố vấn")

    assigned_count = 0
    skipped_count = 0
    errors: list[str] = []

    for uname in payload.student_usernames:
        uname = uname.strip()
        if not uname:
            continue
        sv = db.query(models.User).filter(
            models.User.username == uname, models.User.role == "student"
        ).first()
        if not sv:
            errors.append(f"{uname}: không tìm thấy sinh viên")
            continue
        # 1 SV → 1 advisor. Nếu đã thuộc advisor này → skip; thuộc advisor khác → REPLACE.
        existing = db.query(models.AdvisorAssignment).filter(
            models.AdvisorAssignment.student_id == sv.id,
        ).first()
        if existing and existing.advisor_id == advisor_id:
            skipped_count += 1
            continue
        if existing:
            db.query(models.AdvisorAssignment).filter(
                models.AdvisorAssignment.id == existing.id
            ).delete()
            db.flush()
        db.add(models.AdvisorAssignment(advisor_id=advisor_id, student_id=sv.id))
        existing_conn = db.query(models.UserConnection).filter(
            ((models.UserConnection.from_id == advisor_id) & (models.UserConnection.to_id == sv.id)) |
            ((models.UserConnection.from_id == sv.id) & (models.UserConnection.to_id == advisor_id))
        ).first()
        if not existing_conn:
            db.add(models.UserConnection(from_id=advisor_id, to_id=sv.id, status="accepted"))
        elif existing_conn.status != "accepted":
            existing_conn.status = "accepted"
        assigned_count += 1

    db.commit()
    _log(db, admin, "bulk_assign_students", "advisor_assignment", str(advisor_id),
         f"Bulk assign {assigned_count} SV → cố vấn {adv.username}")
    return schemas.AdminAdvisorBulkAssignOut(
        assigned_count=assigned_count,
        skipped_count=skipped_count,
        errors=errors,
    )


@app.delete("/admin/advisors/{advisor_id}/assign/{student_id}", response_model=schemas.MessageOut)
def admin_unassign_student(
    advisor_id: int,
    student_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    admin = _require_admin(authorization, db)
    asgn = db.query(models.AdvisorAssignment).filter(
        models.AdvisorAssignment.advisor_id == advisor_id,
        models.AdvisorAssignment.student_id == student_id,
    ).first()
    if not asgn:
        raise HTTPException(status_code=404, detail="Không tìm thấy phân công")
    db.delete(asgn)
    db.commit()
    _log(db, admin, "unassign_student", "advisor_assignment", str(asgn.id),
         f"Hủy phân công SV {student_id} khỏi cố vấn {advisor_id}")
    return {"message": "Đã hủy phân công"}


@app.get("/admin/advisors/{advisor_id}/default-password", response_model=schemas.AdminDefaultPasswordOut)
def admin_get_advisor_default_password(
    advisor_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    _require_admin(authorization, db)
    u = db.query(models.User).filter(
        models.User.id == advisor_id, models.User.role == "advisor"
    ).first()
    if not u:
        raise HTTPException(status_code=404, detail="Không tìm thấy cố vấn")
    return schemas.AdminDefaultPasswordOut(
        has_default=bool(u.default_password),
        password=u.default_password,
    )


@app.post("/admin/advisors/{advisor_id}/reset-password", response_model=schemas.AdminResetPasswordOut)
def admin_reset_advisor_password(
    advisor_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    import random as _rnd
    admin = _require_admin(authorization, db)
    u = db.query(models.User).filter(
        models.User.id == advisor_id, models.User.role == "advisor"
    ).first()
    if not u:
        raise HTTPException(status_code=404, detail="Không tìm thấy cố vấn")
    new_pw = str(_rnd.randint(100000, 999999))
    u.password_hash = _hash_temp_password(new_pw)
    u.default_password = new_pw
    u.is_first_login = True
    db.commit()
    _clear_failed_login(u.username, db)  # xóa rate limit để advisor login ngay được
    _log(db, admin, "RESET_PASSWORD", "advisor", u.username, "admin reset advisor pw")
    return schemas.AdminResetPasswordOut(
        message=f"Đã reset mật khẩu cho cố vấn {u.username}",
        password=new_pw,
    )


@app.get("/admin/advisors/next-code")
def admin_next_advisor_code(
    specialization: str | None = None,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Trả về mã GV tiếp theo theo bộ môn (prefix tự động theo specialization)."""
    _require_admin(authorization, db)
    prefix = _SPEC_TC_PREFIX.get(specialization, "GV")
    existing = db.query(models.User.teacher_code).filter(
        models.User.role == "advisor",
        models.User.teacher_code.like(f"{prefix}%"),
    ).all()
    max_seq = 0
    import re as _re
    for (tc,) in existing:
        if tc:
            m = _re.match(rf"^{prefix}(\d{{3}})$", tc.upper())
            if m:
                max_seq = max(max_seq, int(m.group(1)))
    next_code = f"{prefix}{max_seq + 1:03d}"
    return {"next_code": next_code, "prefix": prefix}


@app.get("/admin/students/next-code")
def admin_next_student_code(
    cohort: str,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Trả mã SV tiếp theo cho khóa.

    Format demo: ``sv{KK}{NNN}`` — vd `sv14001` (`sv` + 2 số khoá + 3 số STT).
    Tối đa 999 SV/khoá (đủ demo). Lowercase.

    Backward compat: vẫn nhận diện 2 format cũ để tiếp tục seq đúng:
      - 10 số `{KK}{NNNNNNNN}` (vd `1400000001`)
      - SV + 6 số `SV{KK}{NNNN}` (vd `SV140001`)
    """
    _require_admin(authorization, db)
    import re as _re
    from sqlalchemy import or_
    if not _re.match(r"^\d{2}$", cohort):
        raise HTTPException(status_code=422, detail="cohort phải là 2 chữ số (VD: 14)")
    pat_new = _re.compile(rf"^sv{cohort}(\d{{3}})$", _re.IGNORECASE)
    pat_legacy_10 = _re.compile(rf"^{cohort}(\d{{8}})$")
    pat_legacy_sv = _re.compile(rf"^SV{cohort}(\d{{4}})$", _re.IGNORECASE)
    existing = db.query(models.User.username).filter(
        models.User.role == "student",
        or_(
            models.User.username.ilike(f"sv{cohort}%"),
            models.User.username.ilike(f"{cohort}%"),
            models.User.username.ilike(f"SV{cohort}%"),
        ),
    ).all()
    max_seq = 0
    for (uname,) in existing:
        u = uname or ""
        for pat in (pat_new, pat_legacy_10, pat_legacy_sv):
            m = pat.match(u)
            if m:
                try:
                    max_seq = max(max_seq, int(m.group(1)))
                except (ValueError, IndexError):
                    pass
                break
    next_code = f"sv{cohort}{max_seq + 1:03d}"
    return {"next_code": next_code, "cohort": cohort}


@app.post("/admin/students/auto-classify")
def admin_auto_classify_students(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Tự động phân chuyên ngành cho SV dựa trên môn CN đã học trong bảng điểm.

    Nghiệp vụ:
    - SV chưa học môn CN nào → spec=NULL (đại cương) → advisor = trưởng BM đại cương
    - SV đã học ≥1 môn CN → spec = CN có nhiều môn nhất → advisor = trưởng BM CN đó
    - Nếu không có trưởng BM cho CN đó → giữ ở đại cương

    Idempotent: chạy nhiều lần OK.
    """
    admin = _require_admin(authorization, db)
    from collections import Counter

    students = db.query(models.User).filter(models.User.role == "student").all()
    course_map = {c.course_code: c for c in db.query(models.Course).all()}
    # M2M map: course_code → list[spec]
    m2m_for_course: dict[str, list[str]] = {}
    for r in db.query(models.CourseSpecialization).all():
        m2m_for_course.setdefault(r.course_code, []).append(r.specialization)

    def _unique_spec_for(course) -> str | None:
        """Trả CN duy nhất gắn với môn này, hoặc None nếu không có / nhiều CN.
        Chỉ môn có ĐÚNG 1 CN (single-spec) mới uniquely identify CN của SV.
        Bỏ qua môn HK1-6 (đại cương + cơ sở) — kể cả khi dữ liệu sai vẫn không classify nhầm."""
        if not course.typical_semester or course.typical_semester < 7:
            return None
        specs = set()
        if course.required_specialization:
            specs.add(course.required_specialization)
        for s in m2m_for_course.get(course.course_code, []):
            specs.add(s)
        # Multi-spec courses (vd: pool_c của nhiều CN) không xác định CN — bỏ qua
        if len(specs) != 1:
            return None
        return next(iter(specs))

    classified_to_spec = 0
    classified_to_general = 0
    advisor_changes = 0
    skipped_no_head = 0

    for sv in students:
        grades = db.query(models.UserGrade).filter(models.UserGrade.user_id == sv.id).all()
        spec_counter: Counter = Counter()
        for g in grades:
            c = course_map.get(g.course_code)
            if not c:
                continue
            unique_spec = _unique_spec_for(c)
            if unique_spec:
                spec_counter[unique_spec] += 1

        new_spec = spec_counter.most_common(1)[0][0] if spec_counter else None
        old_spec = sv.specialization

        if new_spec != old_spec:
            sv.specialization = new_spec
            db.flush()
        if new_spec:
            classified_to_spec += 1
        else:
            classified_to_general += 1

        # Re-assign advisor (qua class_group nếu có, fallback round-robin)
        target, warn = assign_advisor_for_student(db, sv.id, new_spec)
        if not target:
            skipped_no_head += 1
            continue
        advisor_changes += 1

    db.commit()
    _log(db, admin, "auto_classify_students", "students", "*",
         f"Auto-classify: {classified_to_spec} CN, {classified_to_general} đại cương, {advisor_changes} đổi GV")
    return {
        "total_students": len(students),
        "classified_to_spec": classified_to_spec,
        "classified_to_general": classified_to_general,
        "advisor_reassignments": advisor_changes,
        "skipped_no_head": skipped_no_head,
        "message": f"Đã phân loại {len(students)} SV ({classified_to_spec} có CN, {classified_to_general} đại cương) · {advisor_changes} thay đổi cố vấn",
    }


@app.post("/admin/students/bulk-set-specialization")
def admin_bulk_set_specialization(
    payload: dict = Body(...),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Admin nhập danh sách 'username → spec_code' (sau khi phòng đào tạo gửi danh sách).

    Input: { items: [{ username: str, specialization: str }, ...] }
      - specialization có thể là code (7480201_07) hoặc null/empty (đại cương)

    Tự động cập nhật spec + reassign advisor sang trưởng BM tương ứng.
    """
    admin = _require_admin(authorization, db)
    items = payload.get("items") or []
    if not isinstance(items, list) or not items:
        raise HTTPException(status_code=422, detail="items phải là list không rỗng")

    updated = 0
    reassigned = 0
    errors: list[str] = []

    for it in items:
        uname = (it.get("username") or "").strip()
        new_spec = (it.get("specialization") or "").strip() or None
        if not uname:
            errors.append("Thiếu username"); continue
        sv = db.query(models.User).filter(
            models.User.username == uname, models.User.role == "student"
        ).first()
        if not sv:
            errors.append(f"{uname}: không tìm thấy SV"); continue

        if sv.specialization != new_spec:
            sv.specialization = new_spec
            db.flush()
            updated += 1

        # Re-assign advisor (qua class_group nếu có, fallback round-robin)
        target, warn = assign_advisor_for_student(db, sv.id, new_spec)
        if not target:
            errors.append(f"{uname}: {warn}")
            continue
        reassigned += 1

    db.commit()
    _log(db, admin, "bulk_set_specialization", "students", "*",
         f"Bulk set spec: updated={updated}, reassigned={reassigned}")
    return {
        "updated_specialization": updated,
        "reassigned_advisor": reassigned,
        "errors": errors,
        "message": f"Cập nhật {updated} CN · phân lại {reassigned} GV · {len(errors)} lỗi",
    }


# ════════════════════════════════════════════════════════════════════════════
# CLASS GROUP MANAGEMENT (Lớp sinh hoạt — GVCN ↔ SV)
# ════════════════════════════════════════════════════════════════════════════

def sync_advisor_for_class(db: Session, class_group: models.ClassGroup) -> int:
    """Re-sync AdvisorAssignment cho mọi SV trong lớp.

    Trigger: sau khi class_group.advisor_id được update, hoặc sau khi SV được
    add/remove khỏi lớp. Đảm bảo invariant: SV trong lớp X có advisor = lớp X
    .advisor_id. KHÔNG commit — caller tự commit.

    Trả về số SV được sync (re-assigned).
    """
    students = db.query(models.User).filter(
        models.User.class_group_id == class_group.id,
        models.User.role == "student",
    ).all()
    student_ids = [s.id for s in students]
    if not student_ids:
        return 0

    # Xoá assignments cũ — sau đó chỉ tạo mới (không UPSERT) để giữ logic đơn giản
    db.query(models.AdvisorAssignment).filter(
        models.AdvisorAssignment.student_id.in_(student_ids)
    ).delete(synchronize_session=False)

    # Bulk insert assignments mới với GVCN của lớp
    for sid in student_ids:
        db.add(models.AdvisorAssignment(
            advisor_id=class_group.advisor_id,
            student_id=sid,
        ))

    db.flush()
    return len(student_ids)


def _class_group_to_out(db: Session, cg: models.ClassGroup) -> schemas.ClassGroupOut:
    """Build ClassGroupOut với joined data (advisor + student count)."""
    advisor = db.query(models.User).filter(models.User.id == cg.advisor_id).first()
    student_count = db.query(models.User).filter(
        models.User.class_group_id == cg.id,
        models.User.role == "student",
    ).count()
    return schemas.ClassGroupOut(
        id=cg.id,
        code=cg.code,
        name=cg.name,
        cohort=cg.cohort,
        specialization=cg.specialization,
        advisor_id=cg.advisor_id,
        advisor_teacher_code=advisor.teacher_code if advisor else None,
        advisor_full_name=advisor.full_name if advisor else None,
        student_count=student_count,
        created_at=cg.created_at,
    )


def _class_default_name(code: str) -> str:
    """Auto-generate tên đẹp từ mã lớp DCCTCT66_07A → 'Lớp K66 CNPM 07A'."""
    try:
        cohort, spec_full, letter = schemas.parse_class_code(code)
        spec_short = _SPEC_DISPLAY.get(spec_full, spec_full)
        return f"Lớp K{cohort} {spec_short} {letter}"
    except Exception:
        return f"Lớp {code}"


@app.get("/admin/classes", response_model=list[schemas.ClassGroupOut])
def admin_list_classes(
    cohort: str | None = None,
    specialization: str | None = None,
    advisor_id: int | None = None,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """List lớp với filters optional."""
    _require_admin(authorization, db)
    q = db.query(models.ClassGroup)
    if cohort:
        q = q.filter(models.ClassGroup.cohort == cohort)
    if specialization:
        q = q.filter(models.ClassGroup.specialization == specialization)
    if advisor_id:
        q = q.filter(models.ClassGroup.advisor_id == advisor_id)
    classes = q.order_by(models.ClassGroup.code).all()
    return [_class_group_to_out(db, c) for c in classes]


@app.post("/admin/classes", response_model=schemas.ClassGroupOut)
def admin_create_class(
    payload: schemas.AdminCreateClassIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Tạo 1 lớp thủ công. GVCN bắt buộc."""
    admin = _require_admin(authorization, db)

    code = payload.code  # đã validate format trong schema
    cohort, spec, _letter = schemas.parse_class_code(code)

    # Validate advisor tồn tại + role=advisor
    tc = payload.advisor_teacher_code.strip().upper()
    advisor = db.query(models.User).filter(
        models.User.teacher_code == tc,
        models.User.role == "advisor",
    ).first()
    if not advisor:
        raise HTTPException(
            status_code=400,
            detail=f"Không tìm thấy GV với mã '{tc}'. Tạo GV trước khi tạo lớp."
        )

    # Check trùng code
    if db.query(models.ClassGroup).filter(models.ClassGroup.code == code).first():
        raise HTTPException(status_code=409, detail=f"Mã lớp '{code}' đã tồn tại")

    name = (payload.name or "").strip() or _class_default_name(code)
    cg = models.ClassGroup(
        code=code,
        name=name,
        cohort=cohort,
        specialization=spec,
        advisor_id=advisor.id,
    )
    db.add(cg)
    db.commit()
    db.refresh(cg)

    _log(db, admin, "CREATE_CLASS", "class_group", str(cg.id),
         f"code={code} advisor={advisor.username}")

    return _class_group_to_out(db, cg)


@app.get("/admin/classes/{class_id}", response_model=schemas.ClassGroupOut)
def admin_get_class(
    class_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    _require_admin(authorization, db)
    cg = db.query(models.ClassGroup).filter(models.ClassGroup.id == class_id).first()
    if not cg:
        raise HTTPException(status_code=404, detail="Không tìm thấy lớp")
    return _class_group_to_out(db, cg)


@app.patch("/admin/classes/{class_id}", response_model=schemas.ClassGroupOut)
def admin_update_class(
    class_id: int,
    payload: schemas.AdminUpdateClassIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Cập nhật tên hoặc GVCN. Đổi GVCN → tự sync AdvisorAssignment."""
    admin = _require_admin(authorization, db)
    cg = db.query(models.ClassGroup).filter(models.ClassGroup.id == class_id).first()
    if not cg:
        raise HTTPException(status_code=404, detail="Không tìm thấy lớp")

    changes: list[str] = []

    if payload.name is not None:
        new_name = (payload.name or "").strip()
        if new_name and new_name != cg.name:
            cg.name = new_name
            changes.append(f"name='{new_name}'")

    advisor_changed = False
    if payload.advisor_teacher_code is not None:
        tc = payload.advisor_teacher_code.strip().upper()
        if not tc:
            raise HTTPException(status_code=400, detail="Mã GVCN không được để trống")
        advisor = db.query(models.User).filter(
            models.User.teacher_code == tc,
            models.User.role == "advisor",
        ).first()
        if not advisor:
            raise HTTPException(
                status_code=400,
                detail=f"Không tìm thấy GV với mã '{tc}'"
            )
        if advisor.id != cg.advisor_id:
            old_advisor_id = cg.advisor_id
            cg.advisor_id = advisor.id
            db.flush()  # commit FK change trước khi sync
            advisor_changed = True
            changes.append(f"advisor: {old_advisor_id}→{advisor.id}")

    synced = 0
    if advisor_changed:
        synced = sync_advisor_for_class(db, cg)
        changes.append(f"synced {synced} SV")

    db.commit()
    db.refresh(cg)

    if changes:
        _log(db, admin, "UPDATE_CLASS", "class_group", str(class_id),
             " · ".join(changes))

    return _class_group_to_out(db, cg)


@app.delete("/admin/classes/{class_id}", response_model=schemas.MessageOut)
def admin_delete_class(
    class_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Xoá lớp. Chặn nếu lớp còn SV — admin phải chuyển SV sang lớp khác trước."""
    admin = _require_admin(authorization, db)
    cg = db.query(models.ClassGroup).filter(models.ClassGroup.id == class_id).first()
    if not cg:
        raise HTTPException(status_code=404, detail="Không tìm thấy lớp")

    student_count = db.query(models.User).filter(
        models.User.class_group_id == class_id,
        models.User.role == "student",
    ).count()
    if student_count > 0:
        raise HTTPException(
            status_code=409,
            detail=(
                f"Lớp '{cg.code}' còn {student_count} sinh viên. "
                "Hãy chuyển SV sang lớp khác trước khi xoá."
            ),
        )

    code_snapshot = cg.code
    db.delete(cg)
    db.commit()

    _log(db, admin, "DELETE_CLASS", "class_group", str(class_id),
         f"code={code_snapshot}")
    return {"message": f"Đã xoá lớp {code_snapshot}"}


def _parse_classes_csv(data: bytes, filename: str) -> list[dict]:
    """Parse CSV/Excel → list of dicts: code, name, advisor_teacher_code.

    File template 5 cột (chỉ 'Mã lớp' + 'Mã GVCN' bắt buộc):
      | Mã lớp | Tên lớp | Khoá | Chuyên ngành | Mã GVCN |

    Khoá + Chuyên ngành chỉ mang tính tham khảo cho admin — backend tự derive
    từ Mã lớp (parse_class_code). Khi có sự khác biệt, ưu tiên parse từ Mã lớp.
    """
    rows = read_rows_from_upload(filename, data)["rows"]
    if not rows:
        raise ValueError("File rỗng hoặc không đọc được")

    headers = [str(c or "").strip().lower() for c in rows[0]]

    def _col(*keys):
        for k in keys:
            for i, h in enumerate(headers):
                if k in h:
                    return i
        return None

    idx_code = _col("mã lớp", "ma lop", "code", "class_code")
    idx_name = _col("tên lớp", "ten lop", "name", "class_name")
    idx_adv  = _col("mã gvcn", "ma gvcn", "gvcn", "advisor_code", "mã gv chủ nhiệm", "ma gv chu nhiem")

    if idx_code is None:
        raise ValueError("Không tìm thấy cột 'Mã lớp' trong file")
    if idx_adv is None:
        raise ValueError("Không tìm thấy cột 'Mã GVCN' trong file")

    result = []
    for row in rows[1:]:
        if not row or all(_is_missing(c) for c in row):
            continue
        code = str(row[idx_code] or "").strip().upper() if idx_code < len(row) else ""
        if not code:
            continue
        result.append({
            "code": code,
            "name": str(row[idx_name] or "").strip() if idx_name is not None and idx_name < len(row) else "",
            "advisor_teacher_code": str(row[idx_adv] or "").strip().upper() if idx_adv < len(row) else "",
        })
    return result


@app.post("/admin/classes/import", response_model=schemas.AdminClassesImportOut)
def admin_import_classes(
    file: UploadFile,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Import bulk lớp từ Excel/CSV. UPSERT — re-import an toàn.

    Reject row nếu:
      • Mã lớp sai format
      • Mã GVCN trống / không tồn tại / không phải role advisor
    """
    admin = _require_admin(authorization, db)

    data = file.file.read(_MAX_UPLOAD_BYTES + 1)
    if len(data) > _MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File quá lớn. Giới hạn 10 MB.")

    try:
        records = _parse_classes_csv(data, file.filename or "classes.xlsx")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Không đọc được file: {exc}")

    if not records:
        raise HTTPException(status_code=400, detail="Không tìm thấy dữ liệu trong file")

    # Pre-load advisors map (teacher_code → User) để tránh N queries
    advisors_by_tc = {
        a.teacher_code: a
        for a in db.query(models.User).filter(
            models.User.role == "advisor",
            models.User.teacher_code.isnot(None),
        ).all()
    }
    existing_classes = {
        c.code: c for c in db.query(models.ClassGroup).all()
    }

    created_count = 0
    updated_count = 0
    skipped_count = 0
    errors: list[schemas.ClassImportError] = []
    classes_to_sync: list[models.ClassGroup] = []

    for row_idx, rec in enumerate(records, start=2):
        code = rec["code"]
        try:
            cohort, spec, _letter = schemas.parse_class_code(code)
        except ValueError as exc:
            errors.append(schemas.ClassImportError(row=row_idx, code=code, reason=str(exc)))
            continue

        adv_tc = rec["advisor_teacher_code"]
        if not adv_tc:
            errors.append(schemas.ClassImportError(
                row=row_idx, code=code, reason="Thiếu Mã GVCN"
            ))
            continue
        advisor = advisors_by_tc.get(adv_tc)
        if not advisor:
            errors.append(schemas.ClassImportError(
                row=row_idx, code=code,
                reason=f"GV '{adv_tc}' chưa tồn tại — tạo GV trước"
            ))
            continue

        name = rec["name"] or _class_default_name(code)

        existing = existing_classes.get(code)
        if existing:
            # UPSERT: update name + advisor nếu khác
            advisor_changed = False
            if existing.advisor_id != advisor.id:
                existing.advisor_id = advisor.id
                advisor_changed = True
            if existing.name != name:
                existing.name = name
            if advisor_changed:
                classes_to_sync.append(existing)
            updated_count += 1
        else:
            cg = models.ClassGroup(
                code=code,
                name=name,
                cohort=cohort,
                specialization=spec,
                advisor_id=advisor.id,
            )
            db.add(cg)
            existing_classes[code] = cg
            created_count += 1
            # Lớp mới chưa có SV → không cần sync. Sync khi import SV.

    try:
        db.flush()  # cần flush để lớp mới có id trước khi sync
        # Sync các lớp có advisor thay đổi
        for cg in classes_to_sync:
            sync_advisor_for_class(db, cg)
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Lỗi lưu dữ liệu: {exc}")

    _log(db, admin, "BULK_IMPORT_CLASSES", "class_groups", None,
         f"created={created_count} updated={updated_count} errors={len(errors)}")

    return schemas.AdminClassesImportOut(
        created_count=created_count,
        updated_count=updated_count,
        skipped_count=skipped_count,
        errors=errors,
    )


# ════════════════════════════════════════════════════════════════════════════
# END CLASS GROUP
# ════════════════════════════════════════════════════════════════════════════


# ════════════════════════════════════════════════════════════════════════════
# ADMIN REVIEWS DASHBOARD (course rating moderation)
# ════════════════════════════════════════════════════════════════════════════

@app.get("/admin/reviews", response_model=schemas.AdminReviewListOut)
def admin_list_reviews(
    course_code: str | None = None,
    student_username: str | None = None,
    min_rating: int | None = None,
    max_rating: int | None = None,
    include_hidden: bool = True,
    limit: int = 50,
    offset: int = 0,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """List tất cả reviews với filters cho admin moderation.

    Filters:
        course_code: lọc theo môn
        student_username: lọc theo MSSV
        min_rating / max_rating: lọc theo số sao
        include_hidden: True (default) → cả review đã ẩn; False → chỉ hiển thị
    """
    _require_admin(authorization, db)

    safe_limit = max(1, min(int(limit), 200))
    safe_offset = max(0, int(offset))

    q = db.query(models.CourseRating)
    if course_code:
        q = q.filter(models.CourseRating.course_code == course_code)
    if min_rating is not None:
        q = q.filter(models.CourseRating.rating >= int(min_rating))
    if max_rating is not None:
        q = q.filter(models.CourseRating.rating <= int(max_rating))
    if not include_hidden:
        q = q.filter(models.CourseRating.hidden == False)  # noqa: E712

    if student_username:
        sv = db.query(models.User).filter(
            models.User.username == student_username.strip().lower()
        ).first()
        if sv:
            q = q.filter(models.CourseRating.user_id == sv.id)
        else:
            return schemas.AdminReviewListOut(total=0, reviews=[])

    total = q.count()
    rows = q.order_by(models.CourseRating.id.desc()).offset(safe_offset).limit(safe_limit).all()

    # Pre-load related users + courses
    user_ids = {r.user_id for r in rows}
    hidden_by_ids = {r.hidden_by for r in rows if r.hidden_by}
    course_codes = {r.course_code for r in rows}
    users_map = {
        u.id: u for u in db.query(models.User).filter(
            models.User.id.in_(user_ids | hidden_by_ids)
        ).all()
    } if (user_ids or hidden_by_ids) else {}
    courses_map = {
        c.course_code: c for c in db.query(models.Course).filter(
            models.Course.course_code.in_(course_codes)
        ).all()
    } if course_codes else {}

    items = []
    for r in rows:
        u = users_map.get(r.user_id)
        course = courses_map.get(r.course_code)
        hidden_by_user = users_map.get(r.hidden_by) if r.hidden_by else None
        items.append(schemas.AdminReviewItem(
            id=r.id,
            user_id=r.user_id,
            student_username=u.username if u else f"unknown:{r.user_id}",
            student_full_name=u.full_name if u else None,
            course_code=r.course_code,
            course_name=course.course_name if course else None,
            rating=r.rating,
            review=r.review,
            is_anonymous=bool(r.is_anonymous),
            admin_feedback=r.admin_feedback,
            hidden=bool(r.hidden),
            hidden_by_username=hidden_by_user.username if hidden_by_user else None,
            hidden_at=r.hidden_at,
            created_at=r.created_at,
        ))

    return schemas.AdminReviewListOut(total=total, reviews=items)


@app.delete("/admin/reviews/{review_id}", response_model=schemas.MessageOut)
def admin_hide_review(
    review_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Soft delete review — set hidden=True. Audit qua hidden_by + hidden_at."""
    admin = _require_admin(authorization, db)
    r = db.query(models.CourseRating).filter(models.CourseRating.id == review_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Không tìm thấy review")
    if r.hidden:
        return {"message": "Review đã được ẩn từ trước"}

    from datetime import datetime as _dt
    r.hidden = True
    r.hidden_by = admin.id
    r.hidden_at = _dt.utcnow()
    db.commit()

    _log(db, admin, "HIDE_REVIEW", "course_rating", str(review_id),
         f"course={r.course_code} rating={r.rating}")
    return {"message": "Đã ẩn review"}


@app.post("/admin/reviews/{review_id}/restore", response_model=schemas.MessageOut)
def admin_restore_review(
    review_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Khôi phục review đã bị ẩn — set hidden=False, clear hidden_by + hidden_at."""
    admin = _require_admin(authorization, db)
    r = db.query(models.CourseRating).filter(models.CourseRating.id == review_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Không tìm thấy review")
    if not r.hidden:
        return {"message": "Review chưa bị ẩn"}

    r.hidden = False
    r.hidden_by = None
    r.hidden_at = None
    db.commit()

    _log(db, admin, "RESTORE_REVIEW", "course_rating", str(review_id),
         f"course={r.course_code}")
    return {"message": "Đã khôi phục review"}


@app.get("/admin/courses/{course_code}/rating-summary",
         response_model=schemas.AdminRatingSummaryOut)
def admin_course_rating_summary(
    course_code: str,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Aggregate rating cho 1 môn — phục vụ tab Đánh giá trong side panel admin.

    - `avg_rating`: trung bình trên review CHƯA ẩn (None nếu 0)
    - `breakdown`: histogram 1★→5★ trên review CHƯA ẩn
    - `total` / `visible_count` / `hidden_count`: counts để admin biết bao nhiêu đã moderate
    """
    _require_admin(authorization, db)
    course = db.query(models.Course).filter(
        models.Course.course_code == course_code
    ).first()
    if not course:
        raise HTTPException(status_code=404, detail=f"Không tìm thấy môn {course_code}")

    rows = db.query(models.CourseRating).filter(
        models.CourseRating.course_code == course_code
    ).all()

    total = len(rows)
    hidden_count = sum(1 for r in rows if r.hidden)
    visible = [r for r in rows if not r.hidden]
    visible_count = len(visible)

    breakdown = {str(s): 0 for s in range(1, 6)}
    for r in visible:
        key = str(int(r.rating))
        if key in breakdown:
            breakdown[key] += 1

    avg = round(sum(r.rating for r in visible) / visible_count, 1) if visible_count else None

    return schemas.AdminRatingSummaryOut(
        course_code=course_code,
        course_name=course.course_name,
        avg_rating=avg,
        total=total,
        visible_count=visible_count,
        hidden_count=hidden_count,
        breakdown=breakdown,
    )


# ════════════════════════════════════════════════════════════════════════════
# END ADMIN REVIEWS
# ════════════════════════════════════════════════════════════════════════════


@app.post("/admin/advisors/auto-assign")
def admin_auto_assign(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Tự động phân công SV chưa có cố vấn theo nhóm lớp (chuyên ngành + khóa).
    Mỗi lớp (spec+cohort) được phân cho 1 cố vấn, giúp cố vấn đóng vai trò GVCN."""
    import re as _re
    from sqlalchemy import func as sqlfunc, or_ as _or
    admin = _require_admin(authorization, db)

    # Students without any advisor assignment
    assigned_ids = {row[0] for row in db.query(models.AdvisorAssignment.student_id).all()}
    unassigned = db.query(models.User).filter(
        models.User.role == "student",
        ~models.User.id.in_(assigned_ids) if assigned_ids else True,
    ).all()

    if not unassigned:
        return {"assigned_count": 0, "skipped_count": 0, "class_groups": [],
                "message": "Không có sinh viên nào cần phân công."}

    def _cohort(uname: str) -> str:
        uname = (uname or "").upper()
        if uname.startswith("SV") and len(uname) >= 4:
            return uname[2:4]
        return uname[:2] if len(uname) >= 2 else ""

    # Group students by (specialization, cohort) = "lớp"
    from collections import defaultdict
    class_groups: dict[tuple, list] = defaultdict(list)
    for sv in unassigned:
        spec = sv.specialization or ""
        cohort = _cohort(sv.username)
        class_groups[(spec, cohort)].append(sv)

    # Build advisor map by specialization
    advisors_by_spec: dict[str, list] = defaultdict(list)
    for adv in db.query(models.User).filter(models.User.role == "advisor").all():
        advisors_by_spec[adv.managed_specialization or ""].append(adv)

    # Count current assignments per advisor
    current_load: dict[int, int] = defaultdict(int)
    for row in db.query(models.AdvisorAssignment.advisor_id,
                        sqlfunc.count(models.AdvisorAssignment.id)).group_by(
            models.AdvisorAssignment.advisor_id).all():
        current_load[row[0]] = row[1]

    assigned_count = 0
    skipped_count = 0
    result_groups = []

    for (spec, cohort), students in sorted(class_groups.items()):
        advisors = advisors_by_spec.get(spec, []) or advisors_by_spec.get("", [])
        if not advisors:
            skipped_count += len(students)
            result_groups.append({"spec": spec, "cohort": cohort,
                                   "count": len(students), "advisor": None,
                                   "skipped": True})
            continue

        # Pick the advisor with smallest current load
        best_adv = min(advisors, key=lambda a: current_load[a.id])

        for sv in students:
            db.add(models.AdvisorAssignment(advisor_id=best_adv.id, student_id=sv.id))
            existing_conn = db.query(models.UserConnection).filter(
                ((models.UserConnection.from_id == best_adv.id) & (models.UserConnection.to_id == sv.id)) |
                ((models.UserConnection.from_id == sv.id) & (models.UserConnection.to_id == best_adv.id))
            ).first()
            if not existing_conn:
                db.add(models.UserConnection(from_id=best_adv.id, to_id=sv.id, status="accepted"))
            elif existing_conn.status != "accepted":
                existing_conn.status = "accepted"
            current_load[best_adv.id] += 1
            assigned_count += 1

        result_groups.append({
            "spec": spec, "cohort": cohort,
            "count": len(students),
            "advisor": best_adv.full_name or best_adv.username,
            "skipped": False,
        })

    db.commit()
    _log(db, admin, "auto_assign", "advisor_assignment", "",
         f"Auto-assign {assigned_count} SV theo lớp, bỏ qua {skipped_count}")
    return {
        "assigned_count": assigned_count,
        "skipped_count": skipped_count,
        "class_groups": result_groups,
        "message": f"Đã phân công {assigned_count} sinh viên theo {len(result_groups)} lớp, bỏ qua {skipped_count}."
    }


# ── Department-level views (admin-only sau khi bỏ concept Trưởng bộ môn) ─────
# Trước đây 2 endpoints dưới chỉ trưởng BM xài. Sau refactor, mọi GV cùng spec
# đều có thể view (collaboration). Dùng _require_advisor + filter theo spec của
# advisor đang đăng nhập.

def _require_head_advisor(authorization: str | None, db: Session) -> models.User:
    """Compat wrapper — trả về advisor đang đăng nhập (không còn check head).

    Giữ lại để minimize impact lên 2 endpoints `/advisor/my-department-*`.
    """
    return _require_advisor(authorization, db)


@app.get("/advisor/my-department-students", response_model=list[schemas.DeptStudentItem])
def advisor_dept_students(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    head = _require_head_advisor(authorization, db)
    spec = head.managed_specialization

    if spec:
        students = db.query(models.User).filter(
            models.User.role == "student",
            models.User.specialization == spec,
        ).order_by(models.User.username).all()
    else:
        # NULL → SV chưa xác định CN (năm 1-2)
        students = db.query(models.User).filter(
            models.User.role == "student",
            models.User.specialization == None,  # noqa: E711
        ).order_by(models.User.username).all()

    result = []
    for sv in students:
        asgn = db.query(models.AdvisorAssignment).filter(
            models.AdvisorAssignment.student_id == sv.id
        ).first()
        adv_id = adv_name = None
        if asgn:
            adv = db.query(models.User).filter(models.User.id == asgn.advisor_id).first()
            if adv:
                adv_id = adv.id
                adv_name = adv.full_name or adv.username
        try:
            prog = build_progress_snapshot(db, sv.id)
        except Exception:
            prog = {}
        result.append(schemas.DeptStudentItem(
            id=sv.id,
            username=sv.username,
            full_name=sv.full_name,
            specialization=sv.specialization,
            earned_credits=prog.get("earned_credits"),
            avg_score4=prog.get("avg_score4"),
            current_advisor_id=adv_id,
            current_advisor_name=adv_name,
        ))
    return result


@app.get("/advisor/my-department-advisors", response_model=list[schemas.DeptAdvisorItem])
def advisor_dept_advisors(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    head = _require_head_advisor(authorization, db)
    advisors = db.query(models.User).filter(
        models.User.role == "advisor",
        models.User.managed_specialization == head.managed_specialization,
    ).order_by(models.User.id).all()

    result = []
    for adv in advisors:
        count = db.query(models.AdvisorAssignment).filter(
            models.AdvisorAssignment.advisor_id == adv.id
        ).count()
        class_count = db.query(models.ClassGroup).filter(
            models.ClassGroup.advisor_id == adv.id
        ).count()
        result.append(schemas.DeptAdvisorItem(
            id=adv.id,
            username=adv.username,
            full_name=adv.full_name,
            student_count=count,
            class_count=class_count,
        ))
    return result


@app.post("/advisor/students/{student_id}/reassign-advisor", response_model=schemas.MessageOut)
def advisor_reassign_student(
    student_id: int,
    body: schemas.ReassignAdvisorIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    head = _require_head_advisor(authorization, db)

    sv = db.query(models.User).filter(
        models.User.id == student_id,
        models.User.role == "student",
    ).first()
    if not sv:
        raise HTTPException(status_code=404, detail="Không tìm thấy sinh viên")

    # SV phải thuộc bộ môn của trưởng bộ môn
    if sv.specialization != head.managed_specialization:
        raise HTTPException(status_code=403, detail="Sinh viên không thuộc bộ môn của bạn")

    new_adv = db.query(models.User).filter(
        models.User.id == body.new_advisor_id,
        models.User.role == "advisor",
        models.User.managed_specialization == head.managed_specialization,
    ).first()
    if not new_adv:
        raise HTTPException(status_code=400, detail="Cố vấn không hợp lệ hoặc không cùng bộ môn")

    # Xóa assignment cũ, thêm mới
    db.query(models.AdvisorAssignment).filter(
        models.AdvisorAssignment.student_id == student_id
    ).delete(synchronize_session=False)
    db.add(models.AdvisorAssignment(advisor_id=new_adv.id, student_id=student_id))

    # Auto-accept connection để chat được ngay
    existing_conn = db.query(models.UserConnection).filter(
        ((models.UserConnection.from_id == new_adv.id) & (models.UserConnection.to_id == student_id)) |
        ((models.UserConnection.from_id == student_id) & (models.UserConnection.to_id == new_adv.id))
    ).first()
    if not existing_conn:
        db.add(models.UserConnection(from_id=new_adv.id, to_id=student_id, status="accepted"))
    elif existing_conn.status != "accepted":
        existing_conn.status = "accepted"

    db.commit()
    return {"message": f"Đã phân công {sv.username} cho cố vấn {new_adv.full_name or new_adv.username}"}


# ── Admin bulk-import assignments ────────────────────────────────────────────

def _parse_assignment_file(data: bytes, filename: str) -> list[dict]:
    """Parse xlsx/csv file with columns Mã SV | Mã GV → list of {student_code, teacher_code}."""
    import io as _io
    rows: list[dict] = []
    fname = (filename or "").lower()

    if fname.endswith(".xlsx") or fname.endswith(".xls"):
        import openpyxl as _xl
        wb = _xl.load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        header = None
        for row in ws.iter_rows(values_only=True):
            cells = [str(c or "").strip() for c in row]
            if not any(cells):
                continue
            if header is None:
                header = [c.lower().replace(" ", "_").replace("ã", "a").replace("ư", "u") for c in cells]
                continue
            rows.append(dict(zip(header, cells)))
    else:
        import csv as _csv
        text = data.decode("utf-8-sig", errors="replace")
        reader = _csv.DictReader(_io.StringIO(text))
        for r in reader:
            rows.append({k.strip(): str(v or "").strip() for k, v in r.items()})

    # Normalize column names → student_code, teacher_code
    def _find(d: dict, *keys: str) -> str:
        for k in d:
            kn = k.lower().strip()
            for kk in keys:
                if kk in kn:
                    return str(d[k] or "").strip()
        return ""

    result = []
    for r in rows:
        sc = _find(r, "ma_sv", "masv", "ma sv", "student", "sinh_vien", "sv")
        tc = _find(r, "ma_gv", "magv", "ma gv", "teacher", "giao_vien", "gv")
        if sc or tc:
            result.append({"student_code": sc, "teacher_code": tc.upper()})
    return result


@app.post("/admin/assignments/bulk-import", response_model=schemas.AssignmentBulkImportOut)
def admin_bulk_import_assignments(
    file: UploadFile,
    auto_create: bool = Form(default=False),
    new_teachers_json: str = Form(default="[]"),
    dry_run: bool = Form(default=False),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Import phân công advisor từ file xlsx/csv.
    - dry_run=true: phân tích không ghi DB
    - auto_create=true + new_teachers_json: tạo GV mới trước khi import
    """
    import json as _json
    admin = _require_admin(authorization, db)

    data = file.file.read(_MAX_UPLOAD_BYTES + 1)
    if len(data) > _MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File quá lớn. Giới hạn 10 MB.")

    try:
        rows = _parse_assignment_file(data, file.filename or "import.xlsx")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Không đọc được file: {exc}")

    if not rows:
        raise HTTPException(status_code=400, detail="Không tìm thấy dữ liệu phân công trong file")

    # Parse new_teachers from JSON string
    try:
        raw_new_teachers = _json.loads(new_teachers_json or "[]")
        new_teacher_objs = [schemas.NewTeacherInfo(**t) for t in raw_new_teachers]
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"new_teachers_json không hợp lệ: {exc}")

    errors: list[schemas.AssignmentImportError] = []
    missing_map: dict[str, list[str]] = {}  # teacher_code → [student_codes]
    created_teachers: list[schemas.CreatedTeacherInfo] = []

    # ── Step 1: create new teachers if requested ──────────────────────────────
    if auto_create and new_teacher_objs and not dry_run:
        import random as _rnd
        for t in new_teacher_objs:
            tc = validate_teacher_code(t.teacher_code)
            existing = db.query(models.User).filter(
                (models.User.teacher_code == tc) | (models.User.username == tc)
            ).first()
            if existing:
                continue  # already exists, skip creation
            plain_pw = str(_rnd.randint(100000, 999999))
            new_adv = models.User(
                username=tc,
                teacher_code=tc,
                password_hash=_hash_temp_password(plain_pw),
                full_name=normalize_vietnamese_name(t.full_name),
                role="advisor",
                managed_specialization=t.managed_specialization or None,
                default_password=plain_pw,
                is_first_login=True,
            )
            db.add(new_adv)
            created_teachers.append(schemas.CreatedTeacherInfo(
                teacher_code=tc,
                full_name=normalize_vietnamese_name(t.full_name),
                password_plain=plain_pw,
            ))
        if created_teachers:
            db.commit()

    # ── Step 2: build lookup maps ─────────────────────────────────────────────
    all_students = {u.username: u for u in db.query(models.User).filter(models.User.role == "student").all()}
    all_advisors_by_tc = {u.teacher_code: u for u in db.query(models.User).filter(
        models.User.role == "advisor", models.User.teacher_code != None  # noqa: E711
    ).all()}
    all_advisors_by_un = {u.username: u for u in db.query(models.User).filter(models.User.role == "advisor").all()}

    existing_assignments = {
        a.student_id: a for a in db.query(models.AdvisorAssignment).all()
    }

    # ── Step 3: process each row ──────────────────────────────────────────────
    created_count = 0
    updated_count = 0
    skipped_count = 0

    for idx, row in enumerate(rows, start=2):
        sc = row["student_code"]
        tc = row["teacher_code"]

        if not sc:
            errors.append(schemas.AssignmentImportError(row=idx, reason="Thiếu Mã SV"))
            continue
        if not tc:
            errors.append(schemas.AssignmentImportError(row=idx, student_code=sc, reason="Thiếu Mã GV"))
            continue

        student = all_students.get(sc)
        if not student:
            errors.append(schemas.AssignmentImportError(row=idx, student_code=sc, teacher_code=tc,
                          reason=f"SV {sc} chưa tồn tại trong hệ thống"))
            continue

        # Find advisor by teacher_code first, fallback to username
        advisor = all_advisors_by_tc.get(tc) or all_advisors_by_un.get(tc)

        if not advisor:
            # Missing teacher — collect for response
            if tc not in missing_map:
                missing_map[tc] = []
            missing_map[tc].append(sc)
            skipped_count += 1
            continue

        if dry_run:
            existing = existing_assignments.get(student.id)
            if existing and existing.advisor_id == advisor.id:
                skipped_count += 1
            elif existing:
                updated_count += 1
            else:
                created_count += 1
            continue

        existing = existing_assignments.get(student.id)
        if existing:
            if existing.advisor_id == advisor.id:
                skipped_count += 1
                continue
            existing.advisor_id = advisor.id
            updated_count += 1
        else:
            db.add(models.AdvisorAssignment(advisor_id=advisor.id, student_id=student.id))
            created_count += 1

    if not dry_run and (created_count + updated_count) > 0:
        db.commit()
        _log(db, admin, "BULK_IMPORT_ASSIGNMENTS", "assignments", None,
             f"created={created_count} updated={updated_count} skipped={skipped_count}")

    missing_teachers = [
        schemas.MissingTeacherInfo(
            teacher_code=tc,
            student_count=len(svs),
            affected_students=svs,
        )
        for tc, svs in missing_map.items()
    ]

    return schemas.AssignmentBulkImportOut(
        total=len(rows),
        created=created_count,
        updated=updated_count,
        skipped=skipped_count,
        missing_teachers=missing_teachers,
        created_teachers=created_teachers,
        errors=errors,
        dry_run=dry_run,
    )


# ---------------------------------------------------------------------------
# DEV / DEMO seed endpoint — chỉ hoạt động khi DEBUG=True
# ---------------------------------------------------------------------------
@app.post("/dev/seed-demo")
def dev_seed_demo(db: Session = Depends(get_db)):
    if os.getenv("DEBUG", "").lower() not in ("1", "true", "yes"):
        raise HTTPException(status_code=403, detail="Endpoint chỉ khả dụng khi DEBUG=True")

    results = []

    def _upsert(username, password, full_name, role, specialization=None):
        user = db.query(models.User).filter(models.User.username == username).first()
        if user:
            results.append({"username": username, "role": role, "status": "skipped"})
            return user
        user = models.User(
            username=username,
            password_hash=_hash_password(password),
            full_name=full_name,
            role=role,
            specialization=specialization,
        )
        db.add(user)
        db.flush()
        results.append({"username": username, "role": role, "status": "created"})
        return user

    admin = _upsert("demo_admin", "Demo@2025", "Admin Demo", "admin")
    advisor = _upsert("demo_advisor", "Demo@2025", "Nguyễn Thị Cố Vấn", "advisor")
    student = _upsert("demo_sv", "Demo@2025", "Sinh Viên Demo", "student", "7480201_07")

    # Phân công advisor → student nếu chưa có
    existing = db.query(models.AdvisorAssignment).filter(
        models.AdvisorAssignment.advisor_id == advisor.id,
        models.AdvisorAssignment.student_id == student.id,
    ).first()
    if existing:
        results.append({"action": "assign demo_advisor → demo_sv", "status": "skipped"})
    else:
        db.add(models.AdvisorAssignment(advisor_id=advisor.id, student_id=student.id))
        results.append({"action": "assign demo_advisor → demo_sv", "status": "created"})

    db.commit()
    return {"results": results}


# ════════════════════════════════════════════════════════════════════════════
# NHÓM 1 — TIN NHẮN TRỰC TIẾP 1-1
# ════════════════════════════════════════════════════════════════════════════

_VALID_MSG_CATEGORIES = {
    "course", "gpa", "roadmap", "graduation",
    "scholarship", "internship", "other"
}

def _fmt_msg(m: models.DirectMessage, me_id: int) -> dict:
    return {
        "id": m.id,
        "sender_id": m.sender_id,
        "receiver_id": m.receiver_id,
        "content": m.content,
        "category": m.category,
        "attachment_filename": m.attachment_filename,
        "attachment_url": f"/uploads/messages/{m.attachment_path}" if m.attachment_path else None,
        "attachment_type": m.attachment_type,
        "attachment_size": m.attachment_size,
        "created_at": m.created_at.isoformat(),
        "read_at": m.read_at.isoformat() if m.read_at else None,
        "is_mine": m.sender_id == me_id,
    }


@app.get("/messages/direct/unread/count")
def direct_unread_count(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Tổng số tin nhắn 1-1 chưa đọc gửi đến tôi."""
    me = _get_user_by_token(authorization, db)
    count = db.query(models.DirectMessage).filter(
        models.DirectMessage.receiver_id == me.id,
        models.DirectMessage.read_at.is_(None),
    ).count()
    # Cũng trả kèm per-sender để frontend có thể cập nhật badge từng cuộc trò chuyện
    from sqlalchemy import func as sqlfunc
    rows = (
        db.query(
            models.DirectMessage.sender_id,
            sqlfunc.count(models.DirectMessage.id).label("cnt"),
        )
        .filter(
            models.DirectMessage.receiver_id == me.id,
            models.DirectMessage.read_at.is_(None),
        )
        .group_by(models.DirectMessage.sender_id)
        .all()
    )
    return {
        "total": count,
        "per_sender": {str(r.sender_id): r.cnt for r in rows},
    }


@app.get("/me/advisor")
def get_my_advisor(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """SV → trả về cố vấn được phân công (assigned advisor).

    Dùng cho:
      - Floating-chat dual-tab (AI + Cố vấn) — biết advisor_id để mở DM.
      - Sidebar widget "Cố vấn của bạn".
      - Contextual CTAs ("Hỏi cố vấn về môn", "Trao đổi về kỳ này"...).

    Trả {"advisor": null} nếu SV chưa được phân công (admin chưa assign).
    Reject nếu role != student.
    """
    me = _get_user_by_token(authorization, db)
    if me.role != "student":
        raise HTTPException(status_code=403, detail="Endpoint này chỉ dành cho sinh viên")

    assignment = db.query(models.AdvisorAssignment).filter(
        models.AdvisorAssignment.student_id == me.id
    ).first()
    if not assignment:
        return {"advisor": None}

    advisor = db.query(models.User).filter(models.User.id == assignment.advisor_id).first()
    if not advisor:
        return {"advisor": None}

    return {
        "advisor": {
            "id": advisor.id,
            "username": advisor.username,
            "full_name": advisor.full_name or advisor.username,
            "teacher_code": advisor.teacher_code,
            "managed_specialization": advisor.managed_specialization,
        }
    }


@app.get("/messages/direct/conversations")
def list_direct_conversations(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Danh sách người dùng đã trao đổi tin nhắn 1-1, sắp xếp theo tin nhắn mới nhất."""
    me = _get_user_by_token(authorization, db)

    sent_ids = {r[0] for r in db.query(models.DirectMessage.receiver_id).filter(
        models.DirectMessage.sender_id == me.id
    ).distinct().all()}
    recv_ids = {r[0] for r in db.query(models.DirectMessage.sender_id).filter(
        models.DirectMessage.receiver_id == me.id
    ).distinct().all()}
    other_ids = sent_ids | recv_ids

    result = []
    for uid in other_ids:
        other = db.query(models.User).filter(models.User.id == uid).first()
        if not other:
            continue
        last_msg = db.query(models.DirectMessage).filter(
            ((models.DirectMessage.sender_id == me.id) & (models.DirectMessage.receiver_id == uid)) |
            ((models.DirectMessage.sender_id == uid) & (models.DirectMessage.receiver_id == me.id))
        ).order_by(models.DirectMessage.created_at.desc()).first()
        unread = db.query(models.DirectMessage).filter(
            models.DirectMessage.sender_id == uid,
            models.DirectMessage.receiver_id == me.id,
            models.DirectMessage.read_at.is_(None),
        ).count()
        result.append({
            "user": {"id": other.id, "username": other.username, "full_name": other.full_name, "role": other.role},
            "last_message_at": last_msg.created_at.isoformat() if last_msg else None,
            "last_message_preview": last_msg.content[:60] if last_msg else None,
            "unread_count": unread,
        })

    result.sort(key=lambda x: x["last_message_at"] or "", reverse=True)
    return {"conversations": result}


@app.get("/messages/direct/{user_id}")
def direct_history(
    user_id: int,
    limit: int = 50,
    before_id: int | None = None,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Lịch sử chat 1-1 với user_id (gồm cả 2 chiều gửi/nhận). Chỉ advisor ↔ student."""
    me = _get_user_by_token(authorization, db)
    if me.role == "admin":
        raise HTTPException(status_code=403, detail="Admin không có quyền xem tin nhắn trực tiếp")
    other = db.query(models.User).filter(models.User.id == user_id).first()
    if not other:
        raise HTTPException(status_code=404, detail="Người dùng không tồn tại")
    roles = {me.role, other.role}
    if roles != {"advisor", "student"}:
        raise HTTPException(status_code=403, detail="Chỉ được nhắn tin giữa sinh viên và cố vấn")

    q = db.query(models.DirectMessage).filter(
        (
            (models.DirectMessage.sender_id == me.id) &
            (models.DirectMessage.receiver_id == user_id)
        ) | (
            (models.DirectMessage.sender_id == user_id) &
            (models.DirectMessage.receiver_id == me.id)
        )
    )
    if before_id:
        q = q.filter(models.DirectMessage.id < before_id)
    messages = q.order_by(models.DirectMessage.created_at.desc()).limit(limit).all()
    messages.reverse()

    # Đánh dấu đã đọc tất cả tin từ đối phương gửi cho mình
    db.query(models.DirectMessage).filter(
        models.DirectMessage.sender_id == user_id,
        models.DirectMessage.receiver_id == me.id,
        models.DirectMessage.read_at.is_(None),
    ).update({"read_at": datetime.utcnow()})
    db.commit()

    return {
        "messages": [_fmt_msg(m, me.id) for m in messages],
        "other_user": {
            "id": other.id,
            "username": other.username,
            "full_name": other.full_name,
            "role": other.role,
        },
    }


@app.post("/messages/direct/{user_id}", status_code=201)
def direct_send(
    user_id: int,
    body: dict = Body(...),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Gửi tin nhắn 1-1 đến user_id. Chỉ advisor ↔ student được phép."""
    me = _get_user_by_token(authorization, db)
    # Admin không được dùng direct message
    if me.role == "admin":
        raise HTTPException(status_code=403, detail="Admin không có quyền gửi tin nhắn trực tiếp")
    content = (body.get("content") or "").strip()
    if not content:
        raise HTTPException(status_code=422, detail="Nội dung tin nhắn không được rỗng")
    if len(content) > 4000:
        raise HTTPException(status_code=422, detail="Tin nhắn tối đa 4000 ký tự")

    other = db.query(models.User).filter(models.User.id == user_id).first()
    if not other:
        raise HTTPException(status_code=404, detail="Người dùng không tồn tại")
    if other.id == me.id:
        raise HTTPException(status_code=422, detail="Không thể tự nhắn tin cho mình")

    # Chỉ cho phép advisor ↔ student; không cho student↔student hay advisor↔advisor
    roles = {me.role, other.role}
    if roles != {"advisor", "student"}:
        raise HTTPException(status_code=403, detail="Chỉ được nhắn tin giữa sinh viên và cố vấn")

    # Kiểm tra quan hệ phụ trách
    if me.role == "student":
        assigned = db.query(models.AdvisorAssignment).filter(
            models.AdvisorAssignment.student_id == me.id,
            models.AdvisorAssignment.advisor_id == other.id,
        ).first()
    else:
        assigned = db.query(models.AdvisorAssignment).filter(
            models.AdvisorAssignment.advisor_id == me.id,
            models.AdvisorAssignment.student_id == other.id,
        ).first()
    if not assigned:
        raise HTTPException(status_code=403, detail="Bạn không có quan hệ cố vấn với người dùng này")

    # Optional category
    category = (body.get("category") or "").strip().lower() or None
    if category and category not in _VALID_MSG_CATEGORIES:
        category = "other"

    msg = models.DirectMessage(
        sender_id=me.id,
        receiver_id=other.id,
        content=content,
        category=category,
    )
    db.add(msg)
    db.commit()
    db.refresh(msg)
    return _fmt_msg(msg, me.id)


# Đường dẫn lưu file attachment
import uuid as _uuid_mod
_ALLOWED_MIMES = {
    "image/png", "image/jpeg", "image/gif", "image/webp",
    "application/pdf",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "text/plain",
}
_MAX_FILE_BYTES = 10 * 1024 * 1024  # 10 MB

@app.post("/messages/direct/{user_id}/with-file", status_code=201)
async def direct_send_with_file(
    user_id: int,
    file: UploadFile = File(...),
    content: str = Form(""),
    category: str = Form(""),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Gửi tin nhắn 1-1 kèm file attachment (image/document).
    Multipart: file (required), content (optional), category (optional).
    """
    me = _get_user_by_token(authorization, db)
    if me.role == "admin":
        raise HTTPException(status_code=403, detail="Admin không có quyền gửi tin nhắn")
    text_content = (content or "").strip()
    if len(text_content) > 4000:
        raise HTTPException(status_code=422, detail="Tin nhắn tối đa 4000 ký tự")

    other = db.query(models.User).filter(models.User.id == user_id).first()
    if not other:
        raise HTTPException(status_code=404, detail="Người dùng không tồn tại")
    if other.id == me.id:
        raise HTTPException(status_code=422, detail="Không thể tự nhắn cho mình")

    roles = {me.role, other.role}
    if roles != {"advisor", "student"}:
        raise HTTPException(status_code=403, detail="Chỉ giữa SV và cố vấn")

    # Quan hệ
    if me.role == "student":
        assigned = db.query(models.AdvisorAssignment).filter(
            models.AdvisorAssignment.student_id == me.id,
            models.AdvisorAssignment.advisor_id == other.id,
        ).first()
    else:
        assigned = db.query(models.AdvisorAssignment).filter(
            models.AdvisorAssignment.advisor_id == me.id,
            models.AdvisorAssignment.student_id == other.id,
        ).first()
    if not assigned:
        raise HTTPException(status_code=403, detail="Không có quan hệ cố vấn")

    # Validate file
    mime = file.content_type or "application/octet-stream"
    if mime not in _ALLOWED_MIMES:
        raise HTTPException(status_code=422, detail=f"Định dạng không hỗ trợ: {mime}")

    raw = await file.read()
    if len(raw) > _MAX_FILE_BYTES:
        raise HTTPException(status_code=422, detail=f"File quá lớn (tối đa {_MAX_FILE_BYTES//(1024*1024)} MB)")
    if len(raw) == 0:
        raise HTTPException(status_code=422, detail="File rỗng")

    # Save file: uploads/messages/<yyyy-mm>/<uuid>.<ext>
    from datetime import datetime as _dtnow
    ymdir = _dtnow.utcnow().strftime("%Y-%m")
    save_dir = _UPLOADS_DIR / "messages" / ymdir
    save_dir.mkdir(parents=True, exist_ok=True)
    orig_name = (file.filename or "file").replace("\\", "/").split("/")[-1][:200]
    ext = ""
    if "." in orig_name:
        ext = "." + orig_name.rsplit(".", 1)[1].lower()[:8]
    safe_name = _uuid_mod.uuid4().hex + ext
    save_path = save_dir / safe_name
    save_path.write_bytes(raw)

    rel_path = f"{ymdir}/{safe_name}"

    cat = (category or "").strip().lower() or None
    if cat and cat not in _VALID_MSG_CATEGORIES:
        cat = "other"

    msg = models.DirectMessage(
        sender_id=me.id,
        receiver_id=other.id,
        content=text_content or f"[Đã gửi: {orig_name}]",
        attachment_filename=orig_name,
        attachment_path=rel_path,
        attachment_type=mime,
        attachment_size=len(raw),
        category=cat,
    )
    db.add(msg)
    db.commit()
    db.refresh(msg)
    return _fmt_msg(msg, me.id)


@app.patch("/messages/direct/{msg_id}/read")
def direct_mark_read(
    msg_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Đánh dấu một tin nhắn cụ thể là đã đọc (chỉ người nhận mới được gọi)."""
    me = _get_user_by_token(authorization, db)
    msg = db.query(models.DirectMessage).filter(models.DirectMessage.id == msg_id).first()
    if not msg:
        raise HTTPException(status_code=404, detail="Tin nhắn không tồn tại")
    if msg.receiver_id != me.id:
        raise HTTPException(status_code=403, detail="Bạn không phải người nhận tin nhắn này")
    if msg.read_at is None:
        msg.read_at = datetime.utcnow()
        db.commit()
    return {"id": msg.id, "read_at": msg.read_at.isoformat()}


# Đọc tất cả tin nhắn từ một người gửi cụ thể (dùng khi mở cửa sổ chat)
@app.patch("/messages/direct/read-all/{sender_id}")
def direct_read_all(
    sender_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Đánh dấu toàn bộ tin nhắn từ sender_id là đã đọc."""
    me = _get_user_by_token(authorization, db)
    updated = db.query(models.DirectMessage).filter(
        models.DirectMessage.sender_id == sender_id,
        models.DirectMessage.receiver_id == me.id,
        models.DirectMessage.read_at.is_(None),
    ).update({"read_at": datetime.utcnow()})
    db.commit()
    return {"marked_read": updated}


# ════════════════════════════════════════════════════════════════════════════
# NHÓM 2 — NHÓM CHAT
# ════════════════════════════════════════════════════════════════════════════

def _is_group_member(group_id: int, user_id: int, db: Session) -> bool:
    return db.query(models.ChatGroupMember).filter(
        models.ChatGroupMember.group_id == group_id,
        models.ChatGroupMember.user_id == user_id,
    ).first() is not None


def _fmt_group(g: models.ChatGroup, me_id: int, db: Session) -> dict:
    members = db.query(models.ChatGroupMember).filter(
        models.ChatGroupMember.group_id == g.id
    ).all()
    muted_by = g.muted_by if isinstance(g.muted_by, list) else []
    # Lấy tin nhắn mới nhất
    last_msg = (
        db.query(models.GroupMessage)
        .filter(models.GroupMessage.group_id == g.id)
        .order_by(models.GroupMessage.created_at.desc())
        .first()
    )
    return {
        "id": g.id,
        "name": g.name,
        "created_by": g.created_by,
        "created_at": g.created_at.isoformat(),
        "member_count": len(members),
        "is_muted": me_id in muted_by,
        "last_message": {
            "content": last_msg.content,
            "sender_id": last_msg.sender_id,
            "created_at": last_msg.created_at.isoformat(),
        } if last_msg else None,
    }


@app.get("/messages/groups")
def list_groups(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Danh sách nhóm mà tôi là thành viên."""
    me = _get_user_by_token(authorization, db)
    memberships = db.query(models.ChatGroupMember).filter(
        models.ChatGroupMember.user_id == me.id
    ).all()
    group_ids = [m.group_id for m in memberships]
    groups = db.query(models.ChatGroup).filter(
        models.ChatGroup.id.in_(group_ids)
    ).order_by(models.ChatGroup.created_at.desc()).all()
    return {"groups": [_fmt_group(g, me.id, db) for g in groups]}


@app.post("/messages/groups", status_code=201)
def create_group(
    body: dict = Body(...),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Tạo nhóm chat mới. Body: {name, member_ids: [int]}"""
    me = _get_user_by_token(authorization, db)
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=422, detail="Tên nhóm không được rỗng")
    if len(name) > 100:
        raise HTTPException(status_code=422, detail="Tên nhóm tối đa 100 ký tự")

    member_ids: list[int] = body.get("member_ids") or []
    # Luôn thêm người tạo
    all_ids = list({me.id, *member_ids})

    group = models.ChatGroup(name=name, created_by=me.id, muted_by=[])
    db.add(group)
    db.flush()  # lấy group.id

    for uid in all_ids:
        db.add(models.ChatGroupMember(group_id=group.id, user_id=uid))

    db.commit()
    db.refresh(group)
    return _fmt_group(group, me.id, db)


@app.post("/messages/groups/{group_id}/members", status_code=201)
def add_group_member(
    group_id: int,
    body: dict = Body(...),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Thêm thành viên vào nhóm. Chỉ thành viên hiện tại mới được thêm."""
    me = _get_user_by_token(authorization, db)
    group = db.query(models.ChatGroup).filter(models.ChatGroup.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Nhóm không tồn tại")
    if not _is_group_member(group_id, me.id, db):
        raise HTTPException(status_code=403, detail="Bạn không phải thành viên nhóm này")

    user_id: int = body.get("user_id")
    if not user_id:
        raise HTTPException(status_code=422, detail="Thiếu user_id")
    target = db.query(models.User).filter(models.User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Người dùng không tồn tại")
    if _is_group_member(group_id, user_id, db):
        raise HTTPException(status_code=409, detail="Người dùng đã là thành viên nhóm")

    db.add(models.ChatGroupMember(group_id=group_id, user_id=user_id))
    db.commit()
    return {"group_id": group_id, "user_id": user_id, "joined": True}


@app.get("/messages/groups/{group_id}/messages")
def group_history(
    group_id: int,
    limit: int = 50,
    before_id: int | None = None,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Lịch sử tin nhắn nhóm (chỉ thành viên mới xem được)."""
    me = _get_user_by_token(authorization, db)
    group = db.query(models.ChatGroup).filter(models.ChatGroup.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Nhóm không tồn tại")
    if not _is_group_member(group_id, me.id, db):
        raise HTTPException(status_code=403, detail="Bạn không phải thành viên nhóm này")

    q = db.query(models.GroupMessage).filter(models.GroupMessage.group_id == group_id)
    if before_id:
        q = q.filter(models.GroupMessage.id < before_id)
    messages = q.order_by(models.GroupMessage.created_at.desc()).limit(limit).all()
    messages.reverse()

    # Lấy tên các sender
    sender_ids = list({m.sender_id for m in messages})
    senders = {
        u.id: {"full_name": u.full_name, "username": u.username}
        for u in db.query(models.User).filter(models.User.id.in_(sender_ids)).all()
    }

    return {
        "group": _fmt_group(group, me.id, db),
        "messages": [
            {
                "id": m.id,
                "group_id": m.group_id,
                "sender_id": m.sender_id,
                "sender_name": senders.get(m.sender_id, {}).get("full_name") or senders.get(m.sender_id, {}).get("username", "?"),
                "content": m.content,
                "created_at": m.created_at.isoformat(),
                "is_mine": m.sender_id == me.id,
            }
            for m in messages
        ],
    }


@app.post("/messages/groups/{group_id}/messages", status_code=201)
def group_send(
    group_id: int,
    body: dict = Body(...),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Gửi tin nhắn vào nhóm."""
    me = _get_user_by_token(authorization, db)
    group = db.query(models.ChatGroup).filter(models.ChatGroup.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Nhóm không tồn tại")
    if not _is_group_member(group_id, me.id, db):
        raise HTTPException(status_code=403, detail="Bạn không phải thành viên nhóm này")

    content = (body.get("content") or "").strip()
    if not content:
        raise HTTPException(status_code=422, detail="Nội dung tin nhắn không được rỗng")
    if len(content) > 4000:
        raise HTTPException(status_code=422, detail="Tin nhắn tối đa 4000 ký tự")

    msg = models.GroupMessage(group_id=group_id, sender_id=me.id, content=content)
    db.add(msg)
    db.commit()
    db.refresh(msg)
    return {
        "id": msg.id,
        "group_id": msg.group_id,
        "sender_id": msg.sender_id,
        "content": msg.content,
        "created_at": msg.created_at.isoformat(),
        "is_mine": True,
    }


@app.patch("/messages/groups/{group_id}/mute")
def toggle_group_mute(
    group_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Toggle tắt/bật thông báo nhóm cho bản thân."""
    me = _get_user_by_token(authorization, db)
    group = db.query(models.ChatGroup).filter(models.ChatGroup.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Nhóm không tồn tại")
    if not _is_group_member(group_id, me.id, db):
        raise HTTPException(status_code=403, detail="Bạn không phải thành viên nhóm này")

    muted_by: list = list(group.muted_by) if isinstance(group.muted_by, list) else []
    if me.id in muted_by:
        muted_by.remove(me.id)
        is_muted = False
    else:
        muted_by.append(me.id)
        is_muted = True

    group.muted_by = muted_by
    db.commit()
    return {"group_id": group_id, "is_muted": is_muted}


# ════════════════════════════════════════════════════════════════════════════
# NHÓM 3 — KẾT NỐI NGƯỜI DÙNG
# ════════════════════════════════════════════════════════════════════════════

def _fmt_connection(c: models.UserConnection, me_id: int, db: Session) -> dict:
    other_id = c.to_id if c.from_id == me_id else c.from_id
    other = db.query(models.User).filter(models.User.id == other_id).first()
    return {
        "id": c.id,
        "from_id": c.from_id,
        "to_id": c.to_id,
        "status": c.status,
        "created_at": c.created_at.isoformat(),
        "other_user": {
            "id": other.id,
            "username": other.username,
            "full_name": other.full_name,
            "role": other.role,
            "specialization": other.specialization,
        } if other else None,
        "direction": "sent" if c.from_id == me_id else "received",
    }


@app.post("/connections/request/{user_id}", status_code=201)
def send_connection_request(
    user_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Gửi lời mời kết nối đến user_id."""
    me = _get_user_by_token(authorization, db)
    if user_id == me.id:
        raise HTTPException(status_code=422, detail="Không thể kết nối với chính mình")

    target = db.query(models.User).filter(models.User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Người dùng không tồn tại")

    # Kiểm tra đã có kết nối / lời mời chưa (cả 2 chiều)
    existing = db.query(models.UserConnection).filter(
        (
            (models.UserConnection.from_id == me.id) &
            (models.UserConnection.to_id == user_id)
        ) | (
            (models.UserConnection.from_id == user_id) &
            (models.UserConnection.to_id == me.id)
        )
    ).first()
    if existing:
        if existing.status == "accepted":
            raise HTTPException(status_code=409, detail="Hai bạn đã kết nối rồi")
        if existing.status == "pending":
            raise HTTPException(status_code=409, detail="Đã có lời mời đang chờ xử lý")
        # rejected → cho phép gửi lại
        existing.status = "pending"
        existing.from_id = me.id
        existing.to_id = user_id
        db.commit()
        db.refresh(existing)
        return _fmt_connection(existing, me.id, db)

    conn = models.UserConnection(from_id=me.id, to_id=user_id, status="pending")
    db.add(conn)
    db.commit()
    db.refresh(conn)
    return _fmt_connection(conn, me.id, db)


@app.patch("/connections/{conn_id}/accept")
def accept_connection(
    conn_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Chấp nhận lời mời kết nối (chỉ người nhận mới được gọi)."""
    me = _get_user_by_token(authorization, db)
    conn = db.query(models.UserConnection).filter(models.UserConnection.id == conn_id).first()
    if not conn:
        raise HTTPException(status_code=404, detail="Lời mời không tồn tại")
    if conn.to_id != me.id:
        raise HTTPException(status_code=403, detail="Bạn không phải người nhận lời mời này")
    if conn.status != "pending":
        raise HTTPException(status_code=409, detail=f"Lời mời đã ở trạng thái '{conn.status}'")
    conn.status = "accepted"
    db.commit()
    return _fmt_connection(conn, me.id, db)


@app.patch("/connections/{conn_id}/reject")
def reject_connection(
    conn_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Từ chối lời mời kết nối (chỉ người nhận mới được gọi)."""
    me = _get_user_by_token(authorization, db)
    conn = db.query(models.UserConnection).filter(models.UserConnection.id == conn_id).first()
    if not conn:
        raise HTTPException(status_code=404, detail="Lời mời không tồn tại")
    if conn.to_id != me.id:
        raise HTTPException(status_code=403, detail="Bạn không phải người nhận lời mời này")
    if conn.status != "pending":
        raise HTTPException(status_code=409, detail=f"Lời mời đã ở trạng thái '{conn.status}'")
    conn.status = "rejected"
    db.commit()
    return _fmt_connection(conn, me.id, db)


@app.get("/connections/pending")
def list_pending_connections(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Lời mời đang chờ — cả gửi đi lẫn nhận về."""
    me = _get_user_by_token(authorization, db)
    rows = db.query(models.UserConnection).filter(
        models.UserConnection.status == "pending",
        (
            (models.UserConnection.from_id == me.id) |
            (models.UserConnection.to_id == me.id)
        ),
    ).order_by(models.UserConnection.created_at.desc()).all()
    return {
        "pending": [_fmt_connection(c, me.id, db) for c in rows],
        "received_count": sum(1 for c in rows if c.to_id == me.id),
    }


@app.get("/connections")
def list_connections(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Danh sách kết nối:
    - UserConnection accepted (friend-request flow)
    - AdvisorAssignment (admin gán SV ↔ GV) — tự động coi là accepted để
      messaging hoạt động ngay sau khi admin phân công, không cần SV/GV tự
      gửi friend request.
    """
    me = _get_user_by_token(authorization, db)

    # 1. Friend-request connections (status=accepted)
    rows = db.query(models.UserConnection).filter(
        models.UserConnection.status == "accepted",
        (
            (models.UserConnection.from_id == me.id) |
            (models.UserConnection.to_id == me.id)
        ),
    ).order_by(models.UserConnection.created_at.desc()).all()
    out = [_fmt_connection(c, me.id, db) for c in rows]

    # 2. AdvisorAssignment-based connections — wrap as virtual accepted entries
    # để frontend (messaging.html) render giống nhau.
    # Tránh duplicate nếu đã có UserConnection giữa cùng 2 user.
    seen_user_ids = {c.get("other_user", {}).get("id") for c in out if c.get("other_user")}

    if me.role == "student":
        # Student → list their advisor(s)
        assigns = db.query(models.AdvisorAssignment).filter_by(student_id=me.id).all()
        for a in assigns:
            if a.advisor_id in seen_user_ids:
                continue
            adv = db.query(models.User).filter_by(id=a.advisor_id).first()
            if not adv:
                continue
            out.append({
                "id": -a.id,  # negative to differentiate from UserConnection.id
                "status": "accepted",
                "source": "advisor_assignment",
                "created_at": a.assigned_at.isoformat() if a.assigned_at else None,
                "other_user": {
                    "id": adv.id,
                    "username": adv.username,
                    "full_name": adv.full_name,
                    "role": adv.role,
                    "teacher_code": getattr(adv, "teacher_code", None),
                    "specialization": adv.managed_specialization,
                },
            })
    elif me.role == "advisor":
        # Advisor → list their assigned students
        assigns = db.query(models.AdvisorAssignment).filter_by(advisor_id=me.id).all()
        for a in assigns:
            if a.student_id in seen_user_ids:
                continue
            std = db.query(models.User).filter_by(id=a.student_id).first()
            if not std:
                continue
            out.append({
                "id": -a.id,
                "status": "accepted",
                "source": "advisor_assignment",
                "created_at": a.assigned_at.isoformat() if a.assigned_at else None,
                "other_user": {
                    "id": std.id,
                    "username": std.username,
                    "full_name": std.full_name,
                    "role": std.role,
                    "specialization": std.specialization,
                    "cohort": std.cohort,
                },
            })

    return {"connections": out}


@app.get("/users/search")
def search_users(
    q: str = Query(default=""),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Tìm kiếm người dùng theo username/tên (cho tính năng kết nối)."""
    me = _get_user_by_token(authorization, db)
    if not q or len(q) < 2:
        return {"users": []}
    results = (
        db.query(models.User)
        .filter(
            models.User.id != me.id,
            (
                models.User.username.ilike(f"%{q}%") |
                models.User.full_name.ilike(f"%{q}%")
            ),
        )
        .limit(10)
        .all()
    )
    return {
        "users": [
            {
                "id": u.id,
                "username": u.username,
                "full_name": u.full_name,
                "role": u.role,
                "specialization": u.specialization,
            }
            for u in results
        ]
    }


# ════════════════════════════════════════════════════════════════════════════
# NHÓM 4 — MÔN TỰ CHỌN ĐÃ PLAN (lộ trình tùy chỉnh)
# ════════════════════════════════════════════════════════════════════════════

def _elective_pool_codes(user: models.User, db: Session) -> set[str]:
    """Trả về set course_code thuộc pool tự chọn B/C của chuyên ngành user."""
    if not user.specialization:
        return set()
    rows = db.query(models.CourseElectiveGroup).filter(
        (models.CourseElectiveGroup.specialization == user.specialization) |
        (models.CourseElectiveGroup.specialization == "Chung"),
    ).all()
    return {r.course_code for r in rows}


def _fmt_planned(row: models.PlannedElective, course: models.Course | None) -> dict:
    return {
        "term_label": row.term_label,
        "slot_id": row.slot_id,
        "course_code": row.course_code,
        "course_name": course.course_name if course else None,
        "credits": float(course.credits) if course and course.credits else None,
    }


@app.get("/roadmap/planned-electives/me")
def get_planned_electives(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Lấy toàn bộ môn tự chọn đã plan của sinh viên đang đăng nhập."""
    me = _get_user_by_token(authorization, db)
    rows = (
        db.query(models.PlannedElective)
        .filter(models.PlannedElective.user_id == me.id)
        .order_by(models.PlannedElective.term_label, models.PlannedElective.slot_id)
        .all()
    )
    # Lấy thông tin môn học kèm theo
    codes = {r.course_code for r in rows}
    course_map = {
        c.course_code: c
        for c in db.query(models.Course).filter(models.Course.course_code.in_(codes)).all()
    }
    return {
        "planned": [_fmt_planned(r, course_map.get(r.course_code)) for r in rows]
    }


@app.put("/roadmap/planned-electives/me")
def put_planned_electives(
    body: dict = Body(...),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """
    Lưu toàn bộ plan môn tự chọn (replace — không merge).
    Body: { "planned": [ {"term_label": str, "slot_id": str, "course_code": str} ] }
    """
    me = _get_user_by_token(authorization, db)
    items: list[dict] = body.get("planned") or []

    if not isinstance(items, list):
        raise HTTPException(status_code=422, detail="'planned' phải là mảng")
    if len(items) > 50:
        raise HTTPException(status_code=422, detail="Tối đa 50 slot tự chọn")

    # Lấy pool hợp lệ để validate
    pool = _elective_pool_codes(me, db)

    seen_slots: set[tuple] = set()
    validated: list[dict] = []
    for item in items:
        term = (item.get("term_label") or "").strip()
        slot = (item.get("slot_id") or "").strip()
        code = (item.get("course_code") or "").strip()
        if not term or not slot or not code:
            raise HTTPException(
                status_code=422,
                detail=f"Mỗi item cần có term_label, slot_id và course_code. Lỗi: {item}",
            )
        # Validate course_code thuộc pool của chuyên ngành (bỏ qua nếu SV chưa chọn CN)
        if pool and code not in pool:
            raise HTTPException(
                status_code=422,
                detail=f"Môn '{code}' không thuộc pool tự chọn của chuyên ngành bạn",
            )
        # Validate course tồn tại trong DB
        if not db.query(models.Course).filter(models.Course.course_code == code).first():
            raise HTTPException(status_code=422, detail=f"Môn '{code}' không tồn tại trong hệ thống")
        key = (term, slot)
        if key in seen_slots:
            raise HTTPException(
                status_code=422,
                detail=f"Trùng slot ({term}, {slot}) trong cùng một lần gửi",
            )
        seen_slots.add(key)
        validated.append({"term_label": term, "slot_id": slot, "course_code": code})

    # Replace toàn bộ: xóa cũ rồi insert mới trong 1 transaction
    db.query(models.PlannedElective).filter(
        models.PlannedElective.user_id == me.id
    ).delete()
    for v in validated:
        db.add(models.PlannedElective(
            user_id=me.id,
            term_label=v["term_label"],
            slot_id=v["slot_id"],
            course_code=v["course_code"],
        ))
    db.commit()

    # Trả lại plan mới kèm thông tin môn học
    codes = {v["course_code"] for v in validated}
    course_map = {
        c.course_code: c
        for c in db.query(models.Course).filter(models.Course.course_code.in_(codes)).all()
    }
    rows = (
        db.query(models.PlannedElective)
        .filter(models.PlannedElective.user_id == me.id)
        .order_by(models.PlannedElective.term_label, models.PlannedElective.slot_id)
        .all()
    )
    return {
        "saved": len(validated),
        "planned": [_fmt_planned(r, course_map.get(r.course_code)) for r in rows],
    }


# ════════════════════════════════════════════════════════════════════════════
# NHÓM 5 — CỐ VẤN XEM LỘ TRÌNH SINH VIÊN
# ════════════════════════════════════════════════════════════════════════════

@app.get("/advisor/students/{student_id}/career")
def advisor_get_student_career(
    student_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Cố vấn xem định hướng nghề SV đã chọn + tiến độ skill."""
    advisor = _require_advisor(authorization, db)
    _check_advisor_has_student(advisor, student_id, db)

    choice = db.query(models.UserCareerChoice).filter(
        models.UserCareerChoice.user_id == student_id
    ).first()
    if not choice or not choice.primary_path_id:
        return {"chosen": False}

    primary = db.query(models.CareerPath).filter(
        models.CareerPath.id == choice.primary_path_id
    ).first()
    if not primary:
        return {"chosen": False}

    skills = db.query(models.CareerSkill).filter(
        models.CareerSkill.path_id == primary.id
    ).order_by(models.CareerSkill.priority.asc(), models.CareerSkill.id.asc()).all()

    progress_rows = db.query(models.UserCareerSkillProgress).filter(
        models.UserCareerSkillProgress.user_id == student_id
    ).all()
    progress_map = {r.skill_id: r for r in progress_rows}

    primary_detail = _format_career(primary, skills=skills)
    for s in primary_detail["skills"]:
        pr = progress_map.get(s["id"])
        s["user_status"] = pr.status if pr else "planned"
    completed = sum(1 for s in primary_detail["skills"] if s.get("user_status") == "completed")
    primary_detail["progress"] = {
        "completed": completed,
        "total": len(primary_detail["skills"]),
        "percent": round(100.0 * completed / len(primary_detail["skills"]), 1) if primary_detail["skills"] else 0.0,
    }

    secondary = None
    if choice.secondary_path_id:
        sec = db.query(models.CareerPath).filter(
            models.CareerPath.id == choice.secondary_path_id
        ).first()
        if sec:
            secondary = _format_career(sec)

    return {
        "chosen": True,
        "primary": primary_detail,
        "secondary": secondary,
        "chosen_at": choice.chosen_at.isoformat() if choice.chosen_at else None,
    }


@app.get("/advisor/students/{student_id}/roadmap")
def advisor_get_student_roadmap(
    student_id: int,
    max_credits: int | None = None,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """
    Cố vấn xem lộ trình học tập + môn tự chọn đã plan của sinh viên.
    Yêu cầu role=advisor (hoặc admin) và SV phải thuộc advisor đó.
    """
    advisor = _require_advisor(authorization, db)
    _check_advisor_has_student(advisor, student_id, db)

    student = db.query(models.User).filter(models.User.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Sinh viên không tồn tại")
    if student.role != "student":
        raise HTTPException(status_code=422, detail="ID này không phải sinh viên")

    # Lấy roadmap (engine tự tính max_credits từ remaining workload)
    roadmap = build_semester_roadmap(db, student_id, max_credits_per_term=max_credits)

    # Lấy planned electives của SV
    pe_rows = (
        db.query(models.PlannedElective)
        .filter(models.PlannedElective.user_id == student_id)
        .order_by(models.PlannedElective.term_label, models.PlannedElective.slot_id)
        .all()
    )
    codes = {r.course_code for r in pe_rows}
    course_map = {
        c.course_code: c
        for c in db.query(models.Course).filter(models.Course.course_code.in_(codes)).all()
    }

    # Lấy custom roadmap mà SV tự lưu (nếu có)
    custom_plan = db.query(models.StudyPlan).filter(
        models.StudyPlan.user_id == student_id,
        models.StudyPlan.plan_name == "roadmap_custom",
    ).first()
    custom_semesters: list[dict] = []
    custom_updated_at = None
    if custom_plan:
        custom_updated_at = custom_plan.updated_at or custom_plan.created_at
        custom_items = db.query(models.StudyPlanItem).filter(
            models.StudyPlanItem.plan_id == custom_plan.id
        ).all()
        _all_custom_codes = {it.course_code for it in custom_items}
        _custom_course_map = {
            c.course_code: c
            for c in db.query(models.Course).filter(models.Course.course_code.in_(_all_custom_codes)).all()
        }
        from collections import defaultdict as _dd
        _sem_map: dict[str, list[dict]] = _dd(list)
        for it in custom_items:
            c = _custom_course_map.get(it.course_code)
            _sem_map[it.term_label or ""].append({
                "course_code": it.course_code,
                "course_name": c.course_name if c else it.course_code,
                "credits": float(c.credits) if c and c.credits is not None else 0.0,
                "required_specialization": c.required_specialization if c else None,
            })
        custom_semesters = [
            {"semester_label": k, "courses": v}
            for k, v in _sem_map.items()
        ]
        custom_semesters.sort(key=lambda s: s["semester_label"])

    return {
        "student": {
            "id": student.id,
            "username": student.username,
            "full_name": student.full_name,
            "specialization": student.specialization,
        },
        "roadmap": roadmap,
        "planned_electives": [_fmt_planned(r, course_map.get(r.course_code)) for r in pe_rows],
        "custom_roadmap": {
            "saved": custom_plan is not None,
            "updated_at": custom_updated_at.isoformat() if custom_updated_at else None,
            "semesters": custom_semesters,
        },
    }


@app.get("/advisor/students/{student_id}/grades")
def advisor_get_student_grades(
    student_id: int,
    include_self: bool = Query(default=False, description="Bao gồm điểm SV tự khai (source='self')"),
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Cố vấn xem bảng điểm của SV thuộc danh sách quản lý.

    Approach C: mặc định CHỈ trả điểm `source='admin'` (đã xác minh từ phòng đào tạo).
    Pass `?include_self=true` để bao gồm cả SV tự khai (cần khi SV chưa có
    bản admin import). Mỗi grade trả về kèm `source` để frontend gắn badge.
    """
    advisor = _require_advisor(authorization, db)
    _check_advisor_has_student(advisor, student_id, db)

    student = db.query(models.User).filter(
        models.User.id == student_id, models.User.role == "student"
    ).first()
    if not student:
        raise HTTPException(status_code=404, detail="Sinh viên không tồn tại")

    # Sau refactor 2026-05-05: không còn source field. include_self ignored.
    q = db.query(models.UserGrade).filter(models.UserGrade.user_id == student_id)
    grades = q.order_by(models.UserGrade.term, models.UserGrade.course_code).all()

    course_codes = {g.course_code for g in grades}
    course_map = {
        c.course_code: c
        for c in db.query(models.Course).filter(models.Course.course_code.in_(course_codes)).all()
    }

    total_admin = 0  # Deprecated
    total_self = len(grades)

    return {
        "student_id": student_id,
        "student_name": student.full_name,
        "grades": [
            {
                "course_code": g.course_code,
                "course_name": course_map[g.course_code].course_name if g.course_code in course_map else None,
                "credits": float(course_map[g.course_code].credits) if g.course_code in course_map else None,
                "score10": float(g.score10) if g.score10 is not None else None,
                "score4": float(g.score4) if g.score4 is not None else None,
                "letter": g.letter,
                "passed": g.passed,
                "term": g.term,
                "source": g.source or "self",
            }
            for g in grades
        ],
        "total": len(grades),
        "total_admin": total_admin,
        "total_self": total_self,
        "has_admin_grades": total_admin > 0,
        "include_self": include_self,
    }


# ════════════════════════════════════════════════════════════════════════════
# V2 — Lộ trình tích hợp: Career Blueprint + Integrated Roadmap
# ════════════════════════════════════════════════════════════════════════════

@app.get("/v2/careers")
def v2_list_careers(db: Session = Depends(get_db)):
    """Danh sách tất cả nghề mục tiêu cho picker — không cần auth."""
    paths = db.query(models.CareerPath).order_by(models.CareerPath.name).all()
    out = []
    for p in paths:
        skill_count = db.query(models.CareerSkill).filter(
            models.CareerSkill.path_id == p.id
        ).count()
        out.append({
            "id": p.id,
            "code": p.code,
            "name": p.name,
            "short_description": p.short_description,
            "icon": p.icon,
            "color": p.color,
            "has_blueprint": skill_count > 0,
            "last_blueprint_at": p.last_blueprint_at.isoformat() if p.last_blueprint_at else None,
        })
    return out


@app.get("/v2/careers/my-choice")
def v2_get_my_career_choice(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Lấy nghề SV đã chọn (primary + secondary)."""
    user = _get_user_by_token(authorization, db)
    choice = db.query(models.UserCareerChoice).filter(
        models.UserCareerChoice.user_id == user.id
    ).first()
    if not choice:
        return {"primary": None, "secondary": None}

    def _serialize_path(path_id):
        if not path_id:
            return None
        p = db.query(models.CareerPath).filter(models.CareerPath.id == path_id).first()
        if not p:
            return None
        return {"id": p.id, "code": p.code, "name": p.name, "icon": p.icon, "color": p.color}

    return {
        "primary": _serialize_path(choice.primary_path_id),
        "secondary": _serialize_path(choice.secondary_path_id),
        "chosen_at": choice.chosen_at.isoformat() if choice.chosen_at else None,
    }


@app.post("/v2/careers/my-choice")
def v2_set_my_career_choice(
    payload: schemas.CareerChoiceIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Chọn/đổi nghề mục tiêu — tạo/update UserCareerChoice.

    Khi ĐỔI nghề (path_id mới khác cũ), wipe TC đã chọn — buộc SV chọn lại
    cho phù hợp với career mới. Lý do: TC priority/recommendation phụ thuộc
    nặng vào career_goal (track_match, skill alignment).
    Lần CHỌN ĐẦU TIÊN không wipe (chưa có TC để wipe).
    """
    user = _get_user_by_token(authorization, db)

    primary = db.query(models.CareerPath).filter(
        models.CareerPath.id == payload.primary_path_id
    ).first()
    if not primary:
        raise HTTPException(status_code=404, detail="Nghề mục tiêu không tồn tại")

    secondary = None
    if payload.secondary_path_id:
        secondary = db.query(models.CareerPath).filter(
            models.CareerPath.id == payload.secondary_path_id
        ).first()

    choice = db.query(models.UserCareerChoice).filter(
        models.UserCareerChoice.user_id == user.id
    ).first()
    old_primary_id = choice.primary_path_id if choice else None
    is_career_change = old_primary_id is not None and old_primary_id != payload.primary_path_id

    if choice:
        choice.primary_path_id = payload.primary_path_id
        choice.secondary_path_id = payload.secondary_path_id
        choice.updated_at = datetime.utcnow()
    else:
        choice = models.UserCareerChoice(
            user_id=user.id,
            primary_path_id=payload.primary_path_id,
            secondary_path_id=payload.secondary_path_id,
        )
        db.add(choice)

    # Note: legacy users.career_goal đã DROP 2026-05-05.
    # Career choice giờ chỉ track qua bảng UserCareerChoice (primary/secondary path).
    db.commit()

    # Nếu user ĐỔI nghề (không phải lần đầu chọn) → wipe TC để re-pick
    if is_career_change:
        tc_codes = {
            row[0] for row in db.query(models.CourseElectiveGroup.course_code).distinct().all()
        }
        if tc_codes:
            (
                db.query(models.PlannedElective)
                .filter(
                    models.PlannedElective.user_id == user.id,
                    models.PlannedElective.course_code.in_(tc_codes),
                )
                .delete(synchronize_session=False)
            )
            custom_plan = db.query(models.StudyPlan).filter(
                models.StudyPlan.user_id == user.id,
                models.StudyPlan.plan_name == "roadmap_custom",
            ).first()
            if custom_plan:
                (
                    db.query(models.StudyPlanItem)
                    .filter(
                        models.StudyPlanItem.plan_id == custom_plan.id,
                        models.StudyPlanItem.course_code.in_(tc_codes),
                    )
                    .delete(synchronize_session=False)
                )
            db.commit()

    blueprint_ready = db.query(models.CareerSkill).filter(
        models.CareerSkill.path_id == payload.primary_path_id
    ).count() > 0

    return {
        "ok": True,
        "primary": {"id": primary.id, "code": primary.code, "name": primary.name},
        "secondary": {"id": secondary.id, "code": secondary.code, "name": secondary.name} if secondary else None,
        "blueprint_ready": blueprint_ready,
    }


def _compute_blueprint_fit(user_id: int, path_id: int, db: Session) -> dict:
    """Tính Career Fit % và thống kê tiến độ skill cho 1 path."""
    skills = db.query(models.CareerSkill).filter(
        models.CareerSkill.path_id == path_id
    ).all()
    if not skills:
        return {"fit_percent": 0, "total_skills": 0, "completed_skills": 0, "in_progress_skills": 0, "estimated_external_hours": 0}

    prog_rows = db.query(models.UserCareerSkillProgress).filter(
        models.UserCareerSkillProgress.user_id == user_id,
        models.UserCareerSkillProgress.skill_id.in_([s.id for s in skills]),
    ).all()
    prog_map = {r.skill_id: r.status for r in prog_rows}

    passed_codes: set[str] = {g.course_code for g in db.query(models.UserGrade).filter(
        models.UserGrade.user_id == user_id, models.UserGrade.passed == True
    ).all()}

    completed = in_progress = ext_hours_remaining = 0
    for s in skills:
        status = prog_map.get(s.id, "planned")
        if s.school_covered and s.school_courses:
            if any(c in passed_codes for c in s.school_courses) and status == "planned":
                status = "completed"
        if status == "completed":
            completed += 1
        elif status == "in_progress":
            in_progress += 1
            ext_hours_remaining += max(0, (s.estimated_hours or 0) // 2)
        elif status == "planned" and not s.school_covered:
            ext_hours_remaining += s.estimated_hours or 0

    total = len(skills)
    fit_pct = round((completed / total) * 100) if total > 0 else 0
    return {
        "fit_percent": min(100, fit_pct),
        "total_skills": total,
        "completed_skills": completed,
        "in_progress_skills": in_progress,
        "estimated_external_hours": ext_hours_remaining,
    }


@app.get("/v2/career/{path_id}/detail")
def v2_get_career_detail(
    path_id: int,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Chi tiết nghề: mô tả công việc, kỹ năng cần, tố chất, lương, cơ hội.

    Cache trong CareerPath.long_description (JSON). Lần đầu gọi sẽ AI-generate
    rồi lưu lại; lần sau trả từ cache (instant).

    Body shape:
      { "detail": {overview, responsibilities[], required_hard_skills[],
                   required_soft_skills[], personality_traits[],
                   salary_vn{junior,mid,senior}, growth_paths[], is_suitable_for},
        "from_cache": bool, "provider": str|null }
    """
    _get_user_by_token(authorization, db)  # auth guard
    path = db.query(models.CareerPath).filter(models.CareerPath.id == path_id).first()
    if not path:
        raise HTTPException(status_code=404, detail="Career path không tồn tại")

    # Try cached JSON in long_description
    if path.long_description:
        try:
            cached = json.loads(path.long_description)
            if isinstance(cached, dict) and cached.get("overview"):
                return {
                    "detail": cached,
                    "from_cache": True,
                    "provider": path.blueprint_model,
                    "career": {"id": path.id, "code": path.code, "name": path.name,
                               "short_description": path.short_description,
                               "icon": path.icon, "color": path.color},
                }
        except Exception:
            pass  # not JSON → ignore, regenerate

    # Generate via AI
    from backend.core.ai_advisor import generate_career_detail
    detail, provider = generate_career_detail(
        career_name=path.name,
        short_description=path.short_description or path.name,
    )
    if not detail:
        raise HTTPException(
            status_code=503,
            detail="AI không phản hồi. Vui lòng thử lại sau vài giây."
        )

    # Save back to cache (override any plain-text long_description)
    path.long_description = json.dumps(detail, ensure_ascii=False)
    db.commit()

    return {
        "detail": detail,
        "from_cache": False,
        "provider": provider,
        "career": {"id": path.id, "code": path.code, "name": path.name,
                   "short_description": path.short_description,
                   "icon": path.icon, "color": path.color},
    }


@app.get("/v2/career/blueprint/me")
def v2_get_blueprint_me(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Toàn bộ blueprint + tiến độ SV cho nghề đã chọn."""
    user = _get_user_by_token(authorization, db)

    choice = db.query(models.UserCareerChoice).filter(
        models.UserCareerChoice.user_id == user.id
    ).first()
    if not choice or not choice.primary_path_id:
        return {"path": None, "fit_percent": 0, "total_skills": 0,
                "completed_skills": 0, "in_progress_skills": 0,
                "estimated_external_hours": 0, "groups": []}

    path = db.query(models.CareerPath).filter(
        models.CareerPath.id == choice.primary_path_id
    ).first()
    if not path:
        return {"path": None, "fit_percent": 0, "total_skills": 0,
                "completed_skills": 0, "in_progress_skills": 0,
                "estimated_external_hours": 0, "groups": []}

    skills = db.query(models.CareerSkill).filter(
        models.CareerSkill.path_id == path.id
    ).order_by(models.CareerSkill.priority, models.CareerSkill.id).all()

    prog_rows = db.query(models.UserCareerSkillProgress).filter(
        models.UserCareerSkillProgress.user_id == user.id,
        models.UserCareerSkillProgress.skill_id.in_([s.id for s in skills]) if skills else False,
    ).all()
    prog_map = {r.skill_id: r for r in prog_rows}

    passed_codes = {g.course_code for g in db.query(models.UserGrade).filter(
        models.UserGrade.user_id == user.id, models.UserGrade.passed == True
    ).all()}
    active_codes = {g.course_code for g in db.query(models.UserGrade).filter(
        models.UserGrade.user_id == user.id, models.UserGrade.passed == None
    ).all()}

    from collections import defaultdict
    groups_map: dict[str, list] = defaultdict(list)
    for s in skills:
        row = prog_map.get(s.id)
        status = row.status if row else "planned"
        scheduled_term = row.scheduled_term if row else None

        if s.school_covered and s.school_courses:
            if any(c in passed_codes for c in s.school_courses):
                status = "completed"
            elif any(c in active_codes for c in s.school_courses) and status == "planned":
                status = "in_progress"

        groups_map[s.skill_group].append({
            "id": s.id,
            "skill_group": s.skill_group,
            "skill_name": s.skill_name,
            "skill_type": s.skill_type,
            "level": s.level,
            "priority": s.priority,
            "school_covered": s.school_covered,
            "school_courses": s.school_courses,
            "source_type": s.source_type,
            "source_name": s.source_name,
            "source_url": s.source_url,
            "description": s.description,
            "estimated_hours": s.estimated_hours,
            "status": status,
            "scheduled_term": scheduled_term,
        })

    groups_out = []
    for group_name, skill_list in groups_map.items():
        completed_c = sum(1 for s in skill_list if s["status"] == "completed")
        inprog_c = sum(1 for s in skill_list if s["status"] == "in_progress")
        groups_out.append({
            "group_name": group_name,
            "total": len(skill_list),
            "completed": completed_c,
            "in_progress": inprog_c,
            "skills": skill_list,
        })

    stats = _compute_blueprint_fit(user.id, path.id, db)
    return {
        "path": {
            "id": path.id, "code": path.code, "name": path.name,
            "short_description": path.short_description,
            "icon": path.icon, "color": path.color,
            "has_blueprint": len(skills) > 0,
            "last_blueprint_at": path.last_blueprint_at.isoformat() if path.last_blueprint_at else None,
        },
        "fit_percent": stats["fit_percent"],
        "total_skills": stats["total_skills"],
        "completed_skills": stats["completed_skills"],
        "in_progress_skills": stats["in_progress_skills"],
        "estimated_external_hours": stats["estimated_external_hours"],
        "groups": groups_out,
    }


@app.post("/v2/career/blueprint/regenerate")
def v2_regenerate_blueprint(
    payload: schemas.BlueprintRegenerateIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """AI sinh lại blueprint cho 1 career path."""
    from backend.core.ai_advisor import generate_career_blueprint

    user = _get_user_by_token(authorization, db)

    path = db.query(models.CareerPath).filter(
        models.CareerPath.id == payload.path_id
    ).first()
    if not path:
        raise HTTPException(status_code=404, detail="Career path không tồn tại")

    # Student chỉ regenerate path mình đang chọn
    if user.role == "student":
        choice = db.query(models.UserCareerChoice).filter(
            models.UserCareerChoice.user_id == user.id,
            models.UserCareerChoice.primary_path_id == payload.path_id,
        ).first()
        if not choice:
            raise HTTPException(status_code=403, detail="Bạn chưa chọn nghề này")

    existing_count = db.query(models.CareerSkill).filter(
        models.CareerSkill.path_id == payload.path_id
    ).count()
    if existing_count > 0 and not payload.force:
        return {"ok": True, "generated": False,
                "reason": "Blueprint đã tồn tại. Dùng force=true để sinh lại.",
                "skill_count": existing_count}

    # Lấy CTĐT làm context cho AI
    all_courses = db.query(models.Course).filter(
        models.Course.count_toward_credits == True
    ).order_by(models.Course.typical_semester).all()
    school_ctx = [
        {"code": c.course_code, "name": c.course_name,
         "credits": float(c.credits or 0), "semester": c.typical_semester}
        for c in all_courses
    ]

    skills_data, provider = generate_career_blueprint(
        career_name=path.name,
        career_description=path.short_description or path.name,
        school_courses=school_ctx,
    )
    if not skills_data:
        raise HTTPException(status_code=503,
                            detail="AI không phản hồi. Thử lại sau vài giây.")

    db.query(models.CareerSkill).filter(models.CareerSkill.path_id == payload.path_id).delete()
    for s in skills_data:
        db.add(models.CareerSkill(
            path_id=payload.path_id,
            skill_group=s["skill_group"],
            skill_name=s["skill_name"],
            skill_type=s["skill_type"],
            level=s.get("level"),
            priority=s.get("priority", 2),
            school_covered=s.get("school_covered", False),
            school_courses=s.get("school_courses"),
            source_type=s.get("source_type"),
            source_name=s.get("source_name"),
            source_url=s.get("source_url"),
            description=s.get("description"),
            estimated_hours=s.get("estimated_hours"),
        ))
    path.last_blueprint_at = datetime.utcnow()
    path.blueprint_model = provider or "unknown"
    db.commit()

    return {"ok": True, "generated": True, "skill_count": len(skills_data), "provider": provider}


@app.put("/v2/career/blueprint/skill/{skill_id}/status")
def v2_update_skill_status(
    skill_id: int,
    payload: schemas.BlueprintSkillStatusIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Cập nhật trạng thái học 1 mục (planned/in_progress/completed/skipped)."""
    user = _get_user_by_token(authorization, db)
    if payload.status not in {"planned", "in_progress", "completed", "skipped"}:
        raise HTTPException(status_code=400, detail="status không hợp lệ")

    skill = db.query(models.CareerSkill).filter(models.CareerSkill.id == skill_id).first()
    if not skill:
        raise HTTPException(status_code=404, detail="Skill không tồn tại")

    row = db.query(models.UserCareerSkillProgress).filter(
        models.UserCareerSkillProgress.user_id == user.id,
        models.UserCareerSkillProgress.skill_id == skill_id,
    ).first()
    now = datetime.utcnow()
    if row:
        row.status = payload.status
        row.note = payload.note
        row.completed_at = now if payload.status == "completed" else None
    else:
        db.add(models.UserCareerSkillProgress(
            user_id=user.id, skill_id=skill_id, status=payload.status,
            note=payload.note,
            completed_at=now if payload.status == "completed" else None,
        ))
    db.commit()
    return {"ok": True, "skill_id": skill_id, "status": payload.status}


@app.put("/v2/career/blueprint/skill/{skill_id}/schedule")
def v2_schedule_skill(
    skill_id: int,
    payload: schemas.BlueprintSkillScheduleIn,
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Gán mục ngoài trường vào học kỳ (drag-drop từ UI)."""
    if payload.scheduled_term is not None:
        import re as _re
        term = payload.scheduled_term.strip()
        if not _re.match(r'^HK[1-9](/\d{4}-\d{4})?$', term):
            raise HTTPException(
                status_code=400,
                detail=f"scheduled_term phải có format 'HKn' hoặc 'HKn/YYYY-YYYY' (n=1-9). Nhận được: {payload.scheduled_term!r}",
            )
        payload.scheduled_term = term

    user = _get_user_by_token(authorization, db)

    skill = db.query(models.CareerSkill).filter(models.CareerSkill.id == skill_id).first()
    if not skill:
        raise HTTPException(status_code=404, detail="Skill không tồn tại")

    row = db.query(models.UserCareerSkillProgress).filter(
        models.UserCareerSkillProgress.user_id == user.id,
        models.UserCareerSkillProgress.skill_id == skill_id,
    ).first()
    if row:
        row.scheduled_term = payload.scheduled_term
    else:
        db.add(models.UserCareerSkillProgress(
            user_id=user.id, skill_id=skill_id,
            status="planned", scheduled_term=payload.scheduled_term,
        ))
    db.commit()
    return {"ok": True, "skill_id": skill_id, "scheduled_term": payload.scheduled_term}


@app.get("/v2/career/fit/me")
def v2_career_fit_me(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Career Fit % + thống kê cho KPI dashboard."""
    user = _get_user_by_token(authorization, db)

    choice = db.query(models.UserCareerChoice).filter(
        models.UserCareerChoice.user_id == user.id
    ).first()
    if not choice or not choice.primary_path_id:
        return {"fit_percent": 0, "career_name": None, "career_code": None,
                "total_skills": 0, "completed_skills": 0,
                "in_progress_skills": 0, "estimated_external_hours": 0,
                "has_blueprint": False}

    path = db.query(models.CareerPath).filter(
        models.CareerPath.id == choice.primary_path_id
    ).first()
    if not path:
        return {"fit_percent": 0, "career_name": None, "career_code": None,
                "total_skills": 0, "completed_skills": 0,
                "in_progress_skills": 0, "estimated_external_hours": 0,
                "has_blueprint": False}

    stats = _compute_blueprint_fit(user.id, path.id, db)
    return {
        **stats,
        "career_name": path.name,
        "career_path_name": path.name,
        "career_code": path.code,
        "career_icon": path.icon or "work",
        "career_color": path.color or "indigo",
        "has_blueprint": stats["total_skills"] > 0,
    }


@app.get("/v2/integrated-roadmap/me")
def v2_integrated_roadmap(
    authorization: str | None = Header(default=None),
    db: Session = Depends(get_db),
):
    """Lộ trình tích hợp: môn học trong trường + mục ngoài trường theo HK."""
    from backend.core.academic_engine import CURRICULUM_ORDER
    from collections import defaultdict

    user = _get_user_by_token(authorization, db)

    # ── TRONG TRƯỜNG ──────────────────────────────────────────────────────
    spec = user.specialization
    all_courses = db.query(models.Course).all()
    course_map = {c.course_code: c for c in all_courses}

    grades = db.query(models.UserGrade).filter(
        models.UserGrade.user_id == user.id
    ).all()
    grade_best: dict[str, models.UserGrade] = {}
    for g in grades:
        prev = grade_best.get(g.course_code)
        if not prev or (float(g.score10 or 0)) > float(prev.score10 or 0):
            grade_best[g.course_code] = g

    def _sem(code: str) -> int:
        c = course_map.get(code)
        if c and c.typical_semester:
            return c.typical_semester
        return CURRICULUM_ORDER.get(code, 9)

    def _letter(s10: float | None) -> str | None:
        if s10 is None:
            return None
        if s10 >= 9.0: return "A+"
        if s10 >= 8.5: return "A"
        if s10 >= 8.0: return "B+"
        if s10 >= 7.0: return "B"
        if s10 >= 6.5: return "C+"
        if s10 >= 5.5: return "C"
        if s10 >= 5.0: return "D+"
        if s10 >= 4.0: return "D"
        return "F"

    # Tất cả tiên quyết
    prereqs_map: dict[str, list[str]] = defaultdict(list)
    for p in db.query(models.CoursePrerequisite).all():
        prereqs_map[p.course_code].append(p.prerequisite_code)

    school_by_sem: dict[int, list] = defaultdict(list)
    for code, course in course_map.items():
        if not course.count_toward_credits:
            continue
        if course.required_specialization and spec and course.required_specialization != spec:
            continue
        sem = _sem(code)
        if sem < 1 or sem > 9:
            continue

        g = grade_best.get(code)
        if g:
            status = "passed" if g.passed else "current"
            score10 = float(g.score10) if g.score10 else None
            grade_letter = _letter(score10) if g.passed else None
        else:
            prereq_list = prereqs_map.get(code, [])
            all_clear = all(
                grade_best.get(pr) and grade_best[pr].passed
                for pr in prereq_list
            )
            status = "upcoming" if all_clear else "locked"
            score10 = None
            grade_letter = None

        school_by_sem[sem].append({
            "kind": "school",
            "title": course.course_name,
            "subtitle": f"{code} · {float(course.credits or 0):.0f} TC",
            "status": status,
            "course_code": code,
            "credits": float(course.credits or 0),
            "grade_letter": grade_letter,
            "score10": score10,
        })

    # ── NGOÀI TRƯỜNG ──────────────────────────────────────────────────────
    external_by_sem: dict[int, list] = defaultdict(list)
    unscheduled_external: list = []
    choice = db.query(models.UserCareerChoice).filter(
        models.UserCareerChoice.user_id == user.id
    ).first()

    if choice and choice.primary_path_id:
        skills = db.query(models.CareerSkill).filter(
            models.CareerSkill.path_id == choice.primary_path_id
        ).order_by(models.CareerSkill.priority, models.CareerSkill.id).all()

        prog_map: dict[int, models.UserCareerSkillProgress] = {}
        if skills:
            for r in db.query(models.UserCareerSkillProgress).filter(
                models.UserCareerSkillProgress.user_id == user.id,
                models.UserCareerSkillProgress.skill_id.in_([s.id for s in skills]),
            ).all():
                prog_map[r.skill_id] = r

        passed_set = {c for c, g in grade_best.items() if g.passed}

        for s in skills:
            row = prog_map.get(s.id)
            status = row.status if row else "planned"
            scheduled_term = row.scheduled_term if row else None

            if s.school_covered and s.school_courses:
                if any(c in passed_set for c in s.school_courses):
                    status = "completed"

            if status == "completed":
                continue  # không hiển thị mục đã xong

            item = {
                "kind": "external",
                "title": s.skill_name,
                "subtitle": s.source_name,
                "status": status,
                "skill_id": s.id,
                "skill_group": s.skill_group,
                "estimated_hours": s.estimated_hours,
                "source_url": s.source_url,
                "source_type": s.source_type,
            }
            if scheduled_term:
                try:
                    sem_num = int(scheduled_term.split("/")[0].replace("HK", "").strip())
                    if 1 <= sem_num <= 9:
                        external_by_sem[sem_num].append(item)
                    else:
                        unscheduled_external.append(item)
                except Exception:
                    unscheduled_external.append(item)
            else:
                unscheduled_external.append(item)

    # ── BUILD TERMS ───────────────────────────────────────────────────────
    all_sem_nums = sorted(set(list(school_by_sem.keys()) + [1, 2, 3, 4, 5, 6, 7, 8]))
    terms_out = []
    for sem in all_sem_nums:
        s_items = sorted(school_by_sem.get(sem, []), key=lambda x: x["title"])
        e_items = external_by_sem.get(sem, [])
        terms_out.append({
            "term_label": f"HK{sem}",
            "semester_index": sem,
            "school_credits": sum(i.get("credits", 0) for i in s_items),
            "external_hours": sum((i.get("estimated_hours") or 0) for i in e_items),
            "school_items": s_items,
            "external_items": e_items,
        })

    # ── FIT ───────────────────────────────────────────────────────────────
    fit_data: dict = {"fit_percent": 0, "career_name": None, "career_code": None,
                      "career_icon": "work", "career_color": "indigo"}
    if choice and choice.primary_path_id:
        path = db.query(models.CareerPath).filter(
            models.CareerPath.id == choice.primary_path_id
        ).first()
        if path:
            stats = _compute_blueprint_fit(user.id, path.id, db)
            fit_data = {
                "fit_percent": stats["fit_percent"],
                "career_name": path.name,
                "career_code": path.code,
                "career_icon": path.icon or "work",
                "career_color": path.color or "indigo",
            }

    return {**fit_data, "terms": terms_out, "unscheduled_external": unscheduled_external}
